# -*- coding: utf-8 -*-
"""round-16 엽서북떡메 webadmin 복붙용 import.xlsx 빌더.

입력 = /tmp/pcb_export/*.psv (라이브 t_prc_* read-only 실측 덤프, 2026-06-13).
출력 = postcard-book-memo-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[중요] 이 시트(엽서북떡메)는 라이브에 이미 적재됨 → 그릇은 '재현(RU)'(재적재 아님·대조용).
가격사슬 단절 2건은 별 시트(FIX)로 분리 — 그릇은 단절 교정 제안을 담되 본 RU 그릇과 구분.
"""
import csv

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

EXPORT = "/tmp/pcb_export"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "postcard-book-memo/postcard-book-memo-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # 단절/교정 시트 표식(주황)
RU_FILL = PatternFill("solid", fgColor="E2EFDA")  # 재현(라이브 기존) 표식(초록)


def read_psv(name: str) -> list[list[str]]:
    rows: list[list[str]] = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def empty_to_none(v: str):
    return None if v == "" else v


def write_sheet(
    wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None
):
    """columns=영문 DB컬럼, labels=한글라벨, data_rows=list of list(컬럼순)."""
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(28, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
ws = wb.create_sheet("0_README")
readme = [
    ["round-16 엽서북떡메 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 엽서북떡메 (index 12)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "2개 논리블록 모두 순수 수량구간 단가 MATRIX (합가/세트 아님)"],
    ["", ""],
    ["[핵심 발견] 핸드오프 가설 '떡메=합가형 세트' → 반증", ""],
    ["  · 엽서북·떡메 둘 다 라이브 prc_typ_cd=PRICE_TYPE.01 단가형 (장당가 × 수량)", ""],
    ["  · 스티커(MATRIX)와 동형이되 페이지축(20P/30P)·권당장수축(50/100장)이 추가된 다차원", ""],
    ["  · '세트/고정형 신규 구조'는 이 시트로 발견되지 않음 (다음 후보=박/제본/봉투)", ""],
    ["", ""],
    ["[경고] 라이브 기존 적재 — 재적재 금지", "이미 라이브 COMMIT 580행. 이 그릇은 재현(대조용)."],
    ["  · 초록 시트(_RU) = 라이브 기존 재현 — 신규 적재 아님", ""],
    ["  · 주황 시트(_FIX) = 가격사슬 단절 교정 제안 — 인간 승인 후", ""],
    ["", ""],
    ["[가격사슬 단절 2건 발견] 단가행 적재됐으나 엔진 조회 불가", ""],
    ["  단절1: PRF_PCB_FIXED 배선이 20P 2개(S1_20P·S2_20P)만 — 30P 2개(S1_30P·S2_30P) 미배선", ""],
    ["         → 30P 단가행 234행 적재됐으나 30P 선택 시 엔진이 못 찾음", ""],
    ["  단절2: PRD_000097(떡메모지) → PRF_TTEOKME_FIXED 바인딩 0행", ""],
    ["         → 떡메 공식·구성요소·단가행 112행 다 있는데 상품 바인딩 없어 엔진 조회 불가", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_RU", "공식정의 2 (라이브 재현)"],
    ["  1b_product_price_formulas_RU", "상품↔공식 바인딩 1 (라이브 재현·엽서북만)"],
    ["  2_formula_components_RU", "공식 배선 3 (라이브 재현·30P 미배선 상태)"],
    ["  3_price_components_RU", "구성요소 정의 5 (라이브 재현·전건 단가형)"],
    ["  4_component_prices_RU", "단가행 580 (이 시트 원천·라이브 재현·무손실)"],
    ["  8_FIX_wiring_chain", "가격사슬 단절1 교정 — 30P 배선 +2행 제안"],
    ["  9_FIX_binding_chain", "가격사슬 단절2 교정 — 떡메 바인딩 +1행 제안"],
    ["", ""],
    ["분해 무손실 검산", "가격표 데이터셀 580 (엽서북 39×12=468 + 떡메 28×4=112) = 라이브 580행 일치"],
    ["사이즈 BLOCKED", "없음 — 5 siz 전부 라이브 실재(SIZ_000003/004/124/119/266)"],
]
for ri, (a, b) in enumerate(readme, start=1):
    ws.cell(row=ri, column=1, value=a)
    ws.cell(row=ri, column=2, value=b)
    if ri == 1:
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 72

# ============================================================
# 1. price_formulas (라이브 재현)
# ============================================================
frm = read_psv("formulas.psv")
write_sheet(
    wb,
    "1_price_formulas_RU",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in frm],
    note="[RU] 라이브 t_prc_price_formulas 재현 — frm_typ_cd·prd_cd 컬럼 없음(라이브 실측). 신규 적재 아님.",
    fill=RU_FILL,
    col_comments={
        "frm_cd": "공식 식별자 PK. PRF_PCB_FIXED=엽서북 / PRF_TTEOKME_FIXED=떡메모지",
        "use_yn": "Y=노출. 둘 다 Y",
    },
)

# ============================================================
# 1b. product_price_formulas (바인딩, 라이브 재현)
# ============================================================
bind = read_psv("binding.psv")
write_sheet(
    wb,
    "1b_product_price_formulas_RU",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일(yyyy-MM-dd)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in bind],
    note="[RU] 라이브 t_prd_product_price_formulas 재현 — 별 테이블(공식정의와 분리). 엽서북만 바인딩됨, 떡메(PRD_000097)는 바인딩 0 → 단절2(9_FIX 참조).",
    fill=RU_FILL,
    col_comments={
        "prd_cd": "상품 식별자. PRD_000094=엽서북. PRD_000097(떡메)는 여기 없음=단절2",
        "apply_bgn_ymd": "PK 구성요소(prd_cd, apply_bgn_ymd). 적용 시작일",
    },
)

# ============================================================
# 2. formula_components (배선, 라이브 재현)
# ============================================================
wire = read_psv("wiring.psv")
write_sheet(
    wb,
    "2_formula_components_RU",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    [[r[0], r[1], int(r[2]) if r[2] else None, r[3]] for r in wire],
    note="[RU] 라이브 t_prc_formula_components 재현 — 3행. PRF_PCB_FIXED는 20P 2개만 배선(30P 미배선=단절1·8_FIX 참조).",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "이 공식이 흡수하는 구성요소. 엽서북=면·페이지별 comp 분리, 떡메=단일 comp",
        "addtn_yn": "라이브 전건 Y. Phase11 엔진은 무시(런타임 선택값으로 1개 comp 매칭)",
    },
)

# ============================================================
# 3. price_components (구성요소 정의, 라이브 재현)
# ============================================================
comp = read_psv("components.psv")
write_sheet(
    wb,
    "3_price_components_RU",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    [[r[0], r[1], r[2], r[3], r[4] if len(r) > 4 else "", r[5] if len(r) > 5 else "Y"] for r in comp],
    note="[RU] 라이브 t_prc_price_components 재현 5종 — 전건 PRICE_TYPE.01 단가형(완제품가). 떡메 가설 '합가형' 반증.",
    fill=RU_FILL,
    col_comments={
        "prc_typ_cd": "전건 PRICE_TYPE.01 단가형(장당가×수량). 수량구간별 장당가 차등(min_qty). 합가형 아님",
        "use_dims": "엽서북=[siz_cd,min_qty](면·페이지는 comp로 분리) / 떡메=[siz_cd,bdl_qty,min_qty](권당장수=bdl_qty)",
        "comp_typ_cd": "PRC_COMPONENT_TYPE.06 = 완제품 단가",
        "comp_cd": "엽서북 4종=면(S1단면/S2양면)×페이지(20P/30P) 분리 / 떡메 1종",
    },
)

# ============================================================
# 4. component_prices (단가행 580, 이 시트 원천·라이브 재현)
# ============================================================
cp = read_psv("component_prices.psv")
# columns: comp_cd|siz_cd|clr_cd|mat_cd|proc_cd|coat_side_cnt|opt_cd|bdl_qty|min_qty|apply_ymd|unit_price
cp_rows = []
for r in cp:
    cp_rows.append([
        r[0],
        empty_to_none(r[1]),
        empty_to_none(r[2]),
        empty_to_none(r[3]),
        empty_to_none(r[4]),
        int(r[5]) if r[5] else None,
        empty_to_none(r[6]),
        int(r[7]) if r[7] else None,
        int(r[8]) if r[8] else None,
        r[9],
        float(r[10]) if r[10] else None,
    ])
write_sheet(
    wb,
    "4_component_prices_RU",
    ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "사이즈(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수(권당장수)", "수량구간시작(min_qty)", "적용일", "단가"],
    cp_rows,
    note="[RU] 라이브 t_prc_component_prices 재현(580행) — 엽서북떡메 가격표=이 단가의 원천. 10차원 자연키. 안 쓰는 차원 NULL(빈칸).",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "엽서북 COMP_PCB_S1/S2_20P/30P(4×117) + 떡메 COMP_TTEOKME(112)",
        "siz_cd": "SIZ_000003=100x150 / SIZ_000124=150x100 / SIZ_000004=135x135 / SIZ_000119=90x90 / SIZ_000266=70x120",
        "bdl_qty": "떡메 전용=권당장수(50/100장 1권). 엽서북은 NULL",
        "min_qty": "수량구간 시작값(상향구간). 주문수량 이하 최대 min_qty 행 매칭",
        "clr_cd": "전건 NULL — 도수 무관(완제품가). 별색 분리 없음",
        "proc_cd": "신설 공정 차원(8→10). 엽서북떡메 미사용 NULL",
        "opt_cd": "신설 옵션 차원(8→10). 엽서북떡메 미사용 NULL",
        "unit_price": "장당가(단가형). 엔진=단가×주문수량",
    },
)

# ============================================================
# 8. FIX wiring chain — 단절1 (30P 배선 누락) 교정 제안
# ============================================================
fix_wire = [
    ["PRF_PCB_FIXED", "COMP_PCB_S1_30P", 3, "Y"],
    ["PRF_PCB_FIXED", "COMP_PCB_S2_30P", 4, "Y"],
]
write_sheet(
    wb,
    "8_FIX_wiring_chain",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    fix_wire,
    note="[FIX·단절1] PRF_PCB_FIXED에 30P 구성요소 2개 배선 추가 제안. 현재 20P만 배선됨→30P 선택 시 엔진 조회 불가. 단가행 234행은 이미 적재됨(배선만 누락). 인간 승인 후 INSERT.",
    fill=BLK_FILL,
)

# ============================================================
# 9. FIX binding chain — 단절2 (떡메 바인딩 누락) 교정 제안
# ============================================================
fix_bind = [
    ["PRD_000097", "PRF_TTEOKME_FIXED", "2026-06-01", "[교정] 떡메모지 바인딩 누락분 — 공식·단가행 적재 완료, 바인딩만 0행이었음"],
]
write_sheet(
    wb,
    "9_FIX_binding_chain",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일", "비고"],
    fix_bind,
    note="[FIX·단절2] PRD_000097(떡메모지)→PRF_TTEOKME_FIXED 바인딩 1행 추가 제안. 공식·구성요소·단가행 112행 다 있으나 바인딩 0→엔진 조회 불가. 인간 승인 후 INSERT.",
    fill=BLK_FILL,
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("component_prices rows:", len(cp_rows))
