# -*- coding: utf-8 -*-
"""할인 매핑 검토 워크북 빌더 (dbm-validator)
입력: load/*.csv 3종 + DB 룩업 TSV(/tmp/dbm_names.tsv, /tmp/dbm_paths.tsv)
출력: discount-mapping-review.xlsx (4시트) + sheets/*.csv
검증 로직(STEP 2): 실재 / 카테고리정합 / 중복 / 누락
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap"
LOAD = f"{BASE}/02_mapping/load"
OUT = f"{BASE}/04_review"
SHEETS = f"{OUT}/sheets"
os.makedirs(SHEETS, exist_ok=True)

# ---------- DB 룩업 로드 ----------
names = {}   # prd_cd -> prd_nm
paths = {}   # prd_cd -> category path string
with open("/tmp/dbm_names.tsv", encoding="utf-8") as f:
    for line in f:
        parts = line.rstrip("\n").split("\t")
        if len(parts) >= 2:
            names[parts[0]] = parts[1]
        elif parts and parts[0]:
            names[parts[0]] = ""
with open("/tmp/dbm_paths.tsv", encoding="utf-8") as f:
    for line in f:
        parts = line.rstrip("\n").split("\t")
        if len(parts) >= 2:
            paths[parts[0]] = parts[1]
        elif parts and parts[0]:
            paths[parts[0]] = ""

# ---------- CSV 입력 ----------
def read_csv(path):
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))

tbls = read_csv(f"{LOAD}/t_dsc_discount_tables.csv")
dets = read_csv(f"{LOAD}/t_dsc_discount_details.csv")
links = read_csv(f"{LOAD}/t_prd_product_discount_tables.csv")

tbl_nm = {t["dsc_tbl_cd"]: t["dsc_tbl_nm"] for t in tbls}

# 그룹별 기대 카테고리 키워드(정합 판정용)
EXPECT = {
    "DSC_ACR_QTY":     (["아크릴"], "아크릴 카테고리"),
    "DSC_ACRCARA_QTY": (["아크릴"], "아크릴카라비너(상품명 분리)"),
    "DSC_FABRIC_QTY":  (["에코백", "파우치", "필통", "레더파우치"], "원단(에코백/파우치/필통)"),
    "DSC_STAT_QTY":    (["문구", "플래너", "노트"], "문구"),
    "DSC_GOODSA_QTY":  (["소품", "라이프", "데스크", "디지털", "여행", "기념품", "패션", "사무"], "굿즈A(소품/라이프 등)"),
    "DSC_GOODSB_QTY":  (["소품", "라이프", "데스크", "디지털", "여행", "기념품", "패션", "사무"], "굿즈B(소품/라이프 등)"),
    "DSC_SQUISHY_QTY": (["말랑", "PVC"], "말랑(PVC고주파)"),
}

# 적용구간 요약 문자열
def bracket_summary(tbl_cd):
    rows = [d for d in dets if d["dsc_tbl_cd"] == tbl_cd]
    rows.sort(key=lambda r: int(r["min_qty"]))
    parts = []
    for r in rows:
        mn = r["min_qty"]
        mx = r["max_qty"].strip()
        rng = f"{mn}~{mx}" if mx else f"{mn}~무제한"
        rate = r["dsc_rate"]
        # 정수면 정수로
        try:
            rf = float(rate)
            rate_s = str(int(rf)) if rf == int(rf) else str(rf)
        except Exception:
            rate_s = rate
        parts.append(f"{rng}:{rate_s}%")
    return " / ".join(parts)

bracket_cache = {t["dsc_tbl_cd"]: bracket_summary(t["dsc_tbl_cd"]) for t in tbls}

# 중복 링크 판정
from collections import Counter
prd_link_count = Counter(l["prd_cd"] for l in links)

# ---------- STEP 2 검증: 링크별 비고 ----------
def verify_link(l):
    prd = l["prd_cd"]
    grp = l["dsc_tbl_cd"]
    exists = prd in names and names[prd] != ""
    path = paths.get(prd, "")
    kws, label = EXPECT.get(grp, ([], grp))
    matched = any(k in path for k in kws) if path else False
    cat_consistent = matched
    dup = prd_link_count[prd] > 1

    notes = []
    # 실재
    notes.append("실재:Y" if exists else "실재:N(DB 미존재)")
    # 카테고리정합
    leaf = path.split(">")[-1].strip().replace("(주)", "") if path else "(카테고리 없음)"
    if cat_consistent:
        notes.append(f"카테고리정합:정합({leaf})")
    else:
        notes.append(f"카테고리정합:불일치(실제 '{path or '없음'}' / 기대 {label})")
    # 중복
    if dup:
        notes.append(f"중복:이중적용({prd_link_count[prd]}개 테이블)")
    else:
        notes.append("중복:없음")
    return "; ".join(notes), exists, cat_consistent, dup

link_results = []
for l in links:
    note, exists, consistent, dup = verify_link(l)
    link_results.append({
        "dsc_tbl_cd": l["dsc_tbl_cd"],
        "dsc_tbl_nm": tbl_nm.get(l["dsc_tbl_cd"], ""),
        "prd_cd": l["prd_cd"],
        "prd_nm": names.get(l["prd_cd"], ""),
        "cat": (paths.get(l["prd_cd"], "") or "(없음)"),
        "bracket": bracket_cache.get(l["dsc_tbl_cd"], ""),
        "note": note,
        "exists": exists,
        "consistent": consistent,
        "dup": dup,
    })

# ---------- 누락 분석(카테고리단위 그룹) ----------
# DB 실측 멤버십(이 스크립트는 미리 계산된 사실을 상수로 사용 — DB 쿼리 결과 반영)
acrylic_members = ["PRD_000160","PRD_000161","PRD_000162","PRD_000163","PRD_000164",
                   "PRD_000165","PRD_000166","PRD_000167","PRD_000168","PRD_000169",
                   "PRD_000170","PRD_000171"]  # CAT_000009 서브트리 전체 12
fabric_members = [f"PRD_{i:06d}" for i in range(230, 280)]  # 230~279 = 50 (280 부자재 제외)

linked_prds = set(l["prd_cd"] for l in links)
acr_missing = [p for p in acrylic_members if p not in linked_prds]
fab_missing = [p for p in fabric_members if p not in linked_prds]

# ---------- 워크북 작성 ----------
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"

def write_sheet(ws, headers, rows, widths, csv_name):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            cell.alignment = LEFT
    # CSV
    with open(f"{SHEETS}/{csv_name}", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)

# 시트1 할인테이블
ws1 = wb.active
ws1.title = "할인테이블"
rows1 = [[t["dsc_tbl_cd"], t["dsc_tbl_nm"], t["note"], t["use_yn"]] for t in tbls]
write_sheet(ws1, ["dsc_tbl_cd", "할인테이블명", "비고", "사용여부"],
            rows1, [20, 34, 50, 10], "01_할인테이블.csv")

# 시트2 할인구간 (그룹 → min_qty 정렬)
ws2 = wb.create_sheet("할인구간")
order = {t["dsc_tbl_cd"]: i for i, t in enumerate(tbls)}
dets_sorted = sorted(dets, key=lambda d: (order.get(d["dsc_tbl_cd"], 99), int(d["min_qty"])))
rows2 = []
for d in dets_sorted:
    mx = d["max_qty"].strip()
    try:
        rf = float(d["dsc_rate"]); rate_s = str(int(rf)) if rf == int(rf) else str(rf)
    except Exception:
        rate_s = d["dsc_rate"]
    rows2.append([d["dsc_tbl_cd"], d["apply_ymd"], d["min_qty"],
                  (mx if mx else "무제한"), d["dsc_typ_cd"], rate_s])
write_sheet(ws2, ["dsc_tbl_cd", "적용일자", "최소수량", "최대수량", "할인유형", "할인율(%)"],
            rows2, [20, 12, 10, 10, 14, 11], "02_할인구간.csv")

# 시트3 상품별 적용현황 (dsc_tbl_cd → prd_cd 정렬)
ws3 = wb.create_sheet("상품별적용현황")
lr_sorted = sorted(link_results, key=lambda r: (order.get(r["dsc_tbl_cd"], 99), r["prd_cd"]))
rows3 = [[r["dsc_tbl_cd"], r["dsc_tbl_nm"], r["prd_cd"], r["prd_nm"],
          r["cat"], r["bracket"], r["note"]] for r in lr_sorted]
write_sheet(ws3, ["dsc_tbl_cd", "할인테이블명", "prd_cd", "상품명", "카테고리", "적용구간요약", "비고(검증결과)"],
            rows3, [20, 26, 14, 24, 30, 56, 60], "03_상품별적용현황.csv")
# 시트3 행 강조: 불일치/이중적용 = WARN
for idx, r in enumerate(lr_sorted, start=2):
    if (not r["exists"]) or (not r["consistent"]) or r["dup"]:
        for c in range(1, 8):
            ws3.cell(row=idx, column=c).fill = WARN_FILL

# 시트4 요약 (테이블별)
ws4 = wb.create_sheet("요약")
rows4 = []
for t in tbls:
    cd = t["dsc_tbl_cd"]
    grp_dets = [d for d in dets if d["dsc_tbl_cd"] == cd]
    nbr = len(grp_dets)
    rates = []
    for d in grp_dets:
        try: rates.append(float(d["dsc_rate"]))
        except Exception: pass
    rate_rng = f"{int(min(rates))}~{int(max(rates))}%" if rates else "-"
    grp_links = [r for r in link_results if r["dsc_tbl_cd"] == cd]
    nprd = len(grp_links)
    # 카테고리단위/행별
    if cd in ("DSC_ACR_QTY", "DSC_FABRIC_QTY"):
        unit = "카테고리단위"
    elif cd == "DSC_ACRCARA_QTY":
        unit = "행별(상품명 분리)"
    else:
        unit = "행별(마스터)"
    # 누락
    if cd == "DSC_ACR_QTY":
        miss = f"{len(acr_missing)}건" + (f" {acr_missing}" if acr_missing else " (없음)")
    elif cd == "DSC_FABRIC_QTY":
        miss = f"{len(fab_missing)}건" + (f" {fab_missing}" if fab_missing else " (없음, PRD_000280 부자재 의도적 제외)")
    else:
        miss = "해당없음(행별)"
    # 검증결과
    bad = [r for r in grp_links if (not r["exists"]) or (not r["consistent"]) or r["dup"]]
    verdict = "통과" if not bad else f"주의({len(bad)}건)"
    rows4.append([t["dsc_tbl_nm"], nbr, rate_rng, nprd, unit, miss, verdict])
write_sheet(ws4, ["할인테이블명", "구간수", "할인율범위", "상품수", "카테고리단위/행별", "누락여부", "검증결과"],
            rows4, [34, 9, 13, 9, 18, 46, 14], "04_요약.csv")
for idx, r4 in enumerate(rows4, start=2):
    fill = OK_FILL if r4[6] == "통과" else WARN_FILL
    ws4.cell(row=idx, column=7).fill = fill

wb.save(f"{OUT}/discount-mapping-review.xlsx")

# ---------- 콘솔 요약(반환값 근거) ----------
n_exist_n = sum(1 for r in link_results if not r["exists"])
n_inconsistent = [r for r in link_results if not r["consistent"]]
n_dup = sum(1 for r in link_results if r["dup"])
print("=== VERIFY SUMMARY ===")
print(f"links={len(links)} tables={len(tbls)} details={len(dets)}")
print(f"실재N={n_exist_n} 불일치={len(n_inconsistent)} 이중적용={n_dup}")
print(f"acr_missing={acr_missing} fab_missing={fab_missing}")
if n_inconsistent:
    for r in n_inconsistent:
        print("  불일치:", r["prd_cd"], r["dsc_tbl_cd"], r["cat"])
print("WROTE:", f"{OUT}/discount-mapping-review.xlsx")
