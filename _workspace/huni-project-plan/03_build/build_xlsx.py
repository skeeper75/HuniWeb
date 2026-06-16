# -*- coding: utf-8 -*-
"""후니프린팅 프로젝트 일정관리 통합 IA 워크북 빌더.
기존 입력 xlsx(5시트) 보강 + 신규 6시트. 실측 근거(01_research/*)만 반영, 가상 금지.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "docs/huni/후니프린팅_통합IA_일정_역할분담_260616.xlsx"
OUT = "docs/huni/후니프린팅_프로젝트일정관리_통합IA_260616.xlsx"

# ---------- 공통 스타일 ----------
THIN = Side(style="thin", color="C9CDD2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="2C3E50")
TITLE = Font(name="맑은 고딕", size=14, bold=True, color="1A2733")
SUB = Font(name="맑은 고딕", size=10, italic=True, color="566573")
CELL = Font(name="맑은 고딕", size=10, color="1A2733")
CELLB = Font(name="맑은 고딕", size=10, bold=True, color="1A2733")
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

PRI = {"P0": "FADBD8", "P1": "FCF3CF", "P2": "EAECEE"}
STATE = {"완료": "D5F5E3", "진행중": "D6EAF8", "대기": "FDEBD0", "미착수": "F4F6F7"}
OWNER = {"인쇄개발": "EBF5FB", "쇼핑개발": "F5EEF8", "PM": "E8F8F5", "협업": "FEF9E7"}
SECT = {"외부연동": "FDEBD0", "사이트내부": "FFFFFF"}
SECTC = {"외부": "FDEBD0", "내부": "EBF5FB", "내부+외부": "FEF9E7"}


def hrow(ws, headers, r=1, fill=HFILL):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HFONT
        cell.fill = fill
        cell.alignment = CTR
        cell.border = BORDER
    ws.row_dimensions[r].height = 30


def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w


def title_block(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = TITLE; t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=sub)
    s.font = SUB; s.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 18


# ============================================================
wb = Workbook()
wb.remove(wb.active)

# ---------- 입력 원본 IA 로드 ----------
src = openpyxl.load_workbook(SRC, data_only=True)
ia = src["02_IA마스터"]
ia_rows = []  # dict per feature
hdr = None
for row in ia.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    if hdr is None and row[0] == "No":
        continue
    if str(row[0]).strip() == "No":
        continue
    no = row[0]
    if not isinstance(no, int):
        continue
    ia_rows.append({
        "no": no, "sys": row[1] or "", "area": row[2] or "", "feat": row[3] or "",
        "pri": (row[4] or "").strip(), "owner": (row[5] or "").strip(),
        "size": (row[6] or "").strip(), "state": (row[7] or "").strip(),
        "phase": (row[8] or "").strip(), "pre": row[9] or "", "note": row[10] or "",
    })

# ---------- 담당자 매핑 ----------
OWNER_NAME = {"쇼핑개발": "김동학", "인쇄개발": "서희항·신우진", "PM": "신우진", "협업": "신우진·김동학"}
# ---------- 외부연동 기능 No (사이트내부/외부 구분) ----------
EXT_NO = {3, 7, 14, 15, 54, 55, 80, 82, 125, 138, 144}

# ---------- Phase1 P0 주차 그룹 ----------
W12 = {33, 49, 66, 100, 104, 105, 106, 107, 108}                     # 진행중·상품가격
W13 = {1, 2, 4, 5, 6, 8, 21, 22, 23, 24, 29, 31, 84, 85, 94, 123}    # 회원·기초·약관
W35 = {32, 34, 35, 36, 37, 38, 52, 53, 60, 101, 102, 103}            # 옵션위젯·파일·카탈로그
W57 = {42, 43, 48, 61, 62, 63, 64, 75, 76}                          # 제본·제약·주문전환
W79 = {9, 10, 77, 78, 79, 80, 81, 136, 137, 139, 140}               # 결제·배송·주문관리


def week_of(d):
    ph = d["phase"]
    no = d["no"]
    if ph.startswith("Phase 3"):
        return "3차 (W21~)"
    if ph.startswith("Phase 2"):
        return "2차 (W11~20)"
    # Phase 1
    if no in W12:
        return "1차 W1-2"
    if no in W13:
        return "1차 W1-3"
    if no in W35:
        return "1차 W3-5"
    if no in W57:
        return "1차 W5-7"
    if no in W79:
        return "1차 W7-9"
    return "1차 W1-10"


def sect_of(d):
    return "외부연동" if d["no"] in EXT_NO else "사이트내부"


for d in ia_rows:
    d["person"] = OWNER_NAME.get(d["owner"], d["owner"])
    d["week"] = week_of(d)
    d["sect"] = sect_of(d)

# ---------- 신규 보강행(실측 근거 — Shopby갭/MES/주문런타임) ----------
NEW = [
    # no, sys, area, feat, pri, owner, size, state, phase, pre, note, person, week, sect
    (145,"쇼핑몰","회원","간편로그인(네이버/카카오)","P1","쇼핑개발","M","미착수","Phase 2","소셜 OAuth 앱등록·검수 (PM)","Shopby 표준·인쇄몰 전환율 핵심","김동학","2차 (W11~20)","외부연동"),
    (146,"쇼핑몰","회원","회원등급(VIP·B2B 등급가)","P1","인쇄개발","M","미착수","Phase 2","등급별 할인·등급가 정책 (PM)","B2B/단골 재구매 핵심","신우진·김동학","2차 (W11~20)","사이트내부"),
    (147,"쇼핑몰","회원","휴면회원 전환/해제","P1","쇼핑개발","S","미착수","Phase 2","개인정보보호법 1년 미접속 의무","법적 필수(Shopby 표준)","김동학","2차 (W11~20)","사이트내부"),
    (148,"쇼핑몰","결제","무통장입금/가상계좌+입금확인","P1","쇼핑개발","M","대기","Phase 2","이니시스 가상계좌 계약 (PM)","B2B 계좌이체 선호","김동학","2차 (W11~20)","외부연동"),
    (149,"쇼핑몰","결제","간편결제(네이버페이)","P1","쇼핑개발","M","대기","Phase 2","네이버페이 가맹 (PM)","전환율 향상","김동학","2차 (W11~20)","외부연동"),
    (150,"쇼핑몰","주문","도서산간 추가배송비 데이터화","P0","쇼핑개발","M","미착수","Phase 1","권역·배송비 정책 확정 (PM)","정확청구·CS분쟁방지 (IA 비고→P0 승격)","김동학","1차 W7-9","사이트내부"),
    (151,"쇼핑몰","주문","배송비 템플릿/조건부 무료배송","P1","쇼핑개발","M","미착수","Phase 2","상품군별 배송비 정책","가격정책 일부","김동학","2차 (W11~20)","사이트내부"),
    (152,"쇼핑몰","정보","세금계산서 발행 플로우(B2B)","P1","쇼핑개발","M","미착수","Phase 2","국세청/팝빌 연동·사업자정보 동의 (PM)","B2B 결제완결성","김동학","2차 (W11~20)","외부연동"),
    (153,"쇼핑몰","상품","전시 카테고리 트리 운영툴","P0","쇼핑개발","M","미착수","Phase 1","카테고리 구조 확정","다depth 진열 운영툴 (승격)","김동학","1차 W3-5","사이트내부"),
    (154,"쇼핑몰","상품","상품 검색·필터 엔진","P1","쇼핑개발","M","미착수","Phase 2","검색 대상·필터축 정의","탐색성","김동학","2차 (W11~20)","사이트내부"),
    (155,"쇼핑몰","마케팅","베스트/기획전 진열 운영툴","P2","쇼핑개발","M","미착수","Phase 3","진열 운영정책","마케팅 진열","김동학","3차 (W21~)","사이트내부"),
    (156,"쇼핑몰","마이페이지","첫구매/생일 자동쿠폰","P2","쇼핑개발","S","미착수","Phase 3","자동발급 룰 정의","재구매 유도","김동학","3차 (W21~)","사이트내부"),
    (157,"관리자","주문관리","생산정보 전달(주문→생산BOM 환원)","P0","인쇄개발","L","미착수","Phase 1","MES↔Front 통신 규격 (PM)","결제완료시 생산정보 전달 핵심","서희항·신우진","1차 W9-10","외부연동"),
    (158,"관리자","상품관리","MES_ITEM_CD 전상품 채번","P0","인쇄개발","M","진행중","Phase 1","라이브 16/275만 채번됨(실측)","생산브릿지·94% 미채번","서희항·신우진","1차 W9-10","사이트내부"),
    (159,"관리자","주문관리","생산파일명 일관화(Rename 자동)","P0","인쇄개발","M","미착수","Phase 1","파일명 규칙 확정","품목_사이즈_소재_고객_번호_수량","서희항·신우진","1차 W9-10","사이트내부"),
    (160,"관리자","주문관리","접수파일 썸네일 자동생성","P1","인쇄개발","M","미착수","Phase 2","인쇄공정 파일 기준","발주 자동화","서희항·신우진","2차 (W11~20)","사이트내부"),
    (161,"쇼핑몰","주문","주문상태 추적(고객 마이페이지)","P0","쇼핑개발","M","미착수","Phase 1","주문상태머신 정의","미입금→제작중→출고완료","김동학","1차 W7-9","사이트내부"),
    (162,"관리자","주문관리","주문/주문상세 런타임 그릇 구축","P0","쇼핑개발","XL","미착수","Phase 1","라이브 주문테이블 전무(실측·to-be)","커머스 백엔드 그릇","김동학","1차 W5-7","사이트내부"),
]
for n in NEW:
    ia_rows.append({"no": n[0], "sys": n[1], "area": n[2], "feat": n[3], "pri": n[4],
                    "owner": n[5], "size": n[6], "state": n[7], "phase": n[8],
                    "pre": n[9], "note": n[10], "person": n[11], "week": n[12], "sect": n[13]})

# ============================================================
# 00_읽는법
# ============================================================
ws = wb.create_sheet("00_읽는법")
title_block(ws, "이 표 읽는 법 (실무진 안내)", "후니프린팅 리뉴얼 — 프로젝트 일정관리 통합 IA · 3인 체제 + 고객사 협조", 3)
hrow(ws, ["컬럼", "뜻", "값 설명"], r=4)
legend = [
    ("우선순위", "언제 만드나", "P0=런칭에 꼭 필요 · P1=런칭 직후(안정화) · P2=나중에(확장)"),
    ("담당", "어느 팀이", "PM=기획·정책·계약 · 인쇄개발=상품·가격·옵션·위젯 · 쇼핑개발=쇼핑몰화면·주문·페이지빌더 · 협업=두 팀 함께"),
    ("담당자", "누가", "신우진=PM 겸 인쇄개발 · 서희항=인쇄개발 · 김동학=쇼핑개발 · (고객) 최숙진 실장=디자인·콘텐츠·계약 협조"),
    ("개발규모", "작업량 가늠", "S=작음 · M=보통 · L=큼 · XL=매우 큼"),
    ("진행상태", "지금 상태", "완료 · 진행중 · 대기(외부계약·결정 기다림) · 미착수"),
    ("Phase / 주차", "단계와 시점", "Phase 1=런칭 · 2=안정화 · 3=확장 / 주차(W)=착수일(W1) 기준 상대 일정. 실제 날짜는 착수일 확정 후 환산"),
    ("구분", "사이트 안/밖", "사이트내부=우리가 만드는 화면·기능 · 외부연동=에디쿠스·이니시스·알림톡·MES·택배 등 바깥 서비스와 연결"),
    ("확인·결정 필요사항", "먼저 풀 것", "계약·정책 결정·데이터 정리 등 시작 전 챙길 선행조건 (쉬운 말로 기재)"),
    ("비고", "적용 상품/메모", "해당 기능이 어떤 상품에 들어가는지 또는 근거 메모"),
]
r = 5
for a, b, c in legend:
    ws.cell(row=r, column=1, value=a).font = CELLB
    ws.cell(row=r, column=2, value=b).font = CELL
    ws.cell(row=r, column=3, value=c).font = CELL
    for col in range(1, 4):
        cc = ws.cell(row=r, column=col); cc.alignment = WRAP; cc.border = BORDER
    ws.row_dimensions[r].height = 30
    r += 1
notes = [
    "※ 이 워크북은 후니프린팅에서 논의된 IA·정책문서를 기반으로 작성됐으며, 프로젝트 진행 중 상호보완·갱신됩니다.",
    "※ 시트 안내: 00 읽는법 / 01 역할정의 / 02 IA마스터(전 기능목록) / 03 페이즈일정(주차·병렬레인) / 04 진행현황 / 05 위젯·에디쿠스 기능목록 / 06 주문→생산(MES) 흐름 / 07 고객 준비물·디자인지원 / 08 외부연동·계약 / 09 용어집 / 10 개발자 상세(위젯·에디쿠스·DB)",
    "※ 개발자용 기술 상세(레드프린팅 SDK 함수·에디쿠스 메시지·t_* 스키마)는 '10_개발자상세' 시트에 별도 정리.",
    "※ 실제 라이브 DB·레드프린팅 역공학·Shopby API 표준을 실측해 기능목록을 보강했습니다(가상 기능 없음).",
]
for n in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cc = ws.cell(row=r, column=1, value=n); cc.font = SUB; cc.alignment = WRAP
    ws.row_dimensions[r].height = 30
    r += 1
widths(ws, {"A": 22, "B": 16, "C": 95})
ws.freeze_panes = "A5"

# ============================================================
# 01_역할정의
# ============================================================
ws = wb.create_sheet("01_역할정의")
title_block(ws, "역할 분담 & 진행현황 (실 담당자 반영)",
            "3인 체제 · 신우진(PM+인쇄)/서희항(인쇄)/김동학(쇼핑) + 고객사 협조(최숙진 실장) · 기준 IA 162개 기능", 6)
hrow(ws, ["역할", "담당자", "핵심 책임범위", "담당 IA 영역", "현재 진행상황", "주요 산출물"], r=4)
roles = [
    ("PM", "신우진", "일정/우선순위, 운영정책 확정, 외부계약(이니시스·알림톡·에디터·도메인), QA, 실무진 테스트 조율, 콘텐츠(약관·인쇄가이드)", "전 영역 의사결정·정책·계약·검수", "운영정책 5건 + 에디터 도입(D3/D4) 결정 대기", "일정표·정책 확정서·계약 완료·QA"),
    ("인쇄개발", "신우진·서희항", "관리자 상품관리(상품·사이즈·소재·용지·가격) + 고객 인쇄상품 옵션위젯·실시간가격(3엔진)·가이드·견적·CPQ 엔진·생산정보 전달", "관리자 상품관리 14종 + 인쇄상품 옵션/가격/가이드 24종 + 생산정보·MES 브릿지", "진행중: 상품·가격 DB 스키마 완료→데이터 매핑·시뮬레이터 검수 / 가격엔진(C1)·옵션API(A2)", "검증된 상품·가격 DB·CPQ 3엔진·임베드 위젯·생산정보 환원"),
    ("쇼핑개발", "김동학", "상세페이지 빌더, 파일업로드·주문전환·플로팅·장바구니, 쇼핑몰 사용자 전반, 관리자 커머스 운영, 결제·배송·주문상태", "쇼핑몰 사용자 + 인쇄상품 파일/주문 11종 + 관리자 커머스 33종 + 주문 런타임 신규구축", "진행중: 상세 콘텐츠 빌더(G1)→실무진 테스트 전달", "상세페이지 빌더·스토어프론트·주문/결제 플로우·관리자 운영툴"),
    ("(고객) 후니프린팅", "최숙진 실장", "AI 페이지 디자인 산출물 지원, 콘텐츠·가이드 원고, 이니시스 가맹관리자, 도메인, 문자/알림톡, 사업자·서류, 상품/가격 데이터 확정", "디자인·콘텐츠·계약·서류 협조(개발 외)", "AI로 페이지 디자인 진행 중 — 디자인 산출물·가이드·계약 협조 대기", "디자인 산출물·콘텐츠 원고·계약 서류·데이터 확정"),
]
r = 5
for role in roles:
    for c, v in enumerate(role, 1):
        cc = ws.cell(row=r, column=c, value=v)
        cc.font = CELLB if c <= 2 else CELL
        cc.alignment = WRAP; cc.border = BORDER
        if role[0] in OWNER:
            cc.fill = PatternFill("solid", fgColor=OWNER[role[0]])
        elif role[0].startswith("(고객)"):
            cc.fill = PatternFill("solid", fgColor="FDEBD0")
    ws.row_dimensions[r].height = 58
    r += 1
r += 1
collab = [
    "⟡ 협업 포인트",
    "· 인쇄상품 옵션·가격 위젯(인쇄개발) ↔ 상세페이지 임베드·주문전환(쇼핑개발) — 웹 컴포넌트 위에 결합",
    "· 온라인 에디터 에디쿠스(Edicus) D3·D4 — 협업/외부 도입 검토, 런칭 후순위(P2)",
    "· B15 옵션 제약 ↔ 폴리모픽 스키마(ref_key+ref_dim_cd) — 데이터로 표현돼야 동적 위젯 성립",
    "· 생산정보 전달(인쇄개발) ↔ 주문 런타임·MES 연동 — 결제완료 시 생산 BOM 환원이 두 팀 경계면",
    "⚠ 부하 주의: 쇼핑개발(김동학 1명) 공수 점수가 인쇄개발의 2배 이상 — 병목. AI 에이전트 병행 가속 권장(03_페이즈일정 참조)",
]
for line in collab:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cc = ws.cell(row=r, column=1, value=line)
    cc.font = CELLB if line.startswith("⟡") or line.startswith("⚠") else CELL
    cc.alignment = WRAP
    ws.row_dimensions[r].height = 22
    r += 1
widths(ws, {"A": 16, "B": 14, "C": 42, "D": 30, "E": 34, "F": 30})
ws.freeze_panes = "A5"

# ============================================================
# 02_IA마스터
# ============================================================
ws = wb.create_sheet("02_IA마스터")
title_block(ws, "통합 IA 기능목록 (162개 · 실측 보강)",
            "기존 144 + 보강 18(Shopby 표준 갭·주문 런타임·MES 생산브릿지). 색: 우선순위(빨강P0/노랑P1/회색P2)", 14)
heads = ["No", "시스템", "영역", "기능", "우선순위", "담당", "담당자", "규모", "진행상태", "Phase/주차", "구분", "확인·결정 필요사항(선행조건)", "비고"]
hrow(ws, heads, r=4)
r = 5
for d in sorted(ia_rows, key=lambda x: x["no"]):
    vals = [d["no"], d["sys"], d["area"], d["feat"], d["pri"], d["owner"], d["person"],
            d["size"], d["state"], d["week"], d["sect"], d["pre"], d["note"]]
    for c, v in enumerate(vals, 1):
        cc = ws.cell(row=r, column=c, value=v)
        cc.font = CELL; cc.border = BORDER
        cc.alignment = CTR if c in (1, 5, 6, 8, 9, 11) else WRAP
    # 색칠: 우선순위
    ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=PRI.get(d["pri"], "FFFFFF"))
    ws.cell(row=r, column=9).fill = PatternFill("solid", fgColor=STATE.get(d["state"], "FFFFFF"))
    ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=OWNER.get(d["owner"], "FFFFFF"))
    ws.cell(row=r, column=11).fill = PatternFill("solid", fgColor=SECT.get(d["sect"], "FFFFFF"))
    # 보강행 강조(No>=145): No 셀 굵게
    if d["no"] >= 145:
        ws.cell(row=r, column=1).font = Font(name="맑은 고딕", size=10, bold=True, color="B03A2E")
    ws.row_dimensions[r].height = 30
    r += 1
widths(ws, {"A": 5, "B": 8, "C": 12, "D": 34, "E": 8, "F": 9, "G": 12, "H": 7, "I": 9, "J": 13, "K": 11, "L": 34, "M": 30})
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A4:M{r-1}"

# ============================================================
# 03_페이즈일정
# ============================================================
ws = wb.create_sheet("03_페이즈일정")
title_block(ws, "페이즈 일정 — 주차 기반 상대일정 · 병렬 레인",
            "전 작업 병렬 진행. 주차(W)=착수일(W1) 기준 상대. 실제 날짜는 착수일 확정 후 환산. ⚠ 쇼핑개발 병목→AI 병행 가속", 6)
hrow(ws, ["주차(상대)", "PM (신우진)", "인쇄개발 (서희항·신우진)", "쇼핑개발 (김동학)", "AI 에이전트 병행", "단계 게이트 / 산출"], r=4)
lanes = [
    ("1차 W1-2\n(착수·진행중)",
     "이니시스·도메인/SSL·통신판매 신고 착수 / 운영정책 5건 정리",
     "가격엔진(C1)·옵션API(A2) 진행 지속 / 상품·가격 DB 매핑·시뮬레이터 검수 / 상품·가격 관리자 셋업",
     "상세 콘텐츠 빌더(G1) 완성 / 회원·로그인·약관 기초 화면",
     "DB 매핑 교정·가격 시뮬 검증(dbm-*) / 위젯 명세 동기화",
     "기반작업 마무리 게이트: 가격엔진 검수 GO · 상세빌더 실무진 테스트"),
    ("1차 W3-5\n(옵션위젯·파일)",
     "약관·인쇄가이드 콘텐츠 확정 / 에디터 도입(D3) 결정",
     "옵션위젯 규격·수량·종이·도수(B1~B5)·작업사이즈(E4) / 사이즈·소재·종이 팝업",
     "파일업로드(D1)·미리보기(D2) / 카탈로그·상세 진입(A1) / 전시 카테고리 운영툴",
     "위젯 빌드·Red 동등성 검증(hw-*) / 파일검증 로직 보조",
     "옵션위젯 1차 동작 게이트: 규격~도수 선택→가격 표시"),
    ("1차 W5-7\n(제본·제약·주문전환)",
     "배송비·도서산간 정책 확정 / 인증·배송비 정책",
     "제본(B9·B10)·옵션 제약(B15) / 실시간가격(C1) 안정화",
     "주문전환(F1~F4)·플로팅 주문박스·장바구니 / 주문/주문상세 런타임 그릇",
     "옵션 제약 데이터화(폴리모픽) 보조 / 위젯-주문 경계 QA",
     "주문전환 게이트: 옵션→장바구니 성립 · 제약 동적 노출"),
    ("1차 W7-9\n(결제·배송·주문관리)",
     "이니시스 계약 완료 / 배송사 계약 / 본인확인 정책",
     "관리자 상품/가격 셋업 안정화 / 도서산간 배송비 데이터",
     "결제(이니시스)·주문완료·배송정보 / 관리자 주문관리·파일확인·주문서출력 / 주문상태 추적",
     "결제·주문 플로우 검증 / 회원·주문 경계 QA",
     "결제 게이트: 실결제 1건 성공 · 주문→관리자 노출"),
    ("1차 W9-10\n(생산정보·MES·통합QA)",
     "실무진 통합 테스트 조율 / 런칭 체크리스트",
     "생산정보 전달(주문→생산BOM 환원) / MES_ITEM_CD 전상품 채번 / 파일명 일관화",
     "주문관리 상태변경·주문서출력 마감 / 런칭 회귀 테스트",
     "MES_ITEM_CD 채번 보조 / 생산정보 환원 검증 / 통합 회귀",
     "★ 1차 런칭 게이트(P0 57+7건): 주문→결제→생산정보 전달 종단 성공"),
    ("2차 W11~20\n(안정화·P1)",
     "알림톡·PG·증빙 계약 / 증빙 정책 / 대량견적 IA 보완",
     "코팅·커팅·박/형압·캘린더가공(B6~B13)·가격표/견적(C2·C3)·가이드(E1~E3)·굿즈/포장 등록 / 썸네일 자동생성",
     "쿠폰·머니·증빙·세금계산서·고객센터·옵션보관함·거래처/원장·통계·후불·알림톡·간편로그인/결제·휴면회원",
     "옵션·가격 확장 자동 적재(dbm-*) / 운영툴 검증 / 회귀",
     "안정화 게이트: 재구매·B2B·운영툴 가동"),
    ("3차 W21~\n(확장·P2)",
     "에디터 도입·수작 존치 결정 / 마케팅 기획",
     "수작/디자인 등록·개별포장(B14)",
     "온라인 에디터(에디쿠스 D3·D4)·리뷰·체험단·마케팅·고급통계·기획전 진열",
     "에디쿠스 연동 구현·QA / 마케팅 페이지 빌드",
     "확장 게이트: 미결정 의존 기능 순차 오픈"),
]
r = 5
phase_fill = {"1차": "EAF2F8", "2차": "FEF9E7", "3차": "F4ECF7"}
for lane in lanes:
    key = lane[0][:2]
    for c, v in enumerate(lane, 1):
        cc = ws.cell(row=r, column=c, value=v)
        cc.font = CELLB if c == 1 else CELL
        cc.alignment = WRAP; cc.border = BORDER
        if c == 1:
            cc.fill = PatternFill("solid", fgColor=phase_fill.get(key, "FFFFFF"))
        if c == 5:
            cc.fill = PatternFill("solid", fgColor="EBF5FB")
        if c == 6 and lane[5].startswith("★"):
            cc.fill = PatternFill("solid", fgColor="D5F5E3")
    ws.row_dimensions[r].height = 78
    r += 1
r += 1
delay = [
    "⚠ 일정 지연 대응 — AI 에이전트 병행 완성도 확보 방안",
    "1) 병목 식별: 쇼핑개발(김동학 1명) 공수 점수 201 vs 인쇄개발 92 — 쇼핑 단독 병목이 1차 일정의 임계경로(critical path).",
    "2) AI 병행 가속: ① DB 매핑·가격 시뮬 검증(dbm-* 하네스) ② 위젯 빌드·Red 동등성 검증(hw-* 하네스) ③ 반복 화면(목록/폼/관리자 CRUD) 코드 생성 — 사람은 설계·검수에 집중.",
    "3) Shopby 표준 활용: 회원/주문/결제/쿠폰/배송/증빙은 Shopby 표준 모듈을 외부연동·참조로 흡수해 쇼핑개발 자체구현 공수 절감(직접 종속 X).",
    "4) 마스터 선적재 이점: 상품·가격·옵션 마스터는 이미 라이브 적재 완료 → 인쇄개발은 '조립' 중심. 주문 런타임만 신규.",
    "5) 게이트 우선: 각 주차 게이트(가격검수→옵션위젯→주문전환→결제→생산정보)를 통과해야 다음 진행 — 거짓 진척 방지.",
]
for line in delay:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cc = ws.cell(row=r, column=1, value=line)
    cc.font = CELLB if line.startswith("⚠") else CELL
    cc.alignment = WRAP
    ws.row_dimensions[r].height = 30
    r += 1
widths(ws, {"A": 16, "B": 26, "C": 34, "D": 38, "E": 26, "F": 30})
ws.freeze_panes = "A5"

# ============================================================
# 04_진행현황 (재계산)
# ============================================================
ws = wb.create_sheet("04_진행현황")
title_block(ws, "진행현황 (162 기능 기준 재집계)", "보강 18건 반영. 공수: S=1·M=2·L=4·XL=8", 5)
# ① 담당×우선순위
owners = ["PM", "인쇄개발", "쇼핑개발", "협업"]
pri_list = ["P0", "P1", "P2"]
cnt = {o: {p: 0 for p in pri_list} for o in owners}
for d in ia_rows:
    if d["owner"] in cnt and d["pri"] in pri_list:
        cnt[d["owner"]][d["pri"]] += 1
r = 4
ws.cell(row=r, column=1, value="① 담당 × 우선순위 (건수)").font = CELLB
r += 1
hrow(ws, ["담당 \\ 우선순위", "P0", "P1", "P2", "합계"], r=r)
r += 1
tot = {p: 0 for p in pri_list}
for o in owners:
    row = [o, cnt[o]["P0"], cnt[o]["P1"], cnt[o]["P2"], sum(cnt[o].values())]
    for c, v in enumerate(row, 1):
        cc = ws.cell(row=r, column=c, value=v); cc.font = CELL if c == 1 else CELL
        cc.alignment = CTR if c > 1 else LEFT; cc.border = BORDER
        if c == 1:
            cc.fill = PatternFill("solid", fgColor=OWNER.get(o, "FFFFFF"))
    for p in pri_list:
        tot[p] += cnt[o][p]
    r += 1
row = ["합계", tot["P0"], tot["P1"], tot["P2"], sum(tot.values())]
for c, v in enumerate(row, 1):
    cc = ws.cell(row=r, column=c, value=v); cc.font = CELLB; cc.alignment = CTR if c > 1 else LEFT; cc.border = BORDER
r += 2
# ② 진행상태
ws.cell(row=r, column=1, value="② 진행상태 분포").font = CELLB
r += 1
states = ["완료", "진행중", "대기", "미착수"]
scnt = {s: 0 for s in states}
for d in ia_rows:
    if d["state"] in scnt:
        scnt[d["state"]] += 1
hrow(ws, states, r=r)
r += 1
for c, s in enumerate(states, 1):
    cc = ws.cell(row=r, column=c, value=scnt[s]); cc.font = CELL; cc.alignment = CTR; cc.border = BORDER
    cc.fill = PatternFill("solid", fgColor=STATE[s])
r += 2
# ③ 공수
ws.cell(row=r, column=1, value="③ 담당별 가중 공수 (S=1·M=2·L=4·XL=8)").font = CELLB
r += 1
SZ = {"S": 1, "M": 2, "L": 4, "XL": 8}
eff = {o: [0, 0] for o in owners}  # total, p0
for d in ia_rows:
    if d["owner"] in eff:
        w = SZ.get(d["size"], 0)
        eff[d["owner"]][0] += w
        if d["pri"] == "P0":
            eff[d["owner"]][1] += w
hrow(ws, ["담당", "공수점수", "P0 공수점수"], r=r)
r += 1
for o in owners:
    row = [o, eff[o][0], eff[o][1]]
    for c, v in enumerate(row, 1):
        cc = ws.cell(row=r, column=c, value=v); cc.font = CELL; cc.alignment = CTR if c > 1 else LEFT; cc.border = BORDER
        if c == 1:
            cc.fill = PatternFill("solid", fgColor=OWNER.get(o, "FFFFFF"))
    r += 1
r += 1
foot = ("※ 인쇄상품 상세 + Shopby 표준 갭(보강 18건) 반영. 쇼핑개발 공수점수가 인쇄개발의 2배 이상 — "
        "쇼핑개발(김동학 1명) 병목. 주문 런타임은 라이브 부재로 전부 신규. AI 에이전트 병행 + Shopby 표준 흡수로 가속 권장.")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
cc = ws.cell(row=r, column=1, value=foot); cc.font = SUB; cc.alignment = WRAP
ws.row_dimensions[r].height = 44
widths(ws, {"A": 22, "B": 12, "C": 14, "D": 10, "E": 10})

print("sheets so far:", wb.sheetnames)
wb.save(OUT)
print("PART1 saved")
