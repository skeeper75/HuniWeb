# -*- coding: utf-8 -*-
"""
round-16 코팅 webadmin 복붙용 import.xlsx 빌더.
입력 = /tmp/coat_export/*.psv (라이브 read-only 실측 덤프).
출력 = coating-import.xlsx (테이블별 시트, 1행=DB컬럼, 2행=한글라벨).
모든 데이터는 라이브 기존 적재 재현(재적재 아님) — 신규 0.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

EXPORT = "/tmp/coat_export"
OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/coating/coating-import.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
RU_FILL = PatternFill("solid", fgColor="E2EFDA")  # 재현(라이브 기존) 표식


def read_psv(name):
    rows = []
    path = f"{EXPORT}/{name}"
    if not os.path.exists(path):
        return rows
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def write_sheet(wb, title, columns, labels, data_rows, note=None, col_comments=None):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="2E7D32", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(col) + 4)
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    note = "※ 라이브 t_prc_* 재현(신규 적재 0) — 코팅 가격사슬 이미 완결. webadmin 복붙 시 DB 컬럼 1:1."

    # --- 시트 1: price_formulas (코팅 포함 공식 3종) ---
    frm = read_psv("formulas.psv")  # frm_cd|frm_nm
    write_sheet(
        wb, "price_formulas",
        ["frm_cd", "frm_nm", "use_yn"],
        ["공식코드", "공식명(쉬운설명)", "사용여부"],
        [[r[0], r[1], "Y"] for r in frm],
        note="※ 코팅은 디지털인쇄 공식 PRF_DGP_A/D/E의 후가공 구성요소(독립 공식 아님). 아래 3공식에 배선됨.",
        col_comments={
            "frm_cd": "공식 식별코드 (PK). 코팅은 신규 공식 없음 — 기존 디지털인쇄 공식 재사용",
            "frm_nm": "공식 한글명",
            "use_yn": "Y=사용",
        },
    )

    # --- 시트 2: formula_components (코팅행만) ---
    fc = read_psv("formula_components.psv")  # frm_cd|comp_cd|disp_seq|addtn_yn
    write_sheet(
        wb, "formula_components",
        ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
        ["공식코드", "구성요소코드", "표시순서", "추가여부"],
        fc,
        note="※ 각 디지털인쇄 공식에 무광/유광 코팅 구성요소가 배선됨(addtn_yn=Y).",
        col_comments={
            "frm_cd": "어느 공식에 속하나",
            "comp_cd": "코팅 구성요소 (무광=COMP_COAT_MATTE, 유광=COMP_COAT_GLOSSY)",
            "disp_seq": "공식 내 표시 순서",
            "addtn_yn": "Y=합산 구성요소 (Phase11 엔진은 참고만)",
        },
    )

    # --- 시트 3: price_components (코팅 2종 정의) ---
    pc = read_psv("components.psv")  # comp_cd|comp_nm|comp_typ_cd|prc_typ_cd|use_dims|use_yn
    write_sheet(
        wb, "price_components",
        ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
        ["구성요소코드", "구성요소명(쉬운설명)", "구성요소유형", "단가/합가구분", "사용차원(JSON)", "사용여부"],
        pc,
        note="※ 무광/유광 = 별개 구성요소(코팅종류는 collapse 금지). 둘 다 단가형(PRICE_TYPE.01).",
        col_comments={
            "comp_cd": "구성요소 식별코드 (PK)",
            "comp_nm": "구성요소 한글명",
            "comp_typ_cd": "구성요소 유형 (PRC_COMPONENT_TYPE.02 = 후가공류)",
            "prc_typ_cd": "PRICE_TYPE.01=단가형(장당가×수량) / .02=합가형(구간총액÷min_qty). 코팅=.01",
            "use_dims": "이 단가표가 실제로 쓰는 차원 컬럼 배열 (코팅=siz_cd,coat_side_cnt,min_qty)",
            "use_yn": "Y=사용",
        },
    )

    # --- 시트 4: component_prices (184 단가행) ---
    cp = read_psv("component_prices.psv")
    # comp_cd|apply_ymd|siz_cd|clr_cd|mat_cd|coat_side_cnt|bdl_qty|min_qty|unit_price|proc_cd|opt_cd
    # 빈 문자열 → None (NULL 와일드카드 명시)
    cp_rows = []
    for r in cp:
        rr = [(v if v != "" else None) for v in r]
        cp_rows.append(rr)
    write_sheet(
        wb, "component_prices",
        ["comp_cd", "apply_ymd", "siz_cd", "clr_cd", "mat_cd",
         "coat_side_cnt", "bdl_qty", "min_qty", "unit_price", "proc_cd", "opt_cd"],
        ["구성요소코드", "적용일자", "출력판형(siz)", "색상(미사용)", "자재(미사용)",
         "코팅면수(1단/2양)", "번들수(미사용)", "수량구간시작", "장당단가", "공정(미사용)", "옵션(미사용)"],
        cp_rows,
        note="※ 184행 = 무광92+유광92. 빈칸=NULL(와일드카드·해당차원 무관). 국4절=SIZ_000499·3절=SIZ_000077.",
        col_comments={
            "comp_cd": "무광=COMP_COAT_MATTE / 유광=COMP_COAT_GLOSSY",
            "apply_ymd": "단가 적용 시작일 (자연키 일부)",
            "siz_cd": "출력판형. 국4절=SIZ_000499, 3절=SIZ_000077 (단가 round-trip 확정). ※siz_nm 라벨 정비 권장(컨펌)",
            "clr_cd": "NULL — 코팅은 색 무관",
            "mat_cd": "NULL — 코팅은 자재 무관",
            "coat_side_cnt": "코팅면수. 단면=1, 양면=2. 양면단가=단면×2",
            "bdl_qty": "NULL — 번들 무관",
            "min_qty": "수량구간 시작값(주문수량 이하 최대 구간 매칭). 23구간, 1000000=상한없음",
            "unit_price": "장당 단가(단가형). 엔진=unit_price×주문수량",
            "proc_cd": "NULL — 코팅종류는 comp_cd로 구분(proc_cd 불필요)",
            "opt_cd": "NULL — 옵션 무관",
        },
    )

    wb.save(OUT)
    # 검증 출력
    print("SAVED:", OUT)
    for ws in wb.worksheets:
        print(f"  sheet={ws.title:22s} rows(data)={ws.max_row}")


if __name__ == "__main__":
    main()
