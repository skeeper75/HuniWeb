# -*- coding: utf-8 -*-
"""round-16 인쇄후가공 webadmin 복붙용 import.xlsx 빌더.

입력 = docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx > 인쇄후가공 (read-only).
출력 = print-finishing-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[전제] 인쇄후가공은 라이브에 완전 적재 + 배선 완결 + 216/216 전건 정합.
       → 신규 구축 아님. RU(라이브 재현) 시트 = 현행 적재 상태를 webadmin에서 확인·재현.

핵심 차단/컨펌:
  · 단가/합가(P4): 가격표 "합가" 명시 vs 라이브 PRICE_TYPE.01 → .02 권장값 + .01 병기
  · 증분룰: DB 미저장(앱 외삽) → C1_increment_rules_REF 보존
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

SRC = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "print-finishing/print-finishing-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
RU_FILL = PatternFill("solid", fgColor="DEEAF6")   # 라이브 재현(파랑)
CFM_FILL = PatternFill("solid", fgColor="FCE4D6")  # 컨펌/충돌(주황)
REF_FILL = PatternFill("solid", fgColor="FFF2CC")  # 참조(앱 입력·노랑)

APPLY_YMD = "2026-06-01"

# ---- 가격표 읽기 ----
sb = openpyxl.load_workbook(SRC, data_only=True)
ws = sb["인쇄후가공"]


def cell(r, c):
    return ws.cell(row=r, column=c).value


# 14 comp의 (comp_cd, comp_nm, min_qty, unit_price) 추출
# 블록 정의: (col_qty, [(comp_cd, comp_nm, col_price)], row_start, row_end)
comp_meta = []  # (comp_cd, comp_nm)
cp_rows = []    # component_prices long-form


def harvest(qty_col, comps, r0, r1):
    """comps = [(comp_cd, comp_nm, price_col), ...]"""
    for cc, cn, _ in comps:
        comp_meta.append((cc, cn))
    for r in range(r0, r1 + 1):
        q = cell(r, qty_col)
        if q is None:
            continue
        for cc, cn, pcol in comps:
            p = cell(r, pcol)
            if p is None:
                continue
            cp_rows.append((cc, int(q), float(p)))


# B1 모서리: A2헤더, A3~A11. B=직각 C=둥근
harvest(1, [
    ("COMP_PP_CORNER_RIGHT", "모서리 직각", 2),
    ("COMP_PP_CORNER_ROUND", "모서리 둥근", 3),
], 3, 11)
# B2 오시: A16헤더, A17~A26. B=1줄 C=2줄 D=3줄
harvest(1, [
    ("COMP_PP_CREASE_1L", "오시 1줄", 2),
    ("COMP_PP_CREASE_2L", "오시 2줄", 3),
    ("COMP_PP_CREASE_3L", "오시 3줄", 4),
], 17, 26)
# B3 미싱: F16헤더, F17~F26. G=1줄 H=2줄 I=3줄
harvest(6, [
    ("COMP_PP_PERF_1L", "미싱 1줄", 7),
    ("COMP_PP_PERF_2L", "미싱 2줄", 8),
    ("COMP_PP_PERF_3L", "미싱 3줄", 9),
], 17, 26)
# B4 가변텍스트: A33헤더, A34~A56. B=1개 C=2개 D=3개
harvest(1, [
    ("COMP_PP_VARTEXT_1EA", "가변텍스트 1개", 2),
    ("COMP_PP_VARTEXT_2EA", "가변텍스트 2개", 3),
    ("COMP_PP_VARTEXT_3EA", "가변텍스트 3개", 4),
], 34, 56)
# B5 가변이미지: F33헤더, F34~F56. G=1개 H=2개 I=3개
harvest(6, [
    ("COMP_PP_VARIMG_1EA", "가변이미지 1개", 7),
    ("COMP_PP_VARIMG_2EA", "가변이미지 2개", 8),
    ("COMP_PP_VARIMG_3EA", "가변이미지 3개", 9),
], 34, 56)

# 증분룰
incr_rules = [
    ("A12", "모서리(직각/둥근)", "100장당 1000원씩 올립니다.",
     "5000매 초과분: 5000매값 + (수량-5000)/100 × 1000 (앱 외삽)"),
    ("A27", "오시", "1,000장당 20,000원씩 올립니다.",
     "5000매 초과분: 5000매값 + (수량-5000)/1000 × 20000 (앱 외삽)"),
    ("F27", "미싱", "1,000장당 20,000원씩 올립니다.",
     "5000매 초과분: 5000매값 + (수량-5000)/1000 × 20000 (앱 외삽)"),
]

sb.close()


# ---- 시트 작성 헬퍼 ----
def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    wsx = wb.create_sheet(title)
    if note:
        wsx.cell(row=1, column=1, value=note)
        wsx.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = wsx.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = wsx.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        wsx.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(32, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            wsx.cell(row=ri, column=ci, value=val)
    wsx.freeze_panes = wsx.cell(row=data_r0, column=1)
    return wsx


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
wsr = wb.create_sheet("0_README")
readme = [
    ["round-16 인쇄후가공 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 인쇄후가공"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "밴드 단가표 6블록 — 수량구간 × 줄수/개수 (매트릭스 아님)"],
    ["", ""],
    ["[전제·중대] 이 시트는 라이브에 이미 완전 적재 + 배선 완결 + 216/216 전건 정합", ""],
    ["  · 파랑 시트(_RU) = 라이브 재현 (신규 적재 아님·현행 확인용)", ""],
    ["  · 주황 시트(_CONFIRM) = 단가/합가 충돌 (가격표=합가 vs 라이브=단가형)", ""],
    ["  · 노랑 시트(_REF) = 증분룰 (앱 외삽·DB 미저장·보존만)", ""],
    ["", ""],
    ["[가격사슬 완결] 아크릴 단절과 정반대 — 엔진 조회 가능", ""],
    ["  후가공 14 comp → PRF_DGP_A(디지털인쇄 합산형) disp_seq 16~29 / PRF_DGP_D 7~20 배선", ""],
    ["  손님이 엽서 등에서 둥근모서리 선택 → 엔진이 CORNER_ROUND 단가행 매칭 → 합산", ""],
    ["", ""],
    ["[🔴 단가/합가 컨펌 Q-PF-1]", ""],
    ["  가격표 = '합가' 명시(A1·A15·F15) + 증분룰(총액 증가) + 세트거동(1매=100매=2000)", ""],
    ["    → 합가형(PRICE_TYPE.02)이 명백", ""],
    ["  라이브 현행 = 14 comp 전건 PRICE_TYPE.01(단가형)", ""],
    ["    → round-14 진단의 '합가형 백필 미완' 디폴트로 추정 (의도 아닐 가능성)", ""],
    ["  import 권장값 = .02, 라이브 현행 .01 병기. 백필 여부는 인간 컨펌 후", ""],
    ["", ""],
    ["[줄수/개수 = comp 분리] opt_cd 차원 아님 — 라이브 현행(CREASE_1L/2L/3L)", ""],
    ["  근거: 가격이 다른 변형 = comp 분리(foil-small STD/SPC 동형). use_dims=['min_qty']", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  2_formula_components_RU", "공식 배선 28 (PRF_DGP_A 14 + PRF_DGP_D 14)"],
    ["  3_price_components_RU", "구성요소 정의 14 (모서리2·오시3·미싱3·가변텍스트3·가변이미지3)"],
    ["  4_component_prices_RU", "단가행 216 (수량구간 × 14 comp)"],
    ["  C1_increment_rules_REF", "증분룰 3 (앱 외삽·DB 미저장·보존)"],
    ["  P4_prc_typ_confirm", "단가/합가 충돌 판정표 (Q-PF-1)"],
    ["", ""],
    ["분해 무손실 검산", f"가격표 가격셀 216 = 단가행 216 = 라이브 216 (mismatch 0)"],
    ["증분룰 3건", "별 참조 시트 보존 (가격 그릇 외·앱 책임)"],
]
for ri, (a, b) in enumerate(readme, start=1):
    wsr.cell(row=ri, column=1, value=a)
    wsr.cell(row=ri, column=2, value=b)
    if ri == 1:
        wsr.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
wsr.column_dimensions["A"].width = 54
wsr.column_dimensions["B"].width = 72

# ============================================================
# 2. formula_components (RU) — 28 배선
# ============================================================
PP_ORDER = [
    "COMP_PP_CREASE_1L", "COMP_PP_CREASE_2L", "COMP_PP_CREASE_3L",
    "COMP_PP_PERF_1L", "COMP_PP_PERF_2L", "COMP_PP_PERF_3L",
    "COMP_PP_VARTEXT_1EA", "COMP_PP_VARTEXT_2EA", "COMP_PP_VARTEXT_3EA",
    "COMP_PP_VARIMG_1EA", "COMP_PP_VARIMG_2EA", "COMP_PP_VARIMG_3EA",
    "COMP_PP_CORNER_RIGHT", "COMP_PP_CORNER_ROUND",
]
fc_rows = []
# PRF_DGP_A: disp_seq 16~29 (라이브 실측 순서)
for i, cc in enumerate(PP_ORDER):
    fc_rows.append(["PRF_DGP_A", cc, 16 + i, "Y"])
# PRF_DGP_D: disp_seq 7~20 (라이브 실측 순서)
for i, cc in enumerate(PP_ORDER):
    fc_rows.append(["PRF_DGP_D", cc, 7 + i, "Y"])

write_sheet(
    wb,
    "2_formula_components_RU",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    fc_rows,
    note="[RU·라이브 재현] 후가공 14 comp가 디지털인쇄 합산형 공식 PRF_DGP_A(16~29)·PRF_DGP_D(7~20)에 배선됨. 후가공은 독립 공식 없음. addtn_yn=Y(Phase11 무시·런타임 선택 후가공만 활성 합산·모서리/오시/미싱은 택1).",
    fill=RU_FILL,
    col_comments={
        "frm_cd": "디지털인쇄 합산형 공식. 후가공은 이 공식의 부품으로 합산됨(독립 공식 아님)",
        "addtn_yn": "Phase11 엔진 무시. 손님이 선택한 후가공만 런타임 합산. 직각/둥근·1줄/2줄/3줄은 택1",
    },
)

# ============================================================
# 3. price_components (RU) — 14
# ============================================================
# dedup comp_meta (순서 보존)
seen = set()
comp_uniq = []
for cc, cn in comp_meta:
    if cc not in seen:
        seen.add(cc)
        comp_uniq.append((cc, cn))
pc_rows = []
for cc, cn in comp_uniq:
    # 권장 prc_typ = .02(합가형) — 가격표 "합가" 근거. 라이브 현행 .01 병기는 note.
    pc_rows.append([cc, cn, "", "PRICE_TYPE.02", '["min_qty"]', "Y",
                    "[P4] 가격표=합가형(.02 권장). 라이브 현행=.01(단가형·미백필 추정)"])
write_sheet(
    wb,
    "3_price_components_RU",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn", "note"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부", "비고"],
    pc_rows,
    note="[RU·라이브 재현 + P4 컨펌] 14 구성요소. prc_typ_cd 권장=.02(합가형, 가격표 '합가' 명시 근거). 라이브 현행=.01(단가형) → Q-PF-1 컨펌 후 백필. use_dims=['min_qty'](수량구간만·라이브 일치).",
    fill=RU_FILL,
    col_comments={
        "prc_typ_cd": "권장 .02(합가형): 가격표 셀=수량구간 총액(둥근모서리 500매=6000원 전체). 라이브 .01과 충돌→Q-PF-1",
        "use_dims": "['min_qty'] — 후가공은 작업사이즈·자재·도수 무관, 수량구간만으로 가격 결정",
        "comp_nm": "줄수/개수(1L/2L/3L·1EA/2EA/3EA)는 comp 분리(opt_cd 아님). 가격 다른 변형=별 comp(foil-small STD/SPC 동형)",
    },
)

# ============================================================
# 4. component_prices (RU) — 216행
# ============================================================
CP_COLS = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt",
           "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price", "note"]
CP_LBL = ["구성요소", "사이즈(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수",
          "옵션(opt)", "묶음수(bdl)", "수량구간(min_qty)", "적용일", "단가(합가:구간총액)", "비고"]
cp_data = []
for cc, q, p in cp_rows:
    cp_data.append([cc, None, None, None, None, None, None, None, q, APPLY_YMD, p, None])

write_sheet(
    wb,
    "4_component_prices_RU",
    CP_COLS, CP_LBL, cp_data,
    note="[RU·라이브 재현] 단가행 216 (모서리18 + 오시30 + 미싱30 + 가변텍스트69 + 가변이미지69). 전 차원 NULL(use_dims=['min_qty']). 합가형: 셀=수량구간 총액. 가격표↔라이브 216/216 전건 일치(mismatch 0).",
    fill=RU_FILL,
    col_comments={
        "min_qty": "수량구간 시작. 주문수량 이하 최대 min_qty 행 매칭. 모서리/오시/미싱=1~5000, 가변=1~9500",
        "unit_price": "합가형: 수량구간 총액(예: 둥근모서리 1000매=11000원 전체). 엔진=총액÷min_qty=장당가 환산 후 ×주문수량. 직각모서리=전구간 0(무료)",
        "siz_cd": "NULL — 후가공은 작업사이즈 무관(수량구간만)",
        "opt_cd": "NULL — 줄수/개수는 comp 분리로 처리(opt_cd 미사용·라이브 실측)",
        "proc_cd": "NULL — 공정 차원 미사용. 귀돌이/오시/미싱 공정 구분은 comp군 분리로 표현",
    },
)

# ============================================================
# C1. 증분룰 (REF·앱 외삽·DB 미저장)
# ============================================================
incr_data = []
for loc, scope, rule, formula in incr_rules:
    incr_data.append([loc, scope, rule, formula])
write_sheet(
    wb,
    "C1_increment_rules_REF",
    ["src_cell", "scope", "rule_text", "app_extrapolation"],
    ["원본셀", "적용범위", "증분룰 원문", "앱 외삽 공식"],
    incr_data,
    note="[REF·앱 책임] 증분룰 3건 — DB component_prices는 이산 수량구간만 담음(최대 5000매). 5000매 초과분은 앱이 증분룰로 외삽 계산(메모리 권위: 중간계산=앱·DB=룩업). 침묵 삭제 금지·보존만.",
    fill=REF_FILL,
    col_comments={
        "app_extrapolation": "DB 단가행 외 수량의 가격 계산식. off-grid ceiling과 동일 철학(런타임)",
    },
)

# ============================================================
# P4. 단가/합가 충돌 판정표
# ============================================================
p4_rows = [
    ["가격표 헤더 라벨", "A1='모서리 (귀돌이, 합가)'·A15='오시 (합가)'·F15='미싱 (합가)'", "합가형(.02)", "명시 표기"],
    ["증분룰 거동", "1000매당 +10000원(둥근모서리)·총액이 수량 비례 증가", "합가형(.02)", "장당가 아닌 총액"],
    ["세트 거동", "둥근모서리 1매=2000·100매=2000(동일)", "합가형(.02)", "구간 정액·장당가면 100배여야"],
    ["라이브 현행", "14 comp 전건 PRICE_TYPE.01(단가형)", "단가형(.01)", "round-14 진단: 합가형 백필 미완 디폴트"],
    ["foil-small 선례", "명함박도 .01로 등록(동형 충돌)", "충돌", "후니 관례 가능성 배제 못함→컨펌"],
    ["판정(권장)", "가격표 3중 근거 → 합가형(.02) 백필 권장", "PRICE_TYPE.02", "인간 컨펌 Q-PF-1 후 적용"],
]
write_sheet(
    wb,
    "P4_prc_typ_confirm",
    ["evidence", "detail", "implies", "note"],
    ["증거", "내용", "함의", "비고"],
    p4_rows,
    note="[Q-PF-1] 단가/합가 충돌 판정. 가격표 '합가' 3중 근거 vs 라이브 .01. 추정 금지 — 백필 적용은 인간 컨펌 후. 엔진이 .01로 계산하면 6000×출력매수=과대(둥근모서리 500매 후가공 300만원).",
    fill=CFM_FILL,
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("formula_components rows:", len(fc_rows), "(expect 28)")
print("price_components rows:", len(pc_rows), "(expect 14)")
print("component_prices rows:", len(cp_data), "(expect 216)")
print("increment_rules:", len(incr_data), "(expect 3)")
# 자연키 중복 검산
keys = [(cc, q) for cc, q, _ in cp_rows]
print("nat_key (comp,min_qty) dup:", len(keys) - len(set(keys)), "(expect 0)")
