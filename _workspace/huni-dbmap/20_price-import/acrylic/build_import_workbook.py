# -*- coding: utf-8 -*-
"""
round-16 아크릴 webadmin 복붙용 import.xlsx 빌더 (면적매트릭스형).
입력 = /tmp/acr_export/*.psv (라이브 read-only 실측 덤프) + 원본 가격표 엑셀(코롯토/카라비너).
출력 = acrylic-import.xlsx (테이블별 시트, 1행=DB컬럼, 2행=한글라벨).
면적매트릭스 = (siz_cd, unit_price) long-form. 라이브 적재분=재현(_RU), 미적재=_GAP/_NEW.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

EXPORT = "/tmp/acr_export"
XL = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx"
OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/acrylic/acrylic-import.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")   # GAP/BLOCKED 표식
RU_FILL  = PatternFill("solid", fgColor="E2EFDA")    # 재현(라이브 기존) 표식
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")    # 신규 후보 표식


def read_psv(name):
    rows = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def n(v):
    return None if v == "" else v


def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, min(28, len(lbl) + 4))
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


# component_prices 10차원 컬럼/라벨/주석 (재사용)
CP_COLS = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"]
CP_LBLS = ["구성요소", "출력규격(siz·가로×세로)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수", "수량구간(min_qty)", "적용일", "단가"]
CP_CMT = {
    "comp_cd": "COMP_ACRYL_CLEAR3T(투명3T)/CLEAR15T(투명1.5T)/MIRROR3T(미러3T). 자재가 구성요소명에 박힘(mat_cd 미사용)",
    "siz_cd": "가로×세로 면적 규격코드(예 20x20→SIZ_000336). 면적매트릭스 차원=siz_cd 단독",
    "clr_cd": "NULL — 매트릭스 '통용 단가'는 도수 무관(양면9도/단면7도 통가)",
    "min_qty": "NULL — 면적매트릭스는 수량축 없음(스티커/디지털과 차이). 수량할인은 t_dsc 별단계",
    "proc_cd": "신설 공정 차원(8→10). 아크릴 본체 미사용 NULL",
    "opt_cd": "신설 옵션 차원(8→10). 후가공은 별 영역 — 본체 매트릭스 NULL",
    "unit_price": "개당 면적단가(단가형). 엔진=단가×주문수량. off-grid=한단계 큰 규격(런타임·DB미저장)",
    "apply_ymd": "적용일(라이브 2026-06-01)",
}


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
ws = wb.create_sheet("0_README")
readme = [
    ["round-16 아크릴 가격표 import 그릇 (webadmin 복붙용) — 면적매트릭스형", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 아크릴 (index 15)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["가격유형", "면적매트릭스형 — 가격 = 가로×세로 좌표 단가(siz_cd 단독 차원·수량축 없음)"],
    ["단가/합가", "전건 단가형(PRICE_TYPE.01·라이브 실측). 합가형 없음"],
    ["", ""],
    ["[중요] 면적매트릭스 = (siz_cd, unit_price) long-form", ""],
    ["  · 매트릭스 [가로][세로] 격자를 행으로 언피벗(면적함수 회귀 금지)", ""],
    ["  · use_dims=[\"siz_cd\"] — 나머지 9차원 전부 NULL(빈칸)", ""],
    ["  · off-grid(격자 부재 크기) = 한단계 큰 규격 단가(런타임·DB 미저장)", ""],
    ["", ""],
    ["[경고] 라이브 기존 적재 — 재적재 금지 / 가격사슬 단절", ""],
    ["  · 초록(_RU) = 라이브 기존 121행 재현(대조용). CLEAR3T 47·CLEAR15T 37·MIRROR3T 37", ""],
    ["  · 🔴 라이브 formula_components 배선 0행 = 단가행만 있고 공식 미연결(엔진 조회불가)", ""],
    ["    → 1_price_formulas / 2_formula_components / 1b_바인딩 = 신규 제안(사슬 완결)", ""],
    ["  · 주황(_GAP) = 미적재 117좌표(siz 미채번 100) — 좌표 채번 후 적재", ""],
    ["  · 노랑(_NEW) = 코롯토/카라비너 신규 후보", ""],
    ["", ""],
    ["[제외] 구간할인 2블록(아크릴 A49:B56·카라비너 D102:E107)", "round-1 t_dsc 영역 — 가격 t_prc 그릇 아님(제외)"],
    ["[참조] 후가공 옵션(B05 키링/뱃지 등)", "CPQ opt 영역(round-6)·7_finishing 참조 보존(component 변환은 컨펌)"],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_NEW", "공식 정의 5 신규 제안(사슬 단절 해소)"],
    ["  1b_product_price_formulas_NEW", "상품↔공식 바인딩 신규(컨펌 후 상품코드 확정)"],
    ["  2_formula_components_NEW", "공식 배선 신규(라이브 0행)"],
    ["  3_price_components", "구성요소 정의 3 라이브 재현 + 2 신규"],
    ["  4_component_prices_RU", "면적매트릭스 단가행 121 라이브 재현"],
    ["  4b_component_prices_GAP", "미적재 117좌표(siz 미채번 100)"],
    ["  5_korotto_NEW", "아크릴코롯토 매트릭스 21조합 신규"],
    ["  6_carabiner_NEW", "아크릴카라비너 고정가 4형상 신규"],
    ["  7_finishing_options", "후가공 옵션 목록(참조·CPQ 영역)"],
    ["  8_excluded_discount", "구간할인 2블록(t_dsc·그릇 제외·보존)"],
    ["", ""],
    ["무손실 검산", "라이브 121행 전건 엑셀 대조 일치(가격불일치 3=비대칭좌표 방향성·값 존재). 미적재 117=GAP 정직표기"],
]
for ri, (a, b) in enumerate(readme, start=1):
    ws.cell(row=ri, column=1, value=a)
    ws.cell(row=ri, column=2, value=b)
    if ri == 1:
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 72

# ============================================================
# 1. price_formulas (신규 제안)
# ============================================================
formulas_new = [
    ["PRF_ACRYL_CLEAR3T", "투명아크릴3T 면적매트릭스 공식", "Y", "신규 제안 — 라이브 comp 배선용(사슬 단절 해소)"],
    ["PRF_ACRYL_CLEAR15T", "투명아크릴1.5T 면적매트릭스 공식", "Y", "신규 제안"],
    ["PRF_ACRYL_MIRROR3T", "미러아크릴3T 면적매트릭스 공식", "Y", "신규 제안"],
    ["PRF_ACRYL_COROTTO", "아크릴코롯토 면적매트릭스 공식", "Y", "신규 제안(comp도 신규)"],
    ["PRF_ACRYL_CARABINER", "아크릴카라비너 고정가 공식", "Y", "신규 제안(고정가형)"],
]
write_sheet(
    wb, "1_price_formulas_NEW",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    formulas_new,
    note="[NEW] 라이브 t_prc_price_formulas에 아크릴 본체 공식 부재(가격사슬 단절). 신규 제안 — 단가행은 이미 라이브에 있으나 어느 공식에도 미배선.",
    fill=NEW_FILL,
    col_comments={"frm_cd": "공식 식별자 PK. 라이브에 아크릴 본체 공식 0개 — 신규 제안"},
)

# ============================================================
# 1b. product_price_formulas (바인딩, 신규·컨펌)
# ============================================================
bind_new = [
    ["(컨펌 후 확정)", "PRF_ACRYL_CLEAR3T", "2026-06-01", "투명3T 사용 아크릴상품(키링/스탠드 등) 바인딩 — 상품코드 컨펌 Q-ACR-6"],
    ["(컨펌 후 확정)", "PRF_ACRYL_MIRROR3T", "2026-06-01", "미러3T 사용 상품 바인딩"],
    ["PRD_000164", "PRF_ACRYL_COROTTO", "2026-06-01", "아크릴코롯토(라이브 prd 실재)"],
    ["PRD_000166", "PRF_ACRYL_CARABINER", "2026-06-01", "아크릴카라비너(라이브 prd 실재)"],
]
write_sheet(
    wb, "1b_product_price_formulas_NEW",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일(yyyy-MM-dd)", "비고"],
    bind_new,
    note="[NEW·컨펌] 상품↔공식 바인딩 신규. 투명/미러 본체를 어느 상품에 바인딩할지 컨펌 Q-ACR-6(키링·스탠드 등 본체 매트릭스 사용 상품군 확정 필요).",
    fill=NEW_FILL,
    col_comments={"prd_cd": "t_prd_products FK. 코롯토(PRD_000164)·카라비너(PRD_000166)는 라이브 실재 확인"},
)

# ============================================================
# 2. formula_components (배선, 신규)
# ============================================================
wire_new = [
    ["PRF_ACRYL_CLEAR3T", "COMP_ACRYL_CLEAR3T", 1, "Y"],
    ["PRF_ACRYL_CLEAR15T", "COMP_ACRYL_CLEAR15T", 1, "Y"],
    ["PRF_ACRYL_MIRROR3T", "COMP_ACRYL_MIRROR3T", 1, "Y"],
    ["PRF_ACRYL_COROTTO", "COMP_ACRYL_COROTTO", 1, "Y"],
    ["PRF_ACRYL_CARABINER", "COMP_ACRYL_CARABINER", 1, "Y"],
]
write_sheet(
    wb, "2_formula_components_NEW",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    wire_new,
    note="[NEW] 라이브 t_prc_formula_components에 COMP_ACRYL_* 배선 0행(가격사슬 단절 핵심). 본체 1구성요소 배선 신규. 후가공 합산 추가는 컨펌 Q-ACR-1.",
    fill=NEW_FILL,
    col_comments={"comp_cd": "본체 단가 구성요소. 후가공 추가단가 합산 여부는 Q-ACR-1"},
)

# ============================================================
# 3. price_components (3 RU + 2 NEW)
# ============================================================
comp_rows = read_psv("components_acryl.psv")  # comp_cd|comp_nm|comp_typ|prc_typ|use_dims|use_yn
comp_data = [[r[0], r[1], r[2], r[3], r[4], r[5]] for r in comp_rows]
comp_data.append(["COMP_ACRYL_COROTTO", "아크릴코롯토 인쇄가공비", "PRC_COMPONENT_TYPE.01", "PRICE_TYPE.01", '["siz_cd"]', "Y"])
comp_data.append(["COMP_ACRYL_CARABINER", "아크릴카라비너 고정가(3T+3T 접합)", "PRC_COMPONENT_TYPE.01", "PRICE_TYPE.01", '["opt_cd"]', "Y"])
write_sheet(
    wb, "3_price_components",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    comp_data,
    note="[RU 3 + NEW 2] 상단 3종(CLEAR3T/15T/MIRROR3T)=라이브 재현(전건 단가형.01·use_dims=[siz_cd]). 하단 2종(코롯토·카라비너)=신규.",
    fill=RU_FILL,
    col_comments={
        "prc_typ_cd": "PRICE_TYPE.01=단가형(개당단가×수량). 아크릴 전건 단가형(라이브 실측·합가형 없음)",
        "use_dims": "면적매트릭스=[siz_cd] 단독(min_qty 무관·도수/자재 NULL). 카라비너=[opt_cd] 고정가(컨펌 Q-ACR-2)",
        "comp_typ_cd": "PRC_COMPONENT_TYPE.01",
    },
)

# ============================================================
# 4. component_prices RU (라이브 121행 재현)
# ============================================================
cp = read_psv("component_prices_acryl.psv")  # comp_cd|siz_cd|clr|mat|proc|coat|opt|bdl|min|apply|price
cp_rows = []
for r in cp:
    cp_rows.append([
        r[0], n(r[1]), n(r[2]), n(r[3]), n(r[4]),
        int(r[5]) if r[5] else None, n(r[6]),
        int(r[7]) if r[7] else None, int(r[8]) if r[8] else None,
        r[9], float(r[10]) if r[10] else None,
    ])
write_sheet(
    wb, "4_component_prices_RU",
    CP_COLS, CP_LBLS, cp_rows,
    note="[RU] 라이브 t_prc_component_prices 재현(121행) — 아크릴 면적매트릭스. (siz_cd, unit_price) long-form. 나머지 9차원 NULL(빈칸). 재적재 금지·대조용.",
    fill=RU_FILL, col_comments=CP_CMT,
)

wb.save(OUT)
print("PHASE1 SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("component_prices RU rows:", len(cp_rows))

# ============================================================
# 4b. component_prices GAP (미적재 117좌표 — siz 미채번 100)
# ============================================================
import csv
def read_csv(path):
    with open(path, encoding="utf-8") as f:
        return list(csv.reader(f))

gap = read_csv("/tmp/acr_build/matrix_GAP.csv")  # comp_cd,canon,garo,sero,price,siz_cd
gap_rows = []
for r in gap[1:]:
    if r[0] == "COMP_ACRYL_COROTTO":
        continue  # 코롯토는 5_korotto_NEW에서 별도
    comp, canon, garo, sero, price, sizcd = r
    gap_rows.append([
        comp, n(sizcd) or f"(미채번:{garo}x{sero})", None, None, None, None, None, None, None,
        "2026-06-01", float(price),
    ])
write_sheet(
    wb, "4b_component_prices_GAP",
    CP_COLS, CP_LBLS, gap_rows,
    note="[GAP] 매트릭스 미적재 좌표(siz 미채번 다수). siz_cd '(미채번:GxS)'=좌표 채번 요청 대상(round-2 권위 '좌표 siz 등록 요청서'). 채번 후 4_RU로 승격. NULL 강제 금지(별 시트).",
    fill=BLK_FILL, col_comments=CP_CMT,
)

# ============================================================
# 5. 코롯토 NEW (매트릭스 21조합 — 신규 comp)
# ============================================================
wbx = openpyxl.load_workbook(XL, data_only=True)
wsx = wbx["아크릴"]

def matrix_unique(r_rows, c_cols, hdr_r):
    cols = {c: int(str(wsx.cell(row=hdr_r, column=c).value).replace("mm", "")) for c in c_cols}
    uniq = {}
    for r in r_rows:
        garo = int(str(wsx.cell(row=r, column=1).value).replace("mm", ""))
        for c in c_cols:
            v = wsx.cell(row=r, column=c).value
            if v is None:
                continue
            a, b = sorted([garo, cols[c]])
            key = f"{a}x{b}"
            if key not in uniq:
                uniq[key] = (garo, cols[c], float(v))
    return uniq

# siz 사전 (canon -> siz_cd) for 코롯토 좌표
siz_nm2cd = {}
with open(f"{EXPORT}/siz_dict.psv", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        cd, nm = line.split("|")
        a, b = nm.split("x")
        siz_nm2cd[f"{min(int(a),int(b))}x{max(int(a),int(b))}"] = cd

kor = matrix_unique(range(92, 98), range(2, 8), 91)
kor_rows = []
for key, (g, s, price) in sorted(kor.items()):
    sizcd = siz_nm2cd.get(key)
    kor_rows.append([
        "COMP_ACRYL_COROTTO", sizcd or f"(미채번:{g}x{s})", None, None, None, None, None, None, None,
        "2026-06-01", price,
    ])
write_sheet(
    wb, "5_korotto_NEW",
    CP_COLS, CP_LBLS, kor_rows,
    note="[NEW] 아크릴코롯토(A90:G97) 면적매트릭스 21조합. 라이브 comp 부재 — COMP_ACRYL_COROTTO 신규. round-2 권위는 코롯토를 ④고정가형 분류했으나 실측은 가로×세로 매트릭스(컨펌 Q-ACR-3).",
    fill=NEW_FILL, col_comments=CP_CMT,
)

# ============================================================
# 6. 카라비너 NEW (고정가 4형상)
# ============================================================
cara_rows = []
for r in range(104, 108):
    opt_nm = wsx.cell(row=r, column=1).value
    price = wsx.cell(row=r, column=2).value
    if opt_nm is None:
        continue
    cara_rows.append([
        "COMP_ACRYL_CARABINER", None, None, None, None, None, f"(형상:{opt_nm})", None, None,
        "2026-06-01", float(price),
    ])
write_sheet(
    wb, "6_carabiner_NEW",
    CP_COLS, CP_LBLS, cara_rows,
    note="[NEW] 아크릴카라비너(투명3T+3T 접합·A102:B107) 고정가 4형상. opt_cd 분기 단가(고정가형·siz 아님). 형상별 치수(40x69 등)는 명칭에 포함. opt_cd 코드 미채번('형상:'표기)·컨펌 Q-ACR-2.",
    fill=NEW_FILL, col_comments=CP_CMT,
)

# ============================================================
# 7. 후가공 옵션 (참조·CPQ 영역)
# ============================================================
fin_rows = []
cur_group = None
for r in range(61, 87):
    a = wsx.cell(row=r, column=1).value
    b = wsx.cell(row=r, column=2).value
    c = wsx.cell(row=r, column=3).value
    d = wsx.cell(row=r, column=4).value
    if a:
        cur_group = a
    if b is not None or c is not None:
        fin_rows.append([cur_group, b, c, d or ""])
write_sheet(
    wb, "7_finishing_options",
    ["opt_group", "opt_value", "add_price", "color_variants"],
    ["후가공옵션그룹", "옵션값", "추가단가", "색상변형(가격무관)"],
    fin_rows,
    note="[참조·CPQ] 후가공 옵션(B05 키링/뱃지/마그넷/집게/명찰/스마트톡/지비츠/볼펜/머리끈/카라비너고리). round-6 CPQ option_group→item 영역. 추가단가가 가격엔진 component 합산인지 CPQ add_price인지 컨펌 Q-ACR-1. 이 시트는 분해 참조용(t_prc 그릇 행 아님).",
    fill=BLK_FILL,
)

# ============================================================
# 8. 제외 구간할인 (t_dsc·보존)
# ============================================================
disc_rows = []
for r in range(51, 57):
    band = wsx.cell(row=r, column=1).value
    rate = wsx.cell(row=r, column=2).value
    disc_rows.append(["아크릴상품(A49:B56)", band, rate])
for r in range(104, 107):
    band = wsx.cell(row=r, column=4).value
    rate = wsx.cell(row=r, column=5).value
    disc_rows.append(["아크릴카라비너(D102:E107)", band, rate])
write_sheet(
    wb, "8_excluded_discount",
    ["discount_scope", "qty_band", "discount_rate"],
    ["할인범위", "수량구간", "할인율"],
    disc_rows,
    note="[제외·보존] 수량구간할인 2블록 — round-1 t_dsc_* 영역(가격 t_prc 그릇 아님). 침묵삭제 금지로 보존. 엔진은 단가×수량 후 별단계로 적용. 그릇 행 아님(참조).",
    fill=BLK_FILL,
)

wb.save(OUT)
print("PHASE2 SAVED:", OUT)
print("Final sheets:", wb.sheetnames)
print("GAP:", len(gap_rows), " korotto:", len(kor_rows), " carabiner:", len(cara_rows), " finishing:", len(fin_rows), " discount:", len(disc_rows))
