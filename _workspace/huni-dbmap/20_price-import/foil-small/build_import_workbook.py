# -*- coding: utf-8 -*-
"""round-16 후가공_박(소형) webadmin 복붙용 import.xlsx 빌더.

입력 = docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx > 후가공_박(소형) (read-only).
출력 = foil-small-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[중요] 소형박 후가공은 라이브 미적재(공식·구성요소·단가행·배선 모두 0) → 신규 구축(NEW).
선결 차단 2건은 별 시트로 분리:
  · 등급 코드값 GRADE_A~E 미등록 → B1_grade_codes_proposal(코드 선적재 제안)
  · 박이 붙는 prd_cd 미확정 → 1b_product_price_formulas_BLOCKED
면적→등급 매핑표는 가격 그릇 외(앱 책임) → A1_area_grade_map(참조 보존).
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

SRC = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "foil-small/foil-small-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
NEW_FILL = PatternFill("solid", fgColor="E2EFDA")  # 신규 구축(초록)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # 차단/제안(주황)
REF_FILL = PatternFill("solid", fgColor="FFF2CC")  # 참조(앱 입력·노랑)

APPLY_YMD = "2026-06-01"

# ---- 가격표 읽기 ----
sb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws = sb["후가공_박(소형)"]


def cell(r, c):
    return ws.cell(row=r, column=c).value


# 등급 헤더 (I~M = col 9~13) → A,B,C,D,E
GRADES = ["A", "B", "C", "D", "E"]


def extract_band(hdr_row, last_row):
    """등급별 수량 가격 밴드(H=수량, I~M=등급가) → list of (grade, min_qty, price)."""
    out = []
    r = hdr_row + 1
    while r <= last_row:
        qty = cell(r, 8)  # H
        if qty is None:
            break
        for gi, g in enumerate(GRADES):
            price = cell(r, 9 + gi)  # I..M
            if price is not None:
                out.append((g, int(qty), float(price)))
        r += 1
    return out


std_band = extract_band(9, 27)   # 일반박 H9 헤더, 데이터 H10~H27
spc_band = extract_band(32, 50)  # 특수박 H32 헤더, 데이터 H33~H50

# 면적→등급 매핑 (가로 헤더 B9~F9, 세로 라벨 A10~A12)
ga_cols = [cell(9, c) for c in range(2, 7)]   # 10/20/40/60/80mm
ga_map = []
for r in (10, 11, 12):
    v_label = cell(r, 1)  # 세로
    for ci, w_label in enumerate(ga_cols):
        grade = cell(r, 2 + ci)
        ga_map.append((v_label, w_label, (grade.strip() if isinstance(grade, str) else grade)))

# 동판비
plate_qty = cell(3, 1)   # A3=1
plate_size = cell(2, 2)  # B2=80x40mm
plate_price = cell(3, 2)  # B3=5000

sb.close()


# ---- 시트 작성 헬퍼 ----
def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    wsx = wb.create_sheet(title)
    if note:
        wsx.cell(row=1, column=1, value=note)
        wsx.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = wsx.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = wsx.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        wsx.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(30, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            wsx.cell(row=ri, column=ci, value=val)
    wsx.freeze_panes = wsx.cell(row=data_r0, column=1)
    return wsx


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
wsr = wb.create_sheet("0_README")
readme = [
    ["round-16 후가공_박(소형) 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 후가공_박(소형)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "2단 룩업 — ①면적→등급(앱 계산) ②등급×수량→총액(DB 단가행)"],
    ["", ""],
    ["[핵심] 박 = MATRIX/합산형 아닌 제3구조 '2단 룩업'", ""],
    ["  · 1단계 면적→등급(A~E) = 앱(런타임) 계산 → DB 미저장 (메모리 권위)", ""],
    ["  · 2단계 등급×수량 → 총액 = DB component_prices (등급=opt_cd 차원)", ""],
    ["  · '세트/고정형 신규구조'는 아님 — 등급 차원 첫 실사용이 신규성의 본질", ""],
    ["", ""],
    ["[라이브 미적재] 소형박 후가공 가격사슬 전면 부재 → 신규 구축(NEW)", ""],
    ["  · 초록 시트(_NEW) = 신규 구축 그릇 (라이브 0행 → 신규 적재 대상)", ""],
    ["  · 주황 시트(_BLOCKED/_PROPOSAL) = 선결 차단 (코드/바인딩 미확정·인간 승인)", ""],
    ["  · 노랑 시트(_REF) = 면적→등급 매핑표 (앱 책임·가격 그릇 외·보존만)", ""],
    ["", ""],
    ["[선결 차단 2건]", ""],
    ["  차단1: 등급 코드값 GRADE_A~E 미등록 → B1_grade_codes_proposal (코드 선적재)", ""],
    ["  차단2: 박이 붙는 prd_cd 미확정(후가공=옵션) → 1b_..._BLOCKED (바인딩 컨펌)", ""],
    ["", ""],
    ["[단가/합가 컨펌] 가격표=수량구간 총액(합가형.02)이나 명함박 라이브=단가형.01", ""],
    ["  → prc_typ_cd 최종 확정은 명함박 선례 의도 컨펌 후 (P4)", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_NEW", "공식정의 1 (PRF_FOIL_SMALL 합산형)"],
    ["  1b_product_price_formulas_BLOCKED", "상품 바인딩 0 (prd_cd 미확정·차단2)"],
    ["  2_formula_components_NEW", "공식 배선 2 (동판셋업+박가공)"],
    ["  3_price_components_NEW", "구성요소 정의 2 (동판셋업비.01 / 박가공비.02합가)"],
    ["  4_component_prices_NEW", "단가행 181 (동판 1 + 일반박 90 + 특수박 90)"],
    ["  A1_area_grade_map_REF", "면적→등급 매핑 28 (앱 입력·가격 그릇 외·보존)"],
    ["  B1_grade_codes_proposal", "등급 코드값 5 선적재 제안 (GRADE_A~E)"],
    ["", ""],
    ["분해 무손실 검산", f"가격표 가격셀 181 (동판1 + 일반90 + 특수90) = 단가행 181 일치"],
    ["면적→등급 매핑 28셀", "별 참조 시트 보존 (가격 그릇 외·앱 책임)"],
]
for ri, (a, b) in enumerate(readme, start=1):
    wsr.cell(row=ri, column=1, value=a)
    wsr.cell(row=ri, column=2, value=b)
    if ri == 1:
        wsr.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
wsr.column_dimensions["A"].width = 52
wsr.column_dimensions["B"].width = 70

# ============================================================
# 1. price_formulas (NEW)
# ============================================================
write_sheet(
    wb,
    "1_price_formulas_NEW",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    [["PRF_FOIL_SMALL", "소형박 후가공 (동판셋업+박가공)", "Y",
      "합산형: 동판비(고정) + 박가공비(등급×수량 합가)"]],
    note="[NEW] 신규 공식 — 라이브 t_prc_price_formulas에 박 공식 0행. frm_typ_cd·prd_cd 컬럼 없음(라이브 실측).",
    fill=NEW_FILL,
    col_comments={
        "frm_cd": "공식 식별자 PK. 소형박 후가공 단일 공식",
        "frm_nm": "동판셋업비 + 등급별 박가공비 합산",
    },
)

# ============================================================
# 1b. product_price_formulas (BLOCKED)
# ============================================================
write_sheet(
    wb,
    "1b_product_formulas_BLOCK",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드(미확정)", "공식코드", "적용시작일", "비고(컨펌)"],
    [["?? (컨펌필요)", "PRF_FOIL_SMALL", APPLY_YMD,
      "[BLOCKED] 박=후가공(옵션/공정). 붙는 prd_cd 미확정. 다수 상품 공용 옵션인지 특정 상품 전용인지 컨펌"]],
    note="[BLOCKED·차단2] 박이 붙는 상품(prd_cd) 미확정 → NULL 강제 금지·별 시트 분리. 컨펌 후 바인딩 INSERT.",
    fill=BLK_FILL,
    col_comments={
        "prd_cd": "박 후가공이 붙는 본체 상품. 가격표만으로 미확정 → 컨펌",
        "apply_bgn_ymd": "PK 구성(prd_cd, apply_bgn_ymd)",
    },
)

# ============================================================
# 2. formula_components (NEW)
# ============================================================
write_sheet(
    wb,
    "2_formula_components_NEW",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    [
        ["PRF_FOIL_SMALL", "COMP_FOIL_SETUP", 1, "Y"],
        ["PRF_FOIL_SMALL", "COMP_FOIL_SMALL_PROC_STD", 2, "Y"],
        ["PRF_FOIL_SMALL", "COMP_FOIL_SMALL_PROC_SPC", 3, "Y"],
    ],
    note="[NEW] 공식 배선 3 — 동판셋업 + 박가공(일반/특수 군 분리). addtn_yn=Y(Phase11 무시·런타임 선택값으로 1개 박가공 comp 매칭).",
    fill=NEW_FILL,
    col_comments={
        "comp_cd": "동판셋업(고정) + 박가공(STD 일반박군/SPC 특수박군). 손님 박종 선택→앱이 STD/SPC 군 결정",
        "addtn_yn": "Phase11 엔진 무시. STD/SPC는 택1(동시 매칭 안 함)",
    },
)

# ============================================================
# 3. price_components (NEW)
# ============================================================
write_sheet(
    wb,
    "3_price_components_NEW",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    [
        ["COMP_FOIL_SETUP", "박 동판셋업비", "", "PRICE_TYPE.01", "[]", "Y"],
        ["COMP_FOIL_SMALL_PROC_STD", "소형박 가공비(일반박)", "", "PRICE_TYPE.02", '["opt_cd","min_qty"]', "Y"],
        ["COMP_FOIL_SMALL_PROC_SPC", "소형박 가공비(특수박)", "", "PRICE_TYPE.02", '["opt_cd","min_qty"]', "Y"],
    ],
    note="[NEW] 구성요소 3. 동판셋업=단가형.01(고정 5000). 박가공=합가형.02(수량구간 총액÷min_qty 환산). [P4 컨펌] 명함박 라이브는 .01 단가형이라 충돌 — 의도 확인 후 prc_typ_cd 확정.",
    fill=NEW_FILL,
    col_comments={
        "prc_typ_cd": "동판셋업=.01(고정). 박가공=.02 합가형(가격표=수량 총액). 명함박 라이브는 .01이라 컨펌 필요",
        "use_dims": "박가공=[opt_cd(등급),min_qty(수량구간)]. 동판셋업=[](차원 무관 고정)",
        "comp_cd": "STD=일반박 7종(금/은/먹유광·청/적/동/펄박) / SPC=특수박 3종(백/홀로그램/트윙클박). 가격 군 단위 차등→comp 분리(명함박 STD/HOLO 선례)",
    },
)

# ============================================================
# 4. component_prices (NEW) — 181행
# ============================================================
CP_COLS = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt",
           "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price", "note"]
CP_LBL = ["구성요소", "사이즈(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수",
          "옵션(opt)=등급", "묶음수(bdl)", "수량구간(min_qty)", "적용일", "단가(총액)", "비고"]
cp_rows = []
# 동판셋업 1행
cp_rows.append(["COMP_FOIL_SETUP", None, None, None, None, None, None, None, None,
                APPLY_YMD, float(plate_price), "동판 1개 80x40mm·아연판(셋업비·고정)"])
# 일반박 90행
for g, qty, price in std_band:
    cp_rows.append(["COMP_FOIL_SMALL_PROC_STD", None, None, None, None, None,
                    f"GRADE_{g}", None, qty, APPLY_YMD, price, None])
# 특수박 90행
for g, qty, price in spc_band:
    cp_rows.append(["COMP_FOIL_SMALL_PROC_SPC", None, None, None, None, None,
                    f"GRADE_{g}", None, qty, APPLY_YMD, price, None])

write_sheet(
    wb,
    "4_component_prices_NEW",
    CP_COLS, CP_LBL, cp_rows,
    note="[NEW] 단가행 181 (동판1 + 일반박 90 + 특수박 90). opt_cd=등급(GRADE_A~E·차단1 코드 선적재 전 적재 보류). siz_cd=NULL(등급은 사이즈 아님·면적→등급은 앱). 안 쓰는 차원 NULL.",
    fill=NEW_FILL,
    col_comments={
        "opt_cd": "등급 차원(신설 opt_cd 첫 실사용). GRADE_A~E. 코드값 미등록→B1_grade_codes_proposal 선적재 후 적재",
        "siz_cd": "NULL — 등급은 사이즈 아님. 작업 가로×세로→등급 변환은 앱(A1_area_grade_map 참조)",
        "min_qty": "수량구간 시작(200~10000 18구간). 주문수량 이하 최대 min_qty 행 매칭",
        "unit_price": "합가형: 수량구간 총액(예: GRADE_A 200매=12,200원 전체). 엔진=총액÷min_qty=장당가",
        "proc_cd": "신설 공정 차원. 박종 군 차등은 comp 분리로 처리→proc_cd NULL",
        "bdl_qty": "박 미사용 NULL",
    },
)

# ============================================================
# A1. 면적→등급 매핑표 (REF·앱 책임·가격 그릇 외)
# ============================================================
ga_rows = []
for v_label, w_label, grade in ga_map:
    ga_rows.append([v_label, w_label, grade if grade is not None else None,
                    "미정의(공란)" if grade is None else None])
write_sheet(
    wb,
    "A1_area_grade_map_REF",
    ["height_mm", "width_mm", "grade", "note"],
    ["세로(mm)", "가로(mm)", "등급(A~E)", "비고"],
    ga_rows,
    note="[REF·앱 책임] 면적→등급 매핑 (2단 룩업 1단계). 가격 그릇(t_prc_*)에 직접 안 들어감 — 앱이 작업 가로×세로→등급 변환. 일반박/특수박 공통. 세로 60/80mm 미정의=off-grid(앱 ceiling 정책). DDL 원할 시 t_prc_foil_area_grade 신설 입력으로 사용.",
    fill=REF_FILL,
    col_comments={
        "grade": "(세로,가로)→등급. 공란=미정의(10x10mm 등 노트). opt_cd로 가격 룩업 시 이 등급 사용",
    },
)

# ============================================================
# B1. 등급 코드값 선적재 제안 (PROPOSAL·차단1)
# ============================================================
grade_code_rows = [
    [f"GRADE_{g}", f"박 면적등급 {g}", "FOIL_GRADE", i + 1, "Y",
     "소형박 후가공 면적등급 opt_cd 차원값(신규)"]
    for i, g in enumerate(GRADES)
]
write_sheet(
    wb,
    "B1_grade_codes_proposal",
    ["cod_cd", "cod_nm", "upr_cod_cd", "disp_seq", "use_yn", "note"],
    ["코드값", "코드명", "상위코드그룹", "표시순서", "사용여부", "비고"],
    grade_code_rows,
    note="[PROPOSAL·차단1] 등급 코드값 GRADE_A~E 신규 등록 제안 (t_cod_base_codes). opt_cd 차원에 담을 등급. 코드 선적재 후 component_prices 적재. 컬럼=라이브 t_cod_base_codes 실측(cod_cd·cod_nm·upr_cod_cd·disp_seq·use_yn·note).",
    fill=BLK_FILL,
    col_comments={
        "cod_cd": "등급 코드값 PK. opt_cd 단가행이 참조",
        "upr_cod_cd": "상위 코드그룹 FOIL_GRADE(신규 그룹) — 코드그룹도 등록 필요",
    },
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("std_band rows:", len(std_band), "spc_band rows:", len(spc_band))
print("component_prices rows:", len(cp_rows), "(expect 181)")
print("area_grade_map rows:", len(ga_rows), "(expect 15 cells; non-null grades vary)")
