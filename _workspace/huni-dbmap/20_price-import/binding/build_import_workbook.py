# -*- coding: utf-8 -*-
"""round-16 제본 webadmin 복붙용 import.xlsx 빌더.

입력 = /tmp/bind_export/*.psv (라이브 t_prc_* read-only 실측 덤프, 2026-06-13).
출력 = binding-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[중요] 이 시트(제본)는 라이브에 이미 적재됨 → 그릇은 '재현(RU)'(대조용·재적재 아님).
제본방식 차원 = comp_cd 분리(proc_cd 미사용·전건 NULL). 박/스티커와 다른 세 번째 패턴.
가격사슬 단절 2건은 별 시트(FIX·주황)로 분리.
"""
import csv

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

EXPORT = "/tmp/bind_export"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "binding/binding-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # 단절/교정 시트 표식(주황)
RU_FILL = PatternFill("solid", fgColor="E2EFDA")  # 재현(라이브 기존) 표식(초록)

# 제본방식 → comp 매핑 (라이브 실측)
METHOD_MAP = {
    "COMP_BIND_JUNGCHEOL": ("중철제본", "PROC_000018"),
    "COMP_BIND_MUSEON": ("무선제본", "PROC_000019"),
    "COMP_BIND_PUR": ("PUR제본", "PROC_000020"),
    "COMP_BIND_TWINRING": ("트윈링제본", "PROC_000021"),
    "COMP_BIND_HC_MUSEON": ("하드커버무선", "PROC_000023"),
    "COMP_BIND_HC_TWINRING": ("하드커버트윈링", "PROC_000024"),
    "COMP_BIND_SSABARI": ("싸바리바인더", "(PROC 트리 부재)"),
    "COMP_BIND_CAL_WALL": ("벽걸이캘린더", "(PROC 트리 부재)"),
    "COMP_BIND_CAL_DESK220": ("탁상220", "(PROC 트리 부재)"),
    "COMP_BIND_CAL_DESK130": ("탁상130", "(PROC 트리 부재)"),
    "COMP_BIND_CAL_DESKMINI": ("탁상미니", "(PROC 트리 부재)"),
}


def read_psv(name: str) -> list[list[str]]:
    rows: list[list[str]] = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def empty_to_none(v: str):
    return None if v == "" else v


def write_sheet(
    wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None
):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(30, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
ws = wb.create_sheet("0_README")
readme = [
    ["round-16 제본 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 제본 (index 13)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "3개 논리블록 모두 '제본방식(열) × 수량구간(행)' 단가 MATRIX (합가/세트 아님)"],
    ["", ""],
    ["[평결] 핸드오프 가설 '제본=세트/고정형 신규 구조' → 반증", ""],
    ["  · 3블록 전부 MATRIX 동형 (스티커·엽서북떡메와 같은 밴드 단가표)", ""],
    ["  · 제본방식 11종 전건 prc_typ_cd=PRICE_TYPE.01 단가형 (권당 장당가 × 권수)", ""],
    ["  · '세트/고정형 신규 구조'는 이 시트로도 발견되지 않음", ""],
    ["", ""],
    ["[제본방식 차원 = comp_cd 분리] proc_cd 아님 (세 번째 패턴)", ""],
    ["  · 제본방식은 의미상 후가공 공정(PROC_000017 자식)이나 가격엔진은 comp 분리로 표현", ""],
    ["  · component_prices.proc_cd 전건 NULL (신설 차원 미사용·0행)", ""],
    ["  · 장점: PROC 트리에 없는 싸바리/캘린더도 comp만 만들면 됨 → BLOCKED 0·NULL강제 회피", ""],
    ["  · 비교: 스티커=opt_cd/coat_side / 박=opt_cd(등급) / 제본=comp_cd 분리", ""],
    ["", ""],
    ["[경고] 라이브 기존 적재 — 재적재 금지", "이미 라이브 COMMIT 74행. 이 그릇은 재현(대조용)."],
    ["  · 초록 시트(_RU) = 라이브 기존 재현 — 신규 적재 아님", ""],
    ["  · 주황 시트(_FIX) = 가격사슬 단절 교정 제안 — 인간 승인 후·컨펌 선결", ""],
    ["", ""],
    ["[🔴 가격사슬 단절 2건] 단가행 74 적재됐으나 엔진 조회 불가", ""],
    ["  단절1(배선): formula_components에 COMP_BIND_JUNGCHEOL 1개만 배선", ""],
    ["         → 무선/PUR/트윈링/하드커버3/캘린더4/싸바리 = 10 component 미배선", ""],
    ["         → 단가행 적재됐으나 해당 제본방식 선택 시 엔진이 comp 못 찾음", ""],
    ["  단절2(바인딩): PRF_BIND_SUM이 책자4종(PRD_000068~071)에만 바인딩", ""],
    ["         → 하드커버책자/캘린더/포토북/다이어리 등 제본 쓰는 상품 미바인딩", ""],
    ["", ""],
    ["[⚠️ 공식 설계 미스매치] 단순 배선 추가로 해결 안 됨", ""],
    ["  · PRF_BIND_SUM 1개에 11 comp 다 배선하면 모든 제본방식 단가가 합산되는 오류(동시매칭)", ""],
    ["  · 올바른 모델: (A) 제본방식별 공식 분리+상품 1:1 바인딩 또는 (B) opt_cd 옵션 선택", ""],
    ["  · 어느 안인지는 상품↔제본방식 권위 확정 후 (컨펌 BIND-C1) — 추정 분해 금지", ""],
    ["", ""],
    ["[가격경계]", ""],
    ["  · 하드커버: '표지비용 따로 계산' = 표지는 별 상품(예 PRD_000073) / 이 시트는 제본비만", ""],
    ["  · 캘린더: '삼각대 포함' = 받침 부속 포함, 별도 부속비 없음 (가격 완결)", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_RU", "공식정의 1 (PRF_BIND_SUM·라이브 재현)"],
    ["  1b_product_price_formulas_RU", "상품↔공식 바인딩 4 (책자4종·라이브 재현)"],
    ["  2_formula_components_RU", "공식 배선 1 (JUNGCHEOL만·라이브 재현·단절1 상태)"],
    ["  3_price_components_RU", "구성요소 정의 11 (제본방식 11종·라이브 재현·전건 단가형)"],
    ["  4_component_prices_RU", "단가행 74 (이 시트 원천·라이브 재현·무손실)"],
    ["  8_FIX_wiring_chain", "단절1 교정 후보 — 미배선 10 component (⚠️ 그대로 INSERT 금지·공식모델 컨펌 선결)"],
    ["  9_FIX_binding_chain", "단절2 교정 후보 — 미바인딩 상품 바인딩 (컨펌 BIND-C1/C2 선결)"],
    ["", ""],
    ["분해 무손실 검산", "시트 데이터셀 74 (B1 4×8=32 + B2 3×6=18 + B3 4×6=24) = 라이브 74행 일치(mismatch 0)"],
    ["제본방식↔comp", "11/11 매핑 (BLOCKED 없음 — comp 분리로 싸바리·캘린더도 수용)"],
]
for ri, (a, b) in enumerate(readme, start=1):
    ws.cell(row=ri, column=1, value=a)
    ws.cell(row=ri, column=2, value=b)
    if ri == 1:
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 76

# ============================================================
# 1. price_formulas (라이브 재현)
# ============================================================
frm = read_psv("formulas.psv")
write_sheet(
    wb,
    "1_price_formulas_RU",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in frm],
    note="[RU] 라이브 t_prc_price_formulas 재현 — frm_typ_cd·prd_cd 컬럼 없음(라이브 실측). 신규 적재 아님.",
    fill=RU_FILL,
    col_comments={
        "frm_cd": "공식 식별자 PK. PRF_BIND_SUM=제본 합산형(책자 원자합산 공식의 제본비 구성요소)",
        "use_yn": "Y=노출",
    },
)

# ============================================================
# 1b. product_price_formulas (바인딩, 라이브 재현)
# ============================================================
bind = read_psv("binding.psv")
write_sheet(
    wb,
    "1b_product_price_formulas_RU",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일(yyyy-MM-dd)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in bind],
    note="[RU] 라이브 t_prd_product_price_formulas 재현 — 책자4종만 바인딩(PRD_000068 중철~071 트윈링). 하드커버/캘린더/포토북 미바인딩=단절2(9_FIX 참조).",
    fill=RU_FILL,
    col_comments={
        "prd_cd": "PRD_000068 중철책자·069 무선책자·070 PUR책자·071 트윈링책자. 하드커버/캘린더 등 부재=단절2",
        "apply_bgn_ymd": "PK 구성요소(prd_cd, frm_cd, apply_bgn_ymd). 적용 시작일",
    },
)

# ============================================================
# 2. formula_components (배선, 라이브 재현)
# ============================================================
wire = read_psv("wiring.psv")
write_sheet(
    wb,
    "2_formula_components_RU",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    [[r[0], r[1], int(r[2]) if r[2] else None, r[3]] for r in wire],
    note="[RU] 라이브 t_prc_formula_components 재현 — 1행만(COMP_BIND_JUNGCHEOL). 나머지 10 component 미배선=단절1(8_FIX 참조).",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "현재 JUNGCHEOL 1개만 배선됨. 무선/PUR/트윈링/하드커버/싸바리/캘린더 미배선",
        "addtn_yn": "Phase11 엔진 무시(런타임 선택값으로 1개 comp 매칭). 공식모델은 컨펌 BIND-C1",
    },
)

# ============================================================
# 3. price_components (구성요소 정의 11, 라이브 재현)
# ============================================================
comp = read_psv("components.psv")
comp_rows = []
for r in comp:
    cc = r[0]
    label = METHOD_MAP.get(cc, ("", ""))[0]
    comp_rows.append(
        [r[0], r[1], r[2], r[3], r[4] if len(r) > 4 else "", r[5] if len(r) > 5 else "Y"]
    )
write_sheet(
    wb,
    "3_price_components_RU",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    comp_rows,
    note="[RU] 라이브 t_prc_price_components 재현 11종 — 제본방식 11종 = comp 분리(proc_cd 아님). 전건 PRICE_TYPE.01 단가형·use_dims=[min_qty].",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "제본방식 11종 각각 1 comp. JUNGCHEOL중철/MUSEON무선/PUR/TWINRING트윈링/HC_*하드커버/SSABARI싸바리/CAL_*캘린더4",
        "prc_typ_cd": "전건 PRICE_TYPE.01 단가형(권당 장당가×권수). 수량구간별 차등. 합가형 아님",
        "use_dims": "전건 [min_qty] — 제본방식은 comp로 분리됨(차원 아님). 사이즈/도수/자재 무관",
        "comp_typ_cd": "제본비(후가공) 구성요소. 책자 원자합산 공식의 한 항",
    },
)

# ============================================================
# 4. component_prices (단가행 74, 이 시트 원천·라이브 재현)
# ============================================================
cp = read_psv("component_prices.psv")
# columns: comp_cd|siz_cd|clr_cd|mat_cd|proc_cd|coat_side_cnt|opt_cd|bdl_qty|min_qty|apply_ymd|unit_price
cp_rows = []
for r in cp:
    cp_rows.append([
        r[0],
        empty_to_none(r[1]),
        empty_to_none(r[2]),
        empty_to_none(r[3]),
        empty_to_none(r[4]),
        int(r[5]) if r[5] else None,
        empty_to_none(r[6]),
        int(r[7]) if r[7] else None,
        int(r[8]) if r[8] else None,
        r[9],
        float(r[10]) if r[10] else None,
    ])
write_sheet(
    wb,
    "4_component_prices_RU",
    ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "사이즈(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수", "수량구간시작(min_qty)", "적용일", "단가"],
    cp_rows,
    note="[RU] 라이브 t_prc_component_prices 재현(74행) — 제본 가격표=이 단가의 원천. 10차원 자연키. 안 쓰는 차원 NULL(빈칸). use_dims=[min_qty]만 값.",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "제본방식 11종 × 수량구간 6~8 = 74행. 제본방식 = comp_cd로 분리(proc_cd 아님)",
        "min_qty": "수량구간 시작값(상향구간). 주문수량(권수) 이하 최대 min_qty 행 매칭. B1=8구간(30·70 포함)/B2·B3=6구간",
        "proc_cd": "신설 공정 차원(8→10). 제본방식이지만 전건 NULL — comp_cd 분리로 표현(라이브 권위)",
        "opt_cd": "신설 옵션 차원(8→10). 제본 미사용 NULL",
        "clr_cd": "전건 NULL — 도수 무관(권당 정액)",
        "siz_cd": "전건 NULL — 사이즈 무관",
        "unit_price": "권당 장당가(단가형). 엔진=단가×주문권수",
    },
)

# ============================================================
# 8. FIX wiring chain — 단절1 (미배선 10 component) 교정 후보
# ============================================================
fix_wire = []
seq = 2
for cc, (label, proc) in METHOD_MAP.items():
    if cc == "COMP_BIND_JUNGCHEOL":
        continue  # 이미 배선됨
    fix_wire.append([f"PRF_BIND_<{label}>?", cc, seq, "Y"])
    seq += 1
write_sheet(
    wb,
    "8_FIX_wiring_chain",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드(미정)", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    fix_wire,
    note="[FIX·단절1·⚠️그대로 INSERT 금지] 미배선 10 component. frm_cd 미정 — PRF_BIND_SUM 1개에 다 배선하면 동시매칭(전 제본방식 단가 합산) 오류. 공식모델 컨펌 BIND-C1(A 상품별공식 vs B opt_cd) 선결 후 frm_cd 확정·INSERT.",
    fill=BLK_FILL,
)

# ============================================================
# 9. FIX binding chain — 단절2 (미바인딩 상품) 교정 후보
# ============================================================
fix_bind = [
    ["PRD_000072?", "PRF_BIND_<하드커버무선>?", "2026-06-01", "[교정후보] 하드커버책자 — 상품↔제본방식 권위 확정 필요(BIND-C2)"],
    ["PRD_000082?", "PRF_BIND_<하드커버트윈링>?", "2026-06-01", "[교정후보] 하드커버 링책자"],
    ["PRD_000108?", "PRF_BIND_<탁상220>?", "2026-06-01", "[교정후보] 탁상형캘린더"],
    ["PRD_000111?", "PRF_BIND_<벽걸이>?", "2026-06-01", "[교정후보] 벽걸이캘린더"],
    ["PRD_000100?", "PRF_BIND_<?>", "2026-06-01", "[교정후보] 포토북 — 제본방식 미확정"],
]
write_sheet(
    wb,
    "9_FIX_binding_chain",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드(추정)", "공식코드(미정)", "적용시작일", "비고"],
    fix_bind,
    note="[FIX·단절2·⚠️추정·컨펌 선결] 제본 쓰는 상품 미바인딩. prd_cd·frm_cd 모두 미확정(상품명 기반 추정만). 상품↔제본방식 1:1 권위(상품마스터·책자 도메인) 확정 후 INSERT. 추정 분해 금지.",
    fill=BLK_FILL,
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("component_prices rows:", len(cp_rows))
print("components:", len(comp_rows), "bindings:", len(bind), "wiring:", len(wire))
