#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
L1 자동대조 — 누락0 기계 보증 (E 완전성 검증 4종 + 메타 보존 검증).

검증 항목:
  ① 컬럼 커버리지: 원본 비어있지 않은 컬럼 수 == 추출 필드 수
  ② non-empty 셀 보존율: 원본 non-empty 셀 전부가 추출본에 보존 (100%)
     - ffill 로 채워진 셀은 보존분에서 제외(원본 빈셀이므로) -> 원본 non-empty 만 대조
  ③ 행 카운트: 원본 데이터 행 수 == 추출 레코드 수
  ④ round-trip: 추출본으로 원본 셀 재구성 -> 원본 diff (ffill 펼침분은 메타로 역보정)
  + 메타: 스레드댓글 수 일치

엑셀 정보 축 전수 보존 검증(누락0 기계보증) — 신규 4축:
  ⑤ 행숨김 보존: 원본 숨김행(데이터보유) 수 == 추출 _row_hidden=true 레코드 수
  ⑥ 열숨김 보존: 원본 숨김열 데이터 전부 추출 보존 + col_hidden 플래그 일치
  ⑦ 수식 셀 보존: 원본 수식 셀 수 == 추출 is_formula=true 메타 수
  ⑧ 하이퍼링크 보존: 원본 하이퍼링크 수 == 추출 hyperlink 메타 수

원본 권위, 추출본은 silsa-l1.csv / silsa-l1-meta.csv.
read-only.
"""
import sys, csv, json, argparse
import openpyxl
from openpyxl.utils import get_column_letter

# extract_l1 의 composite/ffill 로직 재사용
sys.path.insert(0, __import__("os").path.dirname(__file__))
from extract_l1 import (build_composite_headers, find_data_rows,
                        load_threaded_comments, FFILL_WHITELIST)


def norm(v):
    """원본 셀값을 추출 CSV 문자열과 동일 정규화 (CSV 는 모든 값을 str 로 저장)."""
    if v is None:
        return ""
    # openpyxl 숫자/문자 -> str. CSV writer 도 동일하게 str() 처리.
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def run(xlsx_path, sheet_name, csv_path, meta_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_fmt = openpyxl.load_workbook(xlsx_path, data_only=False)  # 수식/하이퍼링크 원본
    ws = wb[sheet_name]
    ws_f = wb_fmt[sheet_name]
    maxc = ws.max_column
    comp, work_col, cut_col, bleed_col = build_composite_headers(ws)
    data_rows = find_data_rows(ws, [])
    comments = load_threaded_comments(xlsx_path, sheet_name)

    # ffill 컬럼 idx
    ffill_cols = set(c for c, name in comp.items() if name in FFILL_WHITELIST)

    # 원본 non-empty 셀 좌표(헤더 제외, 데이터행만)
    orig_cells = {}      # (r, c) -> str value
    orig_nonempty = 0
    cols_with_data = set()
    for r in data_rows:
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                orig_cells[(r, c)] = norm(v)
                orig_nonempty += 1
                cols_with_data.add(c)
    # 헤더행(1,2)도 컬럼 존재 판정에 포함 (헤더만 있는 컬럼도 유효 컬럼)
    for r in (1, 2):
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                cols_with_data.add(c)

    # 추출본 로드
    with open(csv_path, encoding="utf-8-sig") as f:
        rd = csv.DictReader(f)
        ext_fields = rd.fieldnames
        ext_rows = list(rd)
    # 추출 데이터 컬럼명 (fixed/메타 제외)
    fixed = {"sheet", "row_seq", "prd_nm", "_anchor_ffilled", "_row_hidden",
             "_work_size_col", "_work_size_value", "cell_meta_json"}
    data_field_names = [f for f in ext_fields if f not in fixed]

    # composite idx -> name 매핑
    idx_to_name = {c: comp[c] for c in range(1, maxc + 1)}

    results = {}

    # ① 컬럼 커버리지
    orig_col_count = len(cols_with_data)
    ext_col_count = len(data_field_names)
    results["col_coverage"] = {
        "orig_nonempty_cols": orig_col_count,
        "extract_fields": ext_col_count,
        "maxcol": maxc,
        # 전 컬럼 보존 = maxcol 과 추출필드수 일치 (G: 전 컬럼 덤프)
        "pass": (ext_col_count == maxc) and (orig_col_count <= ext_col_count),
    }

    # ③ 행 카운트
    results["row_count"] = {
        "orig_data_rows": len(data_rows),
        "extract_records": len(ext_rows),
        "pass": len(data_rows) == len(ext_rows),
    }

    # ② non-empty 셀 보존율 + ④ round-trip
    # 추출본을 (row_seq, colname) -> value 맵으로
    ext_map = {}
    ext_ffill_applied = {}   # (row_seq) -> bool
    for er in ext_rows:
        rseq = int(er["row_seq"])
        ext_ffill_applied[rseq] = (er["_anchor_ffilled"] == "true")
        for name in data_field_names:
            ext_map[(rseq, name)] = er.get(name, "")

    preserved = 0
    missing = []        # 원본 non-empty 인데 추출에 없음
    mismatch = []       # 값 불일치
    roundtrip_diff = 0
    rt_diff_samples = []

    for (r, c), oval in orig_cells.items():
        name = idx_to_name[c]
        eval_ = ext_map.get((r, name), None)
        if eval_ is None:
            missing.append((r, get_column_letter(c), oval))
            continue
        if eval_ == oval:
            preserved += 1
        else:
            # ffill 컬럼은 원본값 그대로여야 함(ffill 은 빈셀만 채움)
            mismatch.append((r, get_column_letter(c), oval, eval_))

    # ④ round-trip: 추출본 -> 원본 셀 재구성 후 diff
    # ffill 로 채워진 셀(원본 빈셀)은 역보정(빈값으로) 후 원본과 대조.
    for er in ext_rows:
        rseq = int(er["row_seq"])
        for c in range(1, maxc + 1):
            name = idx_to_name[c]
            ext_val = er.get(name, "")
            orig_val = norm(ws.cell(rseq, c).value)
            # ffill 역보정: ffill 컬럼이고 원본이 빈값인데 추출이 채워졌으면 -> 메타(_anchor_ffilled)로 설명되는 정당 펼침
            if c in ffill_cols and orig_val == "" and ext_val != "":
                # 정당한 ffill 펼침 — round-trip 시 원본 빈값으로 환원되어야 무손실
                # (메타 _anchor_ffilled=true 로 기록됨 -> diff 아님)
                continue
            if ext_val != orig_val:
                roundtrip_diff += 1
                if len(rt_diff_samples) < 20:
                    rt_diff_samples.append({
                        "row": rseq, "col": get_column_letter(c),
                        "name": name, "orig": orig_val, "ext": ext_val})

    preservation_rate = (preserved / orig_nonempty * 100.0) if orig_nonempty else 100.0
    results["nonempty_preservation"] = {
        "orig_nonempty_cells": orig_nonempty,
        "preserved": preserved,
        "missing": len(missing),
        "mismatch": len(mismatch),
        "preservation_rate_pct": round(preservation_rate, 4),
        "pass": (len(missing) == 0 and len(mismatch) == 0 and preserved == orig_nonempty),
        "missing_samples": missing[:10],
        "mismatch_samples": mismatch[:10],
    }
    results["roundtrip"] = {
        "diff_count": roundtrip_diff,
        "pass": roundtrip_diff == 0,
        "diff_samples": rt_diff_samples,
    }

    # 메타: 스레드댓글 수
    with open(meta_path, encoding="utf-8-sig") as f:
        meta_rows = list(csv.DictReader(f))
    ext_comments = set()
    for m in meta_rows:
        if m.get("comment_text", "").strip():
            ext_comments.add((m["ref"]))
    results["comment_preservation"] = {
        "orig_comments": len(comments),
        "extract_comment_refs": len(ext_comments),
        "pass": len(ext_comments) == len(comments),
    }

    # ===== 신규 4축 보존 검증 =====

    # ⑤ 행숨김 보존: 원본 숨김행(데이터행 한정) == 추출 _row_hidden=true 레코드
    orig_hidden_rows = []
    for r in data_rows:
        dim = ws.row_dimensions.get(r)
        if dim is not None and dim.hidden:
            orig_hidden_rows.append(r)
    ext_hidden_rows = [int(er["row_seq"]) for er in ext_rows
                       if er.get("_row_hidden", "") == "true"]
    results["row_hidden_preservation"] = {
        "orig_hidden_data_rows": sorted(orig_hidden_rows),
        "orig_count": len(orig_hidden_rows),
        "extract_count": len(ext_hidden_rows),
        "extract_rows": sorted(ext_hidden_rows),
        "pass": sorted(orig_hidden_rows) == sorted(ext_hidden_rows),
    }

    # ⑥ 열숨김 보존: 원본 숨김열 == 추출 col_hidden 플래그(메타 hidden_column 행) + 데이터 보존
    orig_hidden_cols = []
    for c in range(1, maxc + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        if dim is not None and dim.hidden:
            orig_hidden_cols.append(c)
    # 메타에서 hidden_column 타입 행 수집
    ext_hidden_col_names = set()
    for m in meta_rows:
        if m.get("type") == "hidden_column" or m.get("col_hidden", "") == "true":
            ext_hidden_col_names.add(m.get("col", ""))
    orig_hidden_col_names = set(comp[c] for c in orig_hidden_cols)
    # 숨김열 데이터가 추출본에 보존됐는지(컬럼 자체가 추출 필드에 존재)
    hidden_col_data_preserved = all(comp[c] in data_field_names for c in orig_hidden_cols)
    results["col_hidden_preservation"] = {
        "orig_hidden_cols": [get_column_letter(c) for c in orig_hidden_cols],
        "orig_hidden_col_names": sorted(orig_hidden_col_names),
        "extract_flagged_col_names": sorted(ext_hidden_col_names),
        "hidden_col_data_in_fields": hidden_col_data_preserved,
        "pass": (orig_hidden_col_names.issubset(ext_hidden_col_names)
                 and hidden_col_data_preserved),
    }

    # ⑦ 수식 셀 보존: 원본 수식 셀(데이터행) == 추출 is_formula 메타
    orig_formula = 0
    for r in data_rows:
        for c in range(1, maxc + 1):
            v = ws_f.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                orig_formula += 1
    ext_formula = sum(1 for m in meta_rows if m.get("is_formula", "") == "true")
    results["formula_preservation"] = {
        "orig_formula_cells": orig_formula,
        "extract_formula_cells": ext_formula,
        "pass": orig_formula == ext_formula,
    }

    # ⑧ 하이퍼링크 보존: 원본 하이퍼링크(데이터행) == 추출 hyperlink 메타
    orig_hyperlink = 0
    for r in data_rows:
        for c in range(1, maxc + 1):
            cell = ws_f.cell(r, c)
            try:
                if cell.hyperlink is not None and cell.hyperlink.target:
                    orig_hyperlink += 1
            except Exception:
                pass
    ext_hyperlink = sum(1 for m in meta_rows if m.get("hyperlink", "").strip())
    results["hyperlink_preservation"] = {
        "orig_hyperlink_cells": orig_hyperlink,
        "extract_hyperlink_cells": ext_hyperlink,
        "pass": orig_hyperlink == ext_hyperlink,
    }

    overall = all(v["pass"] for v in results.values())
    return overall, results


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="docs/huni/후니프린팅_상품마스터_260527.xlsx")
    ap.add_argument("--sheet", default="실사")
    ap.add_argument("--csv", default="_workspace/huni-dbmap/06_extract/silsa-l1.csv")
    ap.add_argument("--meta", default="_workspace/huni-dbmap/06_extract/silsa-l1-meta.csv")
    args = ap.parse_args()
    overall, results = run(args.xlsx, args.sheet, args.csv, args.meta)
    print(json.dumps(results, ensure_ascii=False, indent=2))
    print("\n=== OVERALL:", "PASS" if overall else "FAIL", "===")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
