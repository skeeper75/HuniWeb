#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
round-10 버전쌍 키 기반 3-way diff (상품마스터 260527 -> 260610).

dbm-change-tracking references/diff-engine.md 패턴 구현.

[설계 핵심 — 후니 상품마스터의 다중행 블록 구조]
 - 한 시트는 "상품 블록"의 연속이다. 블록 헤더행 = ID(col B) 비공백 행.
   ID가 그 상품의 안정 surrogate 키(예: 14625). 같은 블록 내 후속 행(ID 공백)은
   사이즈/가격 변형(variant) 행이다(상품명/구분/ID는 ffill 됨).
 - 따라서 "위치 기반 비교 금지"(V1)를 두 층위로 적용한다:
     ① 상품 식별 = 정규화 ID (없으면 prd_nm)  -> ADDED/REMOVED/유지 상품 판별
     ② 셀 식별   = (상품키, variant_idx, 컬럼) -> 블록 내 변형행 cell-level diff
   variant_idx = 블록 내 0-based 순번(헤더행=0). 행 삽입/삭제가 블록 끝에서
   일어나면 꼬리 정렬이 유지된다. 블록 중간 삽입은 tail-shift 유령 diff를
   만들 수 있으므로, 블록 행수가 바뀐 상품은 'block_resized' 플래그로 표기하고
   그 블록의 cell-diff는 신뢰도 LOW로 분류(사람 판단).

[정규화 — false diff 차단(V2 전제)]
 - ID 타입 드리프트: 260527은 int(14625), 260610은 float(14625.0).
   숫자형 키/값은 정수면 .0 제거해 동일화. (전 시트 공통 — 미정규화 시
   모든 상품이 MODIFIED 로 오분류됨.)
 - 셀 정규화: None->"", 양끝 공백 strip, 숫자 14625.0->14625 / 2700.0->2700,
   순수정수 float는 int 문자열로. 양 버전 동일 적용.

read-only. 원본·DB 무변경. 비밀값 비취급.
"""
import sys
import os
import csv
import json
import argparse
import openpyxl
from openpyxl.utils import get_column_letter

# 전 13시트(상품마스터). 키 컬럼은 모두 동일 구조(B=ID, D=상품명).
SHEETS = [
    "계산공식집초안", "MAP", "디지털인쇄", "스티커", "책자",
    "포토북(가격포함)", "캘린더", "디자인캘린더(가격포함)", "실사",
    "아크릴", "문구(가격포함)", "굿즈파우치(가격포함)", "상품악세사리(가격포함)",
]

ID_COL = 2   # column B = ID (surrogate)
NM_COL = 4   # column D = 상품명
GB_COL = 1   # column A = 구분


def norm_cell(v):
    """양 버전 동일 셀 정규화. 숫자 정수형 float -> int 문자열, 공백 strip.
    [부동소수 노이즈 제거] 7700.000000000001 같은 IEEE 반올림 오차는
    의미상 변경이 아니다(같은 가격). 정수에 충분히 가까우면(|x-round|<1e-6)
    정수로 동일화한다. 진짜 소수값(예: 0.15 할인율)은 round(12)로 안정화.
    미정규화 시 실사 가격 26셀이 유령 변경으로 오분류됨."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float):
        if abs(v - round(v)) < 1e-6:
            return str(int(round(v)))
        return repr(round(v, 12))
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    # "14625.0" 형태 텍스트도 정수면 정규화(타입 드리프트가 텍스트로 새는 경우)
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        return s[:-2]
    # "7700.000000000001" 형태 텍스트 부동소수 노이즈도 정규화
    try:
        f = float(s)
        if abs(f - round(f)) < 1e-6 and ("." in s or "e" in s.lower()):
            return str(int(round(f)))
    except (ValueError, OverflowError):
        pass
    return s


def composite_headers(ws):
    """row1 가로병합 그룹명 + row2 하위명 -> composite 컬럼명 (extract_l1.py와 동일 규칙)."""
    maxc = ws.max_column
    col_group = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row == 1 and mr.max_col > mr.min_col:
            g = ws.cell(1, mr.min_col).value
            for c in range(mr.min_col, mr.max_col + 1):
                col_group[c] = (str(g).strip() if g is not None else "")
    names = {}
    seen = {}
    for c in range(1, maxc + 1):
        r1 = ws.cell(1, c).value
        r2 = ws.cell(2, c).value
        r1 = str(r1).strip() if r1 is not None else ""
        r2 = str(r2).strip() if r2 is not None else ""
        if c in col_group:
            grp = col_group[c]
            sub = r2 if r2 else r1
            name = f"{grp}_{sub}" if sub else grp
        else:
            name = r1 if r1 else r2
            if not name:
                name = get_column_letter(c)
        if name in seen:
            name = f"{name}({get_column_letter(c)})"
        seen[name] = True
        names[c] = name
    return names


def load_blocks(path, sheet, header_rows=2):
    """상품 블록 맵 구성.
    반환: (headers, blocks, dup_ids, blank_id_rows)
      blocks: { product_key: {
          "anchor_row": excel row,
          "prd_nm": str, "gubun": str, "raw_id": str,
          "variants": [ {col_name: norm_value, ...}, ... ]  # 0=헤더행
      } }
    product_key = 정규화 ID(우선), ID 없으면 prd_nm.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    maxc = ws.max_column
    headers = composite_headers(ws)

    # 1) 블록 헤더행 수집.
    #    상품마스터 두 키 레짐 모두 지원:
    #      (a) ID(col B) 비공백  -> ID-keyed 시트(아크릴/디지털인쇄/스티커 등)
    #      (b) ID 전무 시트는 상품명(col D) 비공백을 블록 앵커로(포토북/상품악세사리)
    #    sheet_keys_on_name = 그 시트 전체에 ID 비공백 행이 0인지로 판정.
    has_any_id = any(
        ws.cell(r, ID_COL).value not in (None, "")
        for r in range(header_rows + 1, ws.max_row + 1)
    )
    head_rows = []
    blank_id_rows = []
    for r in range(header_rows + 1, ws.max_row + 1):
        idv = ws.cell(r, ID_COL).value
        nmv = ws.cell(r, NM_COL).value
        nonempty = any(ws.cell(r, c).value is not None for c in range(1, maxc + 1))
        if not nonempty:
            continue
        id_ok = idv is not None and str(idv).strip() != ""
        nm_ok = nmv is not None and str(nmv).strip() != ""
        if has_any_id:
            if id_ok:
                head_rows.append(r)
            else:
                blank_id_rows.append(r)
        else:
            # 상품명 키 레짐: 상품명 비공백 = 블록 앵커
            if nm_ok:
                head_rows.append(r)
            else:
                blank_id_rows.append(r)

    # 2) 블록 경계: 다음 헤더행 직전까지 (단, 비공백 행만 variant)
    blocks = {}
    dup_ids = []
    for i, hr in enumerate(head_rows):
        end = head_rows[i + 1] if i + 1 < len(head_rows) else ws.max_row + 1
        raw_id = norm_cell(ws.cell(hr, ID_COL).value)
        prd_nm = norm_cell(ws.cell(hr, NM_COL).value)
        gubun = norm_cell(ws.cell(hr, GB_COL).value)
        key = raw_id if raw_id else prd_nm
        if not key:
            continue
        variants = []
        for rr in range(hr, end):
            if not any(ws.cell(rr, c).value is not None for c in range(1, maxc + 1)):
                continue
            cells = {}
            for c in range(1, maxc + 1):
                cells[headers[c]] = norm_cell(ws.cell(rr, c).value)
            cells["__excel_row__"] = str(rr)
            variants.append(cells)
        if key in blocks:
            dup_ids.append(key)
        blocks[key] = {
            "anchor_row": hr,
            "prd_nm": prd_nm,
            "gubun": gubun,
            "raw_id": raw_id,
            "variants": variants,
        }
    wb.close()
    return headers, blocks, sorted(set(dup_ids)), blank_id_rows


def diff_sheet(base_path, new_path, sheet):
    hB, B, dupB, blkB = load_blocks(base_path, sheet)
    hN, N, dupN, blkN = load_blocks(new_path, sheet)
    bk, nk = set(B), set(N)
    added = sorted(nk - bk)
    removed = sorted(bk - nk)
    common = sorted(bk & nk)

    changes = []  # cell-level MODIFIED
    block_resized = []
    for key in common:
        vb = B[key]["variants"]
        vn = N[key]["variants"]
        if len(vb) != len(vn):
            block_resized.append({
                "key": key, "prd_nm": N[key]["prd_nm"],
                "base_rows": len(vb), "new_rows": len(vn),
                "delta": len(vn) - len(vb),
            })
        # 위치 정렬 cell diff (variant_idx 기준). 행수 다르면 LOW 신뢰.
        conf = "HIGH" if len(vb) == len(vn) else "LOW"
        maxv = max(len(vb), len(vn))
        for vi in range(maxv):
            cb = vb[vi] if vi < len(vb) else {}
            cn = vn[vi] if vi < len(vn) else {}
            allcols = (set(cb) | set(cn)) - {"__excel_row__"}
            for col in sorted(allcols):
                x = cb.get(col, "")
                y = cn.get(col, "")
                if x != y:
                    changes.append({
                        "key": key,
                        "prd_nm": N[key]["prd_nm"],
                        "variant_idx": vi,
                        "col": col,
                        "before": x,
                        "after": y,
                        "base_cell": cb.get("__excel_row__", ""),
                        "new_cell": cn.get("__excel_row__", ""),
                        "confidence": conf,
                    })
    return {
        "sheet": sheet,
        "added": [{"key": k, "prd_nm": N[k]["prd_nm"], "gubun": N[k]["gubun"],
                   "anchor_row": N[k]["anchor_row"], "variant_count": len(N[k]["variants"])}
                  for k in added],
        "removed": [{"key": k, "prd_nm": B[k]["prd_nm"], "gubun": B[k]["gubun"],
                     "anchor_row": B[k]["anchor_row"], "variant_count": len(B[k]["variants"])}
                    for k in removed],
        "modified_cells": changes,
        "block_resized": block_resized,
        "dup_ids": sorted(set(dupB + dupN)),
        "base_blank_id_rows": len(blkB),
        "new_blank_id_rows": len(blkN),
        "base_product_count": len(B),
        "new_product_count": len(N),
    }


def rename_suspects(res):
    """ADDED + REMOVED 가 동시에 있으면 prd_nm/variant 유사도로 rename 의심쌍 플래그."""
    out = []
    if not res["added"] or not res["removed"]:
        return out
    for a in res["added"]:
        for r in res["removed"]:
            sa, sr = set(a["prd_nm"]), set(r["prd_nm"])
            sim = len(sa & sr) / max(1, len(sa | sr))
            vc_close = abs(a["variant_count"] - r["variant_count"]) <= 1
            if sim >= 0.5 or vc_close:
                out.append({"added_key": a["key"], "added_nm": a["prd_nm"],
                            "removed_key": r["key"], "removed_nm": r["prd_nm"],
                            "name_sim": round(sim, 2), "variant_close": vc_close})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default="docs/huni/후니프린팅_상품마스터_260527.xlsx")
    ap.add_argument("--new", default="docs/huni/후니프린팅_상품마스터_260610.xlsx")
    ap.add_argument("--outdir", default="_workspace/huni-dbmap/14_change-tracking/260527-to-260610")
    args = ap.parse_args()

    diffdir = os.path.join(args.outdir, "diff")
    os.makedirs(diffdir, exist_ok=True)

    summary = []
    for sheet in SHEETS:
        res = diff_sheet(args.base, args.new, sheet)
        res["rename_suspects"] = rename_suspects(res)
        slug = sheet.replace("(가격포함)", "").replace("/", "-")
        # per-sheet changes CSV
        with open(os.path.join(diffdir, f"{slug}-changes.csv"), "w",
                  newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["sheet", "change_type", "key", "prd_nm", "variant_idx",
                        "column", "before", "after", "base_cell", "new_cell", "confidence"])
            for a in res["added"]:
                w.writerow([sheet, "ADDED", a["key"], a["prd_nm"], "",
                            "(전체 상품 신규)", "", "", "", f"{slug}!{a['anchor_row']}", ""])
            for r in res["removed"]:
                w.writerow([sheet, "REMOVED", r["key"], r["prd_nm"], "",
                            "(전체 상품 삭제)", "", "", f"{slug}!{r['anchor_row']}", "", ""])
            for c in res["modified_cells"]:
                w.writerow([sheet, "MODIFIED", c["key"], c["prd_nm"], c["variant_idx"],
                            c["col"], c["before"], c["after"],
                            c["base_cell"], c["new_cell"], c["confidence"]])
        summary.append(res)

    with open(os.path.join(diffdir, "_diff-raw.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=1)

    # console summary
    print("sheet,added,removed,modified_cells,block_resized,dup_ids,rename_suspects,base_prod,new_prod")
    tot = [0, 0, 0, 0]
    for r in summary:
        print(f"{r['sheet']},{len(r['added'])},{len(r['removed'])},"
              f"{len(r['modified_cells'])},{len(r['block_resized'])},"
              f"{len(r['dup_ids'])},{len(r['rename_suspects'])},"
              f"{r['base_product_count']},{r['new_product_count']}")
        tot[0] += len(r["added"]); tot[1] += len(r["removed"])
        tot[2] += len(r["modified_cells"]); tot[3] += len(r["block_resized"])
    print(f"TOTAL,{tot[0]},{tot[1]},{tot[2]},{tot[3]},,,,")


if __name__ == "__main__":
    main()
