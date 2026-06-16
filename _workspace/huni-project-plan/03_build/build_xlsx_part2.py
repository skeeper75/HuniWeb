# -*- coding: utf-8 -*-
"""Part2 — 신규 6시트 추가(위젯·MES·고객준비물·외부계약·용어집·개발자상세)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "docs/huni/후니프린팅_프로젝트일정관리_통합IA_260616.xlsx"

THIN = Side(style="thin", color="C9CDD2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HFONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="2C3E50")
TITLE = Font(name="맑은 고딕", size=14, bold=True, color="1A2733")
SUB = Font(name="맑은 고딕", size=10, italic=True, color="566573")
CELL = Font(name="맑은 고딕", size=10, color="1A2733")
CELLB = Font(name="맑은 고딕", size=10, bold=True, color="1A2733")
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
EXT = PatternFill("solid", fgColor="FDEBD0")
INT = PatternFill("solid", fgColor="EBF5FB")
SZ = {"S": "D5F5E3", "M": "FCF3CF", "L": "FDEBD0", "XL": "FADBD8"}


def hrow(ws, headers, r=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = HFONT; cell.fill = HFILL; cell.alignment = CTR; cell.border = BORDER
    ws.row_dimensions[r].height = 30


def widths(ws, m):
    for col, w in m.items():
        ws.column_dimensions[col].width = w


def title_block(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title); t.font = TITLE; t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=sub); s.font = SUB; s.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 18


def datarow(ws, r, vals, ctr_cols=(), fills=None, h=30):
    for c, v in enumerate(vals, 1):
        cc = ws.cell(row=r, column=c, value=v)
        cc.font = CELL; cc.border = BORDER
        cc.alignment = CTR if (c in ctr_cols) else WRAP
        if fills and c in fills:
            cc.fill = fills[c]
    ws.row_dimensions[r].height = h


wb = openpyxl.load_workbook(OUT)

# ============================================================
# 05_위젯_에디쿠스 기능목록 (쉬운 말 A부)
# ============================================================
ws = wb.create_sheet("05_위젯_에디쿠스")
title_block(ws, "인쇄 자동견적 위젯 + 에디쿠스 기능목록 (쉬운 말)",
            "레드프린팅 역공학 + 후니 위젯 하네스 실측 근거. 위젯=상품 상세 안에 끼우는 견적/주문 박스. 외부의존: 에디쿠스/S3/가격API", 7)
hrow(ws, ["#", "기능명(쉬운 말)", "무엇을 하는지", "적용 상품군", "입력 IA 매핑", "규모", "외부서비스 의존"], r=4)
W = [
 (1,"위젯 띄우기","상품 상세페이지 안에 견적/주문 박스를 끼워 넣어 보여줌(다른 디자인과 안 섞이게 칸막이)","전 상품군","기반","L","없음(자체)"),
 (2,"규격(사이즈) 선택","명함/엽서/책자 등 인쇄물 크기를 고르게 함","전 상품군","B1 규격","S","없음"),
 (3,"수량 선택","몇 부/몇 권 만들지 고르게 함(직접입력+증감)","전 상품군","B3 수량","S","없음"),
 (4,"용지 선택","종이 종류와 평량(두께)을 2단으로 고르게 함","책자·명함·엽서 등","B4 종이/소재","M","없음"),
 (5,"인쇄 도수(색상) 선택","컬러(4도)/단색(1도)을 표지·내지 따로 고르게 함","종이 인쇄류","B5 인쇄도수","S","없음"),
 (6,"자재(재질) 선택","아크릴·굿즈 등에서 본체 재질을 고르게 함","아크릴·굿즈·파우치","B4 종이/소재","M","없음"),
 (7,"내지 장수 입력","책자에서 속지 페이지 수(2~130)를 입력하게 함","책자","B10 내지/표지","S","없음"),
 (8,"후가공 선택","코팅·박·형압 등 마무리 가공을 고르게 함","상품별 종류 다름","B6·B7·B11","M","없음"),
 (9,"제본/링 옵션","링제본·제본방향(좌철/상철)·링 색상을 고르게 함","책자","B9 제본","M","없음"),
 (10,"옵션 자동 잠금(캐스케이드)","어떤 용지/자재를 고르면 안 맞는 후가공을 자동으로 회색처리해 막음","책자 등","B15 옵션 제약","L","없음(서버 규칙)"),
 (11,"실시간 가격 표시","옵션 바꿀 때마다 가격을 다시 계산해 즉시 표시(정가/할인/부가세/배송비 분리)","전 상품군","C1 실시간가격","M","가격 API"),
 (12,"공정별 가격 분해 보기","인쇄비·코팅비·자재비 등을 항목별로 쪼개 보여줌","전 상품군","C2 가격표","M","가격 API"),
 (13,"견적서/가격표 보기","수량 구간별 가격표·견적서를 뽑아줌","전 상품군","C3 견적서","M","가격 API"),
 (14,"인쇄 가이드 보기","작업 사이즈·도련 등 파일 만들 때 지킬 안내선을 보여줌","전 상품군","E2 인쇄영역","M","없음"),
 (15,"파일 업로드(PDF)","완성된 인쇄용 PDF를 끌어다 올리면 클라우드(S3)에 바로 저장","전 상품군","D1 파일업로드","L","S3 업로드"),
 (16,"업로드 파일 검증","올린 파일의 형식·크기·페이지 수 확인, 표지/내지 중복·누락 막음","전 상품군","D1 파일검증","M","S3 메타조회"),
 (17,"온라인 에디터로 디자인","파일이 없어도 화면에서 직접 표지·굿즈를 디자인(템플릿 편집)","에디터 지원 상품(121개)","D3 온라인에디터","XL","에디쿠스"),
 (18,"에디터 저장·미리보기","에디터 디자인을 저장하고 썸네일/미리보기를 받아 주문에 붙임","에디터 지원 상품","D3·D2","L","에디쿠스"),
 (19,"주문 데이터 생성","고른 옵션+가격+업로드/디자인 결과를 묶어 주문 폼에 넘김","전 상품군","F2 주문전환","M","없음(자체)"),
 (20,"주문 가능 여부 검증","필수 옵션·파일이 다 찼는지 확인해 빠지면 주문 막음","전 상품군","F1 주문 전 검증","M","없음"),
 (21,"부자재(봉투 등) 주문 흐름","봉투·스티커 등 부자재류는 책자와 다른 단순 흐름으로 처리","부자재","B13 추가상품","M","가격 API"),
 (22,"한글 라벨 표시","옵션명을 한글로 보기 좋게 표시(서버 코드 → 한글 사전)","전 상품군","UI 기반","S","없음"),
]
r = 5
for w in W:
    fills = {6: PatternFill("solid", fgColor=SZ.get(w[5], "FFFFFF"))}
    if w[6] != "없음" and "자체" not in w[6]:
        fills[7] = EXT
    datarow(ws, r, list(w), ctr_cols=(1, 5, 6), fills=fills)
    r += 1
r += 1
for line in [
 "요약: 외부서비스 의존 = 에디쿠스 2건(17·18)·S3 업로드 2건(15·16)·가격 API 4건(11·12·13·21). 나머지 14건은 위젯 자체 동작.",
 "★ 핵심 결정: 위젯은 후니 DB가 아니라 '정규화 계약'에 의존 → 레드 데이터로 구현·검증 후 후니 어댑터 교체만으로 무손실 전환(위젯 코드 불변). DB 미정이어도 위젯 개발 진행 가능.",
 "★ 가격 불변식: 위젯은 가격을 절대 계산하지 않음(서버 권위). 응답의 최종가/부가세/배송비만 표시. (개발자 상세는 10_개발자상세 시트)",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cc = ws.cell(row=r, column=1, value=line); cc.font = CELLB if line.startswith("★") else SUB; cc.alignment = WRAP
    ws.row_dimensions[r].height = 30; r += 1
widths(ws, {"A": 4, "B": 22, "C": 46, "D": 20, "E": 14, "F": 6, "G": 14})
ws.freeze_panes = "A5"

# ============================================================
# 06_주문_생산_MES 흐름
# ============================================================
ws = wb.create_sheet("06_주문_생산_MES")
title_block(ws, "주문 → 생산(MES) → 포장 → 배송 흐름 (사이트 안/밖 구분)",
            "주문형 생산. 실측: 라이브 DB는 상품·가격 마스터만 존재 / 주문·결제·파일·배송 런타임 0개=전부 신규구축", 6)
# 흐름도
ws.cell(row=4, column=1, value="① 전체 흐름 (단계별 안/밖)").font = CELLB
flow = [
 "① 상품선택 [내부] → ② 옵션선택(위젯) [내부] → ③ 가격확인 [내부] → ④ 파일업로드[내부+S3] / 에디터디자인[외부 에디쿠스]",
 "→ ⑤ 장바구니 [내부·신규] → ⑥ 결제 [외부 이니시스] —(결제완료 트리거)—",
 "→ ⑦ 생산파일 생성 [내부+외부] → ⑧ 생산정보 전달(BOM 환원) [내부→외부 MES] → ⑨ MES 생산관리 [외부]",
 "→ ⑩ 공정(출력·코팅·재단·후가공·제본) [외부 MES] → ⑪ 포장 [외부 MES] → ⑫ 배송 [내부+외부 택배]",
]
r = 5
for f in flow:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cc = ws.cell(row=r, column=1, value=f); cc.font = CELL; cc.alignment = WRAP
    ws.row_dimensions[r].height = 20; r += 1
r += 1
ws.cell(row=r, column=1, value="② 사이트 내부 기능 (우리가 만든다)").font = CELLB; r += 1
hrow(ws, ["단계", "기능(쉬운 말)", "담당", "관련 DB(t_*)", "입력 IA 매핑", "규모"], r=r); r += 1
INTERNAL = [
 ("① 상품선택","상품 목록·상세 보여주기·고르기","쇼핑개발","t_prd_products·categories","F1 상품진입","하"),
 ("② 옵션선택(위젯)","사이즈·소재·도수·수량 고르는 견적 위젯","인쇄개발","option_groups/options/items·sizes","B1~B15","상"),
 ("③ 가격확인","고른 옵션으로 실시간 가격 계산","인쇄개발","t_prc_*·product_prices","C1 실시간가격","상"),
 ("④-1 파일업로드","인쇄용 원본파일(PDF) 올리기 (실측 Y=219상품)","인쇄개발","file_upload_yn","D1","중"),
 ("④-2 에디터 디자인","웹에서 직접 디자인 후 저장 (실측 Y=121상품)","인쇄개발","editor_yn","D3","상"),
 ("⑤ 장바구니","여러 상품 담아 한 번에 주문 (현재 부재→신규)","쇼핑개발","(신규 그릇)","F4","중"),
 ("⑥ 결제 진입/완료","결제수단 선택→PG 호출→승인콜백→주문확정","쇼핑개발","(신규 그릇)","F4·결제","상"),
 ("⑦ 생산파일 생성","주문/파일번호 채번·파일명 Rename·썸네일 자동생성","인쇄개발","products(MES_ITEM_CD)","D1/D2 후속","상"),
 ("⑧ 생산정보 환원","선택옵션→생산 BOM(자재/공정·사이즈·수량) 환원","인쇄개발","product_materials/processes/sizes","F5 후속","상"),
 ("⑫ 주문 상태조회","고객이 미입금→제작중→출고완료 추적","쇼핑개발","(신규 그릇)","F5","중"),
 ("⑫ 배송정보 입력","배송지·방법(택배/퀵/방문)·추가배송비(제주 등)","쇼핑개발","(신규 그릇)","F4/F5","중"),
]
for row in INTERNAL:
    fills = {3: INT}
    datarow(ws, r, list(row), ctr_cols=(6,), fills=fills); r += 1
r += 1
ws.cell(row=r, column=1, value="③ 외부 연동 (바깥 서비스와 연결)").font = CELLB; r += 1
hrow(ws, ["외부서비스", "무엇을 위해", "연동방식", "계약/준비 필요", "PM 선행조건", "비고"], r=r); r += 1
EXTERNAL = [
 ("에디쿠스(Edicus)","웹 에디터 디자인 + 생산용 랜더링","SDK/iframe + 메시지","SDK 라이선스 계약·랜더 결과 규격","편집기 상품(121개) 범위·저장위치(S3) 합의","P2·D3/D4"),
 ("AWS S3","원본·접수·랜더·썸네일 파일 저장","API(presigned URL)","버킷·권한·보관기간 정책","고객파일 보안·생산팀 다운 권한","P0 인프라"),
 ("이니시스(PG)","결제: 카드·가상계좌·에스크로·포인트","API(승인 콜백)","가맹·MID·결제수단 범위","가상계좌 입금확인 자동화·환불정책","P0 계약"),
 ("알림톡/SMS","주문상태별 자동알림(결제/제작중/출고/재업로드)","API(템플릿)","비즈채널·템플릿 사전심사·발신번호","알림 시나리오 12종 템플릿 확정","P1 계약"),
 ("MES(생산관리)","생산정보 전달·공정상태·바코드·발주","API 또는 파일(EXCEL)","MES↔Front 통신규격·매칭키 정합","MES_ITEM_CD 전수채번(실측 16/275만)·파일명 일관화","P0 핵심"),
 ("Shopby(쇼핑엔진)","커머스 표준 참조(장바구니·주문·결제·배송)","참조/벤치마크(직접종속 X)","통합 여부 결정·표준기능 흡수","자체구축 vs Shopby 채택 결정","검토"),
 ("택배사/퀵","송장발급·배송추적·추가배송비","API 또는 화면","택배사 계약·송장양식","한 주문에 여러 송장 부여 지원 결정","P1 계약"),
 ("외주 생산처","외주제작 발주(현수막·합판·박·폰케이스 등)","파일(웹하드/메일/EXCEL)","외주처별 파일포맷·발주서 양식","발주 자동화 범위(현재 100% 수동)","MES 트랙 포함"),
]
for row in EXTERNAL:
    datarow(ws, r, list(row), fills={1: EXT}); r += 1
r += 1
ws.cell(row=r, column=1, value="④ 생산정보 전달 — 결제완료 시 무슨 데이터가 넘어가나").font = CELLB; r += 1
prod = [
 "트리거: 결제완료 → ① 생산을 위한 파일 생성(주문/파일번호 채번·접수파일 변환·썸네일·파일명 Rename) + ② 생산에 필요한 정보 전달(선택옵션→생산 BOM 환원).",
 "파일명 일관화 규칙(예): 발주일_품목명_출력사이즈_양단면_소재명_거래처명_고객명_파일고유번호_수량 → 1002_접지카드_92x54_D_랑240_BIZ_OOO_30146712_00100ea.pdf",
 "옵션→생산 환원: 사이즈→작업/출력사이즈·판형 / 소재→자재 BOM / 후가공→공정 BOM / 도수→별색 / 수량→제작수량.  (재료=이미 마스터 적재 / 부족=주문 런타임 + MES 매칭키)",
 "★ 실측 결론: MES_ITEM_CD가 275상품 중 16건(5.8%)만 채번됨(94% NULL) = 생산정보 전달의 핵심 브릿지 미완 → '전상품 MES_ITEM_CD 채번'이 1차 선행 과제. 주문/결제/파일/배송 테이블은 라이브에 0개 = 커머스 백엔드 그릇부터 신규.",
]
for line in prod:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cc = ws.cell(row=r, column=1, value=line); cc.font = CELLB if line.startswith("★") else CELL; cc.alignment = WRAP
    ws.row_dimensions[r].height = 32; r += 1
widths(ws, {"A": 16, "B": 32, "C": 18, "D": 26, "E": 14, "F": 18})
ws.freeze_panes = "A4"

# ============================================================
# 07_고객준비물_디자인지원
# ============================================================
ws = wb.create_sheet("07_고객준비물_디자인지원")
title_block(ws, "고객사(후니프린팅) 사전 준비물 & 디자인 지원",
            "개발이 막히지 않으려면 고객사가 먼저 챙겨야 할 것. 디자인·콘텐츠·계약·서류 = 최숙진 실장 담당", 6)
hrow(ws, ["구분", "준비 항목", "왜 필요한가(쉬운 말)", "담당", "필요 시점", "관련 IA/연동"], r=4)
PREP = [
 ("디자인","AI 페이지 디자인 산출물 전달","현재 AI로 페이지 디자인 진행 중 → 메인·상세·랜딩 디자인 산출물(이미지/시안)을 개발에 전달해야 화면 구현 가능","최숙진 실장","1차 W1-5","상세빌더 G1·진열"),
 ("디자인","브랜드 가이드(색·로고·폰트)","화면 톤·컴포넌트 통일. 위젯·상세페이지에 일관 적용","최숙진 실장","1차 W1-3","전 화면"),
 ("콘텐츠","약관·개인정보처리방침 원고","법적 필수 페이지. 런칭 전 게시","최숙진 실장","1차 W3-7","No.84·85"),
 ("콘텐츠","인쇄 작업 가이드 11종 원고","모니터색상/칼라설정/재단선/용지 등 작업 유의사항. 파일 사고 방지","최숙진 실장","2차","No.87·E1~E3"),
 ("콘텐츠","FAQ·공지 초기 원고","고객센터 기본 콘텐츠","최숙진 실장","1차~2차","No.67·68"),
 ("계약","이니시스 가맹 + 가맹관리자 계정","결제(PG) 연동 선행. 가맹 심사·MID 발급","최숙진 실장","1차 W1-7","No.80·이니시스"),
 ("계약","도메인 + SSL 인증서","사이트 주소 연결·HTTPS. 런칭 필수 인프라","최숙진 실장","1차 W1-5","전역"),
 ("계약","문자/알림톡(카카오 비즈채널·발신번호)","주문상태 알림·인증 문자. 발신프로필·템플릿 심사","최숙진 실장","1차~2차","No.138·144·알림톡"),
 ("서류","사업자등록·통신판매업 신고","온라인 판매 법적 요건","최숙진 실장","1차 W1-5","법무"),
 ("데이터","상품마스터·가격표 최신 확정본","위젯·가격엔진의 원천. 변경 시 재매핑 필요","최숙진 실장 + 인쇄개발","1차 W1-3","상품/가격 DB"),
 ("데이터","MES 연동 규격 + MES_ITEM_CD 채번 협조","생산정보 전달의 매칭키(현재 94% 미채번). 생산팀과 합의","최숙진 실장 + 인쇄개발","1차 W7-10","MES 브릿지"),
 ("결정","에디쿠스(온라인 에디터) 도입 여부","편집기 상품(121개)의 디자인 기능. 라이선스·계약 결정 = D3/D4 선행","최숙진 실장 + PM","2차~3차","No.54·55·에디쿠스"),
 ("결정","외주 생산처 발주 규격","외주제작(현수막·합판·박 등) 파일/발주서 양식. 현재 수동","최숙진 실장 + 인쇄개발","2차","MES 외주"),
 ("결정","운영정책 5건(증빙·배송비·인증·탈퇴·수작존치)","화면 설계 분기점. 결정돼야 개발 진행","PM(신우진) + 최숙진 실장","1차 W1-7","다수 No 비고"),
]
r = 5
gcol = {"디자인": "F5EEF8", "콘텐츠": "E8F8F5", "계약": "FDEBD0", "서류": "FEF9E7", "데이터": "EBF5FB", "결정": "FADBD8"}
for row in PREP:
    fills = {1: PatternFill("solid", fgColor=gcol.get(row[0], "FFFFFF"))}
    datarow(ws, r, list(row), ctr_cols=(1, 5), fills=fills, h=34); r += 1
r += 1
for line in [
 "※ 디자인·콘텐츠·이니시스 가맹관리자·도메인·문자메시지·기타 서류는 후니프린팅 최숙진 실장이 담당합니다.",
 "※ 고객사 협조 = AI 페이지 디자인 산출물 지원이 핵심. 디자인 시안이 늦으면 상세빌더·진열이 막히므로 1차 초반에 우선 전달 필요.",
]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cc = ws.cell(row=r, column=1, value=line); cc.font = SUB; cc.alignment = WRAP
    ws.row_dimensions[r].height = 26; r += 1
widths(ws, {"A": 10, "B": 28, "C": 44, "D": 18, "E": 12, "F": 18})
ws.freeze_panes = "A5"

# ============================================================
# 08_외부연동_계약
# ============================================================
ws = wb.create_sheet("08_외부연동_계약")
title_block(ws, "외부 연동 / 계약 체크리스트 (1차 런칭 선행)",
            "바깥 서비스와 연결되는 항목 + 계약. P0=런칭 전 필수 · P1=런칭 직후 · 검토=결정 자체가 선행", 6)
hrow(ws, ["#", "항목", "분류", "1차 필수?", "선행 작업", "관련 IA"], r=4)
EXTC = [
 (1,"이니시스(PG) 결제 계약","PG","P0","가맹 심사·계약, 결제수단(카드/계좌이체/가상계좌) 범위 확정","No.80"),
 (2,"도메인 + SSL(인증서)","인프라","P0","도메인 이전/연결, HTTPS 적용","전역"),
 (3,"사업자/통신판매업/개인정보","법무","P0","통신판매업 신고, 약관·개인정보처리방침 게시","No.84·85"),
 (4,"파일 스토리지(대용량 PDF, S3)","인프라","P0","S3 버킷·업로드 용량/포맷 정책·presigned","No.52"),
 (5,"MES↔Front 통신 규격 + 채번","생산","P0","MES_ITEM_CD 전수 채번(16/275), 주문↔파일↔상품 매칭","생산정보(157·158)"),
 (6,"카카오 알림톡 발신프로필+대행사","메시징","P1","발신프로필 승인, 템플릿 사전심사(주문/배송/재업로드)","No.138·144"),
 (7,"SMS 발송 서비스(문자)","메시징","P1","발신번호 사전등록, SMS 충전","No.7·138·144"),
 (8,"휴대폰 본인인증(나이스/PASS)","인증","P1","가입·비번찾기·증빙 본인확인 연동","No.7·23·24"),
 (9,"간편로그인(네이버/카카오)","인증","P1","OAuth 앱등록·검수, 개인정보 제공범위","No.145"),
 (10,"간편결제(네이버페이/카카오페이/토스)","PG","P1","각 사 가맹 심사, 정산주기","No.80·149"),
 (11,"세금계산서 발행(국세청/팝빌)","증빙","P1","B2B 비중 따라 1차 검토, 사업자정보 동의","No.152"),
 (12,"현금영수증 발급(PG 부가)","PG부가","P1","이니시스 현금영수증 옵션, 발급규칙","No.25·142"),
 (13,"배송사 연동(택배 송장/추적)","배송","P1","계약 택배사 API, 송장출력 양식","No.139·140"),
 (14,"Shopby API(선택적 외부모델)","솔루션","검토","직접 종속 X. 표준기능 보강용 외부연동 여부 결정","전역"),
 (15,"온라인 에디터(에디쿠스) 외부 도입","에디터","P2","도입 여부 결정 자체가 선행(협업)","No.54·55"),
]
r = 5
pf = {"P0": "FADBD8", "P1": "FCF3CF", "검토": "EAECEE", "P2": "EAECEE"}
for row in EXTC:
    fills = {4: PatternFill("solid", fgColor=pf.get(row[3], "FFFFFF"))}
    datarow(ws, r, list(row), ctr_cols=(1, 4), fills=fills, h=30); r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
cc = ws.cell(row=r, column=1, value="요약: 1차 필수(P0) 5건 = 이니시스·도메인/SSL·법무·파일스토리지·MES규격. P1 8건. 검토/P2 2건. P0 계약이 지연되면 결제·생산정보 전달이 막혀 런칭 불가.")
cc.font = CELLB; cc.alignment = WRAP; ws.row_dimensions[r].height = 30
widths(ws, {"A": 4, "B": 30, "C": 12, "D": 9, "E": 44, "F": 18})
ws.freeze_panes = "A5"

# ============================================================
# 09_용어집
# ============================================================
ws = wb.create_sheet("09_용어집")
title_block(ws, "용어집 (범례) — 전문용어 쉬운 설명", "본문은 쉬운 말로 작성. 불가피한 용어만 여기서 설명", 3)
hrow(ws, ["용어", "쉬운 설명", "어디에 나오나"], r=4)
TERMS = [
 ("위젯","상품 상세페이지 안에 끼워 넣는 '견적/주문 박스'. 옵션 고르면 가격이 즉시 바뀜","05 위젯"),
 ("CPQ","옵션을 골라(Configure) 가격을 매기고(Price) 주문(Quote)하는 구성기. 주문제조 인쇄의 핵심","02·05"),
 ("옵션 제약(캐스케이드)","어떤 옵션을 고르면 안 맞는 옵션을 자동으로 막거나 숨기는 규칙(B15)","02·05"),
 ("실시간 가격(C1)","옵션 바꿀 때마다 서버가 다시 계산해 즉시 보여주는 가격. 위젯은 계산 안 하고 표시만","02·05"),
 ("에디쿠스(Edicus)","웹에서 직접 디자인하는 외부 온라인 에디터 서비스. 편집기 상품에 연결","05·06·10"),
 ("S3","대용량 파일(인쇄용 PDF)을 올려두는 아마존 클라우드 저장소","06·08·10"),
 ("MES","생산관리 시스템. 결제 후 '무엇을 어떻게 만들지' 생산 현장으로 정보를 넘김","06·08"),
 ("MES_ITEM_CD","우리 상품과 MES 생산품목을 잇는 매칭 코드. 현재 94% 미채번(선행과제)","06"),
 ("이니시스","카드·계좌이체·가상계좌 결제를 처리하는 결제대행(PG) 회사","06·08"),
 ("알림톡","카카오톡으로 보내는 주문/배송 자동 알림. 발신프로필·템플릿 심사 필요","06·08"),
 ("PG","Payment Gateway. 결제대행. 이니시스·네이버페이 등","08"),
 ("Shopby","NHN커머스의 쇼핑몰 솔루션. 직접 쓰지 않고 '표준 기능 체크리스트·외부연동 후보'로만 참조","03·06·08"),
 ("BOM","Bill of Materials. 하나의 상품을 만드는 데 필요한 자재·공정 명세","06"),
 ("도수","인쇄 색상 수. 4도=컬러, 1도=단색. 별색·화이트 추가 가능(B5)","02·05"),
 ("판형/출력용지규격","인쇄 출력에 쓰는 전지(원지) 규격. 작업사이즈와 다름","06·10"),
 ("Phase / 주차(W)","Phase=단계(1런칭·2안정화·3확장). W=착수일 기준 상대 주차(W1=첫 주)","02·03"),
 ("정규화 계약","위젯이 특정 DB에 묶이지 않게 약속한 표준 데이터 형태. 어댑터로 후니/레드를 바꿔 끼움","05·10"),
 ("폴리모픽 스키마","하나의 참조 컬럼이 사이즈·소재·도수 등 여러 종류를 가리키게 한 설계(ref_dim_cd)","02·10"),
]
r = 5
for t in TERMS:
    datarow(ws, r, list(t), ctr_cols=(3,), h=26); r += 1
widths(ws, {"A": 22, "B": 70, "C": 16})
ws.freeze_panes = "A5"

# ============================================================
# 10_개발자상세
# ============================================================
ws = wb.create_sheet("10_개발자상세")
title_block(ws, "개발자 상세 — 위젯·에디쿠스·DB (기술 참조)",
            "레드프린팅 역공학 + 라이브 실측 근거. 비개발자는 05~06 시트 참조. [라이브]=실측 / [정적]=소스", 4)
r = 4


def sec(title):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cc = ws.cell(row=r, column=1, value=title); cc.font = CELLB
    cc.fill = PatternFill("solid", fgColor="D6EAF8"); ws.row_dimensions[r].height = 20; r += 1


def line(txt, bold=False, h=18):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cc = ws.cell(row=r, column=1, value=txt); cc.font = CELLB if bold else CELL; cc.alignment = WRAP
    ws.row_dimensions[r].height = h; r += 1


sec("B-1. 레드 SDK 3계층 아키텍처 [정적]")
line("브릿지: productRedWidgetSDK.js(33KB·jQuery·17 호스트통합 함수) → 후니=얇은 임베드 로더")
line("런타임: widget.js(438KB·Vue3+Pinia 5스토어·Shadow DOM 렌더) → 후니=React-in-Shadow-DOM")
line("에디터: RedEditorSDK.min.js(45 메서드) → 후니=에디쿠스 브리지")
sec("B-2. 핵심 브릿지 함수 17 (호스트 통합 API) [정적]")
line("초기화 sdkInit/fnInitSdk · 옵션변경→가격 sdkOptionChange · 자재 sdkInformMaterials · 에디터 sdkOpenEditor/fnRpEditor · 가이드 sdkPrintAreaGuide · 주문생성 sdkCreatePot · 주문검증 fnPreOrder/fn_order_able · 가격표/견적 fnCalcPriceTable/fnEstimate", h=34)
sec("B-3. 45 에디터 메서드 중 위젯 호출 핵심 [정적]")
line("후니 에디쿠스 브리지 우선순위: createProject → setToken → setPrice → on(이벤트) → save/saveThenClose → checkOrderable → prepareOrder")
sec("B-4. 가격 API 실측 계약 [라이브]")
line("POST /ko/product_price/get_ajax_price_vTmpl · 인증=세션쿠키 · 평문 JSON")
line("요청: dataJson{ ORD_INFO[PDT_CD·CUT/WRK_WDT/HGH·PRN_CNT·PAGE_CNT·CVR/INN_CLR_CNT·CVR/INN_MTRL_CD]·PCS_INFO[PCS_COD·PCS_DTL_COD]·price_gbn·mb_cust_cod }", h=30)
line("응답: result[](공정별 PRICE 분해) + result_sum(PRICE/PRICE_MALL/ORG_PRICE+VAT) + result_log(단가·배송) + book_info(무게·배송비)", h=30)
line("price_gbn 분기[라이브]: 책자 book2025_price · 굿즈 tiered_price · 아크릴 vTmpl_price (레드 고유키, 후니는 후니 가격체계로 분기)", h=30)
line("★ 규칙[8조합 역산]: 수량=구간단가 룩업(선형 아님)·페이지=선형가산(~1,115원/p)·색상=인쇄단가 분기. 클라이언트 재계산 금지·서버 권위(순진 곱셈은 2배+ 오차로 반증).", bold=True, h=30)
sec("B-5. 에디쿠스(Edicus) 연동 6단계 [정적+라이브 핸들러]")
line("1 에디터 띄우기: BFF 토큰체인 → editor-bridge.createProject → iframe src={EDICUS_EDITOR_HOST}/ed#/editor_landing?cmd=create&token={JWT}", h=30)
line("2 핸드셰이크: from-edicus(request-prod-info) → 호스트 send-extra-param(prod_info/options). 큰 데이터는 deferred-param(wait_{name}=true)", h=30)
line("3 편집 doc-changed · 4 저장 save-doc-report(start→end)→docInfo(projectID·tnUrlList·페이지수) · 5 썸네일 tnUrlList · 6 주문확정 goto-cart(projectID·tnUrlList·totalPageCount·case)→close", h=34)
line("라이프사이클: CreatingProject→load-project-report(start)→ready-to-listen→doc-changed→request-prod-info→project-id-created→load-project-report(end)─편집─save-doc-report→goto-cart→close", h=34)
line("★ origin 검증을 페이로드 파싱 전 최우선(라이브 origin=edicusbase.firebaseapp.com). 토큰(JWT ~55분·자동갱신)은 위젯 미보관, BFF/어댑터 책임.", bold=True, h=30)
sec("B-6. .env.local EDICUS_* 키 (이름만·값 비노출)")
line("PARTNER_CODE(파트너식별·BFF주입)·API_KEY·API_HOST·EDITOR_HOST(iframe src)·BASE_HOST(origin 허용)·RESOURCE/ASSET_HOST·RENDER_DPI·FIREBASE_*(에디터/저장 백엔드)·MANAGER_URL/ID/PW(템플릿관리)", h=34)
sec("B-7. S3 presigned 업로드 [라이브 end-to-end]")
line("1 POST /api/aws/presigned-url{file_name·pdt_cod·content_type}→{filename(UUID)·presignedURL(60분)} · 2 PUT presignedURL Body=PDF(S3직접) · 3 POST /ko/product/s3GetObjectJson{file_name}→메타(페이지수/크기) · 4 fileUploadInfo[0]내지/[1]표지", h=44)
line("버킷=*.tempo(임시)→주문확정시 영구이동(추정)·검증=PDF·1GB 제한")
sec("B-8. 미검증/리스크 (라이브 보강 필요·비차단)")
line("S3 헤더셋·회원등급 할인(PRICE_MALL≠PRICE)·비책자 ORD_INFO·고급후가공(박/스코딕스) 가격기여·에디터 case 값종류·ACC 가격 페이로드·Vue3 위젯 25상품 전수스키마(대표3군만 라이브). 전부 어댑터/후속캡처로 흡수 가능→'라이브 보강 캡처' 후속 태스크로 분리.", h=44)
sec("B-9. 라이브 DB 실측 (주문→생산 그릇)")
line("★ 라이브 35 t_* = 상품·가격 마스터만. 주문/결제/파일/배송/장바구니 런타임 테이블 0개(전부 to-be).", bold=True)
line("MES 브릿지 = t_prd_products.MES_ITEM_CD 단 1컬럼. 275행 중 16행만 채번(94.2% NULL). 포맷 NNN-NNNN(예 탁상형캘린더=007-0001).", h=30)
line("파일/에디터 플래그 분포(275행): 둘다Y 118 · 업로드만 101 · 에디터만 3 · 둘다N 53. t_cus_customers(거래처) 0행.", h=30)
sec("B-10. 후니 전환 메모 (huni-widget 하네스 결정)")
line("위젯=정규화 계약 의존(NormalizedEditorConfig/Result/Artifact). 레드 데이터로 구현·검증 후 후니 어댑터 교체→위젯 코어 불변(무손실 컨버전). 가격은 항상 서버 권위(위젯 계산 0). 동등성: Red 라이브와 4차원×대표모델 동등 입증·vitest 150 통과.", h=44)
widths(ws, {"A": 25, "B": 25, "C": 25, "D": 25})
ws.freeze_panes = "A4"

# 시트 순서 정렬
order = ["00_읽는법","01_역할정의","02_IA마스터","03_페이즈일정","04_진행현황",
         "05_위젯_에디쿠스","06_주문_생산_MES","07_고객준비물_디자인지원","08_외부연동_계약","09_용어집","10_개발자상세"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save(OUT)
print("ALL SHEETS:", wb.sheetnames)
print("DONE")
