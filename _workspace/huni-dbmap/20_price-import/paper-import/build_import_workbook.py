# -*- coding: utf-8 -*-
"""
round-16 출력소재(IMPORT) webadmin 복붙용 import.xlsx 빌더.
입력 = /tmp/paper_export/*.psv (라이브 read-only 실측 덤프 + 시트 분해 GAP/BLOCKED).
출력 = paper-import-import.xlsx (테이블별 시트, 1행=DB컬럼, 2행=한글라벨).
RU=라이브 기존 적재 재현(재적재 아님). GAP/BLOCKED=신규 후보(인간 승인 전).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

EXPORT = "/tmp/paper_export"
OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/paper-import/paper-import-import.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # BLOCKED 표식
RU_FILL = PatternFill("solid", fgColor="E2EFDA")   # 재현(라이브 기존)
GAP_FILL = PatternFill("solid", fgColor="FFF2CC")  # GAP(채움 후보)


def read_psv(name):
    rows = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    # column widths
    for ci, col in enumerate(columns, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, min(28, len(col) + 4))
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


def num(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return x


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- 시트 0: 안내 ----
    ws0 = wb.create_sheet("0_안내")
    guide = [
        ["출력소재(IMPORT) → Phase11 가격엔진 그릇 (round-16)", ""],
        ["", ""],
        ["이 시트의 목적지", "용지 소재 마스터(t_mat_materials) + 용지비 단가(t_prc_component_prices, comp_cd=COMP_PAPER)"],
        ["가격 그릇 핵심", "COMP_PAPER 용지비 단가뿐. 합판/실사/아크릴은 소재 마스터만(가격은 각 트랙)."],
        ["unit_price 의미", "= 가격표 '가격(국4절)' 열 = 국4절 1장당 용지비(절가). '연당가'는 DB 미저장."],
        ["use_dims", '["siz_cd","mat_cd"] — 2차원만 사용. 나머지 8차원=NULL 와일드카드.'],
        ["prc_typ_cd", "PRICE_TYPE.01 단가형(절가×출력매수). 합가형 없음."],
        ["가격사슬", "COMP_PAPER는 PRF_DGP_A~F 6공식에 배선 완결 — 단가행만 채우면 가격 조회 가능."],
        ["", ""],
        ["시트 색 범례", "초록=RU(라이브 기존, 재적재 금지·대조용) / 노랑=GAP(채움 후보, 인간 승인) / 주황=BLOCKED(자재 선적재 필요)"],
        ["복붙 방법", "각 시트 1행=DB 컬럼명, 2행=한글 라벨. 3행부터 데이터. webadmin에 1행 헤더 기준 복붙."],
        ["DB 적재", "본 산출은 그릇 준비까지. 실제 INSERT/COMMIT은 인간 승인(round-5)."],
    ]
    for ri, (a, b) in enumerate(guide, start=1):
        ca = ws0.cell(row=ri, column=1, value=a)
        ws0.cell(row=ri, column=2, value=b)
        if ri == 1:
            ca.font = Font(bold=True, size=12, color="1F4E79")
        elif b == "" and a:
            ca.font = Font(bold=True, color="1F4E79")
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 95

    # ---- 시트 1: price_components (RU) ----
    pc = read_psv("price_components_ru.psv")
    write_sheet(
        wb, "1_price_components_RU",
        ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
        ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(01단가형/02합가형)", "사용차원(jsonb)", "사용여부"],
        pc,
        note="[RU] 라이브 기존 — 재적재 금지·대조용. COMP_PAPER=용지비(단가형).",
        fill=RU_FILL,
        col_comments={
            "prc_typ_cd": "PRICE_TYPE.01=단가형(절가×출력매수). 용지비는 합가형 아님.",
            "use_dims": '["siz_cd","mat_cd"] — 용지비가 실제 쓰는 차원. 나머지 8차원 NULL.',
        },
    )

    # ---- 시트 2: formula_components (RU) ----
    fc = read_psv("formula_components_ru.psv")
    write_sheet(
        wb, "2_formula_components_RU",
        ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
        ["공식코드", "구성요소코드", "표시순서", "합산여부"],
        fc,
        note="[RU] 라이브 기존 — COMP_PAPER가 PRF_DGP_A~F 6공식에 배선됨(가격사슬 정상). 신규 배선 0.",
        fill=RU_FILL,
    )

    # ---- 시트 3: t_mat_materials (RU 용지 자재) ----
    mat = read_psv("materials_paper.psv")
    write_sheet(
        wb, "3_materials_paper_RU",
        ["mat_cd", "mat_nm", "mat_typ_cd", "weight", "note"],
        ["자재코드", "자재명(종이명)", "자재유형(01=용지)", "평량(g)", "비고"],
        mat,
        note="[RU] 라이브 기존 용지 자재(MAT_TYPE.01) 107행 — COMP_PAPER 단가행의 FK 전제. 재적재 금지.",
        fill=RU_FILL,
        col_comments={"mat_typ_cd": "MAT_TYPE.01=용지. 같은 종이 다른 평량=별 mat_cd(백모조 100/120/220g)."},
    )

    # ---- 시트 4: component_prices RU (COMP_PAPER 49행) ----
    ru = read_psv("comp_paper_ru.psv")  # comp_cd|siz_cd|mat_cd|mat_nm|unit_price
    ru_rows = [[r[0], r[1], None, r[2], None, None, None, None, None, None, num(r[4]), r[3]] for r in ru]
    cp_cols = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt",
               "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price", "(참고)종이명"]
    cp_lbls = ["구성요소", "사이즈(출력판형)", "색상(NULL)", "자재(용지)", "공정(NULL)", "코팅면수(NULL)",
               "옵션(NULL)", "묶음수(NULL)", "최소수량(NULL)", "적용일", "단가(국4절절가)", "종이명(참고)"]
    cp_comments = {
        "siz_cd": "SIZ_000499=316x467 국4절. 3절은 SIZ_000077.",
        "mat_cd": "용지 자재코드. use_dims 차원.",
        "unit_price": "= 가격표 I열 '가격(국4절)' = 국4절 1장당 용지비(절가). H열 연당가 아님.",
        "min_qty": "용지비는 수량구간 없음 → NULL(절가 고정).",
        "proc_cd": "신설 차원. 용지비는 공정 무관 → NULL.",
        "opt_cd": "신설 차원. 용지비는 옵션 무관 → NULL.",
    }
    write_sheet(
        wb, "4_component_prices_RU",
        cp_cols, cp_lbls, ru_rows,
        note="[RU] 라이브 기존 COMP_PAPER 단가 49행 — 재적재 금지. unit_price=가격표 I열(국4절) 전건 일치(검증 완료).",
        fill=RU_FILL, col_comments=cp_comments,
    )

    # ---- 시트 4b: component_prices GAP (14 채움 후보) ----
    gap = read_psv("gap_price.psv")  # comp_cd|siz|mat_cd|nm|price
    gap_rows = [[r[0], r[1], None, r[2], None, None, None, None, None, "(빈칸=현재일)", num(r[4]), r[3]] for r in gap]
    write_sheet(
        wb, "4b_component_prices_GAP",
        cp_cols, cp_lbls, gap_rows,
        note="[GAP·채움후보] 자재(mat_cd) 있으나 COMP_PAPER 단가 없음 15행(F-PAP-2 재검증: 아이보리 MAT_000149만 GAP. 뉴크라프트 MAT_000150·팬시크라프트 MAT_000151은 이미 COMP_PAPER 적재됨[RU]·시트명에 평량생략으로 명칭매칭만 빗나간 것). 배선 불요(가격사슬 즉시 완결). 인간 승인 후 적재. (Q-PAP-1)",
        fill=GAP_FILL, col_comments=cp_comments,
    )

    # ---- 시트 4c: component_prices 3절 GAP ----
    jeol = read_psv("jeol3_gap.psv")  # row|nm|mat_cd|price3
    jeol_rows = [[
        "COMP_PAPER", "SIZ_000077", None, (r[2] if r[2] else "(자재정체 컨펌)"),
        None, None, None, None, None, "(빈칸=현재일)", num(r[3]), r[1]
    ] for r in jeol]
    write_sheet(
        wb, "4c_component_prices_3jeol_GAP",
        cp_cols, cp_lbls, jeol_rows,
        note="[GAP·3절] 3절가(J열) 보유 5행 → SIZ_000077 단가행. 라이브 현재 3절 용지비 0행. mat_cd 정체 컨펌(Q-PAP-3).",
        fill=GAP_FILL, col_comments=cp_comments,
    )

    # ---- 시트 9: BLOCKED material (11 자재 미등록) ----
    blk = read_psv("blocked_material.psv")  # row|nm|중분류|price
    # 스티커용지 7종은 라이브에 MAT_TYPE.11(스티커)로 존재하나 용지(MAT_TYPE.01) COMP_PAPER 매칭 정체 미해소.
    sticker_codes = {
        "유포스티커": "MAT_000153", "무광코팅스티커": "MAT_000155", "유광코팅스티커": "MAT_000156",
        "수분리스티커": "MAT_000161", "투명스티커": "MAT_000162", "홀로그램스티커": "MAT_000163",
        "크라프트 스티커": "MAT_000164",
    }
    blk_rows = []
    for r in blk:
        nm = r[1]
        if nm in sticker_codes:
            blk_rows.append([sticker_codes[nm], nm, "MAT_TYPE.11", None,
                             f"중분류={r[2]}·국4절가={r[3]}·자재존재하나 MAT_TYPE.11(스티커,≠용지)·매칭정체컨펌", r[0]])
        else:  # 백색모조지(평량없음 모호명)
            blk_rows.append(["(정체미확정)", nm, "MAT_TYPE.01?", None,
                             f"중분류={r[2]}·국4절가={r[3]}·평량없음→MAT_000071/072 중 무엇인지 컨펌", r[0]])
    write_sheet(
        wb, "9_BLOCKED_material",
        ["mat_cd", "mat_nm", "mat_typ_cd", "weight", "note", "(참고)시트행"],
        ["자재코드", "자재명(종이명)", "자재유형", "평량", "비고", "시트행"],
        blk_rows,
        note="[BLOCKED] 8행 — 백색모조지(평량없음 모호명) 1 + 스티커용지 7(MAT_TYPE.11 존재하나 용지 COMP_PAPER 매칭 정체). NULL/오타입 단가행 금지. 정체 컨펌(Q-PAP-2). (F-PAP-2: 11→8, 아이보리/뉴크라프트/팬시크라프트 GAP 이동.)",
        fill=BLK_FILL,
    )

    # ---- 시트 9b: SPECIAL Roll 단가 ----
    roll = read_psv("roll_special.psv")  # row|nm|raw
    roll_rows = [[r[1], r[2], r[0], "단가 단위=롤(절가 아님)·단가형/합가형 판별불가·컨펌"] for r in roll]
    write_sheet(
        wb, "9b_SPECIAL_roll",
        ["mat_nm", "단가표기", "(참고)시트행", "처리"],
        ["자재명", "원본단가표기", "시트행", "처리방향"],
        roll_rows,
        note="[SPECIAL] Roll 단위 단가 2행 — 절가/장당가 환산 불명. 단가형/합가형 판별불가(Q-PAP-4).",
        fill=BLK_FILL,
    )

    wb.save(OUT)
    print("saved:", OUT)
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
