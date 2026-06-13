# -*- coding: utf-8 -*-
"""round-16 커팅타공 webadmin 복붙용 import.xlsx 빌더.

입력 = docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx > 커팅타공 (read-only).
출력 = cutting-drilling-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[핵심] 커팅타공은 라이브 적재 완료 + 디지털인쇄 공식(PRF_DGP_B~F)에 배선됨 → 신규 적재 0.
산출 가치 = 시트↔라이브 round-trip 무손실 + B3 min_qty stale 교정 + prc_typ_cd 컨펌.

블록:
  B1 커팅(완칼) A1:B38  → COMP_CUT_FULL_DIECUT 36행(siz=SIZ_000499 국4절·단가형.01 정당)
  B2 타공(단가) A42:C66 → COMP_CUT_PERF_1H6 23행(전부 0원·미사용 placeholder)
  B3 타공(합가) F42:H54 → COMP_CUT_FULL_PERF_1H6/_2H6 각 9행(합가형이 맞으나 라이브 .01 오등록)
                          🔴 가격사슬 단절(배선 0) + min_qty 라이브 stale(시트 권위)
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

SRC = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "cutting-drilling/cutting-drilling-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
LIVE_FILL = PatternFill("solid", fgColor="E2EFDA")   # 라이브 정합(초록)
FIX_FILL = PatternFill("solid", fgColor="FCE4D6")    # 교정/차단(주황)
REF_FILL = PatternFill("solid", fgColor="FFF2CC")    # 노트/참조(노랑)

APPLY_YMD = "2026-06-01"
SIZ_GUK4 = "SIZ_000499"  # 국4절 316x467 (라이브 실측)

# ---- 가격표 읽기 ----
sb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws = sb["커팅타공"]


def cell(r, c):
    return ws.cell(row=r, column=c).value


# B1 커팅(완칼): A3..A38 수량, B 가격
b1 = [(int(cell(r, 1)), float(cell(r, 2))) for r in range(3, 39) if cell(r, 1) is not None]
# B2 타공(단가): A44..A66 제작수량, B(1구)=전부 0
b2 = [(int(cell(r, 1)), (float(cell(r, 2)) if cell(r, 2) is not None else 0.0))
      for r in range(44, 67) if cell(r, 1) is not None]
# B3 타공(합가): F44..F52 제작수량, G(1구)/H(2구)
b3_1 = [(int(cell(r, 6)), float(cell(r, 7))) for r in range(44, 53) if cell(r, 6) is not None]
b3_2 = [(int(cell(r, 6)), float(cell(r, 8))) for r in range(44, 53) if cell(r, 6) is not None]

# 라벨/노트
note_c1 = cell(1, 3)   # 인쇄배경지
note_c2 = cell(2, 3)   # 인쇄비+소재+커팅
note_a39 = cell(39, 1)  # 1장당 2000원씩
note_d42 = cell(42, 4)  # 헤더택 / 벽걸이캘린더
note_f53 = cell(53, 6)
note_f54 = cell(54, 6)
sb.close()


# ---- 시트 작성 헬퍼 ----
def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    wsx = wb.create_sheet(title)
    if note:
        wsx.cell(row=1, column=1, value=note)
        wsx.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = wsx.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = wsx.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        wsx.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(34, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            wsx.cell(row=ri, column=ci, value=val)
    wsx.freeze_panes = wsx.cell(row=data_r0, column=1)
    return wsx


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
wsr = wb.create_sheet("0_README")
readme = [
    ["round-16 커팅타공 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 커팅타공"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "합산형 후가공 구성요소 (디지털인쇄 PRF_DGP_B~F에 배선) — 독립 공식 아님"],
    ["", ""],
    ["[핵심] 신규 적재 0 — 라이브에 이미 적재됨 + 배선됨", ""],
    ["  · B1 커팅(완칼) = COMP_CUT_FULL_DIECUT 36행 (siz=SIZ_000499 국4절·단가형.01 정당)", ""],
    ["  · B2 타공(단가) = COMP_CUT_PERF_1H6 23행 (전부 0원·미사용 placeholder)", ""],
    ["  · B3 타공(합가) = COMP_CUT_FULL_PERF_1H6/_2H6 각 9행", ""],
    ["", ""],
    ["[신규 발견 3건 — 교정 대상]", ""],
    ["  🔴 가격사슬 단절: 타공합가 comp 2개가 어느 공식에도 미배선 → 엔진 조회 불가", ""],
    ["  🔴 prc_typ_cd 오등록: 헤더 '타공(합가)'인데 라이브 .01 단가형 (합가형.02가 맞음)", ""],
    ["  🔴 min_qty stale: B3 라이브 구간(200/300/400…)이 260527 시트(300/500/1000…)와 다름", ""],
    ["", ""],
    ["[색상 범례]", ""],
    ["  초록(_LIVE) = 라이브 정합 (이미 적재됨·재사용)", ""],
    ["  주황(_FIX/_BLOCKED) = 교정/배선 제안 (prc_typ·min_qty·가격사슬·인간 승인)", ""],
    ["  노랑(_REF) = 노트/증분룰 보존 (가격 그릇 외·무손실)", ""],
    ["", ""],
    ["[단가/합가 판별 — 가격표 헤더가 직접 표기한 드문 케이스]", ""],
    ["  B1 완칼: A39 '1장당 2000원' = 장당가 → 단가형.01 정당", ""],
    ["  B3 타공: 헤더 '합가' + 셀=수량구간 총액 → 합가형.02가 맞음 (라이브 .01 충돌→컨펌-A)", ""],
    ["", ""],
    ["[미해소 컨펌]", ""],
    ["  컨펌-A: B3 prc_typ_cd .01→.02 교정 여부 (헤더 '합가' vs 후니 .01 관례)", ""],
    ["  컨펌-B: 타공합가가 붙는 상품(헤더택/벽걸이캘린더) 공식 배선 여부", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_LIVE", "신규 0 — PRF_DGP_B~F 참조 맵 (커팅/타공 배선 위치)"],
    ["  2_formula_components_LIVE_FIX", "배선 5(완칼·타공0원) + 타공합가 배선 제안 2(현재 단절)"],
    ["  3_price_components_LIVE_FIX", "구성요소 4 (prc_typ_cd 교정 제안 열 포함)"],
    ["  4_component_prices_LIVE_FIX", "단가행 77 (완칼36+타공단가23+타공합가18·min_qty 시트권위)"],
    ["  N1_increment_rules_REF", "증분룰 3 (완칼·1구·2구·이산구간 외 외삽)"],
    ["", ""],
    ["분해 무손실 검산", "가격셀 77 (완칼36 + 타공단가23 + 타공합가 1구9 + 2구9) = 단가행 77 일치"],
]
for ri, (a, b) in enumerate(readme, start=1):
    wsr.cell(row=ri, column=1, value=a)
    wsr.cell(row=ri, column=2, value=b)
    if ri == 1:
        wsr.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
wsr.column_dimensions["A"].width = 56
wsr.column_dimensions["B"].width = 64

# ============================================================
# 1. price_formulas (LIVE 참조 맵) — 신규 0
# ============================================================
write_sheet(
    wb,
    "1_price_formulas_LIVE",
    ["frm_cd", "comp_cd(커팅/타공)", "disp_seq", "비고"],
    ["공식코드(라이브)", "커팅/타공 구성요소", "배선순서", "비고"],
    [
        ["PRF_DGP_B", "COMP_CUT_FULL_DIECUT", 4, "모양엽서·라벨택 — 완칼"],
        ["PRF_DGP_F", "COMP_CUT_FULL_DIECUT", 4, "썬캡 — 완칼"],
        ["PRF_DGP_C", "COMP_CUT_PERF_1H6", 5, "인쇄배경지·헤더택 — 타공(0원)"],
        ["PRF_DGP_D", "COMP_CUT_PERF_1H6", 6, "소량전단지 — 타공(0원)"],
        ["PRF_DGP_E", "COMP_CUT_PERF_1H6", 10, "접지카드·접지리플렛 — 타공(0원)"],
    ],
    note="[LIVE·신규 0] 커팅/타공은 독립 공식 없음. 디지털인쇄 합산형 공식 PRF_DGP_B~F의 후가공 구성요소로 참여(라이브 실재). 라이브 t_prc_price_formulas 컬럼=frm_cd·frm_nm·note·use_yn(frm_typ_cd/prd_cd 부존재).",
    fill=LIVE_FILL,
    col_comments={
        "frm_cd": "디지털인쇄 합산형 공식. 커팅/타공이 구성요소로 합산됨",
        "comp_cd(커팅/타공)": "이 공식에 배선된 커팅/타공 comp. 타공합가(PERF_1H6/2H6)는 미배선=단절(시트2 참조)",
    },
)

# ============================================================
# 2. formula_components (LIVE + 타공합가 배선 제안)
# ============================================================
write_sheet(
    wb,
    "2_formula_components_FIX",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn", "상태", "비고"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부", "상태", "비고"],
    [
        ["PRF_DGP_B", "COMP_CUT_FULL_DIECUT", 4, "Y", "라이브 실재", "완칼·정합"],
        ["PRF_DGP_F", "COMP_CUT_FULL_DIECUT", 4, "Y", "라이브 실재", "완칼·정합"],
        ["PRF_DGP_C", "COMP_CUT_PERF_1H6", 5, "Y", "라이브 실재", "타공0원·정합"],
        ["PRF_DGP_D", "COMP_CUT_PERF_1H6", 6, "Y", "라이브 실재", "타공0원·정합"],
        ["PRF_DGP_E", "COMP_CUT_PERF_1H6", 10, "Y", "라이브 실재", "타공0원·정합"],
        ["?? (헤더택/벽걸이캘린더)", "COMP_CUT_FULL_PERF_1H6", "??", "Y", "🔴단절·제안", "[컨펌-B] 타공합가 1구 미배선→배선 필요"],
        ["?? (헤더택/벽걸이캘린더)", "COMP_CUT_FULL_PERF_2H6", "??", "Y", "🔴단절·제안", "[컨펌-B] 타공합가 2구 미배선→배선 필요"],
    ],
    note="[LIVE+FIX] 완칼·타공0원 배선 5건은 라이브 실재. 🔴 타공합가(PERF_1H6/2H6) 단가행 18개는 적재됐으나 어느 공식에도 배선 0=가격사슬 단절(엔진 조회 불가). 배선 대상 상품(헤더택/벽걸이캘린더) 컨펌-B.",
    fill=FIX_FILL,
    col_comments={
        "상태": "라이브 실재 = 이미 배선됨. 🔴단절 = 단가행 적재됐으나 미배선(엔진 미조회)",
        "comp_cd": "타공합가 1구/2구는 별 comp(가격 2배 차등). opt_cd 아님(라이브 패턴·과분할 방지)",
    },
)

# ============================================================
# 3. price_components (LIVE + prc_typ_cd 교정 제안)
# ============================================================
write_sheet(
    wb,
    "3_price_components_FIX",
    ["comp_cd", "comp_nm", "prc_typ_cd(라이브)", "prc_typ_cd(제안)", "use_dims", "use_yn", "비고"],
    ["구성요소코드", "구성요소명", "단가유형(현재)", "단가유형(교정제안)", "사용차원(jsonb)", "사용여부", "비고(P4)"],
    [
        ["COMP_CUT_FULL_DIECUT", "커팅 합가(완제품가) [완칼]", "PRICE_TYPE.01", "PRICE_TYPE.01", '["siz_cd","min_qty"]', "Y",
         "단가형 정당 — A39 '1장당 2000원'=장당가"],
        ["COMP_CUT_PERF_1H6", "타공비(후가공) [단가0원]", "PRICE_TYPE.01", "PRICE_TYPE.01", '["min_qty"]', "Y",
         "단가형 정당 — 전부 0원·미사용 경로"],
        ["COMP_CUT_FULL_PERF_1H6", "커팅 합가(완제품가) [타공합가 1구]", "PRICE_TYPE.01", "🔴 PRICE_TYPE.02", '["min_qty"]', "Y",
         "[컨펌-A] 헤더 '타공(합가)'+셀=수량구간 총액 → 합가형.02가 맞음. 라이브 .01 충돌"],
        ["COMP_CUT_FULL_PERF_2H6", "커팅 합가(완제품가) [타공합가 2구]", "PRICE_TYPE.01", "🔴 PRICE_TYPE.02", '["min_qty"]', "Y",
         "[컨펌-A] 동일 — 합가형.02 교정 제안"],
    ],
    note="[LIVE+FIX] 구성요소 4. 완칼·타공0원=단가형.01 정당. 🔴 타공합가 1구/2구=헤더 '합가'+셀 총액이라 합가형.02가 맞으나 라이브 .01 오등록(foil/명함박 횡단 패턴). 컨펌-A 후 교정.",
    fill=FIX_FILL,
    col_comments={
        "prc_typ_cd(라이브)": "라이브 실측값. 4개 전부 .01(단가형)로 등록됨",
        "prc_typ_cd(제안)": ".01 단가형=장당가×수량 / .02 합가형=구간총액÷min_qty 환산. 타공합가는 .02가 맞음",
        "comp_nm": "라이브 comp_nm. 완칼/타공합가 모두 '커팅 합가(완제품가)' 명칭이나 거동은 별개",
    },
)

# ============================================================
# 4. component_prices (LIVE + min_qty 시트권위) — 77행
# ============================================================
CP_COLS = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt",
           "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price", "note"]
CP_LBL = ["구성요소", "사이즈(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수",
          "옵션(opt)", "묶음수(bdl)", "수량구간(min_qty)", "적용일", "단가/총액", "비고"]
cp_rows = []
# B1 커팅(완칼) 36행
for qty, price in b1:
    cp_rows.append(["COMP_CUT_FULL_DIECUT", SIZ_GUK4, None, None, None, None,
                    None, None, qty, APPLY_YMD, price, None])
# B2 타공(단가·0원) 23행
for qty, price in b2:
    cp_rows.append(["COMP_CUT_PERF_1H6", None, None, None, None, None,
                    None, None, qty, APPLY_YMD, price, "미사용 0원" if qty == 1 else None])
# B3 타공(합가) 1구 9행 — min_qty=시트 권위
for qty, price in b3_1:
    cp_rows.append(["COMP_CUT_FULL_PERF_1H6", None, None, None, None, None,
                    None, None, qty, APPLY_YMD, price,
                    "🔴 합가형(총액). min_qty=260527 시트권위(라이브 stale)" if qty == 300 else None])
# B3 타공(합가) 2구 9행
for qty, price in b3_2:
    cp_rows.append(["COMP_CUT_FULL_PERF_2H6", None, None, None, None, None,
                    None, None, qty, APPLY_YMD, price, None])

write_sheet(
    wb,
    "4_component_prices_FIX",
    CP_COLS, CP_LBL, cp_rows,
    note="[LIVE+FIX] 단가행 77 (완칼36 + 타공단가0원23 + 타공합가 1구9 + 2구9). 완칼 siz_cd=SIZ_000499(국4절). 🔴 타공합가 min_qty=260527 시트 권위(300/500/1000…) — 라이브(200/300/400…) stale 교정. 안 쓰는 차원 NULL.",
    fill=FIX_FILL,
    col_comments={
        "siz_cd": "완칼=SIZ_000499(국4절 316x467·출력판형). 타공=NULL(사이즈 무관)",
        "min_qty": "수량구간 시작. 타공합가는 260527 시트 권위(라이브 stale). 주문수량 이하 최대 min_qty 매칭",
        "unit_price": "완칼=장당가×수량(단가형). 타공합가=수량구간 총액(합가형→÷min_qty 환산). 타공단가=0원",
        "opt_cd": "전부 NULL — 1구/2구는 comp 분리(opt_cd 아님·라이브 패턴 준수)",
        "proc_cd": "전부 NULL — 완칼(PROC_000053)/타공(PROC_000079) 공정 있으나 가격 차원 미사용",
    },
)

# ============================================================
# N1. 증분룰 (REF·노트 보존)
# ============================================================
write_sheet(
    wb,
    "N1_increment_rules_REF",
    ["대상", "셀", "증분룰", "처리"],
    ["대상 블록", "원본셀", "증분룰(원문)", "처리(이산구간 외)"],
    [
        ["B1 커팅(완칼)", "A39", note_a39, "수량×2000 선형 외삽(최대 1000장 초과)"],
        ["B3 타공(합가) 1구", "F53", note_f53, "100장당 +1000 외삽(최대 5000장 초과)"],
        ["B3 타공(합가) 2구", "F54", note_f54, "100장당 +2000 외삽(최대 5000장 초과)"],
        ["B1 라벨", "C1", note_c1, "완칼 적용 상품 맥락(인쇄배경지)"],
        ["B1 라벨", "C2", note_c2, "완제품가=인쇄비+소재+커팅 합산 설명"],
        ["B2/B3 라벨", "D42", note_d42, "타공 적용 상품(헤더택/벽걸이캘린더)·컨펌-B 배선 대상"],
    ],
    note="[REF·무손실] 증분룰 3 + 라벨노트 3. 가격 그릇(t_prc_*) 행이 아님 — 이산구간 외 수량은 앱이 외삽(off-grid ceiling 철학). 침묵 삭제 금지.",
    fill=REF_FILL,
    col_comments={
        "증분룰": "이산 구간 단가행 외 별도 명시. 최대 구간 초과 주문 외삽 룰",
    },
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("B1 커팅 rows:", len(b1), "(expect 36)")
print("B2 타공단가 rows:", len(b2), "(expect 23)")
print("B3 타공합가 1구 rows:", len(b3_1), "2구 rows:", len(b3_2), "(expect 9/9)")
print("component_prices rows:", len(cp_rows), "(expect 77)")
