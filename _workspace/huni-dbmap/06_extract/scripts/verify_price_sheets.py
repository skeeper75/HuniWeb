#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
가격표 16 단가시트 L1 자동대조 — 무손실 9게이트 (누락0 기계보증).

원본 권위. 추출본 = price-<slug>-l1.csv (셀단위 long) + price-<slug>-l1-meta.csv.
extract_price_sheets.py 가 nonempty 셀을 1셀1레코드로 떴으므로, 대조는 셀좌표 기준 1:1.

게이트:
  ① cell_count        : 원본 nonempty 셀 수 == 추출 value 레코드 수
  ② nonempty_preserve : 원본 nonempty 셀 전부 추출 보존(missing 0·mismatch 0) = 100%
  ③ roundtrip         : 추출본으로 원본 재구성 → diff 0 (셀좌표·값 완전 일치)
  ④ col_coverage      : 원본 nonempty 컬럼 전부 추출에 등장
  ⑤ row_coverage      : 원본 nonempty 행 전부 추출에 등장
  ⑥ merge_preserve    : 원본 머지 수 == 메타 merged_range 고유 수
  ⑦ formula_preserve  : 원본 수식 셀 수 == 메타 is_formula 수
  ⑧ comment_preserve  : 원본 스레드댓글 수 == 메타 comment_text 셀 수
  ⑨ hidden_preserve   : 원본 숨김열/행 == 메타 col_hidden/row_hidden 표기

read-only.
"""
import os
import sys
import csv
import json

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = "/Users/innojini/Dev/HuniWeb"
XLSX = os.path.join(ROOT, "docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx")
OUT = os.path.join(ROOT, "_workspace/huni-dbmap/06_extract")

sys.path.insert(0, os.path.dirname(__file__))
from extract_l1 import load_threaded_comments  # noqa: E402

SHEET_SLUGS = {
    "디지털인쇄비": "digital-print-price",
    "코팅": "coating",
    "접지옵션": "folding",
    "인쇄후가공": "post-process",
    "커팅타공": "cutting",
    "스티커": "sticker-price",
    "합판도무송스티커": "gangpan-sticker",
    "봉투제작": "envelope",
    "명함포토카드": "namecard-photocard",
    "후가공_박(소형)": "foil-small",
    "엽서북떡메": "postcard-book",
    "제본": "binding",
    "후가공_박(대형)": "foil-large",
    "아크릴": "acrylic-price",
    "포스터사인": "poster-sign",
}


def norm(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def col_letter_to_idx(letter):
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def verify(sheet, slug):
    wb_d = openpyxl.load_workbook(XLSX, data_only=True)
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    ws_d, ws_f = wb_d[sheet], wb_f[sheet]
    comments = load_threaded_comments(XLSX, sheet)

    # 데이터 경계 (extract 와 동일 로직)
    last_r = last_c = 0
    for r in range(1, ws_d.max_row + 1):
        for c in range(1, ws_d.max_column + 1):
            if ws_d.cell(r, c).value not in (None, ""):
                last_r = max(last_r, r)
                last_c = max(last_c, c)
    maxr, maxc = last_r, last_c

    # 원본 nonempty 셀 좌표
    orig = {}
    orig_cols = set()
    orig_rows = set()
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = ws_d.cell(r, c).value
            if v not in (None, ""):
                orig[(r, c)] = norm(v)
                orig_cols.add(c)
                orig_rows.add(r)

    # 추출본 로드
    l1 = os.path.join(OUT, f"price-{slug}-l1.csv")
    with open(l1, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    ext = {}
    ext_cols = set()
    ext_rows = set()
    dup = 0
    for er in rows:
        r = int(er["row_seq"])
        c = col_letter_to_idx(er["col"])
        key = (r, c)
        if key in ext:
            dup += 1
        ext[key] = er["value"]
        ext_cols.add(c)
        ext_rows.add(r)

    # ① cell_count
    g1 = {"orig_nonempty": len(orig), "ext_records": len(rows), "dup_cells": dup,
          "pass": len(orig) == len(rows) and dup == 0}

    # ② nonempty preserve + ③ roundtrip (셀좌표 1:1)
    missing = mismatch = preserved = 0
    miss_s, mis_s = [], []
    for (r, c), ov in orig.items():
        if (r, c) not in ext:
            missing += 1
            if len(miss_s) < 10:
                miss_s.append({"cell": get_column_letter(c) + str(r), "orig": ov})
        elif ext[(r, c)] == ov:
            preserved += 1
        else:
            mismatch += 1
            if len(mis_s) < 10:
                mis_s.append({"cell": get_column_letter(c) + str(r),
                              "orig": ov, "ext": ext[(r, c)]})
    g2 = {"orig": len(orig), "preserved": preserved, "missing": missing,
          "mismatch": mismatch,
          "rate_pct": round(preserved / len(orig) * 100, 4) if orig else 100.0,
          "missing_samples": miss_s, "mismatch_samples": mis_s,
          "pass": missing == 0 and mismatch == 0}
    # ③ roundtrip — 추출 셀이 원본과 어긋나는지(추출→원본 방향)
    rt_diff = 0
    rt_s = []
    for (r, c), ev in ext.items():
        ov = norm(ws_d.cell(r, c).value)
        if ev != ov:
            rt_diff += 1
            if len(rt_s) < 10:
                rt_s.append({"cell": get_column_letter(c) + str(r),
                             "orig": ov, "ext": ev})
    g3 = {"diff": rt_diff, "samples": rt_s, "pass": rt_diff == 0}

    # ④ col coverage ⑤ row coverage
    g4 = {"orig_cols": len(orig_cols), "ext_cols": len(ext_cols),
          "missing_cols": sorted(get_column_letter(c) for c in (orig_cols - ext_cols)),
          "pass": orig_cols.issubset(ext_cols)}
    g5 = {"orig_rows": len(orig_rows), "ext_rows": len(ext_rows),
          "missing_rows": sorted(orig_rows - ext_rows),
          "pass": orig_rows.issubset(ext_rows)}

    # 메타 로드
    meta = os.path.join(OUT, f"price-{slug}-l1-meta.csv")
    with open(meta, encoding="utf-8-sig") as f:
        mrows = list(csv.DictReader(f))
    ext_merges = set()
    ext_formula = 0
    ext_comments = set()
    ext_hidcols = set()
    ext_hidrows = set()
    for mr in mrows:
        if mr["type"] != "cell":
            if mr["type"] == "hidden_column":
                ext_hidcols.add(mr["col"])
            continue
        try:
            mj = json.loads(mr["meta_json"]) if mr["meta_json"] else {}
        except Exception:
            mj = {}
        if mj.get("merged_range"):
            ext_merges.add(mj["merged_range"])
        if mj.get("is_formula"):
            ext_formula += 1
        if mj.get("comment_text"):
            ext_comments.add(mr["ref"])
        if mj.get("col_hidden"):
            ext_hidcols.add(mr["col"])
        if mj.get("row_hidden"):
            ext_hidrows.add(int(mr["row_seq"]))

    # ⑥ merge
    orig_merges = set(str(m) for m in ws_d.merged_cells.ranges
                      if m.min_row <= maxr and m.min_col <= maxc)
    g6 = {"orig_merges": len(orig_merges), "ext_merges": len(ext_merges),
          "pass": orig_merges.issubset(ext_merges)}

    # ⑦ formula (데이터영역 내)
    orig_formula = 0
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            fv = ws_f.cell(r, c).value
            if isinstance(fv, str) and fv.startswith("="):
                orig_formula += 1
    g7 = {"orig": orig_formula, "ext": ext_formula, "pass": orig_formula == ext_formula}

    # ⑧ comment (데이터영역 내 ref만)
    orig_cmt = set()
    for ref in comments:
        m = __import__("re").match(r"([A-Z]+)(\d+)", ref)
        if m:
            cc = col_letter_to_idx(m.group(1))
            rr = int(m.group(2))
            if rr <= maxr and cc <= maxc:
                orig_cmt.add(ref)
    g8 = {"orig": len(orig_cmt), "ext": len(ext_comments),
          "pass": orig_cmt.issubset(ext_comments)}

    # ⑨ hidden col/row
    orig_hidcols = set(get_column_letter(c) for c in range(1, maxc + 1)
                       if (ws_d.column_dimensions.get(get_column_letter(c))
                           and ws_d.column_dimensions[get_column_letter(c)].hidden))
    orig_hidrows = set(r for r in range(1, maxr + 1)
                       if (ws_d.row_dimensions.get(r) and ws_d.row_dimensions[r].hidden))
    # 숨김행은 nonempty 행만 메타에 등장(셀 메타 기반) → orig 숨김행도 nonempty 한정
    orig_hidrows_data = set(r for r in orig_hidrows if r in orig_rows)
    g9 = {"orig_hidden_cols": sorted(orig_hidcols),
          "ext_hidden_cols": sorted(ext_hidcols),
          "orig_hidden_rows_data": sorted(orig_hidrows_data),
          "ext_hidden_rows": sorted(ext_hidrows),
          "pass": orig_hidcols.issubset(ext_hidcols)
                  and orig_hidrows_data.issubset(ext_hidrows)}

    gates = {
        "cell_count": g1, "nonempty_preserve": g2, "roundtrip": g3,
        "col_coverage": g4, "row_coverage": g5, "merge_preserve": g6,
        "formula_preserve": g7, "comment_preserve": g8, "hidden_preserve": g9,
    }
    overall = all(g["pass"] for g in gates.values())
    return overall, gates


def main():
    allout = {}
    npass = 0
    print(f"{'slug':<22} {'cell':>5} {'nonempty':>9} {'rt':>4} {'col':>4} {'row':>4} "
          f"{'mrg':>4} {'fml':>4} {'cmt':>4} {'hid':>4}  OVERALL")
    for sheet, slug in SHEET_SLUGS.items():
        ok, gates = verify(sheet, slug)
        allout[sheet] = {"slug": slug, "overall": ok, "gates": gates}
        if ok:
            npass += 1

        def mk(g):
            return "P" if g["pass"] else "F"
        print(f"{slug:<22} {mk(gates['cell_count']):>5} "
              f"{mk(gates['nonempty_preserve']):>9} {mk(gates['roundtrip']):>4} "
              f"{mk(gates['col_coverage']):>4} {mk(gates['row_coverage']):>4} "
              f"{mk(gates['merge_preserve']):>4} {mk(gates['formula_preserve']):>4} "
              f"{mk(gates['comment_preserve']):>4} {mk(gates['hidden_preserve']):>4}  "
              f"{'PASS' if ok else 'FAIL'}")
    with open(os.path.join(OUT, "_price-sheets-extract-summary.json"), "w",
              encoding="utf-8") as f:
        json.dump({"pass_count": npass, "total": len(SHEET_SLUGS),
                   "sheets": allout}, f, ensure_ascii=False, indent=2)
    print(f"\n=== {npass}/{len(SHEET_SLUGS)} sheets PASS all 9 gates ===")
    sys.exit(0 if npass == len(SHEET_SLUGS) else 1)


if __name__ == "__main__":
    main()
