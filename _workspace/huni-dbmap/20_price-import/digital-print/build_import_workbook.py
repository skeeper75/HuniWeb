# -*- coding: utf-8 -*-
"""
round-16 디지털인쇄비 webadmin 복붙용 import.xlsx 빌더.
입력 = /tmp/dgp_export/*.psv (라이브 read-only 실측 덤프) + round-2 BLOCKED CSV.
출력 = digital-print-import.xlsx (테이블별 시트, 1행=DB컬럼, 2행=한글라벨).
모든 데이터는 라이브 기존 적재 재현(재적재 아님) — 신규 0.
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

EXPORT = "/tmp/dgp_export"
R2 = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/02_mapping/digital-print-engine"
OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/digital-print/digital-print-import.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # BLOCKED 시트 표식
RU_FILL  = PatternFill("solid", fgColor="E2EFDA")  # 재현(라이브 기존) 표식


def read_psv(name):
    rows = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def read_csv(path):
    with open(path, encoding="utf-8") as f:
        return list(csv.reader(f))


def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    """columns=영문 DB컬럼, labels=한글라벨, data_rows=list of list(컬럼순)."""
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, min(26, len(lbl) + 4))
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


def empty_to_none(v):
    return None if v == "" else v


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README 시트
# ============================================================
ws = wb.create_sheet("0_README")
readme = [
    ["round-16 디지털인쇄비 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 디지털인쇄비 (index 2)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["공식유형", "합산형 — 판매가 = Σ components (PRF_DGP_A~F, FRM_TYPE.01)"],
    ["", ""],
    ["[중요] 이 시트(디지털인쇄비)는 상품 1개 공식이 아니라 '인쇄비 구성요소'의 단가 원천", ""],
    ["  · 가격표 도수 7종 중 흑백/CMYK = COMP_PRINT_DIGITAL_S1/S2 (clr_cd 차원)", ""],
    ["  · 별색 5종(화이트/클리어/핑크/금색/은색) = COMP_PRINT_SPOT_* (구성요소 분리, clr_cd=NULL)", ""],
    ["  · 이 인쇄비는 PRF_DGP_A~F 합산형 공식의 한 부품 (+코팅+용지+후가공 등과 합산)", ""],
    ["", ""],
    ["[경고] 라이브 기존 적재 — 재적재 금지", "round-2가 이미 라이브 COMMIT (308행). 이 그릇은 재현(대조용)."],
    ["  · 초록 시트(_RU) = 라이브 기존 재현 — 신규 적재 아님", ""],
    ["  · 주황 시트(_BLOCKED) = 미적재 (siz 미채번 등) — 인간 승인 후", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_RU", "공식 정의 6 (라이브 재현)"],
    ["  1b_product_price_formulas_RU", "상품↔공식 바인딩 19 (라이브 재현)"],
    ["  2_formula_components_RU", "공식 배선 72 (라이브 재현)"],
    ["  3_price_components_print_RU", "인쇄비 구성요소 정의 12 (라이브 재현)"],
    ["  4_component_prices_print_RU", "인쇄비 단가행 954 (이 시트 원천·라이브 재현)"],
    ["  4b_component_prices_print_GAP", "3절 별색 미적재 갭 (가격표에도 없음·정직표기)"],
    ["  9_BLOCKED_binding", "BLOCKED 바인딩 3 (siz 미채번)"],
    ["", ""],
    ["분해 무손실 검산", "가격표 데이터셀 954 (국4절 53×14 + 3절 53×4) = 라이브 954행 일치"],
]
for ri, (a, b) in enumerate(readme, start=1):
    ws.cell(row=ri, column=1, value=a)
    ws.cell(row=ri, column=2, value=b)
    if ri == 1:
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws.column_dimensions["A"].width = 48
ws.column_dimensions["B"].width = 70

# ============================================================
# 1. price_formulas (라이브 재현)
# ============================================================
frm = read_psv("formulas.psv")
write_sheet(
    wb, "1_price_formulas_RU",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in frm],
    note="[RU] 라이브 t_prc_price_formulas 재현 — frm_typ_cd 컬럼 없음(라이브 실측). 신규 적재 아님.",
    fill=RU_FILL,
    col_comments={
        "frm_cd": "공식 식별자 PK. PRF_DGP_A~F = 디지털인쇄 합산형 6공식",
        "use_yn": "Y=노출 / N=미출시(F 썬캡)",
    },
)

# ============================================================
# 1b. product_price_formulas (바인딩, 라이브 재현)
# ============================================================
bind = read_psv("binding.psv")
write_sheet(
    wb, "1b_product_price_formulas_RU",
    ["prd_cd", "frm_cd", "apply_bgn_ymd"],
    ["상품코드", "공식코드", "적용시작일(yyyy-MM-dd)"],
    [[r[0], r[1], r[2]] for r in bind],
    note="[RU] 라이브 t_prd_product_price_formulas 재현 — 별 테이블(공식정의와 분리). 19상품 바인딩.",
    fill=RU_FILL,
    col_comments={
        "prd_cd": "상품 식별자 (t_prd_products FK)",
        "apply_bgn_ymd": "PK 구성요소. 적용 시작일",
    },
)

# ============================================================
# 2. formula_components (배선, 라이브 재현)
# ============================================================
wire = read_psv("wiring.psv")
write_sheet(
    wb, "2_formula_components_RU",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    [[r[0], r[1], int(r[2]) if r[2] else None, r[3]] for r in wire],
    note="[RU] 라이브 t_prc_formula_components 재현 — 합산형 배선 72. addtn_yn은 Phase11 무시(라이브 값 보존).",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "이 공식이 합산하는 구성요소 (t_prc_price_components FK)",
        "addtn_yn": "라이브 전건 Y. Phase11 엔진은 이 플래그 무시(런타임 옵션 활성화로 합산 결정)",
    },
)

# ============================================================
# 3. price_components (인쇄비 구성요소 정의, 라이브 재현)
# ============================================================
comp = read_psv("components_print.psv")
write_sheet(
    wb, "3_price_components_print_RU",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    [[r[0], r[1], r[2], r[3], r[4] if len(r) > 4 else "", r[5] if len(r) > 5 else "Y"] for r in comp],
    note="[RU] 라이브 t_prc_price_components 재현 — 디지털인쇄비 시트가 단가를 채우는 인쇄비 구성요소 12종. 전건 PRICE_TYPE.01 단가형.",
    fill=RU_FILL,
    col_comments={
        "prc_typ_cd": "PRICE_TYPE.01=단가형(장당가×수량). 인쇄비는 전건 단가형 — 수량구간별 장당가 차등(min_qty)",
        "use_dims": "이 구성요소가 실제 쓰는 차원. DIGITAL=[siz,clr,min_qty] / SPOT=[siz,min_qty](clr 무관)",
        "comp_typ_cd": "PRC_COMPONENT_TYPE.01 = 인쇄비",
    },
)

# ============================================================
# 4. component_prices (인쇄비 단가행 954, 이 시트의 원천·라이브 재현)
# ============================================================
cp = read_psv("component_prices_print.psv")
# columns: comp_cd|siz_cd|clr_cd|mat_cd|proc_cd|coat_side_cnt|opt_cd|bdl_qty|min_qty|apply_ymd|unit_price
cp_rows = []
for r in cp:
    cp_rows.append([
        r[0],                       # comp_cd
        empty_to_none(r[1]),        # siz_cd
        empty_to_none(r[2]),        # clr_cd
        empty_to_none(r[3]),        # mat_cd
        empty_to_none(r[4]),        # proc_cd
        int(r[5]) if r[5] else None,  # coat_side_cnt
        empty_to_none(r[6]),        # opt_cd
        int(r[7]) if r[7] else None,  # bdl_qty
        int(r[8]) if r[8] else None,  # min_qty
        r[9],                       # apply_ymd
        float(r[10]) if r[10] else None,  # unit_price
    ])
write_sheet(
    wb, "4_component_prices_print_RU",
    ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "출력판형(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수", "수량구간시작(min_qty)", "적용일", "단가"],
    cp_rows,
    note="[RU] 라이브 t_prc_component_prices 재현(954행) — 디지털인쇄비 시트 = 이 단가의 원천. 10차원 자연키. 안 쓰는 차원 NULL(빈칸).",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "DIGITAL_S1/S2(흑백·CMYK) + SPOT_*S1/S2(별색5종). _S1=단면 _S2=양면",
        "siz_cd": "SIZ_000499=국4절(316x467) / SIZ_000077=3절(300x625)",
        "clr_cd": "CLR_000002=흑백1도 / CLR_000005=CMYK4도. 별색(SPOT)은 NULL(도수 무관)",
        "min_qty": "수량구간 시작값(상향구간). 주문수량 이하 최대 min_qty 행 매칭",
        "proc_cd": "신설 공정 차원(8→10). 인쇄비는 미사용 NULL",
        "opt_cd": "신설 옵션 차원(8→10). 인쇄비는 미사용 NULL",
        "unit_price": "장당가(단가형). 엔진=단가×주문수량",
    },
)

# ============================================================
# 4b. component_prices GAP — 3절 별색 미적재 (정직표기)
# ============================================================
gap_note = "[GAP] 가격표 3절 블록(B62)은 흑백/CMYK만 — 별색 5종 3절 가격 없음. 라이브도 별색=국4절(SIZ_000499)만. 미적재가 정상(가격표 부재). 향후 3절 별색 출시 시 채울 자리."
gap_rows = []
for spot in ["WHITE", "CLEAR", "PINK", "GOLD", "SILVER"]:
    for side in ["S1", "S2"]:
        gap_rows.append([f"COMP_PRINT_SPOT_{spot}_{side}", "SIZ_000077", None, None, None, None, None, None,
                         "(가격표 부재)", "—", None])
write_sheet(
    wb, "4b_component_prices_print_GAP",
    ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "출력판형(siz)", "도수", "자재", "공정", "코팅면수", "옵션", "묶음수", "수량구간", "적용일", "단가"],
    gap_rows,
    note=gap_note,
    fill=BLK_FILL,
)

# ============================================================
# 9. BLOCKED binding (round-2 BLOCKED 3상품)
# ============================================================
blk = read_csv(f"{R2}/t_prd_product_price_formulas_DGP_BLOCKED_siz.csv")
blk_data = [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in blk[1:]]
write_sheet(
    wb, "9_BLOCKED_binding",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일", "차단사유"],
    blk_data,
    note="[BLOCKED] siz 미채번/plate 결함으로 미적재 3상품 (019 투명엽서·030 지그재그엽서·049 와이드접지리플렛). NULL 강제 금지 — 별 시트 분리. siz 채번/plate 교정은 인간 승인.",
    fill=BLK_FILL,
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("component_prices rows:", len(cp_rows))
