#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
namecard-photocard import.xlsx 빌더 (round-16)
- 라이브 t_prc_* 명함/포토카드 그릇을 webadmin 복붙용 엑셀로 재현
- 가격표 "/" collapse 소재를 개별 mat_cd로 전개(소재 개별분해 — sticker-material-axis 동형)
- 6 시트: 1_price_formulas / 1b_product_price_formulas / 2_formula_components
          / 3_price_components / 4_component_prices / 4b_component_prices_BLOCKED
- 입력: /tmp/nc/cp.txt(라이브 단가행 115) /tmp/nc/comps.txt(comp 정의 27)
- DB 미적재. 라이브 읽기전용 덤프를 토대로 파일만 생성.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/namecard-photocard/namecard-photocard-import.xlsx"
APPLY = "2026-01-01"

# ---- 소재명 매핑(라이브 실측) ----
MAT_NM = {
    "MAT_000074":"백색모조지220","MAT_000081":"아트지250","MAT_000082":"아트지300",
    "MAT_000091":"스노우지250","MAT_000092":"스노우지300","MAT_000127":"스타드림다이아240",
    "MAT_000128":"스타드림실버240","MAT_000129":"스타드림골드240","MAT_000130":"스타드림로즈쿼츠240",
    "MAT_000241":"스타드림로츠쿼츠240","MAT_000137":"큐리어스스킨화이트270","MAT_000138":"큐리어스스킨레드270",
    "MAT_000139":"큐리어스스킨다크블루270","MAT_000140":"큐리어스스킨바이올렛270","MAT_000141":"큐리어스스킨블랙270",
    "MAT_000144":"투명PET260","MAT_000147":"반투명PET260","MAT_000109":"몽블랑240",
    "MAT_000101":"랑데뷰240","MAT_000102":"랑데뷰310","MAT_000108":"몽블랑210","MAT_000113":"아코팩250",
    "MAT_000114":"리사이클러스240","MAT_000115":"매쉬멜로우233","MAT_000116":"린넨커버216",
    "MAT_000117":"스타화이트238","MAT_000118":"클래식크러스트270","MAT_000123":"띤또레또200",
    "MAT_000124":"띤또레또250","MAT_000125":"한지170","MAT_000126":"스코트랜드220",
}
SIZ_NM = {"SIZ_000008":"90x50","SIZ_000011":"50x50","SIZ_000012":"55x86"}

# ---- 1. price_formulas (공식정의) ----
# 라이브 2개 + round-16 권고 명함종류별 공식(미적재·권고)은 비고로 표기.
formulas = [
    ("PRF_NAMECARD_FIXED","명함 면/소재/수량별 단가(용지포함)","라이브 적재·STD만 배선(🔴24 comp 미배선)","Y"),
    ("PRF_PHOTOCARD_FIXED","포토카드 세트 고정가","라이브 적재·SET/CLEAR_SET 배선","Y"),
]

# ---- 1b. product_price_formulas (상품바인딩) ----
bindings = [
    ("PRD_000033","PRF_NAMECARD_FIXED",APPLY,"스탠다드명함 ✅"),
    ("PRD_000031","PRF_NAMECARD_FIXED",APPLY,"프리미엄명함 🔴배선X(STD단가 오매칭)"),
    ("PRD_000032","PRF_NAMECARD_FIXED",APPLY,"코팅명함 🔴배선X"),
    ("PRD_000024","PRF_PHOTOCARD_FIXED",APPLY,"포토카드 ✅"),
    ("PRD_000025","PRF_PHOTOCARD_FIXED",APPLY,"투명포토카드 ✅"),
    # round-16 권고 미바인딩(가격사슬 단절 — 인간 승인)
    ("PRD_000034","(PRF_NAMECARD_PEARL 권고)",APPLY,"펄명함 🔴바인딩 부재"),
    ("PRD_000035","(PRF_NAMECARD_SHAPE 권고)",APPLY,"모양명함 🔴바인딩 부재"),
    ("PRD_000036","(PRF_NAMECARD_MINISHAPE 권고)",APPLY,"미니모양명함 🔴바인딩 부재"),
    ("PRD_000037","(PRF_NAMECARD_FOIL 권고)",APPLY,"오리지널박명함 🔴바인딩 부재"),
    ("PRD_000039","(PRF_NAMECARD_CLEAR 권고)",APPLY,"투명명함 🔴바인딩 부재"),
    ("PRD_000040","(PRF_NAMECARD_WHITE 권고)",APPLY,"화이트인쇄명함 🔴바인딩 부재"),
]

# ---- 2. formula_components (배선) ----
wiring = [
    ("PRF_NAMECARD_FIXED","COMP_NAMECARD_STD_S1",1,"N","✅ 라이브 배선"),
    ("PRF_NAMECARD_FIXED","COMP_NAMECARD_STD_S2",2,"N","✅ 라이브 배선"),
    ("PRF_PHOTOCARD_FIXED","COMP_PHOTOCARD_SET",1,"N","✅"),
    ("PRF_PHOTOCARD_FIXED","COMP_PHOTOCARD_CLEAR_SET",2,"N","✅"),
    # round-16 권고 배선(미적재 — 24 고아 comp 복구)
    ("(PRF_NAMECARD_PREMIUM 권고)","COMP_NAMECARD_PREMIUM_S1_MGA",1,"N","🔴 고아 복구권고"),
    ("(PRF_NAMECARD_PREMIUM 권고)","COMP_NAMECARD_PREMIUM_S1_MGB",2,"N","🔴"),
    ("(PRF_NAMECARD_PREMIUM 권고)","COMP_NAMECARD_PREMIUM_S2_MGA",3,"N","🔴"),
    ("(PRF_NAMECARD_PREMIUM 권고)","COMP_NAMECARD_PREMIUM_S2_MGB",4,"N","🔴"),
    ("(PRF_NAMECARD_COAT 권고)","COMP_NAMECARD_COAT_S1",1,"N","🔴"),
    ("(PRF_NAMECARD_COAT 권고)","COMP_NAMECARD_COAT_S2",2,"N","🔴"),
    ("(PRF_NAMECARD_PEARL 권고)","COMP_NAMECARD_PEARL_S1",1,"N","🔴"),
    ("(PRF_NAMECARD_PEARL 권고)","COMP_NAMECARD_PEARL_S2",2,"N","🔴"),
    ("(PRF_NAMECARD_CLEAR 권고)","COMP_NAMECARD_CLEAR_S1",1,"N","🔴"),
    ("(PRF_NAMECARD_WHITE 권고)","COMP_NAMECARD_WHITE_S1W_NOCL",1,"N","🔴"),
    ("(PRF_NAMECARD_WHITE 권고)","COMP_NAMECARD_WHITE_S1W_CL",2,"N","🔴"),
    ("(PRF_NAMECARD_WHITE 권고)","COMP_NAMECARD_WHITE_S2W_NOCL",3,"N","🔴"),
    ("(PRF_NAMECARD_WHITE 권고)","COMP_NAMECARD_WHITE_S2W_CL",4,"N","🔴"),
    ("(PRF_NAMECARD_SHAPE 권고)","COMP_NAMECARD_SHAPE_S1",1,"N","🔴"),
    ("(PRF_NAMECARD_SHAPE 권고)","COMP_NAMECARD_SHAPE_S2",2,"N","🔴"),
    ("(PRF_NAMECARD_MINISHAPE 권고)","COMP_NAMECARD_MINISHAPE_S1",1,"N","🔴"),
    ("(PRF_NAMECARD_MINISHAPE 권고)","COMP_NAMECARD_MINISHAPE_S2",2,"N","🔴"),
    ("(PRF_NAMECARD_FOIL 권고)","COMP_NAMECARD_FOIL_S1_STD",1,"N","🔴 박단가"),
    ("(PRF_NAMECARD_FOIL 권고)","COMP_NAMECARD_FOIL_S1_HOLO",2,"N","🔴"),
    ("(PRF_NAMECARD_FOIL 권고)","COMP_NAMECARD_FOIL_S2_STD",3,"N","🔴"),
    ("(PRF_NAMECARD_FOIL 권고)","COMP_NAMECARD_FOIL_S2_HOLO",4,"N","🔴"),
    ("(PRF_NAMECARD_FOIL 권고)","COMP_NAMECARD_FOIL_SETUP_S1_STD",5,"Y","🔴 동판셋업 합산"),
    ("(PRF_NAMECARD_FOIL 권고)","COMP_NAMECARD_FOIL_SETUP_S2_STD",6,"Y","🔴"),
    ("(PRF_PHOTOCARD_BULK 권고)","COMP_PHOTOCARD_BULK",1,"N","🔴 대량밴드"),
]

# ---- 3. price_components (구성요소 정의) — 라이브 27 읽어옴 ----
comps = []
with open("/tmp/nc/comps.txt") as f:
    for ln in f:
        p = ln.rstrip("\n").split("|")
        if len(p) < 5: continue
        comp_cd, comp_nm, comp_typ, prc_typ, dims = p[0],p[1],p[2],p[3],p[4]
        # round-16 권고는 comp별 차등(decomposition §2):
        #  - 박(FOIL)·대량(BULK) = 명확한 수량구간 총액 → .02 합가형
        #  - 세트(SET) = 세트총액 → .02 합가형
        #  - 단순명함(STD/PREMIUM/COAT/PEARL/CLEAR/WHITE/SHAPE/MINISHAPE) = 100매 고정단위
        #    → .01(100매당 단가)도 정당 가능·.02(총액÷100)도 가능 → Q-NC-1 컨펌 대상
        is_band = any(k in comp_cd for k in ("FOIL","BULK")) and "SETUP" not in comp_cd
        is_set  = "_SET" in comp_cd
        if is_band:
            rec_prc = "PRICE_TYPE.02 (합가형·구간총액)"
            note = f"라이브={prc_typ}(단가형) / round-16=🔴.02 합가형(수량구간 총액÷min_qty)"
        elif is_set:
            rec_prc = "PRICE_TYPE.02 (합가형·세트총액)"
            note = f"라이브={prc_typ} / round-16=.02 합가형(세트총액·Q-NC-2 단위)"
        elif "SETUP" in comp_cd:
            rec_prc = prc_typ  # 동판셋업=1회 고정비, 단가형 그대로 정당
            note = "동판셋업 1회 고정비(수량무관)·단가형 정당"
        else:
            rec_prc = f"{prc_typ} 또는 .02 (Q-NC-1)"
            note = "100매 고정단위 → .01(100매당) 정당 / 위젯 수량단위가 '매'면 .02(총액÷100) — Q-NC-1"
        comps.append((comp_cd, comp_nm, prc_typ, rec_prc, dims, "Y", note))

# ---- 4 / 4b. component_prices ----
# 라이브 115행을 main으로. 단, 소재 collapse 블록은 개별 mat_cd 전개행을 추가.
main_rows = []   # 라이브 적재행(재현)
expand_rows = [] # 가격표 "/" 개별소재 전개(라이브 미적재 → BLOCKED 후보)

# 4-A 라이브행 로드
live = []
with open("/tmp/nc/cp.txt") as f:
    for ln in f:
        p = ln.rstrip("\n").split("|")
        if len(p) < 7: continue
        comp_cd,siz,mat,coat,bdl,mn,up = p
        live.append((comp_cd, siz or None, mat or None, coat or None, bdl or None, mn or None, float(up)))

# [P3 정정] PEARL 로츠쿼츠: 라이브가 MAT_000130(로즈쿼츠·가격표 R24 부재)을 적재 →
#   가격표는 다이아/실버/골드/로츠쿼츠 4종뿐. 130은 가격표 부재 소재이므로 main 재현에서 제외하고
#   정답 MAT_000241(로츠쿼츠)만 expand로 전개(round-5 라이브 교정 대상·Q-NC-7).
SKIP_LIVE = {("COMP_NAMECARD_PEARL_S1","MAT_000130"),("COMP_NAMECARD_PEARL_S2","MAT_000130")}
skipped=0
for comp_cd,siz,mat,coat,bdl,mn,up in live:
    if (comp_cd,mat) in SKIP_LIVE:
        skipped+=1; continue
    main_rows.append((comp_cd,siz,None,mat,None,coat,None,bdl,mn,APPLY,up,
                      SIZ_NM.get(siz,""),MAT_NM.get(mat,""),"라이브 적재행(재현·재적재 금지)"))

# 4-B 소재 개별분해 전개(가격표 "/" 동가 소재 — 라이브 collapse 누락분)
# B01 스탠다드: 백모조220(074·라이브있음)/아트250(081)/스노우250(091) 단면3500·양면4500 ; 아트300(082·라이브)/스노우300(092) 3800·4800
EXP = [
  # (comp_cd, mat, min, unit, 사유)
  ("COMP_NAMECARD_STD_S1","MAT_000081",100,3500,"B01 아트250 동가(라이브 collapse 누락)"),
  ("COMP_NAMECARD_STD_S1","MAT_000091",100,3500,"B01 스노우250 동가(누락)"),
  ("COMP_NAMECARD_STD_S1","MAT_000092",100,3800,"B01 스노우300 동가(누락)"),
  ("COMP_NAMECARD_STD_S2","MAT_000081",100,4500,"B01 아트250 양면 동가(누락)"),
  ("COMP_NAMECARD_STD_S2","MAT_000091",100,4500,"B01 스노우250 양면(누락)"),
  ("COMP_NAMECARD_STD_S2","MAT_000092",100,4800,"B01 스노우300 양면(누락)"),
  # B04 펄: 실버(128)/골드(129) 다이아 동가 9000/10000 ; 로츠쿼츠 라이브오적재(130→241)
  ("COMP_NAMECARD_PEARL_S1","MAT_000128",100,9000,"B04 실버240 동가(누락)"),
  ("COMP_NAMECARD_PEARL_S1","MAT_000129",100,9000,"B04 골드240 동가(누락)"),
  ("COMP_NAMECARD_PEARL_S2","MAT_000128",100,10000,"B04 실버240 양면(누락)"),
  ("COMP_NAMECARD_PEARL_S2","MAT_000129",100,10000,"B04 골드240 양면(누락)"),
  ("COMP_NAMECARD_PEARL_S1","MAT_000241",100,10000,"B04 로츠쿼츠240(라이브 130 로즈쿼츠 오적재 정정·Q-NC-7)"),
  ("COMP_NAMECARD_PEARL_S2","MAT_000241",100,11000,"B04 로츠쿼츠240 양면(정정)"),
]
# B06 화이트: R38 "A: 화이트/레드/다크블루/바이올렛/블랙" 5색 동가 소재그룹.
#   라이브는 화이트(137)만 4 comp 적재 → 나머지 4색(138~141)을 4 comp 전부 대칭 전개.
#   [P3 정정] NOCL_S1만 5색이던 비대칭을 4 comp×4색=16행 대칭으로 보정.
WHITE_COMPS = [
  ("COMP_NAMECARD_WHITE_S1W_NOCL", 14500),  # 화이트(단면)+클리어(없음)
  ("COMP_NAMECARD_WHITE_S1W_CL",   16000),  # 화이트(단면)+클리어(단면)
  ("COMP_NAMECARD_WHITE_S2W_NOCL", 16000),  # 화이트(양면)+클리어(없음)
  ("COMP_NAMECARD_WHITE_S2W_CL",   19000),  # 화이트(양면)+클리어(양면)
]
WHITE_COLORS = ["MAT_000138","MAT_000139","MAT_000140","MAT_000141"]  # 레드·다크블루·바이올렛·블랙
for wcomp, wup in WHITE_COMPS:
    for wmat in WHITE_COLORS:
        EXP.append((wcomp, wmat, 100, wup, f"B06 {MAT_NM.get(wmat,wmat)} 동가(라이브 화이트137만·대칭 전개)"))
for comp_cd,mat,mn,up,reason in EXP:
    expand_rows.append((comp_cd,None,None,mat,None,None,None,None,mn,APPLY,up,
                        "",MAT_NM.get(mat,""),reason))

# 4b BLOCKED: 라이브 mat=∅인 그룹(PREMIUM 14소재·CLEAR 2소재·SHAPE/MINI 몽블랑·FOIL 박종류=비소재)
# 이들은 개별소재 mat_cd 전개가 가격표상 가능하나 라이브 단가행이 mat 미사용/그룹단가 → 적재 전 소재축 결정 컨펌 필요.
BLOCK = []
# B02 프리미엄 A그룹 7소재(단/양면) MGA: 단4500/양5500
PREM_A = ["MAT_000101","MAT_000102","MAT_000108","MAT_000109","MAT_000113","MAT_000114","MAT_000115"]
PREM_B = ["MAT_000116","MAT_000117","MAT_000118","MAT_000123","MAT_000124","MAT_000125","MAT_000126"]
for m in PREM_A:
    BLOCK.append(("COMP_NAMECARD_PREMIUM_S1_MGA",None,m,100,4500,"B02 프리미엄A 단면(라이브 mat=∅ 그룹단가·소재축 컨펌 Q-NC-?)"))
    BLOCK.append(("COMP_NAMECARD_PREMIUM_S2_MGA",None,m,100,5500,"B02 프리미엄A 양면"))
for m in PREM_B:
    BLOCK.append(("COMP_NAMECARD_PREMIUM_S1_MGB",None,m,100,5000,"B02 프리미엄B 단면"))
    BLOCK.append(("COMP_NAMECARD_PREMIUM_S2_MGB",None,m,100,6500,"B02 프리미엄B 양면"))
# B05 투명 2소재(라이브 CLEAR_S1 mat=∅)
BLOCK.append(("COMP_NAMECARD_CLEAR_S1",None,"MAT_000144",100,13500,"B05 투명PET260(라이브 mat=∅·소재분리 컨펌)"))
BLOCK.append(("COMP_NAMECARD_CLEAR_S1",None,"MAT_000147",100,13500,"B05 반투명PET260"))
# B07/B08 모양·미니 몽블랑240(라이브 siz만·mat 미기입 Q-NC-3)
BLOCK.append(("COMP_NAMECARD_SHAPE_S1","SIZ_000008","MAT_000109",100,18000,"B07 모양 몽블랑240(라이브 mat 미기입·Q-NC-3)"))
BLOCK.append(("COMP_NAMECARD_SHAPE_S2","SIZ_000008","MAT_000109",100,19000,"B07 양면"))
BLOCK.append(("COMP_NAMECARD_MINISHAPE_S1","SIZ_000011","MAT_000109",100,16000,"B08 미니 몽블랑240"))
BLOCK.append(("COMP_NAMECARD_MINISHAPE_S2","SIZ_000011","MAT_000109",100,17000,"B08 양면"))

blocked_rows = []
for comp_cd,siz,mat,mn,up,reason in BLOCK:
    blocked_rows.append((comp_cd,siz,None,mat,None,None,None,None,mn,APPLY,up,
                         SIZ_NM.get(siz,""),MAT_NM.get(mat,""),reason))

# ================= 엑셀 작성 =================
wb = openpyxl.Workbook()
wb.remove(wb.active)

HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")
KFILL = PatternFill("solid", fgColor="D9E1F2")
KFONT = Font(italic=True, size=9)
WRAP = Alignment(wrap_text=True, vertical="top")

def sheet(name, cols, krow, rows, comments=None):
    ws = wb.create_sheet(name)
    for j,c in enumerate(cols,1):
        cell=ws.cell(1,j,c); cell.font=HDR; cell.fill=HFILL
        if comments and c in comments:
            cell.comment=Comment(comments[c],"round-16")
    for j,k in enumerate(krow,1):
        cell=ws.cell(2,j,k); cell.font=KFONT; cell.fill=KFILL
    for i,row in enumerate(rows,3):
        for j,v in enumerate(row,1):
            ws.cell(i,j,v).alignment=WRAP
    for j,c in enumerate(cols,1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(12,min(40,len(str(c))+8))
    ws.freeze_panes="A3"
    return ws

# 1
sheet("1_price_formulas",
   ["frm_cd","frm_nm","note","use_yn"],
   ["공식코드","공식명","비고(실무진 쉬운말)","사용여부"],
   formulas,
   {"frm_cd":"공식 식별자(PK). 명함=PRF_NAMECARD_*, 포토카드=PRF_PHOTOCARD_*",
    "use_yn":"Y/N"})

# 1b
sheet("1b_product_price_formulas",
   ["prd_cd","frm_cd","apply_bgn_ymd","note"],
   ["상품코드","공식코드","적용시작일","비고"],
   bindings,
   {"prd_cd":"상품(PRD_*). PK=(prd_cd, apply_bgn_ymd)",
    "note":"🔴=바인딩 부재(가격사슬 단절)·인간 승인 후 적재"})

# 2
sheet("2_formula_components",
   ["frm_cd","comp_cd","disp_seq","addtn_yn","note"],
   ["공식코드","구성요소코드","표시순서","합산여부","비고"],
   wiring,
   {"addtn_yn":"Y=합산(박+동판셋업 등)·Phase11 엔진은 무시·표시용",
    "note":"🔴=라이브 미배선(고아 comp 복구 권고)"})

# 3
sheet("3_price_components",
   ["comp_cd","comp_nm","prc_typ_cd_live","prc_typ_cd_rec","use_dims","use_yn","note"],
   ["구성요소코드","구성요소명","단가유형(라이브)","단가유형(round-16권고)","사용차원","사용여부","비고"],
   comps,
   {"prc_typ_cd_live":"라이브 적재값(전부 .01 단가형)",
    "prc_typ_cd_rec":"round-16 권고(.02 합가형=총액÷단위)·Q-NC-1",
    "use_dims":"단가표가 실제 쓰는 차원 배열(나머지=NULL 와일드카드)"})

CP_COLS=["comp_cd","siz_cd","clr_cd","mat_cd","proc_cd","coat_side_cnt","opt_cd","bdl_qty","min_qty","apply_ymd","unit_price","siz_label","mat_label","note"]
CP_KROW=["구성요소","사이즈","도수","자재(소재별)","공정","코팅면수","옵션","묶음수","수량구간","적용일","단가","[참고]사이즈","[참고]소재","비고"]
CP_CMT={"clr_cd":"명함 가격행 도수 무관 → NULL(별색=공정 규칙)",
        "coat_side_cnt":"인쇄면(단/양면)은 comp 접미사 _S1/_S2로 분리 → NULL",
        "proc_cd":"명함 단가행 공정차원 미사용 → NULL",
        "opt_cd":"옵션차원 미사용 → NULL",
        "mat_cd":"소재 개별분해(가격표 '/' = 개별 mat_cd·collapse 금지)"}

# 4 main = 라이브 재현행 + 소재 개별분해 전개행
sheet("4_component_prices", CP_COLS, CP_KROW, main_rows + expand_rows, CP_CMT)

# 4b BLOCKED
sheet("4b_component_prices_BLOCKED", CP_COLS, CP_KROW, blocked_rows, CP_CMT)

wb.save(OUT)

# round-trip / 동시매칭 자가검증
allrows = main_rows + expand_rows
keys=[(r[0],r[1],r[3],r[5],r[7],r[8]) for r in allrows]  # comp,siz,mat,coat,bdl,min
dup=len(keys)-len(set(keys))
print(f"saved: {OUT}")
print(f"sheets: formulas={len(formulas)} bindings={len(bindings)} wiring={len(wiring)} comps={len(comps)}")
print(f"4_component_prices: live={len(main_rows)} (raw115 - skip{skipped}) expand={len(expand_rows)} total={len(allrows)}")
print(f"4b_BLOCKED: {len(blocked_rows)}")
print(f"동시매칭(중복 자연키) in main+expand: {dup}")
# P3/P5 정정 검증
pearl=[r for r in allrows if "PEARL" in r[0]]
white=[r for r in allrows if "WHITE" in r[0]]
pearl_mats=sorted(set(r[3] for r in pearl))
print(f"PEARL 행수={len(pearl)} mats={pearl_mats} (130 제거·241 포함 확인)")
from collections import Counter
white_by_comp=Counter(r[0] for r in white)
print(f"WHITE 행수={len(white)} comp별={dict(white_by_comp)} (4 comp 대칭 5색=20 기대)")
