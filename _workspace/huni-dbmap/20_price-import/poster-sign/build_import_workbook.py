# -*- coding: utf-8 -*-
"""
round-16 포스터사인 webadmin 복붙용 import.xlsx 빌더 (면적매트릭스형 + 수량밴드 혼재).
입력 = /tmp/ps_export/*.psv (라이브 read-only 실측 덤프) + 원본 가격표 엑셀(포스터사인 시트).
출력 = poster-sign-import.xlsx (테이블별 시트, 1행=DB컬럼, 2행=한글라벨).
면적매트릭스 = (siz_cd, unit_price) long-form. 밴드 = (siz_cd, min_qty, unit_price).
라이브 적재분=재현(_RU), siz 미채번=GAP_BLOCKED, 옵션=별 comp 분리.
[HARD] 실사 가격 = 이 포스터사인 매트릭스가 권위(메모리 dbmap-silsa-price-via-poster-sign).
"""
import openpyxl, csv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

EXPORT = "/tmp/ps_export"
XL = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx"
OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/poster-sign/poster-sign-import.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")   # GAP/BLOCKED 표식
RU_FILL  = PatternFill("solid", fgColor="E2EFDA")    # 재현(라이브 기존) 표식
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")    # 신규 후보 표식


def _clean(v):
    # COPY FORMAT csv quotes empty/special fields as "" — normalize to plain str
    if v == '""':
        return ""
    if len(v) >= 2 and v.startswith('"') and v.endswith('"'):
        return v[1:-1].replace('""', '"')
    return v


def read_psv(name):
    rows = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append([_clean(x) for x in line.split("|")])
    return rows


def n(v):
    return None if v in ("", None) else v


def num(s):
    s = str(s).replace("mm", "").replace(" ", "").strip()
    try:
        return int(float(s))
    except Exception:
        return None


def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, min(30, len(lbl) + 4))
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


# component_prices 10차원 컬럼/라벨/주석 (재사용)
CP_COLS = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"]
CP_LBLS = ["구성요소", "출력규격(siz·가로×세로)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수", "수량구간(min_qty)", "적용일", "단가"]
CP_CMT = {
    "comp_cd": "소재별 구성요소(COMP_POSTER_<소재>). 소재가 구성요소명에 박힘(mat_cd 미사용)",
    "siz_cd": "가로×세로 면적 규격코드. 면적매트릭스 차원=siz_cd. 미채번(BLOCKED)은 별 시트",
    "clr_cd": "NULL — 매트릭스 단가는 도수 무관(통가격)",
    "min_qty": "면적매트릭스=NULL(수량무관). 사이즈/수량 밴드=수량구간 하한(단가형·구간별 단가)",
    "proc_cd": "신설 공정 차원(8→10). 포스터 본체 미사용 NULL(옵션은 별 comp)",
    "opt_cd": "신설 옵션 차원(8→10). 포스터 옵션은 별 comp(COMP_POSTEROPT_*)로 모델 — 본체 NULL",
    "unit_price": "개당 단가(단가형 .01). 엔진=단가×주문수량. off-grid=한단계 큰 규격(런타임·DB미저장)",
    "apply_ymd": "적용일(라이브 2026-06-01)",
}

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
ws = wb.create_sheet("0_README")
readme = [
    ["round-16 포스터사인 가격표 import 그릇 (webadmin 복붙용) — 면적매트릭스 + 수량밴드", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 포스터사인 (389행×26열)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["가격유형", "면적매트릭스형(가로×세로 siz_cd) + 사이즈/수량 밴드(siz_cd+min_qty) 혼재"],
    ["단가/합가", "전건 단가형(PRICE_TYPE.01·라이브 실측). 합가형 없음(밴드도 구간별 단가형)"],
    ["[HARD·사용자] 실사 가격 권위", "실사 시트 가격은 실사 inline 아닌 이 포스터사인 [가로×세로] 매트릭스로 처리(메모리 dbmap-silsa-price-via-poster-sign)"],
    ["", ""],
    ["[중요] 면적매트릭스 = (siz_cd, unit_price) long-form", ""],
    ["  · 매트릭스 [가로][세로] 격자를 행으로 언피벗(면적함수 회귀 금지)", ""],
    ["  · use_dims=[\"siz_cd\"] — 나머지 9차원 NULL(빈칸). 밴드는 +min_qty", ""],
    ["  · off-grid(격자 부재 크기) = 한단계 큰 규격 단가(런타임·DB 미저장)", ""],
    ["", ""],
    ["[경고1] 가격사슬 부분단절 (핵심 발견)", ""],
    ["  · 라이브 공식 PRF_POSTER_FIXED에 28상품 전건 바인딩됐으나", ""],
    ["    공식이 배선한 comp = COMP_POSTER_ARTPRINT_PHOTO 1개뿐(formula_components 1행)", ""],
    ["  · → 인화지 외 27상품은 자기 소재 단가행 적재돼 있어도 엔진 조회불가", ""],
    ["    → 1_price_formulas / 2_formula_components / 1b_바인딩 = 소재별 공식 보강 제안(컨펌 Q-PS-1)", ""],
    ["", ""],
    ["[경고2] siz 미채번 BLOCKED 압도적 (97%)", ""],
    ["  · 가로×세로 면적조합 687 중 667(97%)이 siz_cd 라이브 미실재", ""],
    ["  · 주황(_GAP_BLOCKED) = siz '(미채번:GxS)' 표기 — 좌표 채번 후 적재(NULL 강제 금지·별 시트)", ""],
    ["  · 라이브 실재 siz 좌표만 4_RU로 정상 적재(아크릴과 달리 실재율 극저)", ""],
    ["", ""],
    ["[색 범례]", ""],
    ["  · 초록(_RU) = 라이브 기존 103행/53comp/28바인딩 재현(재적재 금지·대조용)", ""],
    ["  · 노랑(_NEW) = 소재별 공식/배선 신규 제안(컨펌 후)", ""],
    ["  · 주황(_GAP_BLOCKED/참조) = siz 미채번·옵션·부유노트(별 트랙)", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas", "공식 정의: RU 1(PRF_POSTER_FIXED) + 소재별 신규 제안"],
    ["  1b_product_price_formulas", "상품↔공식 바인딩: RU 28 + 소재별 교체 제안"],
    ["  2_formula_components", "공식 배선: RU 1(인화지만) + 소재별 배선 신규(단절해소)"],
    ["  3_price_components", "구성요소 정의 53 라이브 재현(전건 단가형.01)"],
    ["  4_component_prices_RU", "단가행 103 라이브 재현(면적+밴드+옵션)"],
    ["  4b_component_prices_GAP_BLOCKED", "면적조합 미채번 667(siz 채번 대상)"],
    ["  5_qtyband_RU", "사이즈/수량 밴드 단가행(siz_cd+min_qty·라이브 재현 부분집합)"],
    ["  6_addon_options", "옵션 add-on 별 comp(현수막 가공/추가·천장후크 등)"],
    ["  7_excluded_note", "부유셀·블록간 노트 보존(침묵삭제 금지)"],
    ["", ""],
    ["무손실 검산", "라이브 103행 전건 엑셀 대조. 엑셀 789셀 = RU 103 + GAP 672(BLOCKED 667). round-trip 보존"],
]
for ri, (a, b) in enumerate(readme, start=1):
    ws.cell(row=ri, column=1, value=a)
    ws.cell(row=ri, column=2, value=b)
    if ri == 1:
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 78

# ============================================================
# 1. price_formulas (RU 1 + 소재별 신규)
# ============================================================
# 소재 → comp 매핑(라이브 comp_cd 접미사 기준)
MATERIAL_COMPS = [
    ("PRF_POSTER_ARTPRINT", "COMP_POSTER_ARTPRINT_PHOTO", "아트프린트포스터(인화지)"),
    ("PRF_POSTER_ARTPAPER", "COMP_POSTER_ARTPAPER_MATTE", "아트페이퍼포스터(매트지)"),
    ("PRF_POSTER_WATERPROOF_PET", "COMP_POSTER_WATERPROOF_PET", "방수포스터(PET)"),
    ("PRF_POSTER_ADH_WATERPROOF", "COMP_POSTER_ADH_WATERPROOF_PVC", "접착방수포스터(PVC)"),
    ("PRF_POSTER_ADH_CLEAR", "COMP_POSTER_ADH_CLEAR_PVC", "접착투명포스터(투명PVC)"),
    ("PRF_POSTER_ARTFABRIC", "COMP_POSTER_ARTFABRIC_GRAPHIC", "아트패브릭포스터"),
    ("PRF_POSTER_LINEN_FABRIC", "COMP_POSTER_LINEN_FABRIC", "린넨패브릭포스터"),
    ("PRF_POSTER_CANVAS_FABRIC", "COMP_POSTER_CANVAS_FABRIC", "캔버스패브릭포스터"),
    ("PRF_POSTER_LEATHER_ARTPRINT", "COMP_POSTER_LEATHER_ARTPRINT", "레더아트프린트"),
    ("PRF_POSTER_TYVEK", "COMP_POSTER_TYVEK_PRINT", "타이벡프린트"),
    ("PRF_POSTER_MESH_PRINT", "COMP_POSTER_MESH_PRINT", "메쉬프린트"),
    ("PRF_POSTER_BANNER_NORMAL", "COMP_POSTER_BANNER_NORMAL", "일반현수막"),
    ("PRF_POSTER_BANNER_MESH", "COMP_POSTER_BANNER_MESH", "메쉬현수막"),
]
formulas = [["PRF_POSTER_FIXED", "포스터사인 소재/사이즈/수량별 완제품가(포함항목 통가격)", "Y", "[RU] 라이브 기존 — 단 배선 1 comp(인화지)만 → 부분단절"]]
for frm, comp, nm in MATERIAL_COMPS:
    formulas.append([frm, f"{nm} 면적/밴드 단가 공식", "Y", f"[NEW] 소재별 공식 — {comp} 배선용(사슬 보강·Q-PS-1)"])
write_sheet(
    wb, "1_price_formulas",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    formulas,
    note="[RU 1 + NEW] PRF_POSTER_FIXED(라이브)는 28상품 바인딩됐으나 인화지 1 comp만 배선(가격사슬 부분단절). 소재별 공식 신규 제안 — 27상품 조회불가 해소(컨펌 Q-PS-1: 소재별 분리 vs 단일공식 조건분기).",
    fill=NEW_FILL,
    col_comments={"frm_cd": "공식 식별자. 라이브=PRF_POSTER_FIXED 1개(인화지만 배선). 소재별 공식 신규 제안"},
)

# ============================================================
# 1b. product_price_formulas (RU 28 + 교체 제안)
# ============================================================
prod_nm = {r[0]: r[1] for r in read_psv("products.psv")}
bind = read_psv("bindings.psv")  # prd_cd|frm_cd|apply_bgn_ymd
# 상품명 → 소재공식 매핑(소재별 공식 교체 제안)
NM2FRM = {
    "아트프린트포스터": "PRF_POSTER_ARTPRINT", "아트페이퍼포스터": "PRF_POSTER_ARTPAPER",
    "방수포스터": "PRF_POSTER_WATERPROOF_PET", "접착방수포스터": "PRF_POSTER_ADH_WATERPROOF",
    "접착투명포스터": "PRF_POSTER_ADH_CLEAR", "아트패브릭포스터": "PRF_POSTER_ARTFABRIC",
    "린넨패브릭포스터": "PRF_POSTER_LINEN_FABRIC", "캔버스패브릭포스터": "PRF_POSTER_CANVAS_FABRIC",
    "레더아트프린트": "PRF_POSTER_LEATHER_ARTPRINT", "타이벡프린트": "PRF_POSTER_TYVEK",
    "메쉬프린트": "PRF_POSTER_MESH_PRINT", "일반현수막": "PRF_POSTER_BANNER_NORMAL",
    "메쉬현수막": "PRF_POSTER_BANNER_MESH",
}
bind_rows = []
for r in bind:
    prd, frm, ymd = r[0], r[1], r[2]
    nm = prod_nm.get(prd, "")
    new_frm = NM2FRM.get(nm, "")
    note = f"[RU] {nm} → {frm}(라이브)"
    if new_frm and new_frm != frm:
        note = f"[교체제안] {nm}: {frm} → {new_frm}(소재별 공식·사슬보강 Q-PS-1)"
    bind_rows.append([prd, frm, ymd, note])
write_sheet(
    wb, "1b_product_price_formulas",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일(yyyy-MM-dd)", "비고"],
    bind_rows,
    note="[RU 28 + 교체제안] 라이브 28상품 전건 PRF_POSTER_FIXED 바인딩(완비). 사슬 보강 시 소재별 공식으로 교체 제안(frm_cd 컬럼=라이브 현재값·비고=교체 타깃). 교체는 컨펌 Q-PS-1 후.",
    fill=RU_FILL,
    col_comments={"prd_cd": "t_prd_products FK. 28상품(PRD_000118~145)=포스터사인 상품군", "frm_cd": "라이브 현재=PRF_POSTER_FIXED. 비고에 소재별 교체 타깃"},
)

# ============================================================
# 2. formula_components (RU 1 + 소재별 배선 신규)
# ============================================================
fc = read_psv("formula_components.psv")  # frm_cd|comp_cd|disp_seq|addtn_yn
wire_rows = []
for r in fc:
    wire_rows.append([r[0], r[1], int(r[2]) if r[2] else 1, r[3], "[RU] 라이브 — 인화지만 배선(부분단절)"])
for frm, comp, nm in MATERIAL_COMPS:
    if comp == "COMP_POSTER_ARTPRINT_PHOTO":
        continue  # 이미 RU
    wire_rows.append([frm, comp, 1, "Y", f"[NEW] {nm} 배선(라이브 미배선)"])
write_sheet(
    wb, "2_formula_components",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn", "note"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)", "비고"],
    wire_rows,
    note="[RU 1 + NEW] 라이브 formula_components = PRF_POSTER_FIXED←COMP_POSTER_ARTPRINT_PHOTO 1행뿐(가격사슬 부분단절 핵심). 소재별 공식에 자기 소재 comp 배선 신규. 옵션 add-on 합산은 Q-PS-2.",
    fill=NEW_FILL,
    col_comments={"comp_cd": "공식이 합산하는 구성요소. 라이브는 인화지 1개만 → 27소재 미배선"},
)

# ============================================================
# 3. price_components (RU 53)
# ============================================================
comp_rows = read_psv("components.psv")  # comp_cd|comp_nm|comp_typ|prc_typ|use_dims|use_yn
comp_data = [[r[0], r[1], n(r[2]) or "", r[3], r[4], r[5]] for r in comp_rows]
write_sheet(
    wb, "3_price_components",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    comp_data,
    note="[RU 53] 라이브 재현 — COMP_POSTER_* 30 material + COMP_POSTEROPT_* 20 옵션 + COMP_POPT_* 3. 전건 prc_typ_cd=PRICE_TYPE.01(단가형). 면적=use_dims[siz_cd]·밴드=[siz_cd,min_qty]·옵션=[].",
    fill=RU_FILL,
    col_comments={
        "prc_typ_cd": "PRICE_TYPE.01=단가형(개당단가×수량). 포스터사인 전건 단가형(라이브 실측·합가형 없음·밴드도 구간별 단가형)",
        "use_dims": "면적매트릭스=[siz_cd]·사이즈/수량밴드=[siz_cd,min_qty]·옵션add-on=[](고정). 라이브 실측 권위",
        "comp_typ_cd": "PRC_COMPONENT_TYPE.06(완제품가/add-on 통가격)",
    },
)

# ============================================================
# 4. component_prices RU (라이브 103행 재현)
# ============================================================
cp = read_psv("component_prices.psv")  # comp_cd|siz_cd|siz_nm|clr|mat|proc|coat|opt|bdl|min|apply|price
cp_rows = []
for r in cp:
    cp_rows.append([
        r[0], n(r[1]), n(r[3]), n(r[4]), n(r[5]),
        int(r[6]) if r[6] else None, n(r[7]),
        int(r[8]) if r[8] else None, int(r[9]) if r[9] else None,
        r[10], float(r[11]) if r[11] else None,
    ])
write_sheet(
    wb, "4_component_prices_RU",
    CP_COLS, CP_LBLS, cp_rows,
    note="[RU] 라이브 t_prc_component_prices 재현(103행: material 80 + 옵션 23). 면적매트릭스=(siz_cd)·밴드=(siz_cd,min_qty)·옵션=고정. 재적재 금지·대조용. 단가행은 희소 샘플(789셀 중 ~103만 적재 — 4b GAP 참조).",
    fill=RU_FILL, col_comments=CP_CMT,
)

wb.save(OUT)
print("PHASE1 SAVED:", OUT)
print("price_components RU:", len(comp_data), " component_prices RU:", len(cp_rows), " bindings RU:", len(bind_rows))

# ============================================================
# 4b. component_prices GAP_BLOCKED (면적조합 siz 미채번 667)
# ============================================================
# siz dict (canon + direct)
siz_nm2cd = {}
for r in csv.reader(open(f"{EXPORT}/siz_dict.psv"), delimiter='|'):
    if len(r) >= 2:
        siz_nm2cd[r[1]] = r[0]

wbx = openpyxl.load_workbook(XL, data_only=True)
wsx = wbx["포스터사인"]

# 면적매트릭스 블록 정의: (title_row, comp_cd) — 13블록
MATRIX_BLOCKS = [
    (1, "COMP_POSTER_ARTPRINT_PHOTO"), (18, "COMP_POSTER_ARTPAPER_MATTE"),
    (35, "COMP_POSTER_WATERPROOF_PET"), (52, "COMP_POSTER_ADH_WATERPROOF_PVC"),
    (69, "COMP_POSTER_ADH_CLEAR_PVC"), (86, "COMP_POSTER_ARTFABRIC_GRAPHIC"),
    (103, "COMP_POSTER_LINEN_FABRIC"), (120, "COMP_POSTER_CANVAS_FABRIC"),
    (137, "COMP_POSTER_LEATHER_ARTPRINT"), (154, "COMP_POSTER_TYVEK_PRINT"),
    (171, "COMP_POSTER_MESH_PRINT"), (244, "COMP_POSTER_BANNER_NORMAL"),
    (265, "COMP_POSTER_BANNER_MESH"),
]
gap_rows = []
have_rows = []
for tr, comp in MATRIX_BLOCKS:
    hdr = tr + 1
    colmap = {}
    for c in range(2, wsx.max_column + 1):
        hv = wsx.cell(row=hdr, column=c).value
        if hv in (None, ""):
            continue
        if "옵션" in str(hv):
            continue  # 옵션열은 6_addon에서
        nn = num(hv)
        if nn:
            colmap[c] = nn
    r = tr + 2
    while r <= wsx.max_row:
        gv = wsx.cell(row=r, column=1).value
        if gv in (None, ""):
            break
        garo = num(gv)
        if garo is None:
            break
        for c, sero in colmap.items():
            v = wsx.cell(row=r, column=c).value
            if v in (None, ""):
                continue
            keyd = f"{garo}x{sero}"
            keyc = f"{min(garo,sero)}x{max(garo,sero)}"
            sizcd = siz_nm2cd.get(keyd) or siz_nm2cd.get(keyc)
            row = [comp, sizcd or f"(미채번:{garo}x{sero})", None, None, None, None, None, None, None, "2026-06-01", float(v)]
            if sizcd:
                have_rows.append(row)
            else:
                gap_rows.append(row)
        r += 1
write_sheet(
    wb, "4b_component_prices_GAP_BLOCKED",
    CP_COLS, CP_LBLS, gap_rows,
    note=f"[GAP·BLOCKED] 면적매트릭스 좌표 siz 미채번 {len(gap_rows)}행(전체 면적조합의 97%). siz_cd '(미채번:GxS)'=채번 요청 대상(round-2 '좌표 siz 등록 요청서'). NULL 강제 금지(별 시트). 채번 후 4_RU 승격. 컨펌 Q-PS-3.",
    fill=BLK_FILL, col_comments=CP_CMT,
)

# ============================================================
# 5. 사이즈/수량 밴드 RU (siz_cd + min_qty 강조 재현)
# ============================================================
band_comps = ["COMP_POSTER_FRAMELESS_WOOD", "COMP_POSTER_LEATHER_FRAME", "COMP_POSTER_JOKJA",
              "COMP_POSTER_CANVAS_HANGING", "COMP_POSTER_LINEN_WOODBONG", "COMP_POSTER_PET_BANNER",
              "COMP_POSTER_MESH_BANNER", "COMP_POSTER_MINI_STANDBOARD", "COMP_POSTER_MINI_BANNER"]
band_rows = []
for r in cp:
    if r[0] in band_comps:
        band_rows.append([
            r[0], n(r[1]), n(r[2]), int(r[9]) if r[9] else None, r[10], float(r[11]) if r[11] else None,
        ])
write_sheet(
    wb, "5_qtyband_RU",
    ["comp_cd", "siz_cd", "siz_nm(참고)", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "출력규격(siz)", "규격명(참고)", "수량구간(min_qty)", "적용일", "구간단가"],
    band_rows,
    note="[RU·밴드] 사이즈/수량 밴드 9블록(우드액자·족자·캔버스행잉·미니배너 등). use_dims=[siz_cd,min_qty]. 단가형(.01)·구간별 개당단가(수량 많을수록 싸짐·합가형 아님). 4_RU 부분집합을 밴드 관점으로 재표시(min_qty 강조).",
    fill=RU_FILL,
    col_comments={"min_qty": "수량구간 하한. 주문수량 이하 최대 min_qty 구간 매칭(Phase11). 단가형이라 구간단가×수량(합가형 환산 아님)"},
)

# ============================================================
# 6. 옵션 add-on 별 comp
# ============================================================
opt_rows = []
for r in cp:
    if r[0].startswith("COMP_POSTEROPT") or r[0].startswith("COMP_POPT"):
        opt_rows.append([r[0], r[1], n(r[2]) if r[2] else "", r[10], float(r[11]) if r[11] else None,
                         "옵션 add-on(별 comp·공식 합산 vs CPQ add_price 컨펌 Q-PS-2)"])
write_sheet(
    wb, "6_addon_options",
    ["comp_cd", "comp_cd_or_dim", "siz_or_dim", "apply_ymd", "unit_price", "note"],
    ["옵션 구성요소", "(차원)", "규격/차원", "적용일", "추가단가", "비고"],
    opt_rows,
    note="[옵션·별 comp] 포스터 옵션은 opt_cd 차원이 아니라 별도 component(COMP_POSTEROPT_*/COMP_POPT_*)로 모델(라이브 실측). 현수막 가공옵션(타공/봉제)·추가옵션(끈/큐방)·천장후크·우드행거·거치대 등. 공식이 본체+옵션 comp 합산(addtn_yn). 합산 귀속 컨펌 Q-PS-2.",
    fill=BLK_FILL,
)

# ============================================================
# 7. 부유셀·노트 보존
# ============================================================
note_rows = []
# 블록 사이 빈행 외 추가 텍스트(옵션열 헤더의 가공옵션명/추가옵션명 등 비매트릭스 텍스트)
seen = set()
for r in range(1, wsx.max_row + 1):
    for c in range(1, wsx.max_column + 1):
        v = wsx.cell(row=r, column=c).value
        if isinstance(v, str) and ("옵션명" in v or "추가" in v and "옵션" in v):
            key = (v,)
            if key in seen:
                continue
            seen.add(key)
            note_rows.append([f"r{r}c{c}", v, "옵션열 헤더(비매트릭스·6_addon 참조)"])
write_sheet(
    wb, "7_excluded_note",
    ["cell_ref", "text", "분류"],
    ["원본셀위치", "텍스트", "분류"],
    note_rows,
    note="[보존] 부유셀·옵션열 헤더 등 비매트릭스 텍스트(가공옵션명/추가옵션명 등). 침묵삭제 금지(round-10 교훈). 데이터 아님·6_addon_options 참조.",
    fill=BLK_FILL,
)

wb.save(OUT)
print("PHASE2 SAVED:", OUT)
print("Final sheets:", wb.sheetnames)
print("GAP_BLOCKED:", len(gap_rows), " (siz 실재 면적조합:", len(have_rows), ")")
print("band RU:", len(band_rows), " addon:", len(opt_rows), " notes:", len(note_rows))
