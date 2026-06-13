# -*- coding: utf-8 -*-
"""round-16 후가공_박(대형) webadmin 복붙용 import.xlsx 빌더.

입력 = docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx > 후가공_박(대형) (read-only).
출력 = foil-large-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[박소형 대비 차이]
  · 동판비 = 단일 고정(5000) 아님 → 가로×세로 64셀 면적 매트릭스(11000~64000).
  · 수량구간 = 18 아님 → 13구간(10·200·500·1000·2000~10000).
  · 등급 가격 = 5등급 × 13수량 = 65행/박종군 (소형은 5×18=90).
[박소형 동형(재사용)]
  · 면적→등급 매핑 = 앱 책임 + 별 REF 시트(일반/특수 공통, 차이 0).
  · 등급 = opt_cd 차원(GRADE_A~E 코드 미등록 → 선적재 BLOCK).
  · 박가공 = 합가형(.02) 제안이나 명함박 라이브 .01 충돌 → P4 컨펌.
  · 상품 바인딩 prd_cd 미확정 → BLOCKED.
[동판 매트릭스 처리]
  · 64 좌표 중 13개만 siz_cd 등록(51 미등록) → 두 경로 병기:
    (a) 앱 룩업 + 별 REF 시트(권장·메모리 권위 중간계산=앱).
    (b) siz_cd 좌표 직접단가 64행 + 좌표 51개 선적재(BLOCK·DB관리 원할 시).
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

SRC = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "foil-large/foil-large-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
NEW_FILL = PatternFill("solid", fgColor="E2EFDA")  # 신규 구축(초록)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # 차단/제안(주황)
REF_FILL = PatternFill("solid", fgColor="FFF2CC")  # 참조(앱 입력·노랑)

APPLY_YMD = "2026-06-01"

# 64 좌표 중 라이브 siz 등록분(2026-06-13 read-only 실측) — 미등록은 선적재 제안.
REGISTERED_SIZ = {
    "110x110": "SIZ_000391", "110x170": "SIZ_000006", "170x110": "SIZ_000129",
    "30x30": "SIZ_000330", "30x70": "SIZ_000332", "50x30": "SIZ_000493",
    "50x50": "SIZ_000011", "50x70": "SIZ_000061", "50x90": "SIZ_000132",
    "70x50": "SIZ_000062", "70x70": "SIZ_000211", "90x50": "SIZ_000008",
    "90x90": "SIZ_000119",
}

# ---- 가격표 읽기 ----
sb = openpyxl.load_workbook(SRC, data_only=True)
ws = sb["후가공_박(대형)"]


def cell(r, c):
    return ws.cell(row=r, column=c).value


GRADES = ["A", "B", "C", "D", "E"]
# 면적 격자 좌표(가로 헤더 B2~I2 / 세로 라벨 A3~A10) = 30~170mm 8단계
AXIS = [30, 50, 70, 90, 110, 130, 150, 170]


def extract_band(qty_col_start_row, last_row):
    """등급별 수량 가격(K=수량, L~P=A~E) → list of (grade, min_qty, price)."""
    out = []
    for r in range(qty_col_start_row, last_row + 1):
        qty = cell(r, 11)  # K
        if qty is None:
            continue
        for gi, g in enumerate(GRADES):
            price = cell(r, 12 + gi)  # L..P
            if price is not None:
                out.append((g, int(qty), float(price)))
    return out


std_band = extract_band(16, 28)   # 일반박 K16~K28
spc_band = extract_band(36, 48)   # 특수박 K36~K48


def extract_plate():
    """동판비 가로×세로 매트릭스 B3~I10 → list of (w, h, price)."""
    out = []
    for ri, h in enumerate(AXIS):       # 세로(행) A3~A10
        for ci, w in enumerate(AXIS):   # 가로(열) B~I
            v = cell(3 + ri, 2 + ci)
            if v is not None:
                out.append((w, h, float(v)))
    return out


plate_rows = extract_plate()


def extract_area_grade(start_row):
    """면적→등급 매핑 B(start+2)~I(start+9) → list of (h, w, grade)."""
    out = []
    for ri, h in enumerate(AXIS):       # 세로 라벨
        for ci, w in enumerate(AXIS):   # 가로 헤더
            g = cell(start_row + ri, 2 + ci)
            if isinstance(g, str):
                g = g.strip()
            out.append((h, w, g))
    return out


ga_std = extract_area_grade(16)  # 일반박 면적→등급 B16~I23
ga_spc = extract_area_grade(36)  # 특수박 면적→등급 B36~I43
# 일반/특수 매핑 동일성 검산
assert [g for _, _, g in ga_std] == [g for _, _, g in ga_spc], "면적→등급 매핑 불일치"

sb.close()


# ---- 시트 작성 헬퍼 ----
def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None):
    wsx = wb.create_sheet(title)
    if note:
        wsx.cell(row=1, column=1, value=note)
        wsx.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = wsx.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = wsx.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        wsx.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(30, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            wsx.cell(row=ri, column=ci, value=val)
    wsx.freeze_panes = wsx.cell(row=data_r0, column=1)
    return wsx


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
wsr = wb.create_sheet("0_README")
readme = [
    ["round-16 후가공_박(대형) 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 후가공_박(대형)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "2단 룩업(박소형 동형) + 동판비 면적매트릭스(박소형과 다름)"],
    ["", ""],
    ["[박소형 대비 핵심 차이]", ""],
    ["  · 동판비 = 단일 고정(5000) 아님 → 가로×세로 64셀 면적매트릭스(11000~64000)", ""],
    ["  · 수량구간 = 13개(10·200·500·1000·2000~10000), 소형은 18개", ""],
    ["  · 등급 가격 = 5등급×13수량 = 65행/박종군 (소형 5×18=90)", ""],
    ["", ""],
    ["[박소형 동형(재사용)]", ""],
    ["  · 면적→등급 매핑 = 앱 책임 + 별 REF 시트(일반/특수 공통·차이 0셀)", ""],
    ["  · 등급 = opt_cd 차원(GRADE_A~E 코드 미등록→선적재 BLOCK)", ""],
    ["  · 박가공 = 합가형.02 제안이나 명함박 라이브 .01 충돌→P4 컨펌", ""],
    ["  · 상품 바인딩 prd_cd 미확정→BLOCKED", ""],
    ["", ""],
    ["[동판 매트릭스 처리 — 두 경로 병기]", ""],
    ["  · 64좌표 중 13개만 siz_cd 등록(51 미등록)", ""],
    ["  (a) 앱 룩업+별 REF 시트(권장·메모리 권위 중간계산=앱)", ""],
    ["  (b) siz_cd 좌표 직접단가 64행+좌표 51개 선적재(BLOCK·DB관리 원할 시)", ""],
    ["", ""],
    ["[라이브 미적재] 대형박 후가공 가격사슬 전면 부재 → 신규 구축(NEW)", ""],
    ["  · 명함박 comp 6종만 라이브 존재(그마저 배선 0=단절). 대형박 전용 0", ""],
    ["  · 초록 시트(_NEW)=신규 구축 / 주황(_BLOCK/_PROPOSAL)=선결차단 / 노랑(_REF)=앱 참조", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_NEW", "공식정의 1 (PRF_FOIL_LARGE 합산형)"],
    ["  1b_product_formulas_BLOCK", "상품 바인딩 0 (prd_cd 미확정·차단)"],
    ["  2_formula_components_NEW", "공식 배선 3 (동판 + 일반박STD + 특수박SPC)"],
    ["  3_price_components_NEW", "구성요소 정의 3 (동판비 / 일반박가공.02 / 특수박가공.02)"],
    ["  4_component_prices_NEW", "단가행 194 (동판64 + 일반박65 + 특수박65)"],
    ["  A1_area_grade_map_REF", "면적→등급 매핑 64 (앱 입력·가격 그릇 외·보존)"],
    ["  A2_plate_siz_proposal", "동판 좌표 siz 51개 선적재 제안(경로 b·DB관리 원할 시)"],
    ["  B1_grade_codes_proposal", "등급 코드값 5 선적재 제안 (GRADE_A~E)"],
    ["", ""],
    ["분해 무손실 검산", "가격표 가격셀 194 (동판64+일반65+특수65) = 단가행 194 일치"],
    ["면적→등급 매핑 64셀", "별 참조 시트 보존 (가격 그릇 외·앱 책임)"],
]
for ri, (a, b) in enumerate(readme, start=1):
    wsr.cell(row=ri, column=1, value=a)
    wsr.cell(row=ri, column=2, value=b)
    if ri == 1:
        wsr.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
wsr.column_dimensions["A"].width = 54
wsr.column_dimensions["B"].width = 70

# ============================================================
# 1. price_formulas (NEW)
# ============================================================
write_sheet(
    wb,
    "1_price_formulas_NEW",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    [["PRF_FOIL_LARGE", "대형박 후가공 (동판면적매트릭스+박가공)", "Y",
      "합산형: 동판비(가로×세로 면적매트릭스) + 박가공비(등급×수량 합가)"]],
    note="[NEW] 신규 공식 — 라이브 t_prc_price_formulas에 박 공식 0행. frm_typ_cd·prd_cd 컬럼 없음(라이브 실측).",
    fill=NEW_FILL,
    col_comments={
        "frm_cd": "공식 식별자 PK. 대형박 후가공 단일 공식",
        "frm_nm": "동판비(면적매트릭스) + 등급별 박가공비 합산",
    },
)

# ============================================================
# 1b. product_price_formulas (BLOCKED)
# ============================================================
write_sheet(
    wb,
    "1b_product_formulas_BLOCK",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드(미확정)", "공식코드", "적용시작일", "비고(컨펌)"],
    [["?? (컨펌필요)", "PRF_FOIL_LARGE", APPLY_YMD,
      "[BLOCKED] 박=후가공(옵션/공정). 붙는 prd_cd 미확정. 다수 상품 공용 옵션인지 특정 상품 전용인지 컨펌"]],
    note="[BLOCKED·차단] 박이 붙는 상품(prd_cd) 미확정 → NULL 강제 금지·별 시트 분리. 컨펌 후 바인딩 INSERT.",
    fill=BLK_FILL,
    col_comments={
        "prd_cd": "박 후가공이 붙는 본체 상품. 가격표만으로 미확정 → 컨펌",
        "apply_bgn_ymd": "PK 구성(prd_cd, apply_bgn_ymd)",
    },
)

# ============================================================
# 2. formula_components (NEW)
# ============================================================
write_sheet(
    wb,
    "2_formula_components_NEW",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    [
        ["PRF_FOIL_LARGE", "COMP_FOIL_LARGE_PLATE", 1, "Y"],
        ["PRF_FOIL_LARGE", "COMP_FOIL_LARGE_PROC_STD", 2, "Y"],
        ["PRF_FOIL_LARGE", "COMP_FOIL_LARGE_PROC_SPC", 3, "Y"],
    ],
    note="[NEW] 공식 배선 3 — 동판비(면적매트릭스) + 박가공(일반/특수 군 분리). addtn_yn=Y(Phase11 무시·런타임 선택값으로 1개 박가공 comp 매칭).",
    fill=NEW_FILL,
    col_comments={
        "comp_cd": "동판비(가로×세로 매트릭스) + 박가공(STD 일반박군/SPC 특수박군). 손님 박종 선택→앱이 STD/SPC 군 결정",
        "addtn_yn": "Phase11 엔진 무시. STD/SPC는 택1(동시 매칭 안 함)",
    },
)

# ============================================================
# 3. price_components (NEW)
# ============================================================
write_sheet(
    wb,
    "3_price_components_NEW",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    [
        ["COMP_FOIL_LARGE_PLATE", "대형박 동판비(면적매트릭스)", "", "PRICE_TYPE.01",
         '["siz_cd"]', "Y"],
        ["COMP_FOIL_LARGE_PROC_STD", "대형박 가공비(일반박)", "", "PRICE_TYPE.02",
         '["opt_cd","min_qty"]', "Y"],
        ["COMP_FOIL_LARGE_PROC_SPC", "대형박 가공비(특수박)", "", "PRICE_TYPE.02",
         '["opt_cd","min_qty"]', "Y"],
    ],
    note=("[NEW] 구성요소 3. 동판비=단가형.01(가로×세로 좌표 직접단가, 수량 무관 셋업성). "
          "박가공=합가형.02(수량구간 총액÷min_qty 환산). [P4 컨펌] 명함박 라이브는 .01 단가형이라 충돌. "
          "[동판 use_dims] 경로(b) siz_cd 채택 시 ['siz_cd'], 경로(a) 앱룩업 채택 시 동판 comp 자체를 앱처리."),
    fill=NEW_FILL,
    col_comments={
        "prc_typ_cd": "동판비=.01(좌표 직접단가·수량무관). 박가공=.02 합가형(가격표=수량 총액). 명함박 라이브는 .01이라 컨펌",
        "use_dims": "동판비=[siz_cd](좌표 가로×세로). 박가공=[opt_cd(등급),min_qty(수량구간)]",
        "comp_cd": "STD=일반박 6종(금유광/금무광/은유광/은무광/동박/청박) / SPC=특수박 6종(먹유광/백박/홀로그램/트윙클/적박/녹박). 가격 군 단위 차등→comp 분리",
    },
)

# ============================================================
# 4. component_prices (NEW) — 194행
# ============================================================
CP_COLS = ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt",
           "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price", "note"]
CP_LBL = ["구성요소", "사이즈(siz)", "도수(clr)", "자재(mat)", "공정(proc)", "코팅면수",
          "옵션(opt)=등급", "묶음수(bdl)", "수량구간(min_qty)", "적용일", "단가(총액)", "비고"]
cp_rows = []
# 동판비 64행 (siz_cd=좌표·미등록은 ??표기·코드선적재 후 적재)
for w, h, price in plate_rows:
    key = f"{w}x{h}"
    siz = REGISTERED_SIZ.get(key)
    note = None if siz else f"siz 미등록({key})→A2 선적재 후 적재(경로b) 또는 앱룩업(경로a)"
    cp_rows.append(["COMP_FOIL_LARGE_PLATE", siz or f"?? {key}", None, None, None, None,
                    None, None, None, APPLY_YMD, price, note])
# 일반박 65행
for g, qty, price in std_band:
    cp_rows.append(["COMP_FOIL_LARGE_PROC_STD", None, None, None, None, None,
                    f"GRADE_{g}", None, qty, APPLY_YMD, price, None])
# 특수박 65행
for g, qty, price in spc_band:
    cp_rows.append(["COMP_FOIL_LARGE_PROC_SPC", None, None, None, None, None,
                    f"GRADE_{g}", None, qty, APPLY_YMD, price, None])

write_sheet(
    wb,
    "4_component_prices_NEW",
    CP_COLS, CP_LBL, cp_rows,
    note=("[NEW] 단가행 194 (동판64 + 일반박65 + 특수박65). "
          "동판=siz_cd 좌표(미등록 51개 ??표기·A2 선적재). 박가공 opt_cd=등급(GRADE_A~E·B1 선적재). "
          "siz_cd=NULL(박가공·등급은 사이즈 아님). 안 쓰는 차원 NULL."),
    fill=NEW_FILL,
    col_comments={
        "siz_cd": "동판비=좌표 가로×세로(REGISTERED 13개만 코드·51개 ??=A2 선적재). 박가공=NULL(등급은 사이즈 아님)",
        "opt_cd": "박가공 등급 차원(신설 opt_cd 첫 실사용). GRADE_A~E. 코드 미등록→B1 선적재 후 적재",
        "min_qty": "박가공 수량구간 시작(10·200·500·1000·2000~10000=13구간). 주문수량 이하 최대 min_qty 행 매칭. 동판=NULL(수량무관)",
        "unit_price": "동판=좌표별 단가. 박가공=합가형 수량구간 총액(예: 일반박 GRADE_A 1000매=75,000원 전체)→엔진 ÷min_qty",
        "proc_cd": "신설 공정 차원. 박종 군 차등은 comp 분리로 처리→proc_cd NULL",
    },
)

# ============================================================
# A1. 면적→등급 매핑표 (REF·앱 책임·가격 그릇 외) — 64셀
# ============================================================
ga_rows = []
for h, w, grade in ga_std:
    ga_rows.append([h, w, grade if grade is not None else None,
                    "미정의(공란)" if grade is None else None])
write_sheet(
    wb,
    "A1_area_grade_map_REF",
    ["height_mm", "width_mm", "grade", "note"],
    ["세로(mm)", "가로(mm)", "등급(A~E)", "비고"],
    ga_rows,
    note=("[REF·앱 책임] 면적→등급 매핑 64 (2단 룩업 1단계). 가격 그릇(t_prc_*)에 직접 안 들어감 — "
          "앱이 작업 가로×세로→등급 변환. 일반박/특수박 공통(차이 0셀). 8×8 대칭격자(30~170mm). "
          "DDL 원할 시 t_prc_foil_area_grade 신설 입력으로 사용."),
    fill=REF_FILL,
    col_comments={
        "grade": "(세로,가로)→등급. opt_cd로 가격 룩업 시 이 등급 사용. 박소형과 별개 매핑(대형 전용)",
    },
)

# ============================================================
# A2. 동판 좌표 siz 선적재 제안 (경로b·미등록 좌표만)
# ============================================================
plate_siz_rows = []
seen = set()
for w, h, _ in plate_rows:
    key = f"{w}x{h}"
    if key in REGISTERED_SIZ or key in seen:
        continue
    seen.add(key)
    plate_siz_rows.append([f"?? (채번필요)", key, w, h, "Y",
                           "동판 좌표 siz 미등록. 경로(b) DB관리 채택 시 선적재(채번=MAX+1). 경로(a) 앱룩업 시 불필요"])
write_sheet(
    wb,
    "A2_plate_siz_proposal",
    ["siz_cd", "siz_nm", "width_mm", "height_mm", "use_yn", "note"],
    ["사이즈코드(채번)", "사이즈명(가로x세로)", "가로(mm)", "세로(mm)", "사용여부", "비고"],
    plate_siz_rows,
    note=("[PROPOSAL·경로b] 동판 좌표 siz 미등록 51개 선적재 제안 (t_siz_sizes). "
          "동판비를 siz_cd 차원으로 DB관리할 경우만 필요. 권장 경로(a)=앱룩업이면 불필요. "
          "컬럼=라이브 t_siz_sizes 참조(siz_cd 채번·siz_nm 가로x세로 형식)."),
    fill=BLK_FILL,
    col_comments={
        "siz_cd": "좌표 사이즈 PK. 동판 component_prices siz_cd가 참조. 채번=MAX+1(코드전략)",
        "siz_nm": "가로x세로 형식(라이브 73x98 등 동일 컨벤션)",
    },
)

# ============================================================
# B1. 등급 코드값 선적재 제안 (PROPOSAL·차단)
# ============================================================
grade_code_rows = [
    [f"GRADE_{g}", f"박 면적등급 {g}", "FOIL_GRADE", i + 1, "Y",
     "대형박 후가공 면적등급 opt_cd 차원값(신규·소형박과 코드 공유)"]
    for i, g in enumerate(GRADES)
]
write_sheet(
    wb,
    "B1_grade_codes_proposal",
    ["cod_cd", "cod_nm", "upr_cod_cd", "disp_seq", "use_yn", "note"],
    ["코드값", "코드명", "상위코드그룹", "표시순서", "사용여부", "비고"],
    grade_code_rows,
    note=("[PROPOSAL·차단] 등급 코드값 GRADE_A~E 신규 등록 제안 (t_cod_base_codes). "
          "opt_cd 차원에 담을 등급. 소형박 B1과 동일 코드(공유). 코드 선적재 후 component_prices 적재. "
          "컬럼=라이브 t_cod_base_codes 실측(cod_cd·cod_nm·upr_cod_cd·disp_seq·use_yn·note)."),
    fill=BLK_FILL,
    col_comments={
        "cod_cd": "등급 코드값 PK. opt_cd 단가행이 참조. 소형박과 공유",
        "upr_cod_cd": "상위 코드그룹 FOIL_GRADE(신규 그룹) — 코드그룹도 등록 필요",
    },
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("plate rows:", len(plate_rows), "(expect 64)")
print("std_band:", len(std_band), "spc_band:", len(spc_band), "(expect 65 each)")
print("component_prices rows:", len(cp_rows), "(expect 194)")
print("area_grade_map rows:", len(ga_rows), "(expect 64)")
print("plate_siz_proposal rows:", len(plate_siz_rows), "(expect 51)")
