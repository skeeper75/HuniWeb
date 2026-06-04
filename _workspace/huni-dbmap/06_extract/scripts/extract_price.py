#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
가격표 워크북 L1 추출 — 판걸이수 + 출력소재(IMPORT).
상품마스터 13시트와 동일한 무손실·8정보축 원칙. 단 헤더는 단일행(+IMPORT 다단 fragment).

[축 분류 — HARD 순서원칙] 상품정보(상품정보 토대) vs 가격정보(round-2 이연):
  판걸이수:  상품정보=A사이즈옵션명·B재단·C블리드·D작업·E상품·F판걸이·J인쇄가능영역·K원형아이마크·L격자아이마크
            가격정보=G디지털인쇄(국4절)·H디지털인쇄(3절)·I (round-2 이연 라벨)
  IMPORT:   상품정보=A대분류·B중분류·C종이명·D평량·E약어·F구매정보·G전지·K종이사이즈 + L~Z 상품컬럼(●매트릭스)
            가격정보=H연당가·I가격(국4절)·J가격(3절) (round-2 이연 라벨)

무손실: 전 컬럼·전 행·●·단가 모두 보존(단가는 "가격정보" 라벨). 빈셀 보존, 유효0 보존.
A/B 대분류·중분류 ffill(IMPORT ragged 세로블록 상품귀속). 판걸이수 E상품 ffill 금지(행별 상이).

read-only. 산출만 06_extract/.
"""
import os, sys, csv, json, re, zipfile
import openpyxl
from openpyxl.utils import get_column_letter

ROOT = "/Users/innojini/Dev/HuniWeb"
XLSX = os.path.join(ROOT, "docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx")
OUT = os.path.join(ROOT, "_workspace/huni-dbmap/06_extract")

sys.path.insert(0, os.path.dirname(__file__))
from extract_l1 import load_threaded_comments


def cell_meta_for(ws_f, ws, r, c, comments):
    """8축 셀메타 — 배경색/글자색/코멘트/수식/하이퍼링크/★. (행/열숨김은 시트레벨 별도)"""
    ref = get_column_letter(c) + str(r)
    cell = ws_f.cell(r, c)
    fg = cell.fill.fgColor.rgb if (cell.fill and cell.fill.patternType) else None
    if fg is not None and not isinstance(fg, str):
        fg = str(fg)
    if fg in ("00000000", "FFFFFFFF"):
        fg = None
    font_rgb = None
    try:
        if cell.font and cell.font.color and cell.font.color.rgb:
            fr = cell.font.color.rgb
            font_rgb = fr if isinstance(fr, str) else str(fr)
            if font_rgb in ("FF000000", "00000000"):
                font_rgb = None
    except Exception:
        font_rgb = None
    cmt = comments.get(ref)
    val = ws.cell(r, c).value
    has_star = bool(val and isinstance(val, str) and "★" in val)
    fcell = cell.value
    is_formula = bool(isinstance(fcell, str) and fcell.startswith("="))
    formula = fcell if is_formula else ""
    hyperlink = ""
    try:
        if cell.hyperlink is not None and cell.hyperlink.target:
            hyperlink = cell.hyperlink.target
    except Exception:
        hyperlink = ""
    m = {}
    if fg:
        m["fill_rgb"] = fg
    if font_rgb:
        m["font_rgb"] = font_rgb
    if cmt:
        m["comment_author"] = cmt["author"]
        m["comment_text"] = cmt["text"]
    if has_star:
        m["has_constraint_star"] = True
    if is_formula:
        m["is_formula"] = True
        m["formula"] = formula
    if hyperlink:
        m["hyperlink"] = hyperlink
    return m


def col_hidden_map(ws, maxc):
    out = {}
    for c in range(1, maxc + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        out[c] = bool(dim.hidden) if dim is not None else False
    return out


def row_hidden(ws, r):
    dim = ws.row_dimensions.get(r)
    return bool(dim.hidden) if dim is not None else False


def norm(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def cellval(v):
    """CSV 저장값 — verify.norm 과 동일 정규화(정수형 float -> int 표기, 원형 보존).
    공백 None -> ''. 타입 강제 없이 정수형 float 표기만 일치시킴(엑셀 1.0 = 1)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


# ============================================================
# 판걸이수
# ============================================================
def extract_pangeori():
    sheet = "판걸이수"
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    ws, ws_f = wb[sheet], wb_f[sheet]
    comments = load_threaded_comments(XLSX, sheet)
    # 실데이터 컬럼 A..L (12) — M+ 는 헤더/값 전무
    MAXC = 12
    HEADERS = {1: "사이즈옵션명", 2: "재단사이즈", 3: "블리드", 4: "작업사이즈",
               5: "상품", 6: "판걸이", 7: "디지털인쇄(국4절)", 8: "디지털인쇄(3절)",
               9: "col_I", 10: "인쇄가능영역", 11: "원형아이마크영역(완칼)",
               12: "격자아이마크영역(반칼)"}
    # 축분류
    AXIS = {1: "product", 2: "product", 3: "product", 4: "product", 5: "product",
            6: "product", 7: "price", 8: "price", 9: "price",
            10: "product", 11: "product", 12: "product"}
    ch = col_hidden_map(ws, MAXC)
    # 데이터행 = 2..last nonempty
    last = 0
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, MAXC + 1)):
            last = r
    data_rows = list(range(2, last + 1))

    records, meta_rows = [], []
    for r in data_rows:
        cols = {}
        cmeta = {}
        for c in range(1, MAXC + 1):
            name = HEADERS[c]
            v = ws.cell(r, c).value
            cols[name] = cellval(v)
            m = cell_meta_for(ws_f, ws, r, c, comments)
            if ch.get(c):
                m["col_hidden"] = True
            if m:
                cmeta[name] = m
                meta_rows.append((r, get_column_letter(c) + str(r), name, m))
        records.append({
            "row_seq": r,
            "join_key": cols["사이즈옵션명"],
            "_row_hidden": "true" if row_hidden(ws, r) else "false",
            "cols": cols, "cmeta": cmeta,
        })
    # write l1 csv
    colnames = [HEADERS[c] for c in range(1, MAXC + 1)]
    out_csv = os.path.join(OUT, "pangeori-l1.csv")
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row_seq", "join_key", "_row_hidden"] + colnames + ["cell_meta_json"])
        for rec in records:
            row = [sheet, rec["row_seq"], rec["join_key"], rec["_row_hidden"]]
            row += [rec["cols"][n] for n in colnames]
            row.append(json.dumps(rec["cmeta"], ensure_ascii=False) if rec["cmeta"] else "")
            w.writerow(row)
    # meta csv
    out_meta = os.path.join(OUT, "pangeori-l1-meta.csv")
    with open(out_meta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row_seq", "ref", "col", "axis", "meta_json"])
        for c in range(1, MAXC + 1):
            if ch.get(c):
                w.writerow([sheet, "", get_column_letter(c) + "1", HEADERS[c],
                            AXIS[c], json.dumps({"hidden_column": True}, ensure_ascii=False)])
        for (r, ref, name, m) in meta_rows:
            cidx = next(k for k, v in HEADERS.items() if v == name)
            w.writerow([sheet, r, ref, name, AXIS[cidx], json.dumps(m, ensure_ascii=False)])
    # axis summary
    axis_map = {HEADERS[c]: AXIS[c] for c in range(1, MAXC + 1)}
    summary = {
        "sheet": sheet, "data_rows": len(data_rows), "fields": MAXC,
        "first_row": data_rows[0], "last_row": data_rows[-1],
        "hidden_rows": sum(1 for rec in records if rec["_row_hidden"] == "true"),
        "hidden_cols": [HEADERS[c] for c in range(1, MAXC + 1) if ch.get(c)],
        "comments": len(comments), "meta_cells": len(meta_rows),
        "axis_product_cols": [n for n, a in axis_map.items() if a == "product"],
        "axis_price_cols": [n for n, a in axis_map.items() if a == "price"],
    }
    # verify roundtrip + nonempty
    v = verify_simple(ws, data_rows, MAXC, out_csv, colnames)
    return summary, v, axis_map


# ============================================================
# 출력소재(IMPORT)
# ============================================================
def extract_import():
    sheet = "출력소재(IMPORT)"
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    ws, ws_f = wb[sheet], wb_f[sheet]
    comments = load_threaded_comments(XLSX, sheet)
    MAXC = 26  # A..Z
    # 헤더 = row1 (+ row2/row3 fragment 일부) → composite
    HEAD = {}
    for c in range(1, MAXC + 1):
        h1 = ws.cell(1, c).value
        h2 = ws.cell(2, c).value
        h3 = ws.cell(3, c).value
        parts = [str(x).strip() for x in (h1, h2, h3) if x not in (None, "")]
        HEAD[c] = " :: ".join(parts) if parts else get_column_letter(c)
    # 축분류: A-G,K = product(자재); H,I,J = price; L..Z(12..26) = product(상품매트릭스)
    AXIS = {}
    for c in range(1, MAXC + 1):
        if c in (8, 9, 10):           # H,I,J
            AXIS[c] = "price"
        else:
            AXIS[c] = "product"
    PRODUCT_MATRIX_COLS = list(range(12, MAXC + 1))  # L..Z
    PAPER_DESC_COLS = [1, 2, 3, 4, 5, 6, 7, 11]      # A,B,C,D,E,F,G,K
    PRICE_COLS = [8, 9, 10]                          # H,I,J
    FFILL_COLS = [1, 2]                              # 대분류, 중분류 ragged

    ch = col_hidden_map(ws, MAXC)
    last = 0
    for r in range(4, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, MAXC + 1)):
            last = r
    data_rows = list(range(4, last + 1))

    ffill_state = {c: "" for c in FFILL_COLS}
    records, meta_rows, matrix_rows = [], [], []
    for r in data_rows:
        for c in FFILL_COLS:
            raw = ws.cell(r, c).value
            if raw is not None and str(raw).strip() != "":
                ffill_state[c] = raw
        cols, cmeta = {}, {}
        ffilled = False
        for c in range(1, MAXC + 1):
            name = HEAD[c]
            v = ws.cell(r, c).value
            if c in FFILL_COLS:
                if v is not None and str(v).strip() != "":
                    cols[name] = cellval(v)
                else:
                    cols[name] = cellval(ffill_state[c]) if ffill_state[c] != "" else ""
                    if ffill_state[c] != "":
                        ffilled = True
            else:
                cols[name] = cellval(v)
            m = cell_meta_for(ws_f, ws, r, c, comments)
            if ch.get(c):
                m["col_hidden"] = True
            if m:
                cmeta[name] = m
                meta_rows.append((r, get_column_letter(c) + str(r), name, m))
        paper_name = cols[HEAD[3]]  # 종이명
        records.append({
            "row_seq": r, "paper_name": paper_name,
            "_anchor_ffilled": "true" if ffilled else "false",
            "_row_hidden": "true" if row_hidden(ws, r) else "false",
            "cols": cols, "cmeta": cmeta,
        })
        # unpivot ● matrix (long/tidy): paper row × product col
        for c in PRODUCT_MATRIX_COLS:
            v = ws.cell(r, c).value
            if v not in (None, ""):
                matrix_rows.append({
                    "row_seq": r,
                    "daebunryu": cols[HEAD[1]],
                    "jungbunryu": cols[HEAD[2]],
                    "paper_name": paper_name,
                    "pyeongryang": cols[HEAD[4]],
                    "yakeo": cols[HEAD[5]],
                    "product_col": HEAD[c],
                    "mark": str(v),
                })
    # write paper l1
    colnames = [HEAD[c] for c in range(1, MAXC + 1)]
    out_csv = os.path.join(OUT, "import-paper-l1.csv")
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row_seq", "paper_name", "_anchor_ffilled", "_row_hidden"]
                    + colnames + ["cell_meta_json"])
        for rec in records:
            row = [sheet, rec["row_seq"], rec["paper_name"],
                   rec["_anchor_ffilled"], rec["_row_hidden"]]
            row += [rec["cols"][n] for n in colnames]
            row.append(json.dumps(rec["cmeta"], ensure_ascii=False) if rec["cmeta"] else "")
            w.writerow(row)
    # meta csv
    out_meta = os.path.join(OUT, "import-paper-l1-meta.csv")
    with open(out_meta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "row_seq", "ref", "col", "axis", "meta_json"])
        for c in range(1, MAXC + 1):
            if ch.get(c):
                w.writerow([sheet, "", get_column_letter(c) + "1", HEAD[c],
                            AXIS[c], json.dumps({"hidden_column": True}, ensure_ascii=False)])
        for (r, ref, name, m) in meta_rows:
            cidx = next(k for k, v in HEAD.items() if v == name)
            w.writerow([sheet, r, ref, name, AXIS[cidx], json.dumps(m, ensure_ascii=False)])
    # write unpivot matrix
    out_mat = os.path.join(OUT, "import-paper-matrix-long.csv")
    with open(out_mat, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["row_seq", "daebunryu", "jungbunryu", "paper_name",
                    "pyeongryang", "yakeo", "product_col", "mark"])
        for mr in matrix_rows:
            w.writerow([mr["row_seq"], mr["daebunryu"], mr["jungbunryu"], mr["paper_name"],
                        mr["pyeongryang"], mr["yakeo"], mr["product_col"], mr["mark"]])

    axis_map = {HEAD[c]: AXIS[c] for c in range(1, MAXC + 1)}
    product_headers = [HEAD[c] for c in PRODUCT_MATRIX_COLS]
    summary = {
        "sheet": sheet, "data_rows": len(data_rows), "fields": MAXC,
        "first_row": data_rows[0], "last_row": data_rows[-1],
        "hidden_rows": sum(1 for rec in records if rec["_row_hidden"] == "true"),
        "hidden_cols": [HEAD[c] for c in range(1, MAXC + 1) if ch.get(c)],
        "comments": len(comments), "meta_cells": len(meta_rows),
        "mark_count": len(matrix_rows),
        "product_columns": product_headers,
        "product_column_count": len(product_headers),
        "daebunryu_groups": sorted(set(r["daebunryu"] for r in matrix_rows)),
        "paper_desc_cols": [HEAD[c] for c in PAPER_DESC_COLS],
        "price_cols_deferred": [HEAD[c] for c in PRICE_COLS],
    }
    v = verify_simple(ws, data_rows, MAXC, out_csv, colnames,
                      ffill_cols_names=[HEAD[c] for c in FFILL_COLS])
    v["matrix_marks"] = len(matrix_rows)
    return summary, v, axis_map, product_headers


# ============================================================
# 단순 9게이트 verify (단일행 헤더용)
# ============================================================
def verify_simple(ws, data_rows, maxc, csv_path, colnames, ffill_cols_names=None):
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    ws_f = wb_f[ws.title]
    ffill_cols_names = set(ffill_cols_names or [])
    # orig non-empty (strip-aware)
    orig_cells = {}
    orig_nonempty = 0
    for r in data_rows:
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                orig_cells[(r, c)] = norm(v)
                orig_nonempty += 1
    with open(csv_path, encoding="utf-8-sig") as f:
        rd = csv.DictReader(f)
        ext_rows = list(rd)
    ext_map = {}
    for er in ext_rows:
        rseq = int(er["row_seq"])
        for n in colnames:
            ext_map[(rseq, n)] = er.get(n, "")
    preserved = missing = mismatch = 0
    for (r, c), oval in orig_cells.items():
        name = colnames[c - 1]
        ev = ext_map.get((r, name))
        if ev is None:
            missing += 1
        elif ev == oval:
            preserved += 1
        else:
            mismatch += 1
    # roundtrip
    rt_diff = 0
    rt_samples = []
    for er in ext_rows:
        rseq = int(er["row_seq"])
        for c in range(1, maxc + 1):
            name = colnames[c - 1]
            ext_val = er.get(name, "")
            orig_val = norm(ws.cell(rseq, c).value)
            if name in ffill_cols_names and orig_val.strip() == "" and ext_val != "":
                continue
            if ext_val != orig_val:
                rt_diff += 1
                if len(rt_samples) < 10:
                    rt_samples.append({"row": rseq, "col": get_column_letter(c),
                                       "orig": orig_val, "ext": ext_val})
    # hidden/formula/hyperlink originals
    orig_formula = orig_hyper = 0
    for r in data_rows:
        for c in range(1, maxc + 1):
            fv = ws_f.cell(r, c).value
            if isinstance(fv, str) and fv.startswith("="):
                orig_formula += 1
            try:
                if ws_f.cell(r, c).hyperlink and ws_f.cell(r, c).hyperlink.target:
                    orig_hyper += 1
            except Exception:
                pass
    gates = {
        "row_count": {"orig": len(data_rows), "ext": len(ext_rows),
                      "pass": len(data_rows) == len(ext_rows)},
        "col_coverage": {"maxcol": maxc, "ext_fields": len(colnames),
                         "pass": len(colnames) == maxc},
        "nonempty_preservation": {"orig": orig_nonempty, "preserved": preserved,
                                  "missing": missing, "mismatch": mismatch,
                                  "rate_pct": round(preserved / orig_nonempty * 100, 4) if orig_nonempty else 100.0,
                                  "pass": missing == 0 and mismatch == 0},
        "roundtrip": {"diff": rt_diff, "pass": rt_diff == 0, "samples": rt_samples},
        "orig_formula_cells": orig_formula,
        "orig_hyperlink_cells": orig_hyper,
    }
    gates["overall"] = all(g.get("pass", True) for g in gates.values() if isinstance(g, dict))
    return gates


def main():
    out = {}
    ps, pv, pax = extract_pangeori()
    out["판걸이수"] = {"summary": ps, "verify": pv, "axis_map": pax}
    print(f"[pangeori] rows={ps['data_rows']} fields={ps['fields']} "
          f"hidden_cols={ps['hidden_cols']} comments={ps['comments']} "
          f"| VERIFY={'PASS' if pv['overall'] else 'FAIL'}")
    if not pv["overall"]:
        print("   ", json.dumps({k: v for k, v in pv.items() if isinstance(v, dict) and not v.get("pass", True)}, ensure_ascii=False)[:300])

    isum, iv, iax, iprod = extract_import()
    out["출력소재(IMPORT)"] = {"summary": isum, "verify": iv, "axis_map": iax,
                              "product_headers": iprod}
    print(f"[import] rows={isum['data_rows']} fields={isum['fields']} "
          f"marks={isum['mark_count']} product_cols={isum['product_column_count']} "
          f"hidden_cols={isum['hidden_cols']} comments={isum['comments']} "
          f"| VERIFY={'PASS' if iv['overall'] else 'FAIL'}")
    if not iv["overall"]:
        print("   ", json.dumps({k: v for k, v in iv.items() if isinstance(v, dict) and not v.get("pass", True)}, ensure_ascii=False)[:300])

    with open(os.path.join(OUT, "_price-extract-summary.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
