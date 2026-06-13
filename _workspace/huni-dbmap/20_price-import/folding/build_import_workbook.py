# -*- coding: utf-8 -*-
"""
round-16 접지옵션 webadmin 복붙용 import.xlsx 빌더.
입력 = 가격표 접지옵션 시트(/tmp/fold_export/pricetable.json) + 라이브 read-only 덤프(/tmp/fold_export/*.psv).
출력 = folding-import.xlsx (테이블별 시트, 1행=DB컬럼, 2행=한글라벨).
분해 기준 = Phase11 엔진 매칭(헤더 "/" 개별 접지옵션 = proc_cd 차원 개별분해·collapse 금지).
DB 미적재. 라이브 기존 적재(RU 재현) + 개별분해 그릇(proc_cd 명시) 병기.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

EXP = "/tmp/fold_export"
OUT = "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/folding/folding-import.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
RU_FILL = PatternFill("solid", fgColor="E2EFDA")   # 라이브 기존 재현
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")  # 개별분해 신설 그릇
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")  # 가격사슬 단절 경고


def read_psv(name):
    rows = []
    with open(f"{EXP}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if line:
                rows.append(line.split("|"))
    return rows


def write_sheet(wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None, row_fills=None):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
    for ri, row in enumerate(data_rows):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=data_r0 + ri, column=ci, value=val)
            if row_fills and ri < len(row_fills) and row_fills[ri]:
                c.fill = row_fills[ri]
    for ci in range(1, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16
    return ws


# ───────── 매핑 권위 (가격표 헤더 ↔ 라이브 PROC ↔ 라이브 comp) ─────────
# 블록1 카드접지: 단가컬럼 → (라이브 comp, [개별 접지옵션 PROC ...])
B1MAP = [
    ("2단", "COMP_FOLD_CARD_2H", [("PROC_000065", "2단가로접지"), ("PROC_000066", "2단세로접지")]),
    ("3단", "COMP_FOLD_CARD_3H", [("PROC_000067", "3단가로접지"), ("PROC_000068", "3단세로접지")]),
    ("6단", "COMP_FOLD_CARD_6CR", [("PROC_000073", "6단오시접지"), ("PROC_000074", "6단미싱접지")]),
]
# 블록2 리플렛접지: 단일 접지옵션 4개
B2MAP = [
    ("반접지", "COMP_FOLD_LEAF_HALF", "PROC_000056", "접지"),       # 반접지=단순 접지(half)
    ("3단접지", "COMP_FOLD_LEAF_3FOLD", "PROC_000060", "3단접지"),
    ("4단병풍접지", "COMP_FOLD_LEAF_4GATE", "PROC_000071", "병풍접지"),
    ("4단대문접지", "COMP_FOLD_LEAF_4ACC", None, "4단대문접지"),     # 라이브 PROC 없음(대문접지 미등록)
]

pt = json.load(open(f"{EXP}/pricetable.json"))
b1, b2 = pt["b1"], pt["b2"]  # b1: (qty,2단,3단,6단)  b2: (qty,반,3단,4병풍,4대문)
APPLY = "20260601"

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ════════ 시트1 price_formulas ════════
pf = read_psv("price_formulas.psv")
write_sheet(
    wb, "1_price_formulas",
    ["frm_cd", "frm_nm", "note", "use_yn"],
    ["공식코드", "공식명", "비고", "사용여부(Y/N)"],
    pf, fill=RU_FILL,
    note="라이브 기존 재현(녹색=RU). 접지옵션 공급 공식 3종. frm_typ_cd 컬럼은 라이브 부재(공식정의에 유형 없음).",
    col_comments={
        "frm_cd": "공식 식별코드. PRF_FOLD_SUM=접지 단독 합산형 / PRF_DGP_C·E=디지털인쇄 합산형(접지 포함)",
        "use_yn": "Y=활성",
    },
)

# ════════ 시트1b product_price_formulas (바인딩) ════════
ppf = read_psv("product_price_formulas.psv")
write_sheet(
    wb, "1b_product_price_formulas",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일", "비고"],
    ppf, fill=RU_FILL,
    note="라이브 기존 재현. 공식정의(1번)와 별 테이블 = 상품↔공식 바인딩. PK=(prd_cd, apply_bgn_ymd).",
    col_comments={"prd_cd": "접지 사용 상품(접지카드 27/29·인쇄배경지 43~45·접지리플렛 48)"},
)

# ════════ 시트2 formula_components (배선) ════════
fc = read_psv("formula_components.psv")
# 가격사슬 단절 표식: 3H·6CR은 배선 0
warn_rows = []
for r in fc:
    warn_rows.append(None)
write_sheet(
    wb, "2_formula_components",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    fc, fill=RU_FILL, row_fills=warn_rows,
    note="라이브 기존 배선 재현. 🔴 가격사슬 단절: CARD_3H·CARD_6CR은 단가행 적재됐으나 배선 0(어느 공식에도 미연결=엔진 조회불가). 3단접지카드(PRD_000029) 바인딩은 PRF_DGP_E이나 E엔 LEAF만 배선됨 → 카드 3단/6단 미연결.",
    col_comments={
        "comp_cd": "구성요소. CARD_2H만 PRF_DGP_C·PRF_FOLD_SUM에 배선. 3H/6CR/LEAF는 §단절 참조",
        "addtn_yn": "Y=합산(Phase11은 무시·런타임 결정)",
    },
)

# ════════ 시트3 price_components (구성요소 정의) ════════
pc = read_psv("price_components.psv")
write_sheet(
    wb, "3_price_components",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(단가/합가)", "사용차원(jsonb)", "사용여부"],
    pc, fill=RU_FILL,
    note="라이브 기존 재현. 7 구성요소 전건 PRICE_TYPE.01=단가형(장당가). 근거=가격표 수량↑→장당가↓ 체감(round-trip 일치). use_dims=[min_qty]만(siz/clr/mat 미사용).",
    col_comments={
        "prc_typ_cd": "PRICE_TYPE.01=단가형(장당가×주문수량). 합가형(.02) 아님 — 셀이 장당가이지 구간총액이 아님",
        "use_dims": '["min_qty"]=수량구간만 사용. siz/clr/mat/coat/proc/opt 차원 NULL 와일드카드',
        "comp_typ_cd": "PRC_COMPONENT_TYPE.04=후가공",
    },
)

# ════════ 시트4 component_prices — 라이브 재현(RU·collapse 7 comp) ════════
cp = read_psv("component_prices.psv")
# psv 컬럼순: comp_cd,apply_ymd,siz_cd,clr_cd,mat_cd,proc_cd,coat_side_cnt,opt_cd,bdl_qty,min_qty,unit_price
write_sheet(
    wb, "4_component_prices_RU",
    ["comp_cd", "apply_ymd", "siz_cd", "clr_cd", "mat_cd", "proc_cd",
     "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "unit_price"],
    ["구성요소코드", "적용일", "사이즈", "색상", "자재", "공정",
     "코팅면수", "옵션", "묶음수", "최소수량(구간)", "단가(원)"],
    cp, fill=RU_FILL,
    note="라이브 기존 적재 그대로 재현(336행=7 단가컬럼×48구간). 단가컬럼당 1 comp(2H/3H/6CR collapse)·proc_cd 비움. 가격표 round-trip 일치. min_qty만 채움·나머지 NULL.",
    col_comments={
        "min_qty": "수량구간 하한. 주문수량 이하 최대 min_qty 행 매칭. 1~100000 48구간",
        "unit_price": "장당가(원). 수량↑→단가↓ 체감(단가형)",
        "proc_cd": "라이브=NULL(collapse 적재). 개별분해 그릇은 4b 시트 참조",
    },
)

# ════════ 시트4b component_prices — 개별분해 그릇(proc_cd 명시·collapse 금지) ════════
# 헤더 "/" 6+4=10 개별 접지옵션을 proc_cd 차원으로 분해. 같은 단가 공유는 행 복제.
DEC = []        # 분해행
DEC_FILLS = []
# 블록1: 각 단가컬럼 → 2 PROC 복제 (qty,2단,3단,6단) idx 1,2,3
for qrow in b1:
    qty = qrow[0]
    for ci, (lbl, comp, procs) in enumerate(B1MAP, start=1):
        price = qrow[ci]
        for proc_cd, proc_nm in procs:
            DEC.append([comp, APPLY, "", "", "", proc_cd, "", "", "", qty, price])
            DEC_FILLS.append(NEW_FILL)
# 블록2: 각 단가컬럼 → 1 PROC (qty,반,3단,4병풍,4대문) idx 1,2,3,4
for qrow in b2:
    qty = qrow[0]
    for ci, (lbl, comp, proc_cd, proc_nm) in enumerate(B2MAP, start=1):
        price = qrow[ci]
        if proc_cd is None:
            DEC.append([comp, APPLY, "", "", "", "", "", "", "", qty, price])  # 대문=PROC 미등록→NULL
        else:
            DEC.append([comp, APPLY, "", "", "", proc_cd, "", "", "", qty, price])
        DEC_FILLS.append(NEW_FILL)
write_sheet(
    wb, "4b_component_prices_DECOMP",
    ["comp_cd", "apply_ymd", "siz_cd", "clr_cd", "mat_cd", "proc_cd",
     "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "unit_price"],
    ["구성요소코드", "적용일", "사이즈", "색상", "자재", "공정",
     "코팅면수", "옵션", "묶음수", "최소수량(구간)", "단가(원)"],
    DEC, fill=NEW_FILL, row_fills=DEC_FILLS,
    note="개별분해 그릇(노랑=신설 제안). 헤더 '/' 개별 접지옵션을 proc_cd 차원으로 명시(collapse 금지·교훈③). 블록1 6 PROC(2단가로/세로·3단가로/세로·6단오시/미싱)+블록2 4 PROC. 같은 단가 공유는 행 복제+proc_cd 구분. ⚠ 동시매칭 주의: 한 comp 안 같은 (proc_cd,min_qty) 1행만(검증 통과). 4대문접지=라이브 PROC 미등록→proc_cd NULL(컨펌 Q-FOLD-2).",
    col_comments={
        "proc_cd": "개별 접지옵션 PROC. 손님이 '2단세로접지' 선택→proc_cd=PROC_000066 매칭. 라이브 RU는 NULL(collapse)이라 차원 미활용 — 개별분해는 제안 그릇",
    },
)

wb.save(OUT)
print("saved", OUT)
print("RU component_prices rows:", len(cp))
print("DECOMP rows:", len(DEC))
print("sheets:", wb.sheetnames)
