#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
합판도무송스티커 → Phase11 가격엔진 그릇 import.xlsx 빌더 (round-16)
- 테이블당 1시트, 1행=DB컬럼명(영문), 2행=한글 라벨/도메인 주석, 3행~=데이터
- 라이브 t_prc_* 컬럼 1:1 (information_schema 실측 2026-06-13)
- 소재 2축 → 6 개별 mat_cd 분해, 합가형(.02), 형상=siz_cd 37종
- DB 미적재 — 복붙용 그릇만 생성
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SRC = 'docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx'
OUT = '_workspace/huni-dbmap/20_price-import/plywood-domusong/plywood-domusong-import.xlsx'
APPLY = '20260527'   # 가격표 버전 260527
COMP = 'COMP_GANGPAN_PRINT'
FRM = 'PRF_GANGPAN_FIXED'
PRD = 'PRD_000066'

# 소재 2축 → 6 개별 mat_cd (라이브 실측)
MAT_COAT = ['MAT_000084', 'MAT_000155', 'MAT_000156']   # 비코팅/무광코팅/유광코팅 (코팅축, 홀수컬럼)
MAT_DEDRON = ['MAT_000153', 'MAT_000170', 'MAT_000171']  # 유포/투명데드롱/은데드롱 (데드롱축, 짝수컬럼)
MAT_NM = {
    'MAT_000084': '비코팅스티커', 'MAT_000155': '무광코팅스티커', 'MAT_000156': '유광코팅스티커',
    'MAT_000153': '유포스티커', 'MAT_000170': '투명데드롱스티커', 'MAT_000171': '은데드롱스티커',
}

# 형상 헤더 → siz_cd (라이브 siz_nm 1:1 대조)
SIZ = {
    # 원형 11
    '원형 10mm': 'SIZ_000501', '원형 15mm': 'SIZ_000502', '원형 20mm': 'SIZ_000503',
    '원형 25mm': 'SIZ_000504', '원형 30mm': 'SIZ_000505', '원형 35mm': 'SIZ_000422',
    '원형 40mm': 'SIZ_000506', '원형 45mm': 'SIZ_000507', '원형 50mm': 'SIZ_000508',
    '원형 55mm': 'SIZ_000509', '원형 60mm': 'SIZ_000510',
    # 정사각 12
    '정사각 10 x 10mm': 'SIZ_000212', '정사각 15 x 15mm': 'SIZ_000213', '정사각 20 x 20mm': 'SIZ_000214',
    '정사각 25 x 25mm': 'SIZ_000215', '정사각 30 x 30mm': 'SIZ_000216', '정사각 35 x 35mm': 'SIZ_000217',
    '정사각 40 x 40mm': 'SIZ_000218', '정사각 45 x 45mm': 'SIZ_000219', '정사각 50 x 50mm': 'SIZ_000220',
    '정사각 55 x 55mm': 'SIZ_000221', '정사각 60 x 60mm': 'SIZ_000222', '정사각 90 x 90mm': 'SIZ_000223',
    # 직사각 14
    '직사각 35 x 25mm': 'SIZ_000224', '직사각 40 x 30mm': 'SIZ_000226', '직사각 42 x 20mm': 'SIZ_000228',
    '직사각 50 x 20mm': 'SIZ_000230', '직사각 50 x 30mm': 'SIZ_000232', '직사각 55 x 15mm': 'SIZ_000234',
    '직사각 55 x 20mm': 'SIZ_000236', '직사각 55 x 24mm': 'SIZ_000238', '직사각 55 x 33mm': 'SIZ_000240',
    '직사각 90 x 40mm': 'SIZ_000242', '직사각 90 x 50mm': 'SIZ_000244', '직사각 90 x 60mm': 'SIZ_000245',
    '직사각 90 x 70mm': 'SIZ_000247', '직사각 90 x 80mm': 'SIZ_000249',
}

# 블록 정의: (헤더행, 소재축행, [데이터행], [수량])
BLOCKS = [
    (2, 3, [4, 5, 6, 7, 8], [1000, 2000, 3000, 4000, 5000]),
    (13, 14, [15, 16, 17, 18, 19], [1000, 2000, 3000, 4000, 5000]),
    (24, 25, [26, 27, 28, 29, 30], [1000, 2000, 3000, 4000, 5000]),
]


def parse_prices():
    """가격표 → component_prices long-form (6소재 분해)."""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb['합판도무송스티커']
    rows = []
    seen = set()
    for hr, sr, qrows, qtys in BLOCKS:
        # 사이즈 헤더 위치 (2컬럼 간격)
        sizes = []
        c = 2
        while c <= ws.max_column:
            v = ws.cell(hr, c).value
            if v is not None:
                sizes.append((c, str(v).strip()))
                c += 2
            else:
                c += 1
        for c0, sname in sizes:
            siz = SIZ.get(sname)
            assert siz, f'미매핑 사이즈: {sname!r}'
            coat_col = c0      # 코팅축 (홀수컬럼)
            dedron_col = c0 + 1  # 데드롱축 (짝수컬럼)
            for ri, q in zip(qrows, qtys):
                # 코팅축 → 3소재
                v_coat = ws.cell(ri, coat_col).value
                if v_coat is not None:
                    for mat in MAT_COAT:
                        key = (siz, mat, q)
                        assert key not in seen, f'중복키 {key}'
                        seen.add(key)
                        rows.append((siz, mat, q, int(v_coat)))
                # 데드롱축 → 3소재
                v_ded = ws.cell(ri, dedron_col).value
                if v_ded is not None:
                    for mat in MAT_DEDRON:
                        key = (siz, mat, q)
                        assert key not in seen, f'중복키 {key}'
                        seen.add(key)
                        rows.append((siz, mat, q, int(v_ded)))
    return rows


HDR_FILL = PatternFill('solid', fgColor='1F4E78')
SUB_FILL = PatternFill('solid', fgColor='DDEBF7')
HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
SUB_FONT = Font(color='1F4E78', italic=True, size=9)


def write_sheet(wb, name, columns, sublabels, data):
    ws = wb.create_sheet(name)
    for j, col in enumerate(columns, 1):
        cell = ws.cell(1, j, col)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        sub = ws.cell(2, j, sublabels[j - 1])
        sub.fill = SUB_FILL; sub.font = SUB_FONT
        sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(14, len(str(col)) + 2)
    for i, row in enumerate(data, 3):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    ws.freeze_panes = 'A3'
    return ws


def main():
    prices = parse_prices()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1_price_formulas (공식정의) ──
    write_sheet(wb, '1_price_formulas',
        ['frm_cd', 'frm_nm', 'note', 'use_yn'],
        ['공식코드(영문·기존)', '공식명(한글)', '비고', '사용여부(Y/N)'],
        [[FRM, '합판도무송 사이즈/소재/수량별 단가',
          'round-16: 라이브 기존 공식 재현(신규 아님). 소재 6분해·prc_typ .02 정정 반영', 'Y']])

    # ── 1b_product_price_formulas (상품 바인딩) ──
    write_sheet(wb, '1b_product_price_formulas',
        ['prd_cd', 'frm_cd', 'apply_bgn_ymd', 'note'],
        ['상품코드(합판도무송스티커)', '공식코드', '적용시작일(YYYYMMDD)', '비고'],
        [[PRD, FRM, APPLY, '라이브 바인딩 존재(가격사슬 완결)·재현']])

    # ── 2_formula_components (배선) ──
    write_sheet(wb, '2_formula_components',
        ['frm_cd', 'comp_cd', 'disp_seq', 'addtn_yn'],
        ['공식코드', '구성요소코드', '표시순서', '합산여부(Phase11 무시)'],
        [[FRM, COMP, 1, 'Y']])

    # ── 3_price_components (구성요소 정의) ──
    write_sheet(wb, '3_price_components',
        ['comp_cd', 'comp_nm', 'comp_typ_cd', 'prc_typ_cd', 'use_dims', 'use_yn', 'note'],
        ['구성요소코드', '구성요소명(한글)', '구성요소유형', '단가유형(.01단가/.02합가)',
         '사용차원(jsonb)', '사용여부', '비고'],
        [[COMP, '합판도무송 단가(완제품가) [COMP_GANGPAN_PRINT]', None,
          'PRICE_TYPE.02',
          '["siz_cd", "mat_cd", "min_qty"]', 'Y',
          '🔴 round-16: 라이브 .01(단가형) 오적재 → .02(합가형) 정정. 셀=수량구간 총액(장당가 체감 20→12원)']])

    # ── 4_component_prices (단가행 10차원) ──
    cp_cols = ['comp_cd', 'apply_ymd', 'siz_cd', 'clr_cd', 'mat_cd',
               'proc_cd', 'coat_side_cnt', 'opt_cd', 'bdl_qty', 'min_qty', 'unit_price', 'note']
    cp_sub = ['구성요소코드', '적용일', '사이즈(형상)', '색상(NULL=무관)', '소재(6분해)',
              '공정(NULL=자재variant)', '코팅면수(NULL)', '옵션(NULL)', '묶음수(NULL)',
              '제작수량(구간)', '단가(=구간총액)', '비고']
    cp_data = []
    for siz, mat, q, price in prices:
        cp_data.append([COMP, APPLY, siz, None, mat, None, None, None, None, q, price, MAT_NM[mat]])
    write_sheet(wb, '4_component_prices', cp_cols, cp_sub, cp_data)

    wb.save(OUT)

    # ── 검증 출력 ──
    print(f'생성: {OUT}')
    print(f'4_component_prices 행수: {len(cp_data)} (기대 1110)')
    siz_set = {r[0] for r in prices}
    mat_set = {r[1] for r in prices}
    mq_set = {r[2] for r in prices}
    print(f'distinct siz_cd: {len(siz_set)} (기대 37)')
    print(f'distinct mat_cd: {sorted(mat_set)}')
    print(f'distinct min_qty: {sorted(mq_set)}')
    # 자연키 중복 0 검증 (이미 parse에서 assert)
    keys = [(r[0], r[1], r[2]) for r in prices]
    print(f'자연키(siz,mat,mq) 중복: {len(keys) - len(set(keys))} (기대 0)')
    # 무손실: 셀 370 → 1110 = 370×3
    print(f'round-trip: {len(cp_data)} / 3 = {len(cp_data) // 3} 셀 (기대 370)')


if __name__ == '__main__':
    main()
