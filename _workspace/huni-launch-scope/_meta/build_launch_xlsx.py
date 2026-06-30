# -*- coding: utf-8 -*-
"""1차 런칭 개발범위·Shopby 갭·개발방안 최종본 → xlsx (IA 엑셀 스타일)"""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/Users/innojini/Dev/HuniWeb/_workspace/huni-launch-scope"
OUT = "/Users/innojini/Dev/HuniWeb/docs/huni/후니프린팅_1차런칭_개발범위_Shopby갭_개발방안_260630.xlsx"

TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
H = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
B = Font(name="맑은 고딕", size=10)
BBOLD = Font(name="맑은 고딕", size=10, bold=True)
NOTE = Font(name="맑은 고딕", size=10, italic=True, color="555555")
HEADER_FILL = PatternFill("solid", fgColor="305496")
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D6DCE4")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
VERDICT_FILL = {"SOLVED": PatternFill("solid", fgColor="C6EFCE"),
                "PARTIAL": PatternFill("solid", fgColor="FFEB9C"),
                "CUSTOM": PatternFill("solid", fgColor="FFC7CE")}
VERDICT_FONT = {"SOLVED": Font(name="맑은 고딕", size=10, bold=True, color="006100"),
                "PARTIAL": Font(name="맑은 고딕", size=10, bold=True, color="9C6500"),
                "CUSTOM": Font(name="맑은 고딕", size=10, bold=True, color="9C0006")}
PHASE_FILL = {"1차": PatternFill("solid", fgColor="DDEBF7"),
              "2차": PatternFill("solid", fgColor="E2EFDA"),
              "3차": PatternFill("solid", fgColor="EDEDED")}

wb = openpyxl.Workbook()

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.font = H; cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, ncol, r=1):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    c = ws.cell(row=r, column=1, value=text)
    c.font = TITLE; c.fill = TITLE_FILL; c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[r].height = 26

# 00_읽는법
ws = wb.active; ws.title = "00_읽는법"
set_widths(ws, [22, 95])
title_row(ws, "후니프린팅 1차 런칭 개발범위 · Shopby 갭 · 개발방안 (최종본 260630)", 2)
rows = [
 ("문서 목적", "통합IA 162개 기능을 Shopby에 얹을 때 '그대로 해결 / 일부 손봄 / 새로 만듦'을 전수로 가르고, 1차 런칭에 들어갈 것을 골라 '어떻게 만들지'까지 정리한 개발범위 설계서."),
 ("범위 경계", "설계·명세까지. 실제 구현·고객 데이터 이관·DB 저장은 사람 최종 승인 후 별도 팀(위젯·Shopby 연동·DB)에 위임."),
 ("판정 색 범례", "초록=해결(SOLVED·추가개발 없음) / 노랑=부분(PARTIAL·일부 맞춤) / 빨강=미해결(CUSTOM·새로 만듦)"),
 ("단계 색 범례", "파랑=1차(런칭) / 연두=2차(안정) / 회색=3차(확장)"),
 ("시트 안내", "01 한눈요약 · 02 FitGap 162전수 · 03 1차런칭 개발방안 · 04 회원/프린팅머니 마이그레이션 · 05 2·3차 로드맵 · 06 결정필요(우선순위) · 07 근거출처"),
 ("", ""),
 ("용어: Shopby", "NHN 쇼핑몰 솔루션(상품·회원·장바구니·주문·결제·정산 기본 제공 뼈대). 새 사이트는 이 위에 올림."),
 ("용어: 위젯", "상품 페이지에서 손님이 사이즈·종이·수량을 고르고 실시간 가격을 보는 옵션 선택 화면. 후니가 직접 만듦."),
 ("용어: BFF(중계서버)", "위젯·Shopby·후니 가격엔진 사이에서 데이터를 번역·중계하는 후니 서버(Backend For Frontend)."),
 ("용어: 가격엔진", "후니가 이미 가진 가격 계산 프로그램(evaluate_price). 가격 최종 권위는 항상 이 서버. 0원=오류신호."),
 ("용어: webadmin", "후니가 지금 쓰는 상품·가격 등록 관리자(Django). 신규에서도 그대로 유지."),
 ("용어: 마이그레이션", "옛 사이트의 회원·머니 데이터를 새 사이트로 옮기는 작업."),
]
r = 3
for k, v in rows:
    ws.cell(row=r, column=1, value=k).font = BBOLD
    if k: ws.cell(row=r, column=1).fill = SUB_FILL
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=2, value=v).font = B; ws.cell(row=r, column=2).alignment = LEFT
    r += 1

# 01_한눈요약
ws = wb.create_sheet("01_한눈요약")
set_widths(ws, [16, 12, 62, 14])
title_row(ws, "한눈 요약", 4)
r = 3
ws.cell(row=r, column=1, value="① Shopby 해결도 (전체 162기능)").font = BBOLD; r += 1
ws.append(["구분", "개수", "뜻", ""]); style_header(ws, r, 4); r += 1
for k, n, d, key in [("해결(SOLVED)", 48, "Shopby 기본 기능/화면 스킨으로 추가 개발 없이 충족", "SOLVED"),
                     ("부분(PARTIAL)", 53, "Shopby 기본 위에 후니 맞춤을 일부 얹어야 함", "PARTIAL"),
                     ("미해결(CUSTOM)", 61, "Shopby에 없어 새로 만들어야 함(주로 인쇄 옵션·가격·생산)", "CUSTOM")]:
    ws.cell(row=r, column=1, value=k).font = BBOLD; ws.cell(row=r, column=1).fill = VERDICT_FILL[key]
    ws.cell(row=r, column=2, value=n).alignment = CENTER
    ws.cell(row=r, column=3, value=d).font = B; ws.cell(row=r, column=3).alignment = LEFT
    r += 1
ws.cell(row=r, column=1, value="※ 검증 중 9건 재분류: 결제(80)·주문완료(81)·회원관리(123)는 인쇄 가격·파일·생산 연결 없이 못 써 부분으로 하향, 정적 콘텐츠 6건(56·57·58·60·66·151)은 해결로 상향.").font = NOTE
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); ws.cell(row=r, column=1).alignment = LEFT; r += 2
ws.cell(row=r, column=1, value="② 단계 나눔").font = BBOLD; r += 1
ws.append(["단계", "기능 수", "목표", ""]); style_header(ws, r, 4); r += 1
for k, n, d in [("1차(런칭)", 64, "인쇄 자동견적 주문이 끝까지 돌아가는 최소 사이트 + 생산 연결"),
                ("2차(안정)", 73, "회원 편의·쿠폰/머니·통계·증빙·결제수단 확장"),
                ("3차(확장)", 25, "온라인 편집기(Edicus)·체험단·마케팅 랜딩·수작 브랜드")]:
    ws.cell(row=r, column=1, value=k).font = BBOLD; ws.cell(row=r, column=1).fill = PHASE_FILL[k[:2]]
    ws.cell(row=r, column=2, value=n).alignment = CENTER
    ws.cell(row=r, column=3, value=d).font = B; ws.cell(row=r, column=3).alignment = LEFT
    r += 1
r += 1
ws.cell(row=r, column=1, value="③ ★런칭 전 반드시 결정해야 하는 GATE(관문) — 돈·주문 핵심").font = BBOLD
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FCE4D6"); r += 1
ws.append(["GATE", "무엇", "왜 중요", ""]); style_header(ws, r, 4); r += 1
for k, w, d in [
 ("G1 가격 전달 방식", "위젯이 계산한 정확한 가격을 Shopby 장바구니·주문에 손실 없이 싣는 길 확정",
  "Shopby 장바구니엔 임의 가격 직접 입력 칸이 아예 없음(원문 확인). 가로×세로 자유입력·면적/구간 단가 견적형은 안전한 길 미확정 → 1차 최대 위험"),
 ("G2 정산 기준 확정", "정산이 '주문라인 등록가'와 '실제 결제액' 중 무엇을 진짜로 보는지", "가격을 임시값으로 넣으면 정산이 틀어질 수 있음"),
 ("G3 머니 이관 안전장치", "옛 프린팅머니 잔액을 1원도 안 틀리게 옮기는 검증", "고객 돈. 만료 기본값=제한없음·전 회원 잔액 전수 대조 필요"),
 ("G4 외부 계약", "이니시스 결제·알림톡/문자·휴대폰 인증·MES 통신규격·파일 검수기준", "계약/규격이 늦으면 해당 기능 전체가 멈춤")]:
    ws.cell(row=r, column=1, value=k).font = BBOLD
    ws.cell(row=r, column=2, value=w).font = B; ws.cell(row=r, column=2).alignment = LEFT
    ws.cell(row=r, column=3, value=d).font = B; ws.cell(row=r, column=3).alignment = LEFT
    r += 1

# 02_FitGap_162
ws = wb.create_sheet("02_FitGap_162")
with open(f"{BASE}/02_gap/fit-gap-matrix.csv", encoding="utf-8") as f:
    rows = list(csv.reader(f))
hdr = rows[0]
set_widths(ws, [5, 14, 22, 7, 9, 30, 14, 42, 7, 22, 26])
title_row(ws, "FitGap 전수 매트릭스 (162기능 · 해결48/부분53/미해결61)", len(hdr))
ws.append(hdr); style_header(ws, 2, len(hdr))
ri = 3
for row in rows[1:]:
    ws.append(row)
    verdict = row[4].strip(); phase = row[3].strip()
    for ci in range(1, len(hdr)+1):
        cell = ws.cell(row=ri, column=ci); cell.font = B; cell.alignment = LEFT; cell.border = BORDER
    if phase[:2] in PHASE_FILL:
        ws.cell(row=ri, column=4).fill = PHASE_FILL[phase[:2]]; ws.cell(row=ri, column=4).alignment = CENTER
    if verdict in VERDICT_FILL:
        ws.cell(row=ri, column=5).fill = VERDICT_FILL[verdict]; ws.cell(row=ri, column=5).font = VERDICT_FONT[verdict]; ws.cell(row=ri, column=5).alignment = CENTER
    ws.cell(row=ri, column=1).alignment = CENTER; ws.cell(row=ri, column=9).alignment = CENTER
    ri += 1
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr))}{ri-1}"

# 03_1차런칭_개발방안
ws = wb.create_sheet("03_1차런칭_개발방안")
cols = ["구역", "기능(No)", "판정", "개발 위치", "개발 방안", "비고/주의"]
set_widths(ws, [18, 26, 14, 18, 64, 34])
title_row(ws, "1차 런칭(Phase 1) 세부 개발범위 — 위젯+BFF+Shopby+webadmin 4층", len(cols))
ws.append(cols); style_header(ws, 2, len(cols))
plan = [
 ("동적가격", "49 옵션변경 실시간 재계산", "CUSTOM", "위젯+BFF", "옵션 변경마다 BFF POST /price/evaluate 호출 → 가격엔진 정확가 반환 → 위젯은 표시만. 가격 권위=서버.", "0원=항상 오류신호 · 디바운스/캐싱 필요 · 초대형"),
 ("동적가격", "62 선택옵션→주문데이터 변환", "CUSTOM", "위젯+BFF+Shopby", "Shopby 카트엔 임의가격 직접입력 칸 없음(customPrice 0건). 방법A(가격별 옵션 미리등록·고정가/소조합 안전)/B(주문직전 즉석옵션·심사리스크)/C(수량1+총액옵션·정산위험) 중 택.", "★돈 관문 · 권고=고정가·소조합부터 A · G1·G2 선행 · 대형"),
 ("동적가격", "64 장바구니 담기", "PARTIAL", "Shopby+위젯번역", "표준 post-cart로 담되 위젯 선택값을 상품/옵션으로 번역. 직전 get-cart-validate로 구매가능 점검.", "★돈 관문 · 중형"),
 ("동적가격", "(보강) 주문서→최종계산→결제", "—", "Shopby", "post-order-sheet·calculate·reserve 실재 확인 → 위젯→장바구니→주문서→최종계산→결제 전 경로 설계.", "검증 중 발견한 공백 보강"),
 ("옵션위젯", "33 상품별 옵션 목록", "CUSTOM", "위젯+BFF", "상품마다 옵션 구성 상이 → BFF가 옵션데이터 내려주면 위젯이 옵션 트리 렌더.", "대형"),
 ("옵션위젯", "34~38 사이즈/비규격/수량/소재/도수", "CUSTOM", "위젯+가격엔진", "각 선택이 위젯 선택값(siz_cd·가로/세로·수량·mat_cd·도수)으로 → 가격엔진 계산.", "Shopby 옵션 최대 5종 제한(option.mdx:44)으로 못 담음 · 중형"),
 ("옵션위젯", "42·43 제본·내지/표지", "CUSTOM", "위젯+BFF", "책자=여러 부품 합치는 셋트 상품 → 위젯 구성, BFF가 셋트 가격 계산.", "대형/중형"),
 ("옵션위젯", "48 옵션 제약(동시 불가 조합)", "CUSTOM", "위젯", "'이 종이엔 이 코팅 안 됨' 규칙을 위젯이 평가. 규칙 데이터 선행 정리.", "대형 · 제약데이터 선행"),
 ("파일/편집", "52 PDF 파일 업로드", "CUSTOM", "위젯+S3", "파일을 외부저장소(S3) 업로드, 식별값만 주문라인에 텍스트로 부착(원본은 후니 보관).", "대형"),
 ("파일/편집", "53 편집 미리보기·썸네일", "CUSTOM", "위젯", "올린/편집 결과의 작은 미리보기 생성.", "중형 · Edicus 편집기(54)는 3차"),
 ("회원", "5 정보입력·21 회원정보수정", "PARTIAL", "Shopby+extraInfo", "표준 프로필 + 개인·사업자정보·회사주소 2개(30+항목)는 추가항목(extraInfo)/별도 저장 확장.", "중형"),
 ("회원", "22 비번입력·23 비번변경·24 탈퇴", "SOLVED", "Shopby 표준", "Shopby 표준 사용.", "변경/탈퇴 전 본인확인 방법 결정 필요"),
 ("회원", "29 마이페이지 메인", "PARTIAL", "Aurora+BFF", "회원할인율·프린팅머니 잔액·옵션보관 요약 대시보드. 할인순서(상품금액→회원할인→쿠폰→머니).", "소형"),
 ("상품리스트", "153 전시 카테고리 운영툴", "PARTIAL", "Shopby 전시카테고리", "다단계 진열. 옛 단순 나열을 정렬/필터로 개선.", "중형"),
 ("상품리스트", "31 리스트·154 검색/필터", "PARTIAL", "Shopby 검색+인쇄필터", "Shopby 검색에 인쇄 필터축(소재/사이즈/도수) 추가.", ""),
 ("주문/배송/결제", "162 주문 런타임 그릇", "PARTIAL", "Shopby", "옛 사이트엔 주문 테이블 자체가 없음(실측) → Shopby 주문기능을 그릇으로 채택, 후니는 인쇄사양·파일·생산정보만 라인 연결.", "초대형 · ★중요"),
 ("주문/배송/결제", "80 결제(이니시스)", "PARTIAL", "Shopby+이니시스", "Shopby 결제 표준 + 이니시스. 결제검증이 '라인가=후니가' 전제 → G1 가격브리지 선행.", "재분류(가격브리지 종속)"),
 ("주문/배송/결제", "77·78·79·150 배송/배송지/도서산간", "SOLVED/PARTIAL", "Shopby+지역데이터", "Shopby 표준 + 지역 추가배송비 데이터.", ""),
 ("주문/배송/결제", "161 주문상태 추적", "PARTIAL", "Shopby+생산상태", "Shopby 주문조회 + 인쇄 생산 상태(미입금→제작중→출고완료).", "중형"),
 ("생산브릿지", "157 생산정보 전달(BOM 환원)", "CUSTOM", "BFF→MES", "결제완료 주문 선택값을 생산용 자재명세(BOM)로 환원해 MES 전달.", "★주문 관문 · MES 규격 확정 필요 · 대형"),
 ("생산브릿지", "158 전상품 MES 코드 채번", "CUSTOM", "webadmin", "현 라이브 275개 중 16개만 코드(94% 미채번) → 전상품 채번 체계.", "중형"),
 ("생산브릿지", "137 파일확인·159 파일명 일관화", "CUSTOM", "BFF", "인쇄파일 검수 + 자동 이름변경(품목_사이즈_소재_고객_번호_수량).", ""),
 ("상품/가격관리", "100~108 상품/가격 관리", "유지", "webadmin", "현 webadmin 유지(등록 주인). Shopby엔 상품번호만 동기화. Shopby 상품등록은 인쇄옵션/파일 미수용.", "결정사항"),
 ("상품/가격관리", "108 가격관리", "유지", "webadmin", "가격엔진이 먹는 가격공식·구성요소·단가표 + 시뮬레이터(현 라이브 자산).", "초대형"),
]
ri = 3
for row in plan:
    ws.append(row); v = row[2]
    for ci in range(1, len(cols)+1):
        c = ws.cell(row=ri, column=ci); c.font = B; c.alignment = LEFT; c.border = BORDER
    ws.cell(row=ri, column=1).font = BBOLD
    key = "SOLVED" if v.startswith("SOLVED") else ("PARTIAL" if v.startswith("PARTIAL") else ("CUSTOM" if v == "CUSTOM" else None))
    if key:
        ws.cell(row=ri, column=3).fill = VERDICT_FILL[key]; ws.cell(row=ri, column=3).font = VERDICT_FONT[key]
    ws.cell(row=ri, column=3).alignment = CENTER
    ri += 1
ws.freeze_panes = "A3"

# 04_마이그레이션
ws = wb.create_sheet("04_마이그레이션")
set_widths(ws, [18, 38, 26, 46, 40, 32])
title_row(ws, "회원 / 프린팅머니 마이그레이션 설계 (실 이관은 인간 승인 후)", 6)
r = 3
ws.cell(row=r, column=1, value="■ 이관 요약 / 권고").font = BBOLD; ws.cell(row=r, column=1).fill = SUB_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
for k, v in [
 ("회원 — 그대로 옮김", "아이디·이름·이메일·휴대폰·전화·주소·상호·사업자번호·SMS/마케팅 동의값·등급(매핑)·휴면/탈퇴 상태"),
 ("회원 — 추가항목 보존", "대표자명·업태·업종·직업·회사전화·팩스·회사주소1/2 (extraInfo/주소록)"),
 ("회원 — 못 옮김→대체", "비밀번호(단방향 해시·호환불가)→재설정. 기본=첫 로그인 이메일 재설정 / 휴대폰만=임시PW 문자 / 소셜=네이버·카카오 전환. 전 회원 1회 재인증 불가피(공지)."),
 ("회원 — 약관/개인정보", "사이트 이전=처리방식 변경 → 재동의 또는 사전 고지. 마케팅 동의는 동의일시 원본 없으면 '이관일 기록' 금지(날조·법적위험)→보류+재동의 캠페인."),
 ("프린팅머니 — 모델 차이", "후니=미리 충전하는 선불 잔액 / Shopby 적립금=구매 시 쌓이는 포인트. 단 사용(결제 차감)은 동일."),
 ("프린팅머니 — 권고(방안 A)", "회원별 잔액을 Shopby 적립금으로 1건씩 수동 지급(시드)해 이관. 적립금 명칭을 '프린팅머니'로 표시 가능(원문 확인). 향후 새 충전(PG)만 별도 중계서버 보강."),
 ("프린팅머니 — 보존 수준", "현재 잔액만 이관 권고. 과거 거래내역 1:1 재현은 비용·합계 위험→필요 시 옛 사이트 조회로 분리."),
]:
    ws.cell(row=r, column=1, value=k).font = BBOLD; ws.cell(row=r, column=1).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2, value=v).font = B; ws.cell(row=r, column=2).alignment = LEFT
    r += 1
r += 1
ws.cell(row=r, column=1, value="■ ★돈 안전장치 (HARD)").font = BBOLD; ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="FCE4D6")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
for s in [
 "합계 대사: Σ(옛 회원 잔액)=Σ(옮긴 적립금), 차이 0원이어야 통과. 1원이라도 차이=즉시 중단.",
 "전 회원 전수 대조(샘플 부족): 회원별 옛 잔액↔새 잔액 전부 비교(잔액 뒤바뀜 사고 방지).",
 "만료 기본값=제한없음: 옛 잔여 유효기간 모르면 소멸 금지(고객 돈 소멸 방지).",
 "멱등 키: 같은 작업 두 번 돌려도 중복 지급 0(externalKey).",
 "음수 잔액(미수금)은 적립금으로 이관 불가 → 별도 보류·수기 결정.",
]:
    ws.cell(row=r, column=1, value="• "+s).font = B; ws.cell(row=r, column=1).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
r += 1
ws.cell(row=r, column=1, value="■ 필드 매핑표 (구 필드 → Shopby)").font = BBOLD; ws.cell(row=r, column=1).fill = SUB_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 1
with open(f"{BASE}/03_migration/migration-field-mapping.csv", encoding="utf-8") as f:
    mrows = list(csv.reader(f))
ws.append(mrows[0]); style_header(ws, r, len(mrows[0])); hdrrow = r; r += 1
for row in mrows[1:]:
    ws.append(row)
    for ci in range(1, len(row)+1):
        c = ws.cell(row=r, column=ci); c.font = B; c.alignment = LEFT; c.border = BORDER
    ws.cell(row=r, column=1).font = BBOLD
    r += 1
ws.freeze_panes = ws.cell(row=hdrrow+1, column=1)

# 05_로드맵
ws = wb.create_sheet("05_2차3차_로드맵")
set_widths(ws, [10, 24, 78])
title_row(ws, "2차(안정) · 3차(확장) 보완 로드맵", 3)
ws.append(["단계", "분류", "기능"]); style_header(ws, 2, 3)
road = [
 ("2차", "회원 편의", "휴대폰 인증·비번찾기(3·7)·소셜로그인(145)·회원등급/등급가(146)·휴면처리(147)·사업자정보(26·27)·증빙(25·28)"),
 ("2차", "쿠폰/머니", "프린팅머니/충전(14·15·125)·쿠폰관리(12·13·126~128)"),
 ("2차", "상품 옵션 확장", "코팅·커팅·접지·박·캘린더(39·40·41·44·45)·추가상품(46)·수량별 가격표/견적서(50·51)"),
 ("2차", "결제수단", "가상계좌(148)·네이버페이(149)·세금계산서/팝빌(152)"),
 ("2차", "고객센터/게시판/통계", "공지·FAQ·Q&A·상담(67~73·114~122)·상품/매출통계(129~134)"),
 ("2차", "백오피스", "거래처·원장·미수금(95~99)·후불결제(141)·증빙발급관리(142)"),
 ("3차", "온라인 편집기", "Edicus(54·55) — 상품범위 결정 후. 1차에 편집상품 투입 부적절"),
 ("3차", "마케팅/후기", "체험단(19·90·121)·이용후기 메인(89)·마케팅 랜딩 5종(88)"),
 ("3차", "수작 브랜드", "수작 브랜드(82·93·111·113) — 유지 여부 결정 후"),
 ("3차", "기타", "리뷰 사진(18)·서브메인 랜딩(30)·자동쿠폰(156)·수동카드결제(74) 등"),
]
ri = 3
for p, c, d in road:
    ws.append([p, c, d])
    ws.cell(row=ri, column=1).fill = PHASE_FILL[p]; ws.cell(row=ri, column=1).font = BBOLD; ws.cell(row=ri, column=1).alignment = CENTER
    ws.cell(row=ri, column=2).font = BBOLD; ws.cell(row=ri, column=2).alignment = LEFT
    ws.cell(row=ri, column=3).font = B; ws.cell(row=ri, column=3).alignment = LEFT
    for ci in range(1, 4): ws.cell(row=ri, column=ci).border = BORDER
    ri += 1
ws.freeze_panes = "A3"

# 06_결정필요
ws = wb.create_sheet("06_결정필요")
set_widths(ws, [14, 16, 64, 22])
title_row(ws, "결정·확인 필요 사항 (런칭 전 우선순위)", 4)
ws.append(["우선순위", "ID", "결정할 것", "주체"]); style_header(ws, 2, 4)
oq = [
 ("1순위·돈/주문", "OQ-G1", "가격 전달 방식(A/B/C 중)·견적형 무손실 경로", "아키텍트+PM"),
 ("1순위·돈/주문", "OQ-G2", "주문 직전 즉석옵션(방법 B)이 Shopby 상품심사에 막히는지", "아키텍트(라이브 확인)"),
 ("1순위·돈/주문", "OQ-G3", "정산이 라인가/실결제액 중 무엇을 권위로", "PM+아키텍트"),
 ("2순위·외부계약", "OQ-G7~G13", "이니시스 결제·알림톡/SMS·휴대폰 인증·MES 통신규격·파일 검수기준·본인확인·배송권역정책", "PM"),
 ("3순위·이관", "MQ-1~19", "비번 해시방식·가입일 보존·휴면/탈퇴 플래그·동의일시·중복이메일 규모·등급체계·사업자수용범위·회사주소 용도·잔액/원장 스키마·만료정책·음수잔액·동결창", "옛 DB 받은 뒤 확정"),
 ("4순위·상품/정책", "OQ-G14~G22·I6", "카테고리 구조·옵션 제약 데이터·온라인 편집기 도입·수작 브랜드 유지·증빙/세금계산서 정책·사업자정보 1차 가입폼 포함 여부·webadmin↔Shopby 동기화 범위", "PM"),
]
ri = 3
pri_fill = {"1순위·돈/주문": "FFC7CE", "2순위·외부계약": "FFEB9C", "3순위·이관": "DDEBF7", "4순위·상품/정책": "E2EFDA"}
for p, i, d, who in oq:
    ws.append([p, i, d, who])
    ws.cell(row=ri, column=1).fill = PatternFill("solid", fgColor=pri_fill[p]); ws.cell(row=ri, column=1).font = BBOLD; ws.cell(row=ri, column=1).alignment = LEFT
    ws.cell(row=ri, column=2).font = BBOLD; ws.cell(row=ri, column=2).alignment = LEFT
    ws.cell(row=ri, column=3).font = B; ws.cell(row=ri, column=3).alignment = LEFT
    ws.cell(row=ri, column=4).font = B; ws.cell(row=ri, column=4).alignment = LEFT
    for ci in range(1, 5): ws.cell(row=ri, column=ci).border = BORDER
    ri += 1
ws.freeze_panes = "A3"

# 07_근거출처
ws = wb.create_sheet("07_근거출처")
set_widths(ws, [24, 92])
title_row(ws, "근거 · 출처 / 실 적용 경계", 2)
r = 3
for k, v in [
 ("IA 정본(162기능)", "docs/huni/후니프린팅_프로젝트일정관리_통합IA_260616.xlsx (02_IA마스터 144+18=162, 원본 직접 재실측)"),
 ("Shopby 스펙", "docs/shopby/ (OpenAPI 24종·enterprise). 가격주입 불가(customPrice 등 0건)·옵션 5종 제한(option.mdx:44)·적립금 명칭변경(accumulation-setting.mdx:20)·적립금 지급/차감(profile/accumulations) 원문 확인"),
 ("라이브 As-Is", "_workspace/huni-launch-scope/00_live/ (사이트맵·마이페이지 33기능·견적형 goods.asp/카탈로그형 view.asp·주문테이블 부재)"),
 ("갭 분석/개발방안", "_workspace/huni-launch-scope/02_gap/ (fit-gap 162행·dev-plan-launch·roadmap·open-questions 22)"),
 ("마이그레이션", "_workspace/huni-launch-scope/03_migration/ (회원·머니 설계·필드매핑 31·open-questions 24)"),
 ("독립 교차검증(Codex)", "_workspace/huni-launch-scope/04_codex/reconcile.md (합의 8·조사 9·반박 6·디스코핑 5)"),
 ("게이트 판정", "_workspace/huni-launch-scope/05_gate/gate-verdict.md (L1~L7 GO 조건부)"),
]:
    ws.cell(row=r, column=1, value=k).font = BBOLD; ws.cell(row=r, column=1).fill = SUB_FILL; ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=2, value=v).font = B; ws.cell(row=r, column=2).alignment = LEFT
    r += 1
r += 1
ws.cell(row=r, column=1, value="★ 실 적용 경계").font = TITLE; ws.cell(row=r, column=1).fill = TITLE_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1
for s in [
 "이 문서는 설계·명세까지. 다음은 모두 사람 최종 승인 후 각 트랙 위임.",
 "위젯 실제 구현 → 위젯 구현 트랙(§6)",
 "Shopby 연동·가격브리지 실제 구현 → Shopby 연동 트랙(§24)",
 "DB 적재·회원/머니 실제 이관 → DB 작업 트랙(dbmap)",
 "webadmin 코드 직접 수정 금지(현 자산 유지)",
]:
    ws.cell(row=r, column=1, value="• "+s).font = B; ws.cell(row=r, column=1).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 1

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("SAVED:", OUT)
print("SHEETS:", wb.sheetnames)
