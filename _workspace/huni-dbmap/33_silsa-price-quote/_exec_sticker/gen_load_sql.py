#!/usr/bin/env python3
# ============================================================================
# 스티커 누락 채움 SQL 생성기 (gen_load_sql.py) — 재현·provenance-traced
#   입력(권위): 20_price-import/sticker/sticker-import.xlsx (단가 verbatim)
#               + sticker-3axis-design.md (arbiter 판정·mat/siz/comp 권위)
#               + 라이브 실측(SIZ/MAT 실존·apply_ymd='2026-06-01')
#   출력: S1_b01_materials.sql · S2_clear_remap.sql · S3_b3b4_prices.sql · S8_bindings.sql
#   주의: S2(투명 오매핑 170→162)는 단가 입력 없는 UPDATE 라 본 스크립트 미생성(정적·hand SQL).
#         S1/S3 단가는 전부 xlsx verbatim — 손수정 0.
#   apply_ymd='2026-06-01' 고정(단가행 적용일 분기 금지 = 가격 이중계상·sibling 관행).
# ============================================================================
import openpyxl, os, sys

ROOT = '/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap'
XLSX = f'{ROOT}/20_price-import/sticker/sticker-import.xlsx'
OUT  = f'{ROOT}/33_silsa-price-quote/_exec_sticker'
APPLY = '2026-06-01'

def fmt(p):
    p = float(p)
    return str(int(p)) if p == int(p) else str(p)

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ---------------- S1: B01 소재 4미적재 (live SIZ_059/060) ----------------
ws = wb['4_component_prices']; rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]; idx = {h: i for i, h in enumerate(hdr)}
sizmap = {'A5(4판)': 'SIZ_000059', '90*190(6판)': 'SIZ_000060'}     # B01 live sizes
matmap = {'비코팅': 'MAT_000084', '미색': 'MAT_000242',
          '유광코팅': 'MAT_000156', '홀로그램': 'MAT_000163'}      # live-absent materials
matnm  = {v: k for k, v in matmap.items()}
s1 = []
for r in rows[1:]:
    sl, ml = r[idx['siz_label']], r[idx['mat_label']]
    if sl in sizmap and ml in matmap:
        s1.append((sizmap[sl], matmap[ml], int(r[idx['min_qty']]),
                   float(r[idx['unit_price']]), matnm[matmap[ml]]))
s1.sort(key=lambda x: (x[1], x[0], x[2]))

with open(f'{OUT}/S1_b01_materials.sql', 'w') as f:
    f.write("""-- S1 · B01 반칼 소재 4미적재 채움 (GAP-MAT-1) — component_prices INSERT
-- 출처: 20_price-import/sticker/sticker-import.xlsx#4_component_prices (verbatim) + 라이브 실측(SIZ_000059/060 실존)
-- 비코팅(084)·미색(242)·유광코팅(156)·홀로그램(163) × SIZ_000059(124x186)·SIZ_000060(90x190) × 36 min_qty 구간.
-- 가격=가격표 그룹단가 verbatim (비코팅/미색=유포가, 유광=무광가, 홀로=투명가). proc_cd/clr_cd/bdl_qty=NULL.
-- 멱등: 자연키(comp_cd,apply_ymd,siz_cd,mat_cd,min_qty) NOT EXISTS (PK=comp_price_id 시퀀스라 ON CONFLICT 불가).
-- search-before-mint: mat 4종 전부 라이브 실존(신규 0). siz 신규 0.
INSERT INTO t_prc_component_prices
  (comp_cd, apply_ymd, siz_cd, mat_cd, min_qty, unit_price, note, reg_dt)
SELECT 'COMP_STK_PRINT', v.apply_ymd, v.siz_cd, v.mat_cd, v.min_qty, v.unit_price, v.note, now()
FROM (VALUES
""")
    f.write(",\n".join(
        f"  ('{APPLY}','{s}','{m}',{mq},{fmt(p)}::numeric,'B01 {nm}')" for (s, m, mq, p, nm) in s1))
    f.write("""
) AS v(apply_ymd, siz_cd, mat_cd, min_qty, unit_price, note)
WHERE NOT EXISTS (
  SELECT 1 FROM t_prc_component_prices cp
   WHERE cp.comp_cd='COMP_STK_PRINT' AND cp.apply_ymd=v.apply_ymd
     AND cp.siz_cd=v.siz_cd AND cp.mat_cd=v.mat_cd AND cp.min_qty=v.min_qty
);
""")

# ---------------- S3: B4/B3 단가행 (siz 실존·verbatim) ----------------
ws = wb['4b_component_prices_BLOCKED']; rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]; idx = {h: i for i, h in enumerate(hdr)}
sizmapB = {'B4': 'SIZ_000515', 'B3': 'SIZ_000514'}
s3 = []
for r in rows[1:]:
    sl = r[idx['siz_label']]
    if sl in ('B3', 'B4'):
        comp = r[idx['comp_cd']]
        # arbiter §4.2 권위: 단일 comp COMP_STK_PRINT · 투명=162 · 일반=153
        mat = 'MAT_000162' if comp == 'COMP_STK_PRINT_CLEAR' else 'MAT_000153'
        grp = 'B03 투명' if mat == 'MAT_000162' else 'B02 유포'
        s3.append((sizmapB[sl], mat, int(r[idx['min_qty']]),
                   float(r[idx['unit_price']]), f"{grp} {sl}"))
s3.sort(key=lambda x: (x[1], x[0], x[2]))

with open(f'{OUT}/S3_b3b4_prices.sql', 'w') as f:
    f.write("""-- S3 · B4/B3 단가행 채움 (GAP-SIZ-3) — 즉시 GO (siz 라이브 실존·채번 0)
-- 출처: 20_price-import/sticker/sticker-import.xlsx#4b_component_prices_BLOCKED (verbatim) + 라이브 SIZ_000515(B4)·SIZ_000514(B3) 실존
-- arbiter §4.2 권위: 단일 comp COMP_STK_PRINT, B02 일반낱장=mat 153(유포), B03 투명낱장=mat 162(투명스티커).
--   (xlsx 4b 의 mat 165/COMP_STK_PRINT_CLEAR 는 round-16 stale → arbiter 153/162·단일 comp 채택)
-- 멱등: 자연키 NOT EXISTS. search-before-mint: siz/mat 신규 0.
INSERT INTO t_prc_component_prices
  (comp_cd, apply_ymd, siz_cd, mat_cd, min_qty, unit_price, note, reg_dt)
SELECT 'COMP_STK_PRINT', v.apply_ymd, v.siz_cd, v.mat_cd, v.min_qty, v.unit_price, v.note, now()
FROM (VALUES
""")
    f.write(",\n".join(
        f"  ('{APPLY}','{s}','{m}',{mq},{fmt(p)}::numeric,'{nm}')" for (s, m, mq, p, nm) in s3))
    f.write("""
) AS v(apply_ymd, siz_cd, mat_cd, min_qty, unit_price, note)
WHERE NOT EXISTS (
  SELECT 1 FROM t_prc_component_prices cp
   WHERE cp.comp_cd='COMP_STK_PRINT' AND cp.apply_ymd=v.apply_ymd
     AND cp.siz_cd=v.siz_cd AND cp.mat_cd=v.mat_cd AND cp.min_qty=v.min_qty
);
""")

print(f"S1={len(s1)} rows · S3={len(s3)} rows · (S2/S8 = 정적 hand SQL)")
