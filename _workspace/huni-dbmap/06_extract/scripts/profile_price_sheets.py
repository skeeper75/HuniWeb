#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""가격표 단가시트 구조 프로파일러 (round-2, 무손실 추출 선행 진단).

목적: 가격표 19시트 각각의 구조(밴드헤더 깊이·머지·수량축·옵션축·데이터영역·숨김)를
무손실로 진단해 시트별 추출 전략을 결정하기 위한 객관적 카탈로그를 만든다.
adapter/transform 없음 — 순수 read-only 구조 사실만. 추출·평면화는 후속 스크립트.

산출: 06_extract/price-sheets-profile.json (구조화) + stdout 요약.
"""
from __future__ import annotations

import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = "/Users/innojini/Dev/HuniWeb"
XLSX = os.path.join(ROOT, "docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx")
OUT = os.path.join(ROOT, "_workspace/huni-dbmap/06_extract")

# 비가격(2) + 제외(1)은 참고 표기만, 매핑대상 16시트가 본 대상
NON_PRICE = {"판걸이수", "굿즈파우치(구간할인)"}
EXCLUDED = {"후가공_박(백업)"}


def cellval(v: object) -> object:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def profile_sheet(ws_d, ws_f) -> dict:
    """한 시트의 구조 사실을 추출(값셀·머지·헤더밴드·축후보·숨김·데이터영역)."""
    max_r, max_c = ws_d.max_row, ws_d.max_column
    merged = [str(m) for m in ws_d.merged_cells.ranges]

    # 상단 헤더영역(최대 8행) 셀 그리드 — 밴드헤더 깊이 진단
    head_grid: list[list] = []
    for r in range(1, min(max_r, 8) + 1):
        row = []
        for c in range(1, min(max_c, 30) + 1):
            v = cellval(ws_d.cell(r, c).value)
            if v is not None:
                row.append(f"{get_column_letter(c)}{r}={v}")
        head_grid.append(row)

    # A열 패턴(수량축 후보) — 처음 40개 비어있지 않은 값
    col_a = []
    for r in range(1, min(max_r, 200) + 1):
        v = cellval(ws_d.cell(r, 1).value)
        if v is not None:
            col_a.append(v)
        if len(col_a) >= 40:
            break

    # 숨김 행/열
    hidden_rows = [r for r in range(1, max_r + 1)
                   if (ws_d.row_dimensions.get(r) and ws_d.row_dimensions[r].hidden)]
    hidden_cols = [get_column_letter(c) for c in range(1, max_c + 1)
                   if (ws_d.column_dimensions.get(get_column_letter(c))
                       and ws_d.column_dimensions[get_column_letter(c)].hidden)]

    # 수식·코멘트 신호 (full 모드)
    formula_cells = 0
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            fv = ws_f.cell(r, c).value
            if isinstance(fv, str) and fv.startswith("="):
                formula_cells += 1

    # 데이터 영역 추정: 숫자(가격/수량)가 처음/마지막 등장하는 행·열
    first_num_rc = None
    num_rows = set()
    num_cols = set()
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws_d.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if first_num_rc is None:
                    first_num_rc = f"{get_column_letter(c)}{r}"
                num_rows.add(r)
                num_cols.add(c)
    data_region = None
    if num_rows:
        data_region = {
            "first_num_cell": first_num_rc,
            "num_row_min": min(num_rows), "num_row_max": max(num_rows),
            "num_col_min": get_column_letter(min(num_cols)),
            "num_col_max": get_column_letter(max(num_cols)),
            "num_row_count": len(num_rows),
        }

    return {
        "dims": ws_d.dimensions,
        "max_row": max_r, "max_col": max_c,
        "merged_count": len(merged),
        "merged_sample": merged[:14],
        "header_grid_top8": head_grid,
        "col_a_pattern": col_a,
        "hidden_rows_count": len(hidden_rows),
        "hidden_cols": hidden_cols,
        "formula_cells": formula_cells,
        "data_region": data_region,
    }


def main() -> None:
    wb_d = load_workbook(XLSX, data_only=True)
    wb_f = load_workbook(XLSX, data_only=False)
    out: dict[str, dict] = {}
    for name in wb_d.sheetnames:
        kind = ("non_price" if name in NON_PRICE
                else "excluded" if name in EXCLUDED else "price_target")
        prof = profile_sheet(wb_d[name], wb_f[name])
        prof["sheet"] = name
        prof["kind"] = kind
        out[name] = prof

    with open(os.path.join(OUT, "price-sheets-profile.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # stdout 요약 — 시트별 1줄
    print(f"{'#':>2} {'kind':<12} {'sheet':<22} {'dims':<12} {'merge':>5} "
          f"{'hidR':>4} {'hidC':>4} {'fml':>4}  data_region")
    for i, (name, p) in enumerate(out.items()):
        dr = p["data_region"]
        drs = (f"{dr['first_num_cell']}→{dr['num_col_max']}{dr['num_row_max']}"
               f"({dr['num_row_count']}r)" if dr else "—")
        print(f"{i:>2} {p['kind']:<12} {name:<22} {p['dims']:<12} "
              f"{p['merged_count']:>5} {p['hidden_rows_count']:>4} "
              f"{len(p['hidden_cols']):>4} {p['formula_cells']:>4}  {drs}")


if __name__ == "__main__":
    main()
