#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
가격표 16 단가시트 무손실 L1 추출 (round-2, 추출+구조분석 단계 전용).

[원칙 — E 베스트프랙티스 / G 기준서 계승]
- 추출(extract)과 해석(interpret) 물리 분리. 본 스크립트는 1단계(무손실 충실추출)만.
  평면화(component_prices long-format)·차원 확정·적재 CSV 생성 금지(다음 단계).
- "어떤 셀도 조용히 버리지 않는다." 빈셀·유효0·★·숨김·머지·수식·코멘트 전부 보존.
- 한 시트에 여러 논리블록이 세로(또는 가로)로 stacking → 블록경계 탐지가 핵심 난도.

[블록 모델]
- 블록 = 제목셀(블록타이틀, A열 또는 임의 컬럼의 비수치 문자열)로 anchored 된 직사각 영역.
- 블록은 세로 stacking + 가로 동거(side-by-side) 둘 다 가능(예: 인쇄후가공 오시|미싱, 박 면적정의|수량별가격).
- 각 블록: block_id · block_title · row_range · col_range.
- 밴드헤더(다단 1~N행): 머지 풀어 각 leaf 컬럼의 밴드경로 합성(예 "칼라(CMYK) > 양면").

[무손실 보증 = 셀단위 long 산출]
- <slug>-l1.csv = 전 nonempty 셀 1행1레코드(sheet·block_id·block_title·row_seq·col·cell_ref·
  band_header_path·row_key·value·cell_meta_json). 블록탐지가 틀려도 셀은 전부 살아있음(안전망).
- <slug>-l1-meta.csv = 셀/시트 메타(머지·숨김·수식·코멘트·★·배경/글자색).
- verify_price_sheets.py 가 9게이트(전 셀 보존·roundtrip diff 0·머지/숨김/수식/코멘트 카운트)로 검증.

read-only. 원본·DB 무변경. 산출만 06_extract/.
"""
import os
import sys
import csv
import json
import re

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = "/Users/innojini/Dev/HuniWeb"
XLSX = os.path.join(ROOT, "docs/huni/후니프린팅_인쇄상품_가격표_260527.xlsx")
OUT = os.path.join(ROOT, "_workspace/huni-dbmap/06_extract")

sys.path.insert(0, os.path.dirname(__file__))
from extract_l1 import load_threaded_comments  # noqa: E402  스레드댓글 XML 직접 파서 재사용

# ── 16 단가시트 슬러그 (출력소재 IMPORT·판걸이수·굿즈파우치구간할인·후가공_박(백업) 제외) ──
SHEET_SLUGS = {
    "디지털인쇄비": "digital-print-price",
    "코팅": "coating",
    "접지옵션": "folding",
    "인쇄후가공": "post-process",
    "커팅타공": "cutting",
    "스티커": "sticker-price",
    "합판도무송스티커": "gangpan-sticker",
    "봉투제작": "envelope",
    "명함포토카드": "namecard-photocard",
    "후가공_박(소형)": "foil-small",
    "엽서북떡메": "postcard-book",
    "제본": "binding",
    "후가공_박(대형)": "foil-large",
    "아크릴": "acrylic-price",
    "포스터사인": "poster-sign",
}
# 스티커 슬러그는 상품마스터 sticker-l1 과 충돌 → 가격표는 -price 접미.
# 아크릴도 상품마스터 acrylic-l1 과 충돌 → acrylic-price.


# ════════════════════════════════════════════════════════════
# 셀 메타 (8축) — extract_price.cell_meta_for 와 동일 규약
# ════════════════════════════════════════════════════════════
def cell_meta_for(ws_f, ws, r, c, comments, col_hidden, row_hidden_flag, merged_anchor):
    """셀 1개의 메타: 배경색·글자색·코멘트·수식·하이퍼링크·★·숨김열·숨김행·머지앵커."""
    ref = get_column_letter(c) + str(r)
    cell = ws_f.cell(r, c)
    # fill 배경색 = 품절/신규/제약 신호. rgb는 ARGB 그대로, theme은 themeN으로 보존
    # (theme/indexed 컬러는 .rgb 접근 시 openpyxl이 예외표현 문자열을 반환하므로 type 분기).
    fg = None
    try:
        if cell.fill and cell.fill.patternType:
            fgc = cell.fill.fgColor
            t = getattr(fgc, "type", None)
            if t == "rgb" and isinstance(fgc.rgb, str):
                fg = fgc.rgb
            elif t == "theme":
                tint = getattr(fgc, "tint", 0) or 0
                fg = f"theme{fgc.theme}" + (f"/t{round(tint, 2)}" if tint else "")
    except Exception:
        fg = None
    if fg in ("00000000", "FFFFFFFF"):
        fg = None
    font_rgb = None
    try:
        # theme/indexed 컬러는 .rgb 접근 시 openpyxl이 예외표현 문자열을 반환하므로
        # type=='rgb'(실제 ARGB 지정)일 때만 채택 — 그 외(theme/auto)는 메타 미생성.
        fcol = cell.font.color if cell.font else None
        if fcol is not None and getattr(fcol, "type", None) == "rgb" and isinstance(fcol.rgb, str):
            fr = fcol.rgb
            if fr not in ("FF000000", "00000000"):
                font_rgb = fr
    except Exception:
        font_rgb = None
    cmt = comments.get(ref)
    val = ws.cell(r, c).value
    has_star = bool(val and isinstance(val, str) and "★" in val)
    fcell = cell.value
    is_formula = bool(isinstance(fcell, str) and fcell.startswith("="))
    formula = fcell if is_formula else ""
    hyperlink = ""
    try:
        if cell.hyperlink is not None and cell.hyperlink.target:
            hyperlink = cell.hyperlink.target
    except Exception:
        hyperlink = ""
    m = {}
    if fg:
        m["fill_rgb"] = fg
    if font_rgb:
        m["font_rgb"] = font_rgb
    if cmt:
        m["comment_author"] = cmt["author"]
        m["comment_text"] = cmt["text"]
    if has_star:
        m["has_constraint_star"] = True
    if is_formula:
        m["is_formula"] = True
        m["formula"] = formula
    if hyperlink:
        m["hyperlink"] = hyperlink
    if col_hidden:
        m["col_hidden"] = True
    if row_hidden_flag:
        m["row_hidden"] = True
    if merged_anchor:
        m["merged_range"] = merged_anchor
    return m


def col_hidden_map(ws, maxc):
    out = {}
    for c in range(1, maxc + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        out[c] = bool(dim.hidden) if dim is not None else False
    return out


def row_hidden_map(ws, maxr):
    out = {}
    for r in range(1, maxr + 1):
        dim = ws.row_dimensions.get(r)
        out[r] = bool(dim.hidden) if dim is not None else False
    return out


def cellval(v):
    """CSV 저장값 — 정수형 float -> int 표기(엑셀 1.0=1). 공백 None -> ''. 타입강제 없음."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def norm(v):
    """원본 셀값을 추출 CSV 문자열과 동일 정규화(roundtrip 대조용)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


# ════════════════════════════════════════════════════════════
# 블록 탐지
# ════════════════════════════════════════════════════════════
NUMRE = re.compile(r"^-?\d+(\.\d+)?$")
DIMRE = re.compile(r"^\d+(\.\d+)?\s*mm$")              # "20mm" 류 면적축값
RANGERE = re.compile(r"^\d+\s*~\s*\d+$")               # "1~49" 구간문자열
CLASSRE = re.compile(r"^[A-E]$")                        # 박 면적정의 분류문자 A~E
WHCOMBO = re.compile(r"^\d+\s*[xX*]\s*\d+")             # "1000x1000 : 20000" note
PAPERSIZE = re.compile(r"^A[0-9]([ /].*)?$")           # "A3"/"A2"/"A1"/"A4"/"A5" 규격사이즈 축값

# 행키헤더/옵션밴드 라벨 (블록 제목 아님) — 이 라벨이 leftmost 에 오면 헤더행 신호이지 제목 아님
BAND_KEY_LABELS = {
    "수량", "제작수량", "옵션", "분류 / 수량", "분류/수량", "가로 / 세로", "가로/세로",
    "소재 / 수량(국4절)", "소재 / 제작수량", "소재 / 수량", "세트제작수량", "총제작수량",
    "주문수량", "접지옵션명/ 제작수량", "옵션/ 제작수량", "제본/수량", "수량(국4절)",
    "수량(3절)", "페이지 / 수량", "사이즈", "인쇄", "단면", "양면", "가격",
    "제본/ 수량", "수량구간", "할인율", "기본가(아연판)",
}


def is_numeric_cell(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def is_axis_or_band_value(s):
    """면적축값(mm)·분류문자(A~E)·구간문자(1~49)·규격사이즈(A3/A2/A1)·band_key 라벨
    — 제목 아님(블록 내부 축/헤더)."""
    if (DIMRE.match(s) or RANGERE.match(s) or CLASSRE.match(s)
            or WHCOMBO.match(s) or PAPERSIZE.match(s)):
        return True
    if s in BAND_KEY_LABELS:
        return True
    return False


def is_title_string(v):
    """제목 '문자열' 1차 필터: 비수치·비축값·비밴드라벨 문자열. (위치 검증은 detect_blocks)."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    if not s or NUMRE.match(s):
        return False
    if is_axis_or_band_value(s):
        return False
    return True


def _row_is_band_header(ws_d, r, maxc):
    """행 r 이 '밴드헤더 행'인가: 행 어딘가에 band_key 라벨(옵션/소재.../제작수량/단면/양면/
    가로 / 세로 등)이 등장하면 그 행 전체는 헤더행으로 본다 → 그 행의 비좌측 문자열은
    밴드leaf(제목 아님). 면적정의 'A~E'·옵션명도 이 행에 묶여 leaf 로 처리됨."""
    for c in range(1, maxc + 1):
        v = ws_d.cell(r, c).value
        if isinstance(v, str) and v.strip() in BAND_KEY_LABELS:
            return True
    return False


def _heads_a_region(ws_d, r, c, maxr, maxc):
    """(r,c) 제목후보가 '영역을 이끄는가' 검증.
    조건:
      (a) 제목 자신의 행 r 은 밴드헤더행이 아니어야 함(헤더행이면 비좌측 문자열은 leaf).
      (b) 같은 컬럼 r+1..r+4 안에 band_key 라벨/수치/mm행키가 등장(영역 인도).
    옆칸 note(단/양면 leaf, 코팅포함가, 인쇄배경지, *아연판) 및 밴드leaf(옵션명/소재명)는 배제."""
    # (a) 제목행이 밴드헤더행이면, 좌측에 band_key 가 있는데 이 셀이 그 우측이면 leaf → 거부.
    if _row_is_band_header(ws_d, r, maxc):
        for cc in range(1, c):
            v = ws_d.cell(r, cc).value
            if isinstance(v, str) and v.strip() in BAND_KEY_LABELS:
                return False  # 이 행은 헤더행, 본 셀은 밴드 leaf
    # (b) 영역 인도: 같은 컬럼 아래 band_key/수치/mm 등장
    for rr in range(r + 1, min(r + 5, maxr + 1)):
        v = ws_d.cell(rr, c).value
        if is_numeric_cell(v):
            return True
        if isinstance(v, str) and v.strip() in BAND_KEY_LABELS:
            return True
        if isinstance(v, str) and DIMRE.match(v.strip()):
            return True
    return False


def _first_numeric_row(ws_d, r_title, c, maxr, maxc):
    """제목행 아래 첫 수치 데이터행(블록 헤더존 끝). 없으면 None."""
    for rr in range(r_title + 1, maxr + 1):
        for cc in range(c, maxc + 1):
            if is_numeric_cell(ws_d.cell(rr, cc).value):
                return rr
    return None


def detect_blocks(ws, ws_d, maxr, maxc, merged_ranges):
    """블록 경계 탐지 (2패스: 1차 앵커 → 헤더존 leaf 제거).

    pass1: 제목후보 = is_title_string AND _heads_a_region.
    pass2: 다른 블록의 '헤더존'(제목행+1 ~ 첫 수치데이터행-1) 안에 있는 비좌측 제목후보는
           밴드 leaf(다단헤더 흑백/칼라, A3/A2/A1, 화이트(단면)+클리어)이므로 제거.
           단 컬럼A 후보(세로 stacking 블록)와 가로 동거 분할 컬럼(머지로 다컬럼 스팬 헤더)은 유지.

    경계:
      - 우경계 c1 = 같은 제목행대(±1행) 우측 다음 제목 컬럼 -1, 없으면 maxc(가로 동거 분리).
      - 행범위 r1 = 컬럼 겹치는 다음 세로 제목행 -1, 없으면 maxr.
    """
    raw = []  # (r, c, title)
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = ws_d.cell(r, c).value
            if is_title_string(v) and _heads_a_region(ws_d, r, c, maxr, maxc):
                raw.append((r, c, str(v).strip()))

    # 컬럼A(또는 행 좌측 최초 데이터컬럼) 후보 = primary 앵커 → 헤더존 산정 기준
    # 각 primary 의 헤더존 = (r_title, first_numeric_row) 구간행 × [c, 우측 동거경계]
    primaries = [(r, c, t) for (r, c, t) in raw if c == 1]
    # 컬럼A가 비어 제목이 우측에 단독으로 있는 시트(없음 — 전 시트 A열 제목 보유)도 대비:
    if not primaries:
        primaries = raw[:]
    header_zones = []  # (r0, r1, c0, c1) 헤더존(이 안의 비좌측 후보는 leaf)
    for (r, c, t) in primaries:
        fn = _first_numeric_row(ws_d, r, c, maxr, maxc)
        zr1 = (fn - 1) if fn else r  # 수치 없으면 헤더존 없음(제목행만)
        header_zones.append((r, zr1, c))

    def in_header_zone(r, c):
        for (zr0, zr1, zc0) in header_zones:
            if zr0 < r <= zr1 and c > zc0:  # 제목행 다음~데이터행 전, 좌측앵커 우측
                return True
        return False

    titles = [(r, c, t) for (r, c, t) in raw
              if c == 1 or not in_header_zone(r, c)]

    titles.sort(key=lambda t: (t[0], t[1]))
    n = len(titles)
    blocks = []
    bid = 0
    for i, (r, c, title) in enumerate(titles):
        # 우경계: 같은 제목행대(|행차|<=1) 우측 다음 제목 컬럼 -1
        c1 = maxc
        for j in range(n):
            if j == i:
                continue
            rj, cj, _ = titles[j]
            if abs(rj - r) <= 1 and cj > c and (cj - 1) < c1:
                c1 = cj - 1
        # 행범위: 컬럼 겹치는 다음 세로 제목행 -1
        r1 = maxr
        for j in range(i + 1, n):
            rj, cj, _ = titles[j]
            if rj > r and c <= cj <= c1:
                r1 = rj - 1
                break
        bid += 1
        blocks.append({
            "block_id": f"B{bid:02d}",
            "title": title,
            "r_title": r,
            "r0": r,
            "r1": r1,
            "c0": c,
            "c1": c1,
        })
    return blocks


def assign_block(blocks, r, c):
    """셀 (r,c) 를 감싸는 가장 좁은(작은 면적) 블록 id 반환. 없으면 ''. """
    best = None
    best_area = None
    for b in blocks:
        if b["r0"] <= r <= b["r1"] and b["c0"] <= c <= b["c1"]:
            area = (b["r1"] - b["r0"] + 1) * (b["c1"] - b["c0"] + 1)
            if best_area is None or area < best_area:
                best_area = area
                best = b
    return best["block_id"] if best else ""


# ════════════════════════════════════════════════════════════
# 밴드헤더 합성
# ════════════════════════════════════════════════════════════
def build_merge_lookup(merged_ranges):
    """(r,c) -> 'A1:O1' 머지범위문자열 (셀이 머지에 속하면). 좌상단 포함 전 셀 매핑."""
    lut = {}
    anchor = {}  # (r,c)->(ar,ac) 좌상단
    for mr in merged_ranges:
        s = str(mr)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                lut[(r, c)] = s
                anchor[(r, c)] = (mr.min_row, mr.min_col)
    return lut, anchor


def synth_band_path(ws_d, block, header_rows, leaf_col, merge_anchor):
    """블록 헤더행들(header_rows)에서 leaf_col 의 밴드경로 합성.
    상위밴드(머지로 여러 컬럼 묶임) 값은 좌상단에만 → merge_anchor 로 끌어옴.
    경로는 위→아래 헤더행 순서, '>' join. 빈 단계는 생략."""
    parts = []
    for hr in header_rows:
        v = ws_d.cell(hr, leaf_col).value
        if v in (None, ""):
            # 머지 상위밴드면 좌상단 값
            a = merge_anchor.get((hr, leaf_col))
            if a is not None:
                v = ws_d.cell(a[0], a[1]).value
        if v not in (None, ""):
            s = str(v).strip()
            if s and (not parts or parts[-1] != s):
                parts.append(s)
    return " > ".join(parts)


# ════════════════════════════════════════════════════════════
# 시트 추출 (셀단위 long + 블록/밴드 주석)
# ════════════════════════════════════════════════════════════
def extract_sheet(sheet, slug):
    wb_d = openpyxl.load_workbook(XLSX, data_only=True)
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    ws_d, ws_f = wb_d[sheet], wb_f[sheet]
    comments = load_threaded_comments(XLSX, sheet)

    # 실제 데이터가 있는 최종 행/열 (트레일링 빈 영역 제외하되, 값 있는 셀은 전부 포함)
    last_r = 0
    last_c = 0
    for r in range(1, ws_d.max_row + 1):
        for c in range(1, ws_d.max_column + 1):
            if ws_d.cell(r, c).value not in (None, ""):
                if r > last_r:
                    last_r = r
                if c > last_c:
                    last_c = c
    maxr, maxc = last_r, last_c

    merged_ranges = list(ws_d.merged_cells.ranges)
    merge_lut, merge_anchor = build_merge_lookup(merged_ranges)
    ch = col_hidden_map(ws_d, maxc)
    rh = row_hidden_map(ws_d, maxr)

    blocks = detect_blocks(ws_f, ws_d, maxr, maxc, merged_ranges)

    # 블록별 헤더행 추정: 제목행 다음 ~ 첫 수치행키 전까지를 헤더밴드로,
    # 행키 컬럼(블록 c0)에서 첫 수치 등장행을 데이터 시작으로 본다.
    block_header_rows = {}
    block_data_start = {}
    for b in blocks:
        c0 = b["c0"]
        hdr = []
        data0 = None
        for r in range(b["r_title"] + 1, b["r1"] + 1):
            v = ws_d.cell(r, c0).value
            if is_numeric_cell(v):
                data0 = r
                break
            # 행키 컬럼이 비수치면 헤더행 후보 (단 완전 빈행은 스킵)
            rowvals = [ws_d.cell(r, c).value for c in range(b["c0"], b["c1"] + 1)]
            if any(x not in (None, "") for x in rowvals):
                hdr.append(r)
        block_header_rows[b["block_id"]] = hdr
        block_data_start[b["block_id"]] = data0
    bmap = {b["block_id"]: b for b in blocks}

    # ── 셀단위 long 레코드 + 메타 ──
    records = []   # 데이터(값 보존) 행
    meta_rows = []  # 셀 메타
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            v = ws_d.cell(r, c).value
            ref = get_column_letter(c) + str(r)
            merged_anchor_str = merge_lut.get((r, c), "")
            m = cell_meta_for(ws_f, ws_d, r, c, comments, ch.get(c, False),
                              rh.get(r, False), merged_anchor_str)
            # 메타가 있으면 항상 기록(머지·숨김·수식·코멘트·★·색)
            if m:
                meta_rows.append((r, c, ref, m))
            # 값 레코드는 nonempty 셀만 (무손실 핵심: 전 nonempty 보존)
            if v in (None, ""):
                continue
            bid = assign_block(blocks, r, c)
            b = bmap.get(bid)
            block_title = b["title"] if b else ""
            # 행키 = 블록 행키컬럼(c0) 값 (이 셀이 행키컬럼이면 자기 자신)
            row_key = ""
            band_path = ""
            if b:
                rkv = ws_d.cell(r, b["c0"]).value
                row_key = cellval(rkv) if rkv not in (None, "") else ""
                hdr = block_header_rows.get(bid, [])
                # leaf 밴드경로 — 헤더행 있고 이 셀이 데이터행/값컬럼일 때
                if hdr and c > b["c0"]:
                    band_path = synth_band_path(ws_d, b, hdr, c, merge_anchor)
            records.append({
                "block_id": bid,
                "block_title": block_title,
                "row_seq": r,
                "col": get_column_letter(c),
                "cell_ref": ref,
                "row_key": row_key,
                "band_header_path": band_path,
                "value": cellval(v),
                "meta": m,
            })

    # ── write l1 csv (cell-level long) ──
    out_csv = os.path.join(OUT, f"price-{slug}-l1.csv")
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "block_id", "block_title", "row_seq", "col", "cell_ref",
                    "row_key", "band_header_path", "value", "cell_meta_json"])
        for rec in records:
            w.writerow([sheet, rec["block_id"], rec["block_title"], rec["row_seq"],
                        rec["col"], rec["cell_ref"], rec["row_key"],
                        rec["band_header_path"], rec["value"],
                        json.dumps(rec["meta"], ensure_ascii=False) if rec["meta"] else ""])

    # ── write meta csv (cell/sheet meta) ──
    out_meta = os.path.join(OUT, f"price-{slug}-l1-meta.csv")
    with open(out_meta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "type", "row_seq", "col", "ref", "block_id", "meta_json"])
        # 숨김열 시트레벨
        for c in range(1, maxc + 1):
            if ch.get(c):
                w.writerow([sheet, "hidden_column", "", get_column_letter(c),
                            get_column_letter(c) + "1", "",
                            json.dumps({"col_hidden": True}, ensure_ascii=False)])
        # 블록 인벤토리 (구조분석 입력)
        for b in blocks:
            inv = {
                "title": b["title"], "r_title": b["r_title"],
                "row_range": [b["r0"], b["r1"]],
                "col_range": [get_column_letter(b["c0"]), get_column_letter(b["c1"])],
                "header_rows": block_header_rows.get(b["block_id"], []),
                "data_start_row": block_data_start.get(b["block_id"]),
            }
            w.writerow([sheet, "block", "", "", "", b["block_id"],
                        json.dumps(inv, ensure_ascii=False)])
        # 셀 메타
        for (r, c, ref, m) in meta_rows:
            bid = assign_block(blocks, r, c)
            w.writerow([sheet, "cell", r, get_column_letter(c), ref, bid,
                        json.dumps(m, ensure_ascii=False)])

    # ── summary ──
    nonempty = sum(1 for r in range(1, maxr + 1) for c in range(1, maxc + 1)
                   if ws_d.cell(r, c).value not in (None, ""))
    formula_cells = sum(1 for (r, c, ref, m) in meta_rows if m.get("is_formula"))
    comment_cells = sum(1 for (r, c, ref, m) in meta_rows if m.get("comment_text"))
    summary = {
        "sheet": sheet, "slug": slug,
        "max_row": maxr, "max_col": maxc,
        "nonempty_cells": nonempty,
        "record_count": len(records),
        "block_count": len(blocks),
        "blocks": [{"block_id": b["block_id"], "title": b["title"],
                    "row_range": [b["r0"], b["r1"]],
                    "col_range": [get_column_letter(b["c0"]), get_column_letter(b["c1"])],
                    "header_rows": block_header_rows.get(b["block_id"], []),
                    "data_start_row": block_data_start.get(b["block_id"])}
                   for b in blocks],
        "merged_count": len(merged_ranges),
        "hidden_cols": [get_column_letter(c) for c in range(1, maxc + 1) if ch.get(c)],
        "hidden_rows": [r for r in range(1, maxr + 1) if rh.get(r)],
        "formula_cells": formula_cells,
        "comment_cells": comment_cells,
        "out_csv": os.path.basename(out_csv),
        "out_meta": os.path.basename(out_meta),
    }
    return summary


def main():
    out = {}
    for sheet, slug in SHEET_SLUGS.items():
        s = extract_sheet(sheet, slug)
        out[sheet] = s
        print(f"[{slug:<20}] rows={s['max_row']:>4} cols={s['max_col']:>3} "
              f"nonempty={s['nonempty_cells']:>5} recs={s['record_count']:>5} "
              f"blocks={s['block_count']:>2} merge={s['merged_count']:>3} "
              f"fml={s['formula_cells']:>3} cmt={s['comment_cells']:>2}")
    with open(os.path.join(OUT, "_price-sheets-extract-raw.json"), "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("\n[written] _price-sheets-extract-raw.json")


if __name__ == "__main__":
    main()
