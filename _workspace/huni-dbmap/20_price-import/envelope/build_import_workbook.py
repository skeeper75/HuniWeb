# -*- coding: utf-8 -*-
"""round-16 봉투제작 webadmin 복붙용 import.xlsx 빌더.

입력 = /tmp/env_export/*.psv (라이브 t_prc_* read-only 실측 덤프, 2026-06-13).
출력 = envelope-import.xlsx (테이블별 시트, 1행=DB컬럼명, 2행=한글라벨).

[중요] 이 시트(봉투제작)는 라이브에 이미 적재됨 → 그릇은 '재현(RU)'(대조용·재적재 아님).
이번 round-16 6시트 중 유일하게 가격사슬 완전 정합(단절 0).
봉투종류 차원 = siz_cd(라이브 모델·opt_cd/별상품 아님).
세트형 부재 — 봉투 자체 제작 단가표(round-16 세트형 부재 결론).
복합셀 "레자크체크 / 레자크줄무늬"의 줄무늬(169) 미적재 = 4b_FIX(주황·컨펌 선결).
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

EXPORT = "/tmp/env_export"
OUT = (
    "/Users/innojini/Dev/HuniWeb/_workspace/huni-dbmap/20_price-import/"
    "envelope/envelope-import.xlsx"
)

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
LBL_FILL = PatternFill("solid", fgColor="DDEBF7")
LBL_FONT = Font(bold=True, color="1F4E79", size=9)
BLK_FILL = PatternFill("solid", fgColor="FCE4D6")  # 단절/교정 시트 표식(주황)
RU_FILL = PatternFill("solid", fgColor="E2EFDA")  # 재현(라이브 기존) 표식(초록)

# 봉투종류 → siz_cd 매핑 (라이브 실측·가격 대조 확정)
SIZ_MAP = {
    "SIZ_000191": "티켓봉투 (225x193)",
    "SIZ_000192": "소봉투 (238x262)",
    "SIZ_000193": "자켓봉투 (262x238)",
    "SIZ_000194": "대봉투 (510x387)",
}
# 소재 → 라이브 mat_cd
MAT_MAP = {
    "MAT_000159": "모조 120g",
    "MAT_000168": "레자크체크백색 110g",
    "MAT_000169": "레자크줄무늬백색 110g",
}


def read_psv(name: str) -> list[list[str]]:
    rows: list[list[str]] = []
    with open(f"{EXPORT}/{name}", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            rows.append(line.split("|"))
    return rows


def empty_to_none(v: str):
    return None if v == "" else v


def write_sheet(
    wb, title, columns, labels, data_rows, note=None, fill=None, col_comments=None
):
    ws = wb.create_sheet(title)
    if note:
        ws.cell(row=1, column=1, value=note)
        ws.cell(row=1, column=1).font = Font(italic=True, color="C00000", size=9)
        hdr_r, lbl_r, data_r0 = 2, 3, 4
    else:
        hdr_r, lbl_r, data_r0 = 1, 2, 3
    for ci, (col, lbl) in enumerate(zip(columns, labels), start=1):
        hc = ws.cell(row=hdr_r, column=ci, value=col)
        hc.fill = fill or HDR_FILL
        hc.font = HDR_FONT
        hc.alignment = Alignment(horizontal="center", wrap_text=True)
        lc = ws.cell(row=lbl_r, column=ci, value=lbl)
        lc.fill = LBL_FILL
        lc.font = LBL_FONT
        lc.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_comments and col in col_comments:
            hc.comment = Comment(col_comments[col], "round-16")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(
            12, min(34, len(lbl) + 4)
        )
    for ri, row in enumerate(data_rows, start=data_r0):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = ws.cell(row=data_r0, column=1)
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# 0. README
# ============================================================
ws = wb.create_sheet("0_README")
readme = [
    ["round-16 봉투제작 가격표 import 그릇 (webadmin 복붙용)", ""],
    ["", ""],
    ["원본 시트", "후니프린팅_인쇄상품_가격표_260527.xlsx > 봉투제작 (index 9)"],
    ["그릇 권위", "라이브 t_prc_* information_schema 실측 (2026-06-13, read-only)"],
    ["구조", "단일 MATRIX 1블록 = (봉투종류 4) × (소재 2) × (수량 5) 단가표"],
    ["", ""],
    ["[평결1·세트형] 핸드오프 가설 '봉투제작=세트형(본품+봉투 1세트)' → 반증", ""],
    ["  · 가격축 = 봉투종류·소재·수량만 (결합 본품 차원 부재)", ""],
    ["  · t_prd_product_sets에 PRD_000050 미참조 (세트 구성요소 아님)", ""],
    ["  · 가격 = 봉투 자체 제작 단가 (장당 정액 선형·96원/매 고정)", ""],
    ["  · [round-16 세트형 부재 결론] 엽서북떡메·박·제본·봉투제작 전건 검사 완료 →", ""],
    ["       세트형(본품+부속 1세트 고정가) 구조는 가격표 6시트에서 발견 안 됨", ""],
    ["", ""],
    ["[평결2·합가형] PRICE_TYPE.02 합가형 (엔진 권위 종결·라이브 .01 오적재)", ""],
    ["  · 가격 = 수량할인 0·선형 비례 (96000→192000→288000)지만 단위가 구간총액", ""],
    ["  · unit_price=구간총액(96000=1000매 전체가) → .01단가형이면 96000×주문수 폭증", ""],
    ["  · .02합가형(총액=구간가)이어야 192,000원 정합 → 라이브 현재 .01 = 오적재(교정 대상)", ""],
    ["  · 3_price_components 시트는 .02로 교정 반영 — 복붙 시 .02·엔진 시뮬 192,000원 회귀 후 인간 승인", ""],
    ["", ""],
    ["[봉투종류 차원 = siz_cd] 별상품/opt_cd 아님 (라이브 모델)", ""],
    ["  · 티켓SIZ_000191·소SIZ_000192·자켓SIZ_000193·대SIZ_000194 (가격대조 확정)", ""],
    ["  · siz_nm이 치수('225x193')로 등록 — 봉투종류명 라벨 GAP (ENV-C2)", ""],
    ["  · 비교: 스티커=opt_cd/coat / 박=opt_cd(등급) / 제본=comp_cd / 봉투=siz_cd", ""],
    ["", ""],
    ["[🟢 가격사슬 완전 정합] round-16 6시트 중 유일한 완전체 (단절 0)", ""],
    ["  공식정의 PRF_ENV_MAKING → 바인딩 PRD_000050 → 배선 COMP_ENV_MAKING", ""],
    ["       → 구성요소 COMP_ENV_MAKING → 단가행 40 (전 사슬 라이브 완결·엔진 조회 가능)", ""],
    ["  · 아크릴(배선0)·제본(배선1/11·바인딩4) 같은 단절 없음", ""],
    ["", ""],
    ["[🔴 복합셀 collapse 결함] '레자크체크 / 레자크줄무늬' 줄무늬 미적재", ""],
    ["  · 가격표 C열 = 2소재 동일가(체크·줄무늬 모두 111000…)", ""],
    ["  · 라이브 = 레자크체크(MAT_000168)만 20행 적재, 레자크줄무늬(MAT_000169) 0행", ""],
    ["  · 손님이 줄무늬 선택 시 엔진 단가행 못 찾음 (부분 단절)", ""],
    ["  · 교정 = 4b_FIX (체크와 동일가 20행) — 단, '줄무늬 선택가능?' 컨펌 ENV-C1 선결", ""],
    ["", ""],
    ["[경고] 라이브 기존 적재 — 재적재 금지", "이미 라이브 COMMIT 40행. 이 그릇은 재현(대조용)."],
    ["  · 초록 시트(_RU) = 라이브 기존 재현 — 신규 적재 아님", ""],
    ["  · 주황 시트(_FIX) = 누락 교정 제안 — 인간 승인 후·컨펌 선결", ""],
    ["", ""],
    ["시트 구성", ""],
    ["  1_price_formulas_RU", "공식정의 1 (PRF_ENV_MAKING·단순형·라이브 재현)"],
    ["  1b_product_price_formulas_RU", "상품↔공식 바인딩 1 (PRD_000050·라이브 재현)"],
    ["  2_formula_components_RU", "공식 배선 1 (COMP_ENV_MAKING·라이브 재현)"],
    ["  3_price_components", "구성요소 정의 1 (봉투제작 완제품가·.02합가형 교정·라이브 .01 오적재)"],
    ["  4_component_prices_RU", "단가행 40 (이 시트 원천·라이브 재현·무손실)"],
    ["  4b_FIX_material_chain", "복합셀 누락 교정 후보 — 레자크줄무늬 20행 (⚠️ 컨펌 ENV-C1 선결)"],
    ["", ""],
    ["분해 무손실 검산", "시트 데이터셀 40 (4봉투종류 × 2소재 × 5수량) = 라이브 40행 일치(mismatch 0)"],
    ["봉투종류↔siz", "4/4 매핑(가격 대조 확정) / 소재 모조·체크 적재·줄무늬 미적재(FIX)"],
]
for ri, (a, b) in enumerate(readme, start=1):
    ws.cell(row=ri, column=1, value=a)
    ws.cell(row=ri, column=2, value=b)
    if ri == 1:
        ws.cell(row=ri, column=1).font = Font(bold=True, size=12, color="1F4E79")
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 76

# ============================================================
# 1. price_formulas (라이브 재현)
# ============================================================
frm = read_psv("formulas.psv")
write_sheet(
    wb,
    "1_price_formulas_RU",
    ["frm_cd", "frm_nm", "use_yn", "note"],
    ["공식코드", "공식명", "사용여부(Y/N)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in frm],
    note="[RU] 라이브 t_prc_price_formulas 재현 — frm_typ_cd·prd_cd 컬럼 없음(라이브 실측). 신규 적재 아님.",
    fill=RU_FILL,
    col_comments={
        "frm_cd": "공식 식별자 PK. PRF_ENV_MAKING=봉투제작 단순형(구성요소 1개로 완결)",
        "use_yn": "Y=노출",
        "note": "단순형: 판매가=[수량행][소재열]. 봉투종류·소재는 component_prices 차원",
    },
)

# ============================================================
# 1b. product_price_formulas (바인딩, 라이브 재현)
# ============================================================
bind = read_psv("binding.psv")
write_sheet(
    wb,
    "1b_product_price_formulas_RU",
    ["prd_cd", "frm_cd", "apply_bgn_ymd", "note"],
    ["상품코드", "공식코드", "적용시작일(yyyy-MM-dd)", "비고"],
    [[r[0], r[1], r[2], r[3] if len(r) > 3 else ""] for r in bind],
    note="[RU] 라이브 t_prd_product_price_formulas 재현 — PRD_000050 봉투제작 1건 정상 바인딩(단절 없음).",
    fill=RU_FILL,
    col_comments={
        "prd_cd": "PRD_000050 봉투제작 (PRD_TYPE.04). 카드봉투(281/282/283)는 별 상품군",
        "apply_bgn_ymd": "PK 구성요소(prd_cd, frm_cd, apply_bgn_ymd). 적용 시작일 2026-06-01",
    },
)

# ============================================================
# 2. formula_components (배선, 라이브 재현)
# ============================================================
wire = read_psv("wiring.psv")
write_sheet(
    wb,
    "2_formula_components_RU",
    ["frm_cd", "comp_cd", "disp_seq", "addtn_yn"],
    ["공식코드", "구성요소코드", "표시순서", "합산여부(Y/N)"],
    [[r[0], r[1], int(r[2]) if r[2] else None, r[3]] for r in wire],
    note="[RU] 라이브 t_prc_formula_components 재현 — 1행 정상 배선(COMP_ENV_MAKING). 단절 없음.",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "COMP_ENV_MAKING 1개로 완결(봉투제작=완제품·합산 항 없음)",
        "addtn_yn": "Phase11 엔진 무시. 구성요소 1개라 합산 무관",
    },
)

# ============================================================
# 3. price_components (구성요소 정의 1, 라이브 재현)
# ============================================================
comp = read_psv("components.psv")
# [P4 보정] prc_typ_cd = .02 합가형으로 교정(엔진 권위 11-CONTEXT L17-18 종결).
#   라이브 unit_price=구간총액(96000=1000매가). .01단가형이면 96000×주문수 폭증 →
#   .02합가형(총액=구간가)이어야 192,000원 정합. 라이브 .01은 오적재(교정 대상·인간 승인).
comp_rows = []
for r in comp:
    prc_typ = "PRICE_TYPE.02"  # 교정값(라이브 현재 .01 오적재 → 복붙 시 .02)
    comp_rows.append(
        [r[0], r[1], r[2] if len(r) > 2 else "", prc_typ,
         r[4] if len(r) > 4 else "", r[5] if len(r) > 5 else "Y"]
    )
write_sheet(
    wb,
    "3_price_components",
    ["comp_cd", "comp_nm", "comp_typ_cd", "prc_typ_cd", "use_dims", "use_yn"],
    ["구성요소코드", "구성요소명", "구성요소유형", "단가유형(.01단가/.02합가)", "사용차원(jsonb)", "사용여부"],
    comp_rows,
    note="[P4 교정] prc_typ_cd=PRICE_TYPE.02 합가형(엔진 권위 종결). ⚠️ 라이브 현재 .01 오적재 — 복붙 시 .02로 교정하고 엔진 시뮬 192,000원 회귀 확인 후 인간 승인 적재. unit_price=구간총액(96000=1000매가)이므로 합가형(총액=구간가) 정합.",
    fill=BLK_FILL,
    col_comments={
        "comp_cd": "COMP_ENV_MAKING — 봉투제작 단일 구성요소(완제품가)",
        "prc_typ_cd": "[교정] .02 합가형(정답). unit_price=구간총액(96000=1000매가)이라 .01단가형이면 ×주문수 폭증 → .02여야 192,000원 정합. 라이브 현재 .01 = 오적재(교정 대상·인간 승인)",
        "use_dims": "[siz_cd,mat_cd,min_qty] 3차원 — 봉투종류(siz)·소재(mat)·수량구간(min_qty). 라이브 실측 일치",
        "comp_typ_cd": "PRC_COMPONENT_TYPE.06(완제품가). 봉투제작은 단일 항",
    },
)

# ============================================================
# 4. component_prices (단가행 40, 이 시트 원천·라이브 재현)
# ============================================================
cp = read_psv("component_prices.psv")
# columns: comp_cd|siz_cd|clr_cd|mat_cd|proc_cd|coat_side_cnt|opt_cd|bdl_qty|min_qty|apply_ymd|unit_price
cp_rows = []
for r in cp:
    cp_rows.append([
        r[0],
        empty_to_none(r[1]),
        empty_to_none(r[2]),
        empty_to_none(r[3]),
        empty_to_none(r[4]),
        int(r[5]) if r[5] else None,
        empty_to_none(r[6]),
        int(r[7]) if r[7] else None,
        int(r[8]) if r[8] else None,
        r[9],
        float(r[10]) if r[10] else None,
    ])
write_sheet(
    wb,
    "4_component_prices_RU",
    ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "봉투종류(siz)", "도수(clr)", "소재(mat)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수", "수량구간시작(min_qty)", "적용일", "단가(구간총액)"],
    cp_rows,
    note="[RU] 라이브 t_prc_component_prices 재현(40행) — 봉투 가격표=이 단가의 원천. 자연키 8차원(siz·clr·mat·proc·coat_side·opt·bdl·min_qty) + comp_cd + apply_ymd. 안 쓰는 차원 NULL(빈칸). use_dims=[siz_cd,mat_cd,min_qty]만 값. 레자크줄무늬(MAT_000169) 미적재=4b_FIX.",
    fill=RU_FILL,
    col_comments={
        "comp_cd": "COMP_ENV_MAKING × (4봉투종류 siz × 2소재 mat × 5수량 min_qty) = 40행",
        "siz_cd": "봉투종류 차원. 191티켓·192소·193자켓·194대 (siz_nm은 치수로 등록)",
        "mat_cd": "소재 차원. 159모조120g·168레자크체크 적재 / 169레자크줄무늬 미적재(4b_FIX)",
        "min_qty": "수량구간 시작값. 1000·2000·3000·4000·5000 (5구간 단조). 주문수량 이하 최대 min_qty 매칭",
        "proc_cd": "신설 공정 차원(8→10). 봉투 미사용 NULL",
        "opt_cd": "신설 옵션 차원(8→10). 봉투 미사용 NULL",
        "clr_cd": "전건 NULL — 도수 무관",
        "coat_side_cnt": "전건 NULL — 코팅 무관",
        "unit_price": "구간총액(96000=1000매 전체가). 합가형(.02)이라 총액=구간가로 192,000원(2000매) 정합. 라이브 prc_typ .01은 오적재(3_price_components .02 교정 참조)",
    },
)

# ============================================================
# 4b. FIX material chain — 레자크줄무늬(169) 미적재 교정 후보
# ============================================================
# 레자크체크(168) 행을 줄무늬(169)로 복제 (가격표 동일가 근거)
fix_mat = []
for r in cp_rows:
    if r[3] == "MAT_000168":  # 레자크체크
        nr = list(r)
        nr[3] = "MAT_000169"  # 레자크줄무늬
        fix_mat.append(nr)
write_sheet(
    wb,
    "4b_FIX_material_chain",
    ["comp_cd", "siz_cd", "clr_cd", "mat_cd", "proc_cd", "coat_side_cnt", "opt_cd", "bdl_qty", "min_qty", "apply_ymd", "unit_price"],
    ["구성요소", "봉투종류(siz)", "도수(clr)", "소재(mat=레자크줄무늬)", "공정(proc)", "코팅면수", "옵션(opt)", "묶음수", "수량구간시작(min_qty)", "적용일", "단가(구간총액)"],
    fix_mat,
    note="[FIX·복합셀 누락·⚠️그대로 INSERT 금지] 레자크줄무늬백색(MAT_000169) 20행 — 가격표 C열 '레자크체크/레자크줄무늬' 동일가라 체크(168) 단가 복제. 단 '봉투제작에서 줄무늬 선택가능?' 상품옵션 권위 확정 후 INSERT(컨펌 ENV-C1). 추정 적재 금지.",
    fill=BLK_FILL,
)

wb.save(OUT)
print("SAVED:", OUT)
print("Sheets:", wb.sheetnames)
print("component_prices RU rows:", len(cp_rows))
print("FIX material rows (레자크줄무늬):", len(fix_mat))
print("formulas:", len(frm), "binding:", len(bind), "wiring:", len(wire), "components:", len(comp))
