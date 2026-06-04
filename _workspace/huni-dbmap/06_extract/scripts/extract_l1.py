#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
L1 충실추출 (faithful extract) — 후니프린팅 상품마스터 엑셀.

G 기준서(_workspace/huni-dbmap/05_method/G-extraction-spec.md) 구현체.
- 전 컬럼(A~maxcol) 행 단위 무손실 보존. 속성별 단일컬럼 평면화 금지.
- 그룹헤더 composite 귀속: row1 가로병합 그룹명을 하위 컬럼에 ffill -> `그룹명_하위명`.
- 세로병합(r1:r2 단일컬럼)은 단순 컬럼.
- 작업/재단사이즈는 row2 헤더 매핑으로 식별(고정 컬럼레터 가정 금지).
- 세로리스트 ffill = 화이트리스트(A 구분·B ID·C MES·D 상품명)만. E 이하 옵션 ffill 금지.
- 빈셀 보존(공백 != 없음), 유효 0 보존(타입 강제 금지).
- 셀메타: 스레드댓글(XML 직접 파싱)·배경색 fill·★제약텍스트 플래그를 셀/행 귀속.

엑셀 정보 축 전수(8축, G §⑨) — 어느 축도 버리지 않음(무손실):
  ①값 ②행숨김(row_hidden) ③열숨김(col_hidden) ④셀코멘트 ⑤배경색fill+글자색font
  ⑥수식여부(is_formula+formula) ⑦하이퍼링크 ⑧병합(composite).
  행/열 숨김은 비활성/품절/참고 신호 — 제외·삭제 금지, hidden 플래그로 보존(L2가 판정).
  숨김열도 컬럼으로 추출하되 col_hidden=true 표기(값+숨김플래그 둘 다 보존).
  수식 셀은 data_only 값과 원본 수식 둘 다 보존(실사 S열 `=SUM(R)*1.1` VAT파생 등).

산출: silsa-l1.csv (tidy 1행1레코드), silsa-l1-meta.csv (셀/시트 메타).
시트명 파라미터화 — 전수 확장 가능.

read-only. 원본·DB 무변경.
"""
import sys, os, csv, json, re, zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter

# ---- ffill 화이트리스트 (G §3-3) : row2/row1 헤더명 기준 ----
FFILL_WHITELIST = {"구분", "ID", "MES ITEM_CD", "상품명"}

# ---- 배경색 의미코드 (G §5-2, footnote 근거 확정분만) ----
FILL_MEANING = {
    "FFFFFF00": "신규",          # 노랑 = 신규상품(MES 미등록) — 전 시트 footnote 확정
    # 그레이 FFD9D9D9 는 문구·굿즈 footnote 에서만 확정 -> 시트 인자로 분기
}
GRAY_CONFIRMED_SHEETS = {"문구(가격포함)", "굿즈파우치(가격포함)"}

NS_TC = "{http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments}"


def col_letters_to_idx(letter):
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx


def load_threaded_comments(xlsx_path, sheet_name):
    """xlsx 압축해제 후 threadedComments XML 직접 파싱 (병합셀 댓글 보존 유일 경로).
    반환: {ref(예 'U3'): {'author':..., 'text':...}}"""
    out = {}
    with zipfile.ZipFile(xlsx_path) as z:
        names = z.namelist()
        # workbook.xml: sheet name -> rId
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rid = None
        for m in re.finditer(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml):
            if m.group(1) == sheet_name:
                rid = m.group(2)
                break
        if rid is None:
            return out
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        mt = re.search(r'Id="%s"[^>]*Target="(worksheets/[^"]+)"' % rid, rels)
        if not mt:
            return out
        sheetfile = mt.group(1)                       # worksheets/sheet9.xml
        base = os.path.basename(sheetfile)            # sheet9.xml
        relpath = "xl/worksheets/_rels/%s.rels" % base
        if relpath not in names:
            return out
        srels = z.read(relpath).decode("utf-8")
        tcm = re.search(r'Target="(\.\./threadedComments/[^"]+)"', srels)
        if not tcm:
            return out
        tcfile = "xl/threadedComments/" + os.path.basename(tcm.group(1))
        # persons map
        pmap = {}
        if "xl/persons/person.xml" in names:
            pxml = z.read("xl/persons/person.xml").decode("utf-8")
            for m in re.finditer(r'<person displayName="([^"]+)" id="\{([^}]+)\}"', pxml):
                pmap[m.group(2)] = m.group(1)
        tcxml = z.read(tcfile).decode("utf-8")
        for m in re.finditer(
                r'<threadedComment ref="([^"]+)"[^>]*?(?:personId="\{([^}]+)\}")?[^>]*>(.*?)</threadedComment>',
                tcxml, re.S):
            ref, pid, body = m.group(1), m.group(2), m.group(3)
            tx = re.search(r'<text>(.*?)</text>', body, re.S)
            txt = tx.group(1) if tx else ""
            txt = (txt.replace("&#10;", "\n").replace("&#xA;", "\n")
                      .replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">"))
            out[ref] = {"author": pmap.get(pid, pid or ""), "text": txt}
    return out


def build_composite_headers(ws):
    """row1 가로병합 그룹명을 하위 컬럼에 귀속해 composite 컬럼명 생성.
    반환: {col_idx: composite_name}, work_col_idx, cut_col_idx, bleed_col_idx, group_spans"""
    maxc = ws.max_column
    # 가로병합 그룹(row1, 여러 컬럼) / 세로병합 단일(r1:r2 동일컬럼)
    h_groups = []   # (min_col, max_col, group_name)
    for mr in ws.merged_cells.ranges:
        if mr.min_row == 1 and mr.max_row == 1 and mr.max_col > mr.min_col:
            h_groups.append((mr.min_col, mr.max_col, ws.cell(1, mr.min_col).value))
    # 컬럼별 그룹 귀속 맵
    col_group = {}
    for a, b, g in h_groups:
        for c in range(a, b + 1):
            col_group[c] = (g or "").strip()
    comp = {}
    work_col = cut_col = bleed_col = None
    for c in range(1, maxc + 1):
        r1 = ws.cell(1, c).value
        r2 = ws.cell(2, c).value
        r1 = (str(r1).strip() if r1 is not None else "")
        r2 = (str(r2).strip() if r2 is not None else "")
        if c in col_group:
            grp = col_group[c]
            sub = r2 if r2 else r1   # 하위명 우선 row2, 없으면 row1
            name = "%s_%s" % (grp, sub) if sub else grp
        else:
            # 단순 컬럼: row1 / row2 중 존재값
            name = r1 if r1 else r2
            if not name:
                name = get_column_letter(c)  # 헤더 전무 시 컬럼레터 fallback
        # 중복 방지
        if name in comp.values():
            name = "%s(%s)" % (name, get_column_letter(c))
        comp[c] = name
        # 작업/재단/블리드 식별 (row2 헤더 텍스트 기준 — G HARD: 고정레터 금지)
        if r2 == "작업사이즈":
            work_col = c
        elif r2 == "재단사이즈":
            cut_col = c
        elif r2 == "블리드":
            bleed_col = c
    return comp, work_col, cut_col, bleed_col


def cell_value(ws, r, c):
    """원본 셀값 그대로 보존 — 타입 강제·trim 금지. 공백 None -> '' 으로만 정규화."""
    v = ws.cell(r, c).value
    if v is None:
        return ""
    return v


def find_data_rows(ws, anchor_cols):
    """데이터 행 = 헤더(1,2) 이후 행 중 어느 컬럼이든 non-empty 인 행.
    footnote/legend 셀은 별도 처리되므로 데이터 덤프에 포함하되 메타로 구분."""
    rows = []
    for r in range(3, ws.max_row + 1):
        nonempty = any(ws.cell(r, c).value is not None
                       for c in range(1, ws.max_column + 1))
        if nonempty:
            rows.append(r)
    return rows


def extract_sheet(xlsx_path, sheet_name, out_csv, out_meta):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_fmt = openpyxl.load_workbook(xlsx_path, data_only=False)  # fill/font + 수식
    ws = wb[sheet_name]
    ws_f = wb_fmt[sheet_name]      # data_only=False -> 수식/하이퍼링크 원본 보유
    maxc = ws.max_column

    comp, work_col, cut_col, bleed_col = build_composite_headers(ws)
    comments = load_threaded_comments(xlsx_path, sheet_name)

    # ---- 축③ 열숨김 (col_hidden) : col_idx -> bool. 숨김열도 컬럼으로 보존, 플래그만 표기 ----
    col_hidden = {}
    for c in range(1, maxc + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        col_hidden[c] = bool(dim.hidden) if dim is not None else False

    # ---- 축② 행숨김 (row_hidden) : row -> bool. 숨김행도 레코드로 보존, 플래그만 표기 ----
    def row_is_hidden(r):
        dim = ws.row_dimensions.get(r)
        return bool(dim.hidden) if dim is not None else False

    # 화이트리스트 ffill 대상 컬럼 idx (composite/단순 이름이 화이트리스트에 포함)
    ffill_cols = [c for c, name in comp.items() if name in FFILL_WHITELIST]

    data_rows = find_data_rows(ws, ffill_cols)

    # footnote 범례셀(초록 FFB6D7A8) 탐지 -> 시트메타
    footnotes = []
    for r in range(3, ws.max_row + 1):
        for c in range(1, maxc + 1):
            cell = ws_f.cell(r, c)
            fg = cell.fill.fgColor.rgb if (cell.fill and cell.fill.patternType) else None
            if fg == "FFB6D7A8":
                v = ws.cell(r, c).value
                footnotes.append({"ref": get_column_letter(c) + str(r),
                                  "fill_rgb": fg,
                                  "text": v if v is not None else ""})

    # ffill 상태 추적
    ffill_state = {c: "" for c in ffill_cols}

    records = []
    meta_rows = []
    for r in data_rows:
        # ffill 적용 (화이트리스트 한정)
        ffilled = False
        for c in ffill_cols:
            raw = ws.cell(r, c).value
            if raw is not None and str(raw).strip() != "":
                ffill_state[c] = raw
            else:
                if ffill_state[c] != "":
                    ffilled = True  # 이 행에서 ffill 로 채움
        prd_nm = ""
        # 상품명 컬럼 식별
        prd_col = next((c for c, name in comp.items() if name == "상품명"), None)
        if prd_col is not None:
            prd_nm = ws.cell(r, prd_col).value
            if (prd_nm is None or str(prd_nm).strip() == "") and prd_col in ffill_state:
                prd_nm = ffill_state[prd_col]
        prd_nm = prd_nm if prd_nm is not None else ""

        cols = {}
        for c in range(1, maxc + 1):
            name = comp[c]
            if c in ffill_cols:
                raw = ws.cell(r, c).value
                if raw is not None and str(raw).strip() != "":
                    cols[name] = raw
                else:
                    # ffill 적용 (단, 진짜 빈값 표시 위해 별도 추적은 메타로)
                    cols[name] = ffill_state[c] if ffill_state[c] != "" else ""
            else:
                cols[name] = cell_value(ws, r, c)

        # 셀메타: fill / comment / ★constraint
        cell_meta = {}
        for c in range(1, maxc + 1):
            ref = get_column_letter(c) + str(r)
            cell = ws_f.cell(r, c)
            fg = cell.fill.fgColor.rgb if (cell.fill and cell.fill.patternType) else None
            font_rgb = None
            try:
                if cell.font and cell.font.color and cell.font.color.rgb and isinstance(cell.font.color.rgb, str):
                    font_rgb = cell.font.color.rgb
            except Exception:
                font_rgb = None
            cmt = comments.get(ref)
            val = ws.cell(r, c).value
            has_star = bool(val and isinstance(val, str) and "★" in val)
            # ---- 축⑥ 수식여부 : data_only=False 워크북(ws_f)에서 식별 ----
            fcell_val = cell.value  # cell == ws_f.cell(r,c) -> 수식이면 '=' 시작 문자열
            is_formula = bool(isinstance(fcell_val, str) and fcell_val.startswith("="))
            formula = fcell_val if is_formula else ""
            # ---- 축⑦ 하이퍼링크 : ws_f.cell.hyperlink ----
            hyperlink = ""
            try:
                if cell.hyperlink is not None and cell.hyperlink.target:
                    hyperlink = cell.hyperlink.target
            except Exception:
                hyperlink = ""
            # ---- 축③ 열숨김 플래그 (해당 셀 컬럼이 숨김열인지) ----
            c_hidden = col_hidden.get(c, False)
            if fg and fg not in ("00000000", "FFFFFFFF"):
                fill_meaning = FILL_MEANING.get(fg)
                if fg == "FFD9D9D9" and sheet_name in GRAY_CONFIRMED_SHEETS:
                    fill_meaning = "품절/준비중"
            else:
                fg = None
                fill_meaning = None
            font_show = font_rgb if (font_rgb and font_rgb not in ("FF000000", "00000000")) else None
            if (fg or cmt or has_star or font_show or is_formula or hyperlink or c_hidden):
                m = {}
                if fg:
                    m["fill_rgb"] = fg
                    m["fill_meaning"] = fill_meaning
                if font_show:
                    m["font_rgb"] = font_show
                if cmt:
                    m["comment_author"] = cmt["author"]
                    m["comment_text"] = cmt["text"]
                if has_star:
                    m["has_constraint_star"] = True
                if is_formula:
                    m["is_formula"] = True
                    m["formula"] = formula
                if hyperlink:
                    m["hyperlink"] = hyperlink
                if c_hidden:
                    m["col_hidden"] = True
                cell_meta[comp[c]] = m
                meta_rows.append({
                    "sheet": sheet_name, "row_seq": r, "ref": ref,
                    "col": comp[c],
                    "fill_rgb": fg or "", "fill_meaning": fill_meaning or "",
                    "font_rgb": (font_show or ""),
                    "comment_author": cmt["author"] if cmt else "",
                    "comment_text": cmt["text"] if cmt else "",
                    "has_constraint_star": "true" if has_star else "",
                    "is_formula": "true" if is_formula else "",
                    "formula": formula,
                    "hyperlink": hyperlink,
                    "col_hidden": "true" if c_hidden else "",
                })

        rec = {
            "sheet": sheet_name,
            "row_seq": r,
            "prd_nm": prd_nm,
            "_anchor_ffilled": "true" if ffilled else "false",
            "_row_hidden": "true" if row_is_hidden(r) else "false",   # 축② 행숨김 메타
            "_work_size_col": (comp[work_col] if work_col else ""),
            "_work_size_value": (cols.get(comp[work_col], "") if work_col else ""),
            "cols": cols,
            "cell_meta": cell_meta,
        }
        records.append(rec)

    # 컬럼레벨 댓글(헤더행 row2 등) -> 시트메타
    col_comments = []
    for ref, cmt in comments.items():
        m = re.match(r'([A-Z]+)(\d+)', ref)
        if m and int(m.group(2)) <= 2:
            cidx = col_letters_to_idx(m.group(1))
            col_comments.append({"ref": ref, "col": comp.get(cidx, ref),
                                 "author": cmt["author"], "text": cmt["text"]})

    # ---- write CSV (tidy: 한 행 = 한 레코드, cols 펼침) ----
    all_colnames = [comp[c] for c in range(1, maxc + 1)]
    fixed = ["sheet", "row_seq", "prd_nm", "_anchor_ffilled", "_row_hidden",
             "_work_size_col", "_work_size_value"]
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(fixed + all_colnames + ["cell_meta_json"])
        for rec in records:
            row = [rec["sheet"], rec["row_seq"], rec["prd_nm"],
                   rec["_anchor_ffilled"], rec["_row_hidden"],
                   rec["_work_size_col"], rec["_work_size_value"]]
            for name in all_colnames:
                v = rec["cols"].get(name, "")
                row.append(v)
            row.append(json.dumps(rec["cell_meta"], ensure_ascii=False) if rec["cell_meta"] else "")
            w.writerow(row)

    # ---- write meta CSV ----
    with open(out_meta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "type", "row_seq", "ref", "col",
                    "fill_rgb", "fill_meaning", "font_rgb",
                    "comment_author", "comment_text", "has_constraint_star",
                    "is_formula", "formula", "hyperlink", "col_hidden", "text"])
        for fn in footnotes:
            w.writerow([sheet_name, "footnote", "", fn["ref"], "", fn["fill_rgb"],
                        "범례마커", "", "", "", "", "", "", "", "", fn["text"]])
        for cc in col_comments:
            w.writerow([sheet_name, "column_comment", "", cc["ref"], cc["col"],
                        "", "", "", cc["author"], cc["text"], "", "", "", "", "", ""])
        # 축③ 열숨김 = 시트레벨 메타로도 1행 기록(컬럼명+숨김플래그 보존)
        for c in range(1, maxc + 1):
            if col_hidden.get(c, False):
                w.writerow([sheet_name, "hidden_column", "", get_column_letter(c) + "1",
                            comp[c], "", "", "", "", "", "", "", "", "", "true", ""])
        for m in meta_rows:
            w.writerow([m["sheet"], "cell", m["row_seq"], m["ref"], m["col"],
                        m["fill_rgb"], m["fill_meaning"], m["font_rgb"],
                        m["comment_author"], m["comment_text"], m["has_constraint_star"],
                        m["is_formula"], m["formula"], m["hyperlink"], m["col_hidden"], ""])

    hidden_row_recs = [rec["row_seq"] for rec in records if rec["_row_hidden"] == "true"]
    hidden_cols = [comp[c] for c in range(1, maxc + 1) if col_hidden.get(c, False)]
    formula_cells = sum(1 for m in meta_rows if m["is_formula"] == "true")
    hyperlink_cells = sum(1 for m in meta_rows if m["hyperlink"])
    summary = {
        "sheet": sheet_name,
        "max_col": maxc,
        "field_count": len(all_colnames),
        "record_count": len(records),
        "work_size_col": comp[work_col] if work_col else None,
        "cut_size_col": comp[cut_col] if cut_col else None,
        "bleed_col": comp[bleed_col] if bleed_col else None,
        "footnote_cells": len(footnotes),
        "column_comments": len(col_comments),
        "cell_meta_records": len(meta_rows),
        "comments_total": len(comments),
        # 신규 4축 요약 (축②③⑥⑦)
        "hidden_rows": sorted(hidden_row_recs),
        "hidden_row_count": len(hidden_row_recs),
        "hidden_cols": hidden_cols,
        "hidden_col_count": len(hidden_cols),
        "formula_cells": formula_cells,
        "hyperlink_cells": hyperlink_cells,
    }
    return summary


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="docs/huni/후니프린팅_상품마스터_260527.xlsx")
    ap.add_argument("--sheet", default="실사")
    ap.add_argument("--out", default="_workspace/huni-dbmap/06_extract/silsa-l1.csv")
    ap.add_argument("--meta", default="_workspace/huni-dbmap/06_extract/silsa-l1-meta.csv")
    args = ap.parse_args()
    s = extract_sheet(args.xlsx, args.sheet, args.out, args.meta)
    print(json.dumps(s, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
